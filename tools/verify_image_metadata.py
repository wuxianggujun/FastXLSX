#!/usr/bin/env python3
"""Local image metadata QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It verifies FastXLSX drawing XML
semantics directly, then uses Python XLSX libraries only as local QA/reference
tools. These libraries are not FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import tempfile
import warnings
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


EXPECTED_PICTURES: list[dict[str, str | None]] = [
    {
        "id": "1",
        "name": 'Logo "A&B<1>\'',
        "description": 'Alt "quoted" & <tag> \'owner\'',
        "edit_as": "oneCell",
        "from": "0,0,111,222",
        "to": "2,2,333,444",
    },
    {
        "id": "2",
        "name": "NamedOnly",
        "description": None,
        "edit_as": "absolute",
        "from": "0,2,0,0",
        "to": "2,4,0,0",
    },
    {
        "id": "3",
        "name": "Picture 3",
        "description": None,
        "edit_as": "twoCell",
        "from": "0,4,0,0",
        "to": "2,6,0,0",
    },
]

NAMESPACES = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def read_zip_bytes(path: Path, name: str) -> bytes:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read()


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def count(text: str, fragment: str) -> int:
    return text.count(fragment)


def relationship_count(xml: str) -> int:
    root = ElementTree.fromstring(xml)
    return len(root.findall("rel:Relationship", NAMESPACES))


def marker_signature(anchor: ElementTree.Element, marker_name: str) -> str:
    marker = anchor.find(f"xdr:{marker_name}", NAMESPACES)
    require(marker is not None, f"missing xdr:{marker_name}")
    parts = []
    for child_name in ["col", "row", "colOff", "rowOff"]:
        child = marker.find(f"xdr:{child_name}", NAMESPACES)
        require(child is not None and child.text is not None,
                f"missing xdr:{marker_name}/xdr:{child_name}")
        parts.append(child.text)
    return ",".join(parts)


def verify_fastxlsx_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/media/image1.png",
        "xl/media/image2.png",
        "xl/media/image3.png",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")
    require("xl/drawings/drawing2.xml" not in names, "unexpected second drawing part")
    require("xl/worksheets/_rels/sheet2.xml.rels" not in names, "unexpected second worksheet rels")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require(
        content_types.count('<Default Extension="png" ContentType="image/png"/>') == 1,
        "PNG content type default mismatch",
    )
    require(
        '<Override PartName="/xl/drawings/drawing1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
        in content_types,
        "drawing content type override missing",
    )

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require(
        '</sheetData><drawing r:id="rId1"/></worksheet>' in worksheet_xml,
        "worksheet drawing relationship id mismatch",
    )

    worksheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(worksheet_rels) == 1, "worksheet relationship count mismatch")
    require(
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        'Target="../drawings/drawing1.xml"'
        in worksheet_rels,
        "worksheet drawing relationship target mismatch",
    )

    drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing1.xml.rels")
    require(relationship_count(drawing_rels) == 3, "drawing relationship count mismatch")
    for index in range(1, 4):
        require(
            f'Id="rId{index}"' in drawing_rels
            and f'Target="../media/image{index}.png"' in drawing_rels,
            f"drawing relationship rId{index} mismatch",
        )

    drawing_xml = read_zip_text(path, "xl/drawings/drawing1.xml")
    require(drawing_xml.count("<xdr:twoCellAnchor") == 3, "drawing anchor count mismatch")
    for edit_as in ["oneCell", "absolute", "twoCell"]:
        require(
            drawing_xml.count(f'editAs="{edit_as}"') == 1,
            f"drawing editAs {edit_as} count mismatch",
        )
    require(
        '<xdr:cNvPr id="1" name="Logo &quot;A&amp;B&lt;1&gt;&apos;" '
        'descr="Alt &quot;quoted&quot; &amp; &lt;tag&gt; &apos;owner&apos;"/>'
        in drawing_xml,
        "escaped image name/description XML mismatch",
    )
    require(
        '<xdr:cNvPr id="2" name="NamedOnly"/>' in drawing_xml,
        "named-only image XML mismatch",
    )
    require(
        '<xdr:cNvPr id="3" name="Picture 3"/>' in drawing_xml,
        "default image name XML mismatch",
    )
    require('name="NamedOnly" descr=' not in drawing_xml, "empty named-only description was emitted")
    require('name="Picture 3" descr=' not in drawing_xml, "empty default description was emitted")

    root = ElementTree.fromstring(drawing_xml)
    anchors = root.findall("xdr:twoCellAnchor", NAMESPACES)
    require(len(anchors) == 3, "parsed drawing anchor count mismatch")

    parsed: list[dict[str, str | None]] = []
    for anchor in anchors:
        properties = anchor.find("xdr:pic/xdr:nvPicPr/xdr:cNvPr", NAMESPACES)
        require(properties is not None, "missing xdr:cNvPr")
        parsed.append(
            {
                "id": properties.attrib.get("id"),
                "name": properties.attrib.get("name"),
                "description": properties.attrib.get("descr"),
                "edit_as": anchor.attrib.get("editAs"),
                "from": marker_signature(anchor, "from"),
                "to": marker_signature(anchor, "to"),
            }
        )
    require(parsed == EXPECTED_PICTURES, f"parsed image metadata mismatch: {parsed!r}")

    media_sizes = {
        name: len(read_zip_bytes(path, name))
        for name in ["xl/media/image1.png", "xl/media/image2.png", "xl/media/image3.png"]
    }
    for name, size in media_sizes.items():
        require(size > 0, f"empty media part: {name}")

    return {
        "verified_entries": required,
        "pictures": parsed,
        "media_sizes": media_sizes,
    }


def verify_metadata_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("ImageMetadata" in workbook.sheetnames, "missing ImageMetadata worksheet")
        worksheet = workbook["ImageMetadata"]
        images = list(getattr(worksheet, "_images", []))
        require(len(images) == 3, f"openpyxl image count mismatch: {len(images)}")
        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "image_count": len(images),
        }
    finally:
        workbook.close()


def verify_basic_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "xl/media/image1.png",
        "xl/media/image2.png",
        "xl/drawings/drawing1.xml",
        "xl/drawings/drawing2.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/drawings/_rels/drawing2.xml.rels",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/tables/table1.xml",
    ]
    for name in required:
        require(name in names, f"basic image sample missing package entry: {name}")
    require("xl/worksheets/_rels/sheet3.xml.rels" not in names,
            "basic image plain sheet should not create worksheet relationships")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require('<Default Extension="png" ContentType="image/png"/>' in content_types,
            "basic image PNG content type default missing")
    require('/xl/drawings/drawing1.xml' in content_types,
            "basic image first drawing content type missing")
    require('/xl/drawings/drawing2.xml' in content_types,
            "basic image second drawing content type missing")

    first_sheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require(
        '</sheetData><hyperlinks><hyperlink ref="A2" r:id="rId1"/></hyperlinks>'
        '<drawing r:id="rId2"/><tableParts count="1"><tablePart r:id="rId3"/></tableParts></worksheet>'
        in first_sheet_xml,
        "basic image first worksheet suffix order mismatch",
    )

    first_sheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(first_sheet_rels) == 3,
            "basic image first worksheet relationship count mismatch")
    require('Id="rId1"' in first_sheet_rels and 'Target="https://example.com/items/widget"' in first_sheet_rels,
            "basic image hyperlink relationship mismatch")
    require('Id="rId2"' in first_sheet_rels and 'Target="../drawings/drawing1.xml"' in first_sheet_rels,
            "basic image drawing relationship mismatch")
    require('Id="rId3"' in first_sheet_rels and 'Target="../tables/table1.xml"' in first_sheet_rels,
            "basic image table relationship mismatch")

    first_drawing_xml = read_zip_text(path, "xl/drawings/drawing1.xml")
    require(count(first_drawing_xml, "<xdr:twoCellAnchor") == 1,
            "basic image first drawing anchor count mismatch")
    require("<xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row>" in first_drawing_xml,
            "basic image first drawing from marker mismatch")
    require("<xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>4</xdr:row>" in first_drawing_xml,
            "basic image first drawing to marker mismatch")
    require('<xdr:cNvPr id="1" name="Picture 1"/>' in first_drawing_xml,
            "basic image first picture name mismatch")

    first_drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing1.xml.rels")
    require(relationship_count(first_drawing_rels) == 1,
            "basic image first drawing relationship count mismatch")
    require('Target="../media/image1.png"' in first_drawing_rels,
            "basic image first drawing media target mismatch")

    second_sheet_xml = read_zip_text(path, "xl/worksheets/sheet2.xml")
    require('</sheetData><drawing r:id="rId1"/></worksheet>' in second_sheet_xml,
            "basic image second worksheet drawing relationship id mismatch")
    second_sheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet2.xml.rels")
    require(relationship_count(second_sheet_rels) == 1,
            "basic image second worksheet relationship count mismatch")
    require('Id="rId1"' in second_sheet_rels and 'Target="../drawings/drawing2.xml"' in second_sheet_rels,
            "basic image second worksheet drawing relationship mismatch")
    second_drawing_xml = read_zip_text(path, "xl/drawings/drawing2.xml")
    require('<xdr:cNvPr id="2" name="Picture 2"/>' in second_drawing_xml,
            "basic image second picture name mismatch")
    second_drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing2.xml.rels")
    require('Target="../media/image2.png"' in second_drawing_rels,
            "basic image second drawing media target mismatch")

    return {
        "verified_entries": required,
        "sheets": {
            "Images": {"shapes": 1, "hyperlinks": 1, "tables": 1},
            "SecondImage": {"shapes": 1},
            "Plain": {"shapes": 0},
        },
    }


def verify_basic_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Images", "SecondImage", "Plain"],
                f"basic image sheet names mismatch: {workbook.sheetnames!r}")
        images = workbook["Images"]
        second = workbook["SecondImage"]
        plain = workbook["Plain"]
        require(len(getattr(images, "_images", [])) == 1, "basic Images image count mismatch")
        require(len(getattr(second, "_images", [])) == 1, "basic SecondImage image count mismatch")
        require(len(getattr(plain, "_images", [])) == 0, "basic Plain image count mismatch")
        hyperlink = images["A2"].hyperlink
        require(hyperlink is not None and hyperlink.target == "https://example.com/items/widget",
                "basic Images A2 hyperlink target mismatch")
        require(list(images.tables.keys()) == ["ImageTable"],
                f"basic Images table names mismatch: {list(images.tables.keys())!r}")
        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "image_counts": {
                "Images": len(getattr(images, "_images", [])),
                "SecondImage": len(getattr(second, "_images", [])),
                "Plain": len(getattr(plain, "_images", [])),
            },
            "images_a2_hyperlink": hyperlink.target,
            "images_tables": list(images.tables.keys()),
        }
    finally:
        workbook.close()


def verify_memory_source_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "xl/media/image1.png",
        "xl/media/image2.jpg",
        "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
    ]
    for name in required:
        require(name in names, f"memory image sample missing package entry: {name}")
    require("xl/drawings/drawing2.xml" not in names,
            "memory image sample should share one drawing part")

    png_bytes = read_zip_bytes(path, "xl/media/image1.png")
    jpeg_bytes = read_zip_bytes(path, "xl/media/image2.jpg")
    require(png_bytes.startswith(b"\x89PNG\r\n\x1a\n"), "memory PNG media signature mismatch")
    require(jpeg_bytes.startswith(b"\xff\xd8") and jpeg_bytes.endswith(b"\xff\xd9"),
            "memory JPEG media signature mismatch")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require('<Default Extension="png" ContentType="image/png"/>' in content_types,
            "memory image PNG content type default missing")
    require('<Default Extension="jpg" ContentType="image/jpeg"/>' in content_types,
            "memory image JPEG content type default missing")
    require('/xl/drawings/drawing1.xml' in content_types,
            "memory image drawing content type missing")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('</sheetData><drawing r:id="rId1"/></worksheet>' in worksheet_xml,
            "memory image worksheet drawing relationship id mismatch")

    worksheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(worksheet_rels) == 1,
            "memory image worksheet relationship count mismatch")
    require('Id="rId1"' in worksheet_rels and 'Target="../drawings/drawing1.xml"' in worksheet_rels,
            "memory image worksheet drawing relationship mismatch")

    drawing_xml = read_zip_text(path, "xl/drawings/drawing1.xml")
    require(count(drawing_xml, "<xdr:twoCellAnchor") == 2,
            "memory image drawing anchor count mismatch")
    require('editAs="oneCell"' in drawing_xml, "memory image oneCell editAs missing")
    require('<xdr:colOff>101</xdr:colOff>' in drawing_xml
            and '<xdr:rowOff>202</xdr:rowOff>' in drawing_xml
            and '<xdr:colOff>303</xdr:colOff>' in drawing_xml
            and '<xdr:rowOff>404</xdr:rowOff>' in drawing_xml,
            "memory image marker offsets mismatch")
    require('<xdr:cNvPr id="1" name="Memory PNG" descr="PNG bytes from memory"/>' in drawing_xml,
            "memory image custom name/description mismatch")
    require('<xdr:cNvPr id="2" name="Picture 2"/>' in drawing_xml,
            "memory image default picture name mismatch")
    require('<a:ext cx="9525" cy="9525"/>' in drawing_xml,
            "memory PNG intrinsic EMU size mismatch")
    require('<a:ext cx="19050" cy="9525"/>' in drawing_xml,
            "memory JPEG intrinsic EMU size mismatch")

    drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing1.xml.rels")
    require(relationship_count(drawing_rels) == 2,
            "memory image drawing relationship count mismatch")
    require('Id="rId1"' in drawing_rels and 'Target="../media/image1.png"' in drawing_rels,
            "memory PNG drawing relationship mismatch")
    require('Id="rId2"' in drawing_rels and 'Target="../media/image2.jpg"' in drawing_rels,
            "memory JPEG drawing relationship mismatch")

    return {
        "verified_entries": required,
        "media_sizes": {
            "xl/media/image1.png": len(png_bytes),
            "xl/media/image2.jpg": len(jpeg_bytes),
        },
        "sheet": "MemoryImages",
        "shapes": 2,
    }


def verify_memory_source_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    with warnings.catch_warnings(record=True) as caught_warnings:
        warnings.simplefilter("always")
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("MemoryImages" in workbook.sheetnames,
                f"memory image sheet names mismatch: {workbook.sheetnames!r}")
        worksheet = workbook["MemoryImages"]
        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "image_count_observed_by_openpyxl": len(getattr(worksheet, "_images", [])),
            "warnings": [
                str(item.message)
                for item in caught_warnings
                if "image" in str(item.message).lower()
            ],
            "note": (
                "openpyxl is a reader-visible smoke check; package XML and Excel COM "
                "remain authoritative for PNG/JPEG media and drawing semantics."
            ),
        }
    finally:
        workbook.close()


def verify_image_hyperlink_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "xl/media/image1.png",
        "xl/media/image2.jpg",
        "xl/drawings/drawing1.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
    ]
    for name in required:
        require(name in names, f"image hyperlink sample missing package entry: {name}")
    require("xl/drawings/drawing2.xml" not in names,
            "image hyperlink sample should share one drawing part")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require('<Default Extension="png" ContentType="image/png"/>' in content_types,
            "image hyperlink PNG content type default missing")
    require('<Default Extension="jpg" ContentType="image/jpeg"/>' in content_types,
            "image hyperlink JPEG content type default missing")
    require('/xl/drawings/drawing1.xml' in content_types,
            "image hyperlink drawing content type missing")
    require("hyperlink" not in content_types,
            "image object hyperlink should not add content type entries")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('</sheetData><drawing r:id="rId1"/></worksheet>' in worksheet_xml,
            "image hyperlink worksheet drawing relationship id mismatch")
    require("<hyperlinks>" not in worksheet_xml,
            "image object hyperlinks should not create worksheet hyperlinks")

    worksheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(worksheet_rels) == 1,
            "image hyperlink worksheet relationship count mismatch")
    require('Id="rId1"' in worksheet_rels and 'Target="../drawings/drawing1.xml"' in worksheet_rels,
            "image hyperlink worksheet drawing relationship mismatch")
    require("relationships/hyperlink" not in worksheet_rels,
            "image object hyperlinks should stay in drawing relationships")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require("relationships/hyperlink" not in workbook_rels,
            "image object hyperlinks should not create workbook relationships")

    drawing_xml = read_zip_text(path, "xl/drawings/drawing1.xml")
    require(count(drawing_xml, "<xdr:twoCellAnchor") == 2,
            "image hyperlink drawing anchor count mismatch")
    require(
        '<xdr:cNvPr id="1" name="Linked Path" descr="Path image link">'
        '<a:hlinkClick r:id="rId3" tooltip="Open &quot;path&quot; &amp; &lt;tag&gt;"/>'
        '</xdr:cNvPr>' in drawing_xml,
        "path image hyperlink XML mismatch or tooltip escape failure",
    )
    require(
        '<xdr:cNvPr id="2" name="Linked Memory"><a:hlinkClick r:id="rId4"/></xdr:cNvPr>'
        in drawing_xml,
        "memory image hyperlink XML mismatch",
    )
    require('<a:blip r:embed="rId1"/>' in drawing_xml,
            "path image media relationship id should remain rId1")
    require('<a:blip r:embed="rId2"/>' in drawing_xml,
            "memory image media relationship id should remain rId2")

    root = ElementTree.fromstring(drawing_xml)
    anchors = root.findall("xdr:twoCellAnchor", NAMESPACES)
    require(len(anchors) == 2, "parsed image hyperlink anchor count mismatch")
    parsed_hyperlinks: list[dict[str, str | None]] = []
    for anchor in anchors:
        properties = anchor.find("xdr:pic/xdr:nvPicPr/xdr:cNvPr", NAMESPACES)
        require(properties is not None, "missing image hyperlink cNvPr")
        hyperlink = properties.find("a:hlinkClick", NAMESPACES)
        require(hyperlink is not None, "missing image hyperlink hlinkClick")
        parsed_hyperlinks.append(
            {
                "name": properties.attrib.get("name"),
                "relationship_id": hyperlink.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"),
                "tooltip": hyperlink.attrib.get("tooltip"),
            }
        )
    require(
        parsed_hyperlinks == [
            {
                "name": "Linked Path",
                "relationship_id": "rId3",
                "tooltip": 'Open "path" & <tag>',
            },
            {
                "name": "Linked Memory",
                "relationship_id": "rId4",
                "tooltip": None,
            },
        ],
        f"parsed image hyperlink metadata mismatch: {parsed_hyperlinks!r}",
    )

    drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing1.xml.rels")
    require(relationship_count(drawing_rels) == 4,
            "image hyperlink drawing relationship count mismatch")
    for rel_id, fragment in [
        ("rId1", 'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.png"'),
        ("rId2", 'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image2.jpg"'),
        ("rId3", 'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="https://example.com/path?a=1&amp;b=2" TargetMode="External"'),
        ("rId4", 'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="mailto:image@example.com" TargetMode="External"'),
    ]:
        require(f'Id="{rel_id}"' in drawing_rels and fragment in drawing_rels,
                f"image hyperlink drawing relationship {rel_id} mismatch")

    return {
        "verified_entries": required,
        "hyperlinks": parsed_hyperlinks,
        "relationship_model": "image relationships remain first; picture hyperlinks are drawing-local",
    }


def verify_image_hyperlink_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    with warnings.catch_warnings(record=True) as caught_warnings:
        warnings.simplefilter("always")
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("ImageLinks" in workbook.sheetnames,
                f"image hyperlink sheet names mismatch: {workbook.sheetnames!r}")
        worksheet = workbook["ImageLinks"]
        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "image_count_observed_by_openpyxl": len(getattr(worksheet, "_images", [])),
            "warnings": [
                str(item.message)
                for item in caught_warnings
                if "image" in str(item.message).lower()
            ],
            "note": (
                "openpyxl is only a reader-visible smoke check here; package XML "
                "and Excel COM are authoritative for picture hyperlink metadata."
            ),
        }
    finally:
        workbook.close()


def verify_mixed_object_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "xl/media/image1.png",
        "xl/media/image2.jpg",
        "xl/media/image3.png",
        "xl/drawings/drawing1.xml",
        "xl/drawings/drawing2.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/drawings/_rels/drawing2.xml.rels",
        "xl/tables/table1.xml",
        "xl/tables/table2.xml",
        "xl/tables/table3.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
    ]
    for name in required:
        require(name in names, f"mixed object sample missing package entry: {name}")
    require("xl/drawings/drawing3.xml" not in names,
            "mixed object plain sheet should not create drawing3.xml")
    require("xl/worksheets/_rels/sheet3.xml.rels" not in names,
            "mixed object plain sheet should not create worksheet relationships")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require('<Default Extension="png" ContentType="image/png"/>' in content_types,
            "mixed object PNG content type default missing")
    require('<Default Extension="jpg" ContentType="image/jpeg"/>' in content_types,
            "mixed object JPG content type default missing")

    first_sheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require(
        '</sheetData><hyperlinks><hyperlink ref="A2" r:id="rId1"/>'
        '<hyperlink ref="B2" r:id="rId2"/></hyperlinks><drawing r:id="rId3"/>'
        '<tableParts count="2"><tablePart r:id="rId4"/><tablePart r:id="rId5"/></tableParts></worksheet>'
        in first_sheet_xml,
        "mixed object first worksheet relationship order mismatch",
    )
    first_sheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(first_sheet_rels) == 5,
            "mixed object first worksheet relationship count mismatch")
    for fragment in [
        ('rId1', 'Target="https://example.com/widget"'),
        ('rId2', 'Target="https://example.com/docs"'),
        ('rId3', 'Target="../drawings/drawing1.xml"'),
        ('rId4', 'Target="../tables/table1.xml"'),
        ('rId5', 'Target="../tables/table2.xml"'),
    ]:
        require(f'Id="{fragment[0]}"' in first_sheet_rels and fragment[1] in first_sheet_rels,
                f"mixed object first worksheet relationship {fragment[0]} mismatch")

    first_drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing1.xml.rels")
    require(relationship_count(first_drawing_rels) == 2,
            "mixed object first drawing relationship count mismatch")
    require('Id="rId1"' in first_drawing_rels and 'Target="../media/image1.png"' in first_drawing_rels,
            "mixed object first drawing PNG relationship mismatch")
    require('Id="rId2"' in first_drawing_rels and 'Target="../media/image2.jpg"' in first_drawing_rels,
            "mixed object first drawing JPEG relationship mismatch")
    require(count(read_zip_text(path, "xl/drawings/drawing1.xml"), "<xdr:twoCellAnchor") == 2,
            "mixed object first drawing anchor count mismatch")

    second_sheet_xml = read_zip_text(path, "xl/worksheets/sheet2.xml")
    require(
        '</sheetData><hyperlinks><hyperlink ref="A2" r:id="rId1"/></hyperlinks>'
        '<drawing r:id="rId2"/><tableParts count="1"><tablePart r:id="rId3"/></tableParts></worksheet>'
        in second_sheet_xml,
        "mixed object second worksheet relationship order mismatch",
    )
    second_sheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet2.xml.rels")
    require(relationship_count(second_sheet_rels) == 3,
            "mixed object second worksheet relationship count mismatch")
    for fragment in [
        ('rId1', 'Target="https://example.com/gadget"'),
        ('rId2', 'Target="../drawings/drawing2.xml"'),
        ('rId3', 'Target="../tables/table3.xml"'),
    ]:
        require(f'Id="{fragment[0]}"' in second_sheet_rels and fragment[1] in second_sheet_rels,
                f"mixed object second worksheet relationship {fragment[0]} mismatch")
    second_drawing_rels = read_zip_text(path, "xl/drawings/_rels/drawing2.xml.rels")
    require(relationship_count(second_drawing_rels) == 1,
            "mixed object second drawing relationship count mismatch")
    require('Id="rId1"' in second_drawing_rels and 'Target="../media/image3.png"' in second_drawing_rels,
            "mixed object second drawing relationship id should reset")

    return {
        "verified_entries": required,
        "relationship_model": "owner-local worksheet and drawing relationship ids",
    }


def verify_mixed_object_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    with warnings.catch_warnings(record=True) as caught_warnings:
        warnings.simplefilter("always")
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Objects", "MoreObjects", "Plain"],
                f"mixed object sheet names mismatch: {workbook.sheetnames!r}")
        objects = workbook["Objects"]
        more = workbook["MoreObjects"]
        plain = workbook["Plain"]
        require(objects["A2"].hyperlink is not None
                and objects["A2"].hyperlink.target == "https://example.com/widget",
                "mixed object Objects A2 hyperlink mismatch")
        require(objects["B2"].hyperlink is not None
                and objects["B2"].hyperlink.target == "https://example.com/docs",
                "mixed object Objects B2 hyperlink mismatch")
        require(more["A2"].hyperlink is not None
                and more["A2"].hyperlink.target == "https://example.com/gadget",
                "mixed object MoreObjects A2 hyperlink mismatch")
        require(list(objects.tables.keys()) == ["ObjectTableOne", "ObjectTableTwo"],
                f"mixed object Objects table names mismatch: {list(objects.tables.keys())!r}")
        require(list(more.tables.keys()) == ["ObjectTableThree"],
                f"mixed object MoreObjects table names mismatch: {list(more.tables.keys())!r}")
        require(list(plain.tables.keys()) == [], "mixed object Plain should not expose tables")
        require(len(getattr(plain, "_images", [])) == 0,
                "mixed object Plain should not expose images")
        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "image_counts_observed_by_openpyxl": {
                "Objects": len(getattr(objects, "_images", [])),
                "MoreObjects": len(getattr(more, "_images", [])),
                "Plain": len(getattr(plain, "_images", [])),
            },
            "tables": {
                "Objects": list(objects.tables.keys()),
                "MoreObjects": list(more.tables.keys()),
                "Plain": list(plain.tables.keys()),
            },
            "hyperlinks": {
                "Objects!A2": objects["A2"].hyperlink.target,
                "Objects!B2": objects["B2"].hyperlink.target,
                "MoreObjects!A2": more["A2"].hyperlink.target,
            },
            "warnings": [
                str(item.message)
                for item in caught_warnings
                if "image" in str(item.message).lower()
            ],
            "note": (
                "openpyxl may skip JPEG image loading; package XML checks remain "
                "authoritative for media and drawing relationship counts."
            ),
        }
    finally:
        workbook.close()


def create_xlsxwriter_reference(media_file: Path, path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("ImageMetadata")
    worksheet.write("A1", "Reference image")
    worksheet.insert_image(
        "A2",
        str(media_file),
        {"description": 'Alt "quoted" & <tag> \'owner\''},
    )
    workbook.close()
    return {"status": "created", "path": str(path)}


def create_xlsxwriter_hyperlink_reference(media_file: Path, path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("ImageLinks")
    worksheet.write("A1", "Reference image hyperlink")
    worksheet.insert_image(
        "A2",
        str(media_file),
        {
            "url": "https://example.com/path?a=1&b=2",
            "tip": 'Open "path" & <tag>',
            "description": "Path image link",
        },
    )
    workbook.close()
    return {"status": "created", "path": str(path)}


def extract_reference_media(input_path: Path, work_dir: Path) -> Path:
    media_path = work_dir / "image1.png"
    media_path.write_bytes(read_zip_bytes(input_path, "xl/media/image1.png"))
    return media_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-image-metadata.xlsx"),
        help="FastXLSX image metadata workbook to verify.",
    )
    parser.add_argument(
        "--basic-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-images.xlsx"),
        help="FastXLSX basic image workbook to verify.",
    )
    parser.add_argument(
        "--mixed-object-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-mixed-object-rels.xlsx"),
        help="FastXLSX mixed object relationship workbook to verify.",
    )
    parser.add_argument(
        "--memory-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-memory-images.xlsx"),
        help="FastXLSX memory-source image workbook to verify.",
    )
    parser.add_argument(
        "--hyperlink-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-image-hyperlinks.xlsx"),
        help="FastXLSX picture hyperlink workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/image-metadata"),
        help="Directory for extracted media, reference files, and the JSON report.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    basic_input_path = args.basic_input.resolve()
    mixed_object_input_path = args.mixed_object_input.resolve()
    memory_input_path = args.memory_input.resolve()
    hyperlink_input_path = args.hyperlink_input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    require(basic_input_path.exists(), f"basic input workbook does not exist: {basic_input_path}")
    require(mixed_object_input_path.exists(),
            f"mixed-object input workbook does not exist: {mixed_object_input_path}")
    require(memory_input_path.exists(), f"memory input workbook does not exist: {memory_input_path}")
    require(hyperlink_input_path.exists(),
            f"hyperlink input workbook does not exist: {hyperlink_input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    package_report = verify_fastxlsx_package(input_path)
    openpyxl_report = verify_metadata_with_openpyxl(input_path)
    basic_package_report = verify_basic_package(basic_input_path)
    basic_openpyxl_report = verify_basic_with_openpyxl(basic_input_path)
    mixed_package_report = verify_mixed_object_package(mixed_object_input_path)
    mixed_openpyxl_report = verify_mixed_object_with_openpyxl(mixed_object_input_path)
    memory_package_report = verify_memory_source_package(memory_input_path)
    memory_openpyxl_report = verify_memory_source_with_openpyxl(memory_input_path)
    hyperlink_package_report = verify_image_hyperlink_package(hyperlink_input_path)
    hyperlink_openpyxl_report = verify_image_hyperlink_with_openpyxl(hyperlink_input_path)

    with tempfile.TemporaryDirectory(dir=work_dir) as temp_dir:
        media_file = extract_reference_media(input_path, Path(temp_dir))
        xlsxwriter_report = create_xlsxwriter_reference(
            media_file, work_dir / "reference-xlsxwriter-image-metadata.xlsx")
    with tempfile.TemporaryDirectory(dir=work_dir) as temp_dir:
        memory_media_file = extract_reference_media(memory_input_path, Path(temp_dir))
        memory_xlsxwriter_report = create_xlsxwriter_reference(
            memory_media_file, work_dir / "reference-xlsxwriter-image-memory-source.xlsx")
    with tempfile.TemporaryDirectory(dir=work_dir) as temp_dir:
        hyperlink_media_file = extract_reference_media(hyperlink_input_path, Path(temp_dir))
        hyperlink_xlsxwriter_report = create_xlsxwriter_hyperlink_reference(
            hyperlink_media_file, work_dir / "reference-xlsxwriter-image-hyperlink.xlsx")

    report = {
        "metadata": {
            "fastxlsx_input": str(input_path),
            "fastxlsx_package": package_report,
            "xlsx_libraries": {
                "openpyxl": openpyxl_report,
                "xlsxwriter": xlsxwriter_report,
            },
        },
        "basic": {
            "fastxlsx_input": str(basic_input_path),
            "fastxlsx_package": basic_package_report,
            "xlsx_libraries": {
                "openpyxl": basic_openpyxl_report,
            },
        },
        "mixed_object_relationships": {
            "fastxlsx_input": str(mixed_object_input_path),
            "fastxlsx_package": mixed_package_report,
            "xlsx_libraries": {
                "openpyxl": mixed_openpyxl_report,
            },
        },
        "memory_source": {
            "fastxlsx_input": str(memory_input_path),
            "fastxlsx_package": memory_package_report,
            "xlsx_libraries": {
                "openpyxl": memory_openpyxl_report,
                "xlsxwriter": memory_xlsxwriter_report,
            },
        },
        "image_hyperlinks": {
            "fastxlsx_input": str(hyperlink_input_path),
            "fastxlsx_package": hyperlink_package_report,
            "xlsx_libraries": {
                "openpyxl": hyperlink_openpyxl_report,
                "xlsxwriter": hyperlink_xlsxwriter_report,
            },
        },
    }
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - local QA helper
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
