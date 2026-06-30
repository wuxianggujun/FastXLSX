#!/usr/bin/env python3
"""Local workbook-editor QA runner for FastXLSX.

This helper is intentionally outside CTest. It drives the opt-in
`fastxlsx_workbook_editor_qa_tool`, then validates the produced workbooks with
ZIP/XML checks, openpyxl readback, and optional XlsxWriter-generated reference
workbooks for generated scenarios. It can also scan caller-provided external
fixture directories for a broad materialized-edit smoke pass.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import posixpath
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from xml.etree import ElementTree


NAMESPACES = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "office_rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

GENERATED_SCENARIOS = [
    "generated_rename_materialized",
    "generated_in_memory_insert_formula",
    "generated_in_memory_delete_column_formula",
    "generated_in_memory_insert_column_formula",
    "generated_in_memory_delete_row_formula",
    "generated_in_memory_stationary_formula_shift",
    "generated_in_memory_clear_erase",
    "generated_in_memory_append_row_formula",
    "generated_in_memory_overwrite_formula_text",
    "generated_in_memory_retry_noop_save",
    "generated_in_memory_retry_path_equivalent_noop_save",
    "generated_in_memory_retry_reopen_modify_noop_save",
    "generated_in_memory_retry_path_equivalent_reopen_modify_noop_save",
    "generated_in_memory_retry_path_equivalent_reopen_modify_post_noop_third_save",
    "generated_in_memory_retry_reopen_modify_post_noop_third_save",
    "generated_in_memory_reopen_modify_save",
    "generated_in_memory_reopen_modify_noop_save",
    "generated_in_memory_reopen_modify_post_noop_third_save",
    "generated_in_memory_multi_sheet_save",
    "generated_in_memory_multi_sheet_noop_save",
    "generated_in_memory_multi_sheet_retry_save",
    "generated_in_memory_multi_sheet_retry_noop_save",
    "generated_in_memory_multi_sheet_retry_path_equivalent_noop_save",
    "generated_in_memory_multi_sheet_retry_reopen_modify_save",
    "generated_in_memory_multi_sheet_retry_reopen_modify_noop_save",
    "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_noop_save",
    "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save",
    "generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save",
    "generated_source_formula_audit",
    "generated_formula_rename_rewrite",
    "generated_formula_rename_escaped_sheet_name",
    "generated_formula_rename_chain_rewrite",
    "generated_formula_rename_defined_names_only",
    "generated_formula_rename_default_audit",
    "generated_shared_formula_materialization",
    "generated_shared_formula_boundary_materialization",
    "generated_shared_formula_office_like_materialization",
    "generated_style_passthrough",
    "generated_image_replace",
    "generated_public_e2e",
    "generated_non_default_style_rejection",
]

GENERATED_CASE_DIRECTORY_ALIASES = {
    "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save":
        "ms-pe-post-noop-third",
}

DEFAULT_XLNT_RENAME_FIXTURES = [
    "2_minimal.xlsx",
    "3_default.xlsx",
    "20_active_sheet.xlsx",
]

DEFAULT_XLNT_STRING_FIXTURES = [
    "Issue445_inline_str.xlsx",
    "Issue494_shared_string.xlsx",
]

EXTERNAL_FIXTURE_SCENARIO = "external_fixture_materialized_smoke"
SOURCE_FORMULA_FIXTURE_SCENARIO = "external_source_formula_fixture_audit_smoke"
FORMULA_FIXTURE_SCENARIO = "external_formula_fixture_materialized_smoke"
DEFINED_NAME_FIXTURE_SCENARIO = "external_defined_name_fixture_smoke"
IMAGE_FIXTURE_SCENARIO = "external_fixture_image_replace_smoke"
FIXTURE_SCENARIOS = [
    "xlnt_fixture_rename_smoke",
    "xlnt_fixture_string_smoke",
    EXTERNAL_FIXTURE_SCENARIO,
    SOURCE_FORMULA_FIXTURE_SCENARIO,
    FORMULA_FIXTURE_SCENARIO,
    DEFINED_NAME_FIXTURE_SCENARIO,
    IMAGE_FIXTURE_SCENARIO,
]


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def read_zip_bytes(path: Path, name: str) -> bytes:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read()


DEFAULT_QA_EXE_RELATIVE_CANDIDATES = [
    Path("windows-nmake-release-minizip/tools/fastxlsx_workbook_editor_qa_tool.exe"),
    Path("windows-nmake-release/tools/fastxlsx_workbook_editor_qa_tool.exe"),
]


def resolve_default_qa_exe(build_root: Path = Path("build")) -> Path | None:
    candidates = [
        build_root / relative_candidate
        for relative_candidate in DEFAULT_QA_EXE_RELATIVE_CANDIDATES
    ]
    existing_candidates = [candidate for candidate in candidates if candidate.exists()]
    if existing_candidates:
        return max(existing_candidates, key=lambda candidate: candidate.stat().st_mtime_ns)

    discovered = list(build_root.rglob("fastxlsx_workbook_editor_qa_tool.exe"))
    if discovered:
        return max(discovered, key=lambda candidate: candidate.stat().st_mtime_ns)
    return None


def resolve_default_excel_script() -> Path:
    return Path(__file__).resolve().with_name("verify_workbook_editor_qa_excel.ps1")


def repository_root() -> Path:
    return Path(__file__).resolve().parents[1]


def repository_asset(relative_path: str) -> Path:
    return repository_root() / Path(relative_path)


def workspace_root() -> Path:
    return Path(__file__).resolve().parents[2]


def workspace_asset(relative_path: str) -> Path:
    return workspace_root() / Path(relative_path)


def choose_unique_sheet_name(existing: list[str], preferred: str = "QA_Renamed") -> str:
    existing_lower = {name.lower() for name in existing}
    candidate = preferred[:31]
    if candidate.lower() not in existing_lower:
        return candidate

    for index in range(1, 1000):
        suffix = f"_{index}"
        candidate = preferred[: 31 - len(suffix)] + suffix
        if candidate.lower() not in existing_lower:
            return candidate

    raise RuntimeError("failed to allocate a unique worksheet name")


def compact_slug_component(text: str, *, max_length: int = 24) -> str:
    safe = "".join(
        ch if ch.isascii() and (ch.isalnum() or ch in "._-") else "_" for ch in text
    )
    if len(safe) > max_length:
        safe = safe[:max_length].rstrip("._-")
    return safe or "fixture"


def fixture_case_slug(fixture_root: Path, fixture_path: Path) -> str:
    try:
        relative = fixture_path.relative_to(fixture_root)
    except ValueError:
        relative = fixture_path.name

    text = str(relative).replace("\\", "/")
    safe = compact_slug_component(fixture_path.stem, max_length=12)
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()[:8]
    return f"{safe}-{digest}"


def formula_fixture_case_slug(
    fixture_root: Path,
    fixture_path: Path,
    sheet_name: str,
) -> str:
    safe_sheet = compact_slug_component(sheet_name, max_length=10)
    digest = hashlib.sha1(sheet_name.encode("utf-8")).hexdigest()[:8]
    return f"{fixture_case_slug(fixture_root, fixture_path)}__{safe_sheet}-{digest}"


def discover_external_fixtures(
    fixture_root: Path,
    globs: list[str],
    limit: int,
) -> list[Path]:
    discovered: dict[Path, None] = {}
    for pattern in globs:
        for fixture_path in fixture_root.glob(pattern):
            if not fixture_path.is_file():
                continue
            if fixture_path.name.startswith("~$"):
                continue
            if fixture_path.suffix.lower() != ".xlsx":
                continue
            discovered[fixture_path.resolve()] = None

    fixtures = sorted(discovered.keys(), key=lambda path: str(path).lower())
    if limit > 0:
        fixtures = fixtures[:limit]
    return fixtures


@dataclass
class ScenarioResult:
    name: str
    report: dict[str, Any]
    zip_xml: dict[str, Any]
    openpyxl: dict[str, Any]
    xlsxwriter_reference: dict[str, Any]
    error: str = ""


@dataclass(frozen=True)
class FormulaWorksheetInfo:
    fixture_path: Path
    sheet_name: str
    worksheet_entry: str
    formula_count: int
    shared_formula_count: int
    shared_definition_count: int
    metadata_only_shared_count: int

    def as_report(self, fixture_root: Path | None = None) -> dict[str, Any]:
        try:
            fixture_name = (
                self.fixture_path.relative_to(fixture_root).as_posix()
                if fixture_root is not None
                else self.fixture_path.name
            )
        except ValueError:
            fixture_name = self.fixture_path.name
        return {
            "fixture": fixture_name,
            "sheet_name": self.sheet_name,
            "worksheet_entry": self.worksheet_entry,
            "formula_count": self.formula_count,
            "shared_formula_count": self.shared_formula_count,
            "shared_definition_count": self.shared_definition_count,
            "metadata_only_shared_count": self.metadata_only_shared_count,
        }


@dataclass(frozen=True)
class ImageFixtureInfo:
    fixture_path: Path
    sheet_name: str
    image_part_name: str

    def as_report(self, fixture_root: Path | None = None) -> dict[str, Any]:
        try:
            fixture_name = (
                self.fixture_path.relative_to(fixture_root).as_posix()
                if fixture_root is not None
                else self.fixture_path.name
            )
        except ValueError:
            fixture_name = self.fixture_path.name
        return {
            "fixture": fixture_name,
            "sheet_name": self.sheet_name,
            "image_part_name": self.image_part_name,
        }


@dataclass(frozen=True)
class DefinedNameFixtureInfo:
    fixture_path: Path
    sheet_name: str
    defined_name_count: int
    formula_like_count: int
    local_sheet_scoped_count: int
    external_reference_count: int
    three_d_reference_count: int
    sample_names: tuple[str, ...]

    def as_report(self, fixture_root: Path | None = None) -> dict[str, Any]:
        try:
            fixture_name = (
                self.fixture_path.relative_to(fixture_root).as_posix()
                if fixture_root is not None
                else self.fixture_path.name
            )
        except ValueError:
            fixture_name = self.fixture_path.name
        return {
            "fixture": fixture_name,
            "sheet_name": self.sheet_name,
            "defined_name_count": self.defined_name_count,
            "formula_like_count": self.formula_like_count,
            "local_sheet_scoped_count": self.local_sheet_scoped_count,
            "external_reference_count": self.external_reference_count,
            "three_d_reference_count": self.three_d_reference_count,
            "sample_names": list(self.sample_names),
        }


def load_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required for workbook-editor QA") from exc
    return openpyxl


def load_xlsxwriter():
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return None
    return xlsxwriter


def openpyxl_image_count(worksheet: Any) -> int:
    return len(getattr(worksheet, "_images", []))


def openpyxl_total_image_count(workbook: Any) -> int:
    return sum(openpyxl_image_count(worksheet) for worksheet in workbook.worksheets)


def xml_local_name(tag: str) -> str:
    if tag.startswith("{"):
        return tag.rsplit("}", 1)[1]
    return tag


def relationship_target_to_entry(owner_entry: str, target: str, target_mode: str = "") -> str | None:
    if target_mode.lower() == "external":
        return None
    if "://" in target or "?" in target or "#" in target:
        return None

    if target.startswith("/"):
        candidate = posixpath.normpath(target.lstrip("/"))
    else:
        candidate = posixpath.normpath(posixpath.join(posixpath.dirname(owner_entry), target))

    if candidate in ("", ".") or candidate == ".." or candidate.startswith("../"):
        return None
    return candidate


def workbook_sheet_entry_map(path: Path) -> dict[str, str]:
    names = zip_names(path)
    if "xl/workbook.xml" not in names or "xl/_rels/workbook.xml.rels" not in names:
        return {}

    workbook_root = ElementTree.fromstring(read_zip_bytes(path, "xl/workbook.xml"))
    rels_root = ElementTree.fromstring(read_zip_bytes(path, "xl/_rels/workbook.xml.rels"))
    rels: dict[str, tuple[str, str]] = {}
    for relationship in rels_root:
        if xml_local_name(relationship.tag) != "Relationship":
            continue
        relationship_id = relationship.attrib.get("Id")
        target = relationship.attrib.get("Target")
        if not relationship_id or not target:
            continue
        rels[relationship_id] = (target, relationship.attrib.get("TargetMode", ""))

    sheet_entries: dict[str, str] = {}
    sheets_element = next(
        (child for child in workbook_root if xml_local_name(child.tag) == "sheets"),
        None,
    )
    if sheets_element is None:
        return sheet_entries

    for sheet in sheets_element:
        if xml_local_name(sheet.tag) != "sheet":
            continue
        sheet_name = sheet.attrib.get("name")
        relationship_id = sheet.attrib.get(f"{{{NAMESPACES['office_rel']}}}id")
        if not sheet_name or not relationship_id or relationship_id not in rels:
            continue
        target, target_mode = rels[relationship_id]
        worksheet_entry = relationship_target_to_entry("xl/workbook.xml", target, target_mode)
        if worksheet_entry:
            sheet_entries[sheet_name] = worksheet_entry
    return sheet_entries


def relationship_entries(path: Path, rels_entry: str) -> list[dict[str, str]]:
    names = zip_names(path)
    require(rels_entry in names, f"relationship entry missing: {rels_entry}")
    rels_root = ElementTree.fromstring(read_zip_bytes(path, rels_entry))
    entries: list[dict[str, str]] = []
    for relationship in rels_root:
        if xml_local_name(relationship.tag) != "Relationship":
            continue
        entries.append(
            {
                "id": relationship.attrib.get("Id", ""),
                "type": relationship.attrib.get("Type", ""),
                "target": relationship.attrib.get("Target", ""),
                "target_mode": relationship.attrib.get("TargetMode", ""),
            }
        )
    return entries


def worksheet_formula_summary(worksheet_xml: bytes) -> dict[str, int]:
    root = ElementTree.fromstring(worksheet_xml)
    summary = {
        "formula_count": 0,
        "shared_formula_count": 0,
        "shared_definition_count": 0,
        "metadata_only_shared_count": 0,
    }
    for element in root.iter():
        if xml_local_name(element.tag) != "f":
            continue
        summary["formula_count"] += 1
        if element.attrib.get("t") != "shared":
            continue
        summary["shared_formula_count"] += 1
        formula_text = "".join(element.itertext()).strip()
        if formula_text:
            summary["shared_definition_count"] += 1
        else:
            summary["metadata_only_shared_count"] += 1
    return summary


def formula_worksheet_infos(path: Path) -> list[FormulaWorksheetInfo]:
    sheet_entries = workbook_sheet_entry_map(path)
    if not sheet_entries:
        return []

    infos: list[FormulaWorksheetInfo] = []
    names = zip_names(path)
    for sheet_name, worksheet_entry in sheet_entries.items():
        if worksheet_entry not in names:
            continue
        summary = worksheet_formula_summary(read_zip_bytes(path, worksheet_entry))
        if summary["formula_count"] == 0:
            continue
        infos.append(
            FormulaWorksheetInfo(
                fixture_path=path,
                sheet_name=sheet_name,
                worksheet_entry=worksheet_entry,
                formula_count=summary["formula_count"],
                shared_formula_count=summary["shared_formula_count"],
                shared_definition_count=summary["shared_definition_count"],
                metadata_only_shared_count=summary["metadata_only_shared_count"],
            )
        )
    return infos


def formula_summary_for_sheet(path: Path, sheet_name: str) -> dict[str, Any]:
    sheet_entries = workbook_sheet_entry_map(path)
    worksheet_entry = sheet_entries.get(sheet_name)
    require(worksheet_entry is not None, f"formula summary: sheet not found: {sheet_name}")
    names = zip_names(path)
    require(worksheet_entry in names,
            f"formula summary: worksheet entry missing for {sheet_name}: {worksheet_entry}")
    summary: dict[str, Any] = dict(worksheet_formula_summary(read_zip_bytes(path, worksheet_entry)))
    summary["sheet_name"] = sheet_name
    summary["worksheet_entry"] = worksheet_entry
    return summary


def worksheet_formula_cells(path: Path, sheet_name: str) -> dict[str, str]:
    sheet_entries = workbook_sheet_entry_map(path)
    worksheet_entry = sheet_entries.get(sheet_name)
    require(worksheet_entry is not None, f"formula cells: sheet not found: {sheet_name}")
    root = ElementTree.fromstring(read_zip_bytes(path, worksheet_entry))
    formulas: dict[str, str] = {}
    for cell in root.iter():
        if xml_local_name(cell.tag) != "c":
            continue
        reference = cell.attrib.get("r")
        if not reference:
            continue
        formula = next(
            (child for child in cell if xml_local_name(child.tag) == "f"),
            None,
        )
        if formula is not None:
            formulas[reference] = "".join(formula.itertext())
    return formulas


def worksheet_cell_elements(path: Path, sheet_name: str) -> dict[str, ElementTree.Element]:
    sheet_entries = workbook_sheet_entry_map(path)
    worksheet_entry = sheet_entries.get(sheet_name)
    require(worksheet_entry is not None, f"worksheet cells: sheet not found: {sheet_name}")
    root = ElementTree.fromstring(read_zip_bytes(path, worksheet_entry))
    cells: dict[str, ElementTree.Element] = {}
    for cell in root.iter():
        if xml_local_name(cell.tag) != "c":
            continue
        reference = cell.attrib.get("r")
        if reference:
            cells[reference] = cell
    return cells


def worksheet_formula_cached_values(path: Path, sheet_name: str) -> dict[str, list[str]]:
    sheet_entries = workbook_sheet_entry_map(path)
    worksheet_entry = sheet_entries.get(sheet_name)
    require(worksheet_entry is not None, f"formula cached values: sheet not found: {sheet_name}")
    root = ElementTree.fromstring(read_zip_bytes(path, worksheet_entry))
    cached_values: dict[str, list[str]] = {}
    for cell in root.iter():
        if xml_local_name(cell.tag) != "c":
            continue
        reference = cell.attrib.get("r")
        if not reference:
            continue
        formula = next(
            (child for child in cell if xml_local_name(child.tag) == "f"),
            None,
        )
        if formula is None:
            continue
        values = [
            "".join(child.itertext())
            for child in cell
            if xml_local_name(child.tag) == "v"
        ]
        if values:
            cached_values[reference] = values
    return cached_values


def sorted_string_mapping(values: dict[str, str]) -> dict[str, str]:
    return {key: values[key] for key in sorted(values)}


def attach_shared_formula_report(
    zip_report: dict[str, Any],
    path: Path,
    sheet_name: str,
    expected_formula_cells: dict[str, str],
    *,
    stale_cached_values_removed: bool,
    excel_ui_smoke: str,
) -> dict[str, Any]:
    formula_summary = formula_summary_for_sheet(path, sheet_name)
    formula_cells = worksheet_formula_cells(path, sheet_name)
    for reference, expected in expected_formula_cells.items():
        require(
            formula_cells.get(reference) == expected,
            f"{sheet_name}: {reference} formula mismatch {formula_cells.get(reference)!r}",
        )
    require(
        formula_summary["formula_count"] == len(expected_formula_cells),
        f"{sheet_name}: formula count mismatch {formula_summary!r}",
    )
    require(
        formula_summary["shared_formula_count"] == 0,
        f"{sheet_name}: shared formula metadata remained {formula_summary!r}",
    )

    formula_summary["ordinary_formula_count"] = formula_summary["formula_count"]
    cached_formula_values = worksheet_formula_cached_values(path, sheet_name)
    require(
        not cached_formula_values,
        f"{sheet_name}: cached formula values remained {cached_formula_values!r}",
    )
    formula_summary["shared_metadata_removed"] = True
    formula_summary["stale_cached_values_removed"] = stale_cached_values_removed
    zip_report["formula_output"] = formula_summary
    zip_report["output_formula_cells"] = sorted_string_mapping(formula_cells)
    zip_report["checked_formula_cells"] = sorted_string_mapping(expected_formula_cells)
    zip_report["cached_formula_values_removed"] = True
    zip_report["shared_metadata_removed"] = True
    zip_report["stale_cached_values_removed"] = stale_cached_values_removed
    zip_report["excel_ui_smoke"] = excel_ui_smoke
    return formula_summary


def openpyxl_formula_cells(worksheet: Any, expected_formula_cells: dict[str, str]) -> dict[str, str]:
    formula_cells: dict[str, str] = {}
    for reference, expected in expected_formula_cells.items():
        value = worksheet[reference].value
        require(
            value == "=" + expected,
            f"openpyxl formula mismatch at {reference}: {value!r}",
        )
        formula_cells[reference] = value
    return sorted_string_mapping(formula_cells)


def discover_formula_fixture_infos(
    fixture_root: Path,
    globs: list[str],
    limit: int,
    *,
    shared_only: bool,
) -> list[FormulaWorksheetInfo]:
    infos: list[FormulaWorksheetInfo] = []
    for fixture_path in discover_external_fixtures(fixture_root, globs, 0):
        try:
            fixture_infos = formula_worksheet_infos(fixture_path)
        except Exception:
            continue
        for info in fixture_infos:
            if shared_only and info.shared_formula_count == 0:
                continue
            infos.append(info)
            if limit > 0 and len(infos) >= limit:
                return infos
    return infos


def workbook_defined_name_records_from_xml(workbook_xml: bytes) -> list[dict[str, Any]]:
    root = ElementTree.fromstring(workbook_xml)
    records: list[dict[str, Any]] = []
    for defined_names_element in root:
        if xml_local_name(defined_names_element.tag) != "definedNames":
            continue
        for element in defined_names_element:
            if xml_local_name(element.tag) != "definedName":
                continue
            attributes = {key: element.attrib[key] for key in sorted(element.attrib)}
            records.append(
                {
                    "name": attributes.get("name", ""),
                    "attributes": attributes,
                    "text": "".join(element.itertext()),
                }
            )
    return records


def workbook_defined_name_records(path: Path) -> list[dict[str, Any]]:
    names = zip_names(path)
    if "xl/workbook.xml" not in names:
        return []
    return workbook_defined_name_records_from_xml(read_zip_bytes(path, "xl/workbook.xml"))


def canonical_defined_name_records(records: list[dict[str, Any]]) -> list[str]:
    return sorted(json.dumps(record, sort_keys=True) for record in records)


def defined_name_has_three_d_reference(text: str) -> bool:
    if "!" not in text:
        return False
    qualifier = text.split("!", 1)[0].strip()
    if qualifier.startswith("="):
        qualifier = qualifier[1:].strip()
    if "]" in qualifier:
        qualifier = qualifier.rsplit("]", 1)[1]
    return ":" in qualifier


def summarize_defined_name_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    texts = [str(record.get("text", "")).strip() for record in records]
    return {
        "defined_name_count": len(records),
        "formula_like_count": sum(
            1 for text in texts if text.startswith("=") or "!" in text or "(" in text
        ),
        "local_sheet_scoped_count": sum(
            1 for record in records if "localSheetId" in record.get("attributes", {})
        ),
        "external_reference_count": sum(1 for text in texts if "[" in text and "]" in text),
        "three_d_reference_count": sum(
            1 for text in texts if defined_name_has_three_d_reference(text)
        ),
        "sample_names": [
            str(record.get("name", "")) for record in records[:5] if record.get("name", "")
        ],
    }


def defined_name_fixture_info(path: Path) -> DefinedNameFixtureInfo | None:
    records = workbook_defined_name_records(path)
    if not records:
        return None

    sheet_names = workbook_sheetnames(path)
    if not sheet_names:
        return None

    summary = summarize_defined_name_records(records)
    return DefinedNameFixtureInfo(
        fixture_path=path,
        sheet_name=sheet_names[0],
        defined_name_count=summary["defined_name_count"],
        formula_like_count=summary["formula_like_count"],
        local_sheet_scoped_count=summary["local_sheet_scoped_count"],
        external_reference_count=summary["external_reference_count"],
        three_d_reference_count=summary["three_d_reference_count"],
        sample_names=tuple(summary["sample_names"]),
    )


def discover_defined_name_fixture_infos(
    fixture_root: Path,
    globs: list[str],
    limit: int,
) -> list[DefinedNameFixtureInfo]:
    infos: list[DefinedNameFixtureInfo] = []
    for fixture_path in discover_external_fixtures(fixture_root, globs, 0):
        try:
            info = defined_name_fixture_info(fixture_path)
        except Exception:
            continue
        if info is None:
            continue
        infos.append(info)
        if limit > 0 and len(infos) >= limit:
            return infos
    return infos


def pick_image_part_name(path: Path) -> str | None:
    names = sorted(zip_names(path))
    for name in names:
        lowered = name.lower()
        if not lowered.startswith("xl/media/"):
            continue
        if lowered.endswith(".png") or lowered.endswith(".jpg") or lowered.endswith(".jpeg"):
            return name
    return None


def discover_image_fixture_infos(
    fixture_root: Path,
    globs: list[str],
    limit: int,
) -> list[ImageFixtureInfo]:
    openpyxl = load_openpyxl()
    infos: list[ImageFixtureInfo] = []
    for fixture_path in discover_external_fixtures(fixture_root, globs, 0):
        image_part_name = pick_image_part_name(fixture_path)
        if image_part_name is None:
            continue

        try:
            workbook = openpyxl.load_workbook(fixture_path, read_only=False, data_only=False)
        except Exception:
            continue
        try:
            image_sheet_name: str | None = None
            for worksheet in workbook.worksheets:
                if openpyxl_image_count(worksheet) > 0:
                    image_sheet_name = worksheet.title
                    break
            if image_sheet_name is None:
                continue
            infos.append(
                ImageFixtureInfo(
                    fixture_path=fixture_path,
                    sheet_name=image_sheet_name,
                    image_part_name=image_part_name,
                )
            )
        finally:
            workbook.close()

        if limit > 0 and len(infos) >= limit:
            return infos
    return infos


def run_tool(
    qa_exe: Path,
    scenario: str,
    case_dir: Path,
    *,
    source: Path | None = None,
    output: Path | None = None,
    sheet_name: str | None = None,
    rename_to: str | None = None,
    replacement_image: Path | None = None,
    image_part_name: str | None = None,
) -> dict[str, Any]:
    case_dir.mkdir(parents=True, exist_ok=True)
    report_path = case_dir / "tool-report.json"
    output_path = output if output is not None else case_dir / "output.xlsx"
    args = [
        str(qa_exe),
        "--scenario",
        scenario,
        "--output",
        str(output_path),
        "--report",
        str(report_path),
    ]
    if source is not None:
        args.extend(["--source", str(source)])
    if sheet_name:
        args.extend(["--sheet", sheet_name])
    if rename_to:
        args.extend(["--rename-to", rename_to])
    if replacement_image is not None:
        args.extend(["--replacement-image", str(replacement_image)])
    if image_part_name:
        args.extend(["--image-part", image_part_name])

    completed = subprocess.run(
        args,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if completed.returncode != 0:
        raise RuntimeError(
            f"{scenario}: tool failed with code {completed.returncode}\n"
            f"stdout:\n{completed.stdout}\n"
            f"stderr:\n{completed.stderr}"
        )

    require(report_path.exists(), f"{scenario}: tool did not write report: {report_path}")
    with report_path.open("r", encoding="utf-8") as stream:
        report = json.load(stream)
    return report


def verify_generated_rename_materialized(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated rename/materialized: missing workbook.xml")
    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    sheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('name="EditedData"' in workbook_xml, "generated rename/materialized: missing renamed sheet")
    require('name="Data"' not in workbook_xml, "generated rename/materialized: old sheet name still present")
    require("materialized-edit" in sheet_xml, "generated rename/materialized: missing edited A1 text")
    require('<c r="B2"><v>42</v></c>' in sheet_xml, "generated rename/materialized: missing edited B2 number")
    zip_report["sheet_xml"] = "checked"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["EditedData", "Untouched"],
                f"generated rename/materialized: unexpected sheetnames {workbook.sheetnames!r}")
        edited = workbook["EditedData"]
        untouched = workbook["Untouched"]
        require(edited["A1"].value == "materialized-edit", "generated rename/materialized: A1 mismatch")
        require(edited["B2"].value == 42, "generated rename/materialized: B2 mismatch")
        require(untouched["A1"].value == "keep-me", "generated rename/materialized: untouched A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "EditedData!A1": edited["A1"].value,
            "EditedData!B2": edited["B2"].value,
            "Untouched!A1": untouched["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_insert_formula(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated in-memory insert formula: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory insert formula: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory insert formula: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="A2"' in data_xml and "inserted-row" in data_xml,
            "generated in-memory insert formula: missing inserted A2 text")
    require('<c r="B2"><v>5</v></c>' in data_xml,
            "generated in-memory insert formula: missing inserted B2 number")
    require('r="A3"' in data_xml and "source-row" in data_xml,
            "generated in-memory insert formula: missing shifted source A3 text")
    require('<c r="B3"><v>3</v></c>' in data_xml,
            "generated in-memory insert formula: missing shifted source B3 number")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("C2") == "B2*2",
            f"generated in-memory insert formula: inserted C2 formula mismatch {formulas!r}")
    require(formulas.get("C3") == "B3*2",
            f"generated in-memory insert formula: shifted C3 formula mismatch {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory insert formula: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "item", "generated in-memory insert formula: A1 mismatch")
        require(data["A2"].value == "inserted-row",
                "generated in-memory insert formula: A2 mismatch")
        require(data["B2"].value == 5, "generated in-memory insert formula: B2 mismatch")
        require(data["C2"].value == "=B2*2",
                f"generated in-memory insert formula: C2 mismatch {data['C2'].value!r}")
        require(data["A3"].value == "source-row",
                "generated in-memory insert formula: A3 mismatch")
        require(data["B3"].value == 3, "generated in-memory insert formula: B3 mismatch")
        require(data["C3"].value == "=B3*2",
                f"generated in-memory insert formula: C3 mismatch {data['C3'].value!r}")
        require(notes["A1"].value == "preserved",
                "generated in-memory insert formula: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A2": data["A2"].value,
            "Data!B2": data["B2"].value,
            "Data!C2": data["C2"].value,
            "Data!A3": data["A3"].value,
            "Data!B3": data["B3"].value,
            "Data!C3": data["C3"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_delete_column_formula(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated in-memory delete column formula: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory delete column formula: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory delete column formula: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require("drop-me" not in data_xml,
            "generated in-memory delete column formula: deleted A1 text remained")
    require('<c r="A1"><v>7</v></c>' in data_xml,
            "generated in-memory delete column formula: missing shifted A1 number")
    require('r="C1"' in data_xml and "tail" in data_xml,
            "generated in-memory delete column formula: missing shifted C1 text")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("B1") == "A1+C1",
            f"generated in-memory delete column formula: shifted B1 formula mismatch {formulas!r}")
    require("C1" not in formulas,
            f"generated in-memory delete column formula: stale C1 formula remained {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory delete column formula: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == 7, "generated in-memory delete column formula: A1 mismatch")
        require(data["B1"].value == "=A1+C1",
                f"generated in-memory delete column formula: B1 mismatch {data['B1'].value!r}")
        require(data["C1"].value == "tail",
                "generated in-memory delete column formula: C1 mismatch")
        require(notes["A1"].value == "preserved",
                "generated in-memory delete column formula: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_insert_column_formula(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated in-memory insert column formula: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory insert column formula: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory insert column formula: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="A1"' in data_xml and "item" in data_xml,
            "generated in-memory insert column formula: missing source A1 text")
    require('r="B1"' in data_xml and "inserted-col" in data_xml,
            "generated in-memory insert column formula: missing inserted B1 text")
    require('<c r="C1"><v>2</v></c>' in data_xml,
            "generated in-memory insert column formula: missing shifted C1 number")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("D1") == "C1*2",
            f"generated in-memory insert column formula: shifted D1 formula mismatch {formulas!r}")
    require("C1" not in formulas,
            f"generated in-memory insert column formula: stale C1 formula remained {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory insert column formula: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "item", "generated in-memory insert column formula: A1 mismatch")
        require(data["B1"].value == "inserted-col",
                "generated in-memory insert column formula: B1 mismatch")
        require(data["C1"].value == 2, "generated in-memory insert column formula: C1 mismatch")
        require(data["D1"].value == "=C1*2",
                f"generated in-memory insert column formula: D1 mismatch {data['D1'].value!r}")
        require(notes["A1"].value == "preserved",
                "generated in-memory insert column formula: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Data!D1": data["D1"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_delete_row_formula(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated in-memory delete row formula: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory delete row formula: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory delete row formula: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require("drop-row" not in data_xml,
            "generated in-memory delete row formula: deleted row text remained")
    require('<c r="A1"><v>4</v></c>' in data_xml,
            "generated in-memory delete row formula: missing shifted A1 number")
    require('<c r="A2"><v>6</v></c>' in data_xml,
            "generated in-memory delete row formula: missing shifted A2 number")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("B1") == "A1+A2",
            f"generated in-memory delete row formula: shifted B1 formula mismatch {formulas!r}")
    require("B2" not in formulas,
            f"generated in-memory delete row formula: stale B2 formula remained {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory delete row formula: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == 4, "generated in-memory delete row formula: A1 mismatch")
        require(data["B1"].value == "=A1+A2",
                f"generated in-memory delete row formula: B1 mismatch {data['B1'].value!r}")
        require(data["A2"].value == 6, "generated in-memory delete row formula: A2 mismatch")
        require(notes["A1"].value == "preserved",
                "generated in-memory delete row formula: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!A2": data["A2"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_stationary_formula_shift(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory stationary formula shift"
    expected_formula = (
        "SUM($A$4,C$4,Data!$A$4,$E$1,$E1,Data!$E$1,#REF!,#REF!,"
        "Data!#REF!,#REF!,Data!#REF!)"
    )
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, f"{label}: missing workbook.xml")
    require("xl/calcChain.xml" not in names, f"{label}: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"{label}: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require("delete-row" not in data_xml, f"{label}: deleted row text remained")
    require("delete-column" not in data_xml, f"{label}: deleted column text remained")
    require('r="A4"' in data_xml and "row-target" in data_xml,
            f"{label}: missing row-shifted A4 source text")
    require('r="C4"' in data_xml and "c3-target" in data_xml,
            f"{label}: missing row-shifted C4 source text")
    require('r="E1"' in data_xml and "<v>4</v>" in data_xml,
            f"{label}: missing column-shifted E1 number")
    require('r="F1"' in data_xml and "survive-column" in data_xml,
            f"{label}: missing surviving F1 text")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("C1") == expected_formula,
            f"{label}: stationary C1 formula mismatch {formulas!r}")
    require(len(formulas) == 1, f"{label}: unexpected formulas {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas
    zip_report["structural_rewrite"] = "checked"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"{label}: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "item", f"{label}: Data!A1 mismatch")
        require(data["B1"].value == "label", f"{label}: Data!B1 mismatch")
        require(data["C1"].value == f"={expected_formula}",
                f"{label}: Data!C1 mismatch {data['C1'].value!r}")
        require(data["D1"].value is None, f"{label}: Data!D1 should be blank")
        require(data["E1"].value == 4, f"{label}: Data!E1 mismatch")
        require(data["F1"].value == "survive-column", f"{label}: Data!F1 mismatch")
        require(data["G1"].value is None, f"{label}: Data!G1 should be deleted")
        require(data["A4"].value == "row-target", f"{label}: Data!A4 mismatch")
        require(data["C4"].value == "c3-target", f"{label}: Data!C4 mismatch")
        require(data["A7"].value is None, f"{label}: Data!A7 should be deleted")
        require(notes["A1"].value == "preserved", f"{label}: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Data!D1": data["D1"].value,
            "Data!E1": data["E1"].value,
            "Data!F1": data["F1"].value,
            "Data!G1": data["G1"].value,
            "Data!A4": data["A4"].value,
            "Data!C4": data["C4"].value,
            "Data!A7": data["A7"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_clear_erase(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated in-memory clear/erase: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory clear/erase: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory clear/erase: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="A1"' in data_xml and "keep-a1" in data_xml,
            "generated in-memory clear/erase: missing preserved A1 text")
    require('<c r="A2"><v>8</v></c>' in data_xml,
            "generated in-memory clear/erase: missing preserved A2 number")
    require('r="D1"' in data_xml and "new-d1" in data_xml,
            "generated in-memory clear/erase: missing new D1 text")
    require("erase-me" not in data_xml,
            "generated in-memory clear/erase: erased C1 text remained")

    formulas = worksheet_formula_cells(path, "Data")
    require("B1" not in formulas,
            f"generated in-memory clear/erase: cleared B1 formula remained {formulas!r}")
    cells = worksheet_cell_elements(path, "Data")
    b1 = cells.get("B1")
    require(b1 is not None,
            f"generated in-memory clear/erase: explicit blank B1 missing {sorted(cells)}")
    require(
        not any(xml_local_name(child.tag) in {"f", "v", "is"} for child in b1),
        "generated in-memory clear/erase: explicit blank B1 kept formula/value payload",
    )
    require("C1" not in cells,
            f"generated in-memory clear/erase: erased C1 remained represented {sorted(cells)}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas
    zip_report["represented_cells"] = sorted(cells)

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory clear/erase: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "keep-a1",
                "generated in-memory clear/erase: A1 mismatch")
        require(data["B1"].value is None,
                f"generated in-memory clear/erase: B1 mismatch {data['B1'].value!r}")
        require(data["C1"].value is None,
                f"generated in-memory clear/erase: C1 mismatch {data['C1'].value!r}")
        require(data["D1"].value == "new-d1",
                "generated in-memory clear/erase: D1 mismatch")
        require(data["A2"].value == 8,
                "generated in-memory clear/erase: A2 mismatch")
        require(notes["A1"].value == "preserved",
                "generated in-memory clear/erase: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Data!D1": data["D1"].value,
            "Data!A2": data["A2"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_append_row_formula(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated in-memory append row formula: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory append row formula: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory append row formula: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="A1"' in data_xml and "item" in data_xml,
            "generated in-memory append row formula: missing source A1 text")
    require('r="A2"' in data_xml and "source-row" in data_xml,
            "generated in-memory append row formula: missing source A2 text")
    require('<c r="B2"><v>10</v></c>' in data_xml,
            "generated in-memory append row formula: missing source B2 number")
    require('r="A3"' in data_xml and "appended-row" in data_xml,
            "generated in-memory append row formula: missing appended A3 text")
    require('<c r="B3"><v>4</v></c>' in data_xml,
            "generated in-memory append row formula: missing appended B3 number")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("C3") == "B3*2",
            f"generated in-memory append row formula: appended C3 formula mismatch {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory append row formula: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "item", "generated in-memory append row formula: A1 mismatch")
        require(data["B1"].value == "value", "generated in-memory append row formula: B1 mismatch")
        require(data["C1"].value == "double", "generated in-memory append row formula: C1 mismatch")
        require(data["A2"].value == "source-row",
                "generated in-memory append row formula: A2 mismatch")
        require(data["B2"].value == 10, "generated in-memory append row formula: B2 mismatch")
        require(data["A3"].value == "appended-row",
                "generated in-memory append row formula: A3 mismatch")
        require(data["B3"].value == 4, "generated in-memory append row formula: B3 mismatch")
        require(data["C3"].value == "=B3*2",
                f"generated in-memory append row formula: C3 mismatch {data['C3'].value!r}")
        require(notes["A1"].value == "preserved",
                "generated in-memory append row formula: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Data!A2": data["A2"].value,
            "Data!B2": data["B2"].value,
            "Data!A3": data["A3"].value,
            "Data!B3": data["B3"].value,
            "Data!C3": data["C3"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_overwrite_formula_text(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names,
            "generated in-memory overwrite formula/text: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory overwrite formula/text: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory overwrite formula/text: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require("old-text" not in data_xml,
            "generated in-memory overwrite formula/text: old A1 text remained")
    require("B1*2" not in data_xml,
            "generated in-memory overwrite formula/text: old C1 formula remained")
    require('r="A1"' in data_xml and "new-text" in data_xml,
            "generated in-memory overwrite formula/text: missing overwritten A1 text")
    require('<c r="B1"><v>5</v></c>' in data_xml,
            "generated in-memory overwrite formula/text: missing overwritten B1 number")
    require('r="A2"' in data_xml and "keep-row-two" in data_xml,
            "generated in-memory overwrite formula/text: missing preserved A2 text")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("C1") == "B1+10",
            f"generated in-memory overwrite formula/text: C1 formula mismatch {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory overwrite formula/text: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "new-text",
                "generated in-memory overwrite formula/text: A1 mismatch")
        require(data["B1"].value == 5,
                "generated in-memory overwrite formula/text: B1 mismatch")
        require(data["C1"].value == "=B1+10",
                f"generated in-memory overwrite formula/text: C1 mismatch {data['C1'].value!r}")
        require(data["A2"].value == "keep-row-two",
                "generated in-memory overwrite formula/text: A2 mismatch")
        require(notes["A1"].value == "preserved",
                "generated in-memory overwrite formula/text: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Data!A2": data["A2"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_in_memory_retry_noop_source_unchanged(
    source_path: Path,
    label: str,
) -> dict[str, Any]:
    source_sheet_entries = workbook_sheet_entry_map(source_path)
    require(source_sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"{label}: unexpected source sheet map {source_sheet_entries!r}")
    source_data_xml = read_zip_text(source_path, source_sheet_entries["Data"])
    require("old-text" in source_data_xml and "new-text" not in source_data_xml,
            f"{label}: source Data text payload was overwritten")
    require("B1*2" in source_data_xml and "B1+10" not in source_data_xml,
            f"{label}: source Data formula payload was overwritten")
    return {
        "source_sheet_entries": source_sheet_entries,
        "source_payload": "checked",
    }


def verify_generated_in_memory_retry_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory retry no-op save"
    require(tool_report.get("status") == "expected_retry_observed",
            f"{label}: unexpected tool status {tool_report!r}")
    require(tool_report.get("error_message"),
            f"{label}: missing rejected save diagnostic")
    mutations = tool_report.get("mutations", [])
    require("save_as(noop-output)" in mutations,
            f"{label}: tool did not report the no-op save stage")
    source_report = verify_in_memory_retry_noop_source_unchanged(
        Path(tool_report["source"]),
        label,
    )

    zip_report, openpyxl_report = verify_generated_in_memory_overwrite_formula_text(path)
    zip_report.update(source_report)
    zip_report["retry_error_message"] = tool_report["error_message"]
    zip_report["retry_noop_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_retry_path_equivalent_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory path-equivalent retry no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_retry_noop_save(
        path,
        tool_report,
    )
    require(
        "save_as(path-equivalent-source) rejected" in tool_report.get("mutations", []),
        f"{label}: tool did not report the path-equivalent rejection stage",
    )
    zip_report["path_equivalent_retry"] = "checked"
    return zip_report, openpyxl_report


def verify_in_memory_retry_reopen_source_unchanged(
    source_path: Path,
    label: str,
) -> dict[str, Any]:
    source_sheet_entries = workbook_sheet_entry_map(source_path)
    require(source_sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"{label}: unexpected source sheet map {source_sheet_entries!r}")
    source_data_xml = read_zip_text(source_path, source_sheet_entries["Data"])
    require("seed-a1" in source_data_xml and "first-edit" not in source_data_xml,
            f"{label}: source Data A1 payload was overwritten")
    require("B1+1" in source_data_xml and "B1+5" not in source_data_xml,
            f"{label}: source Data formula payload was overwritten")
    require("reopened-row" not in source_data_xml,
            f"{label}: source Data appended row was written into the source")
    return {
        "source_sheet_entries": source_sheet_entries,
        "source_payload": "checked",
    }


def verify_generated_in_memory_retry_reopen_modify_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory retry reopen modify no-op save"
    require(tool_report.get("status") == "expected_retry_observed",
            f"{label}: unexpected tool status {tool_report!r}")
    require(tool_report.get("error_message"),
            f"{label}: missing rejected save diagnostic")
    mutations = tool_report.get("mutations", [])
    require("second:open(safe-retry-output)" in mutations,
            f"{label}: tool did not report the retry reopen stage")
    require("second:save_as(noop-output)" in mutations,
            f"{label}: tool did not report the no-op save stage")
    source_report = verify_in_memory_retry_reopen_source_unchanged(
        Path(tool_report["source"]),
        label,
    )

    zip_report, openpyxl_report = verify_generated_in_memory_reopen_modify_save(path)
    zip_report.update(source_report)
    zip_report["retry_error_message"] = tool_report["error_message"]
    zip_report["retry_reopen_noop_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_retry_path_equivalent_reopen_modify_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory retry path-equivalent reopen modify no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_retry_reopen_modify_noop_save(
        path,
        tool_report,
    )
    require("first:save_as(path-equivalent-source) rejected" in tool_report.get("mutations", []),
            f"{label}: tool did not report the path-equivalent rejected save stage")
    zip_report["path_equivalent_retry"] = "checked"
    return zip_report, openpyxl_report


def verify_generated_in_memory_retry_reopen_modify_post_noop_third_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory retry reopen modify post-noop third save"
    zip_report, openpyxl_report = verify_generated_in_memory_retry_reopen_modify_noop_save(
        path,
        tool_report,
    )
    mutations = tool_report.get("mutations", [])
    require("third:worksheet(Data).set_cell(E1,text)" in mutations,
            f"{label}: tool did not report the post-noop Data edit")
    require("third:save_as(third-noop-output)" in mutations,
            f"{label}: tool did not report the final no-op save stage")

    sheet_entries = zip_report["sheet_entries"]
    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="E1"' in data_xml and "third-edit" in data_xml,
            f"{label}: missing post-noop Data!E1 text")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        data = workbook["Data"]
        require(data["E1"].value == "third-edit",
                f"{label}: Data!E1 mismatch {data['E1'].value!r}")
        openpyxl_report["Data!E1"] = data["E1"].value
    finally:
        workbook.close()

    zip_report["retry_reopen_post_noop_third_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_retry_path_equivalent_reopen_modify_post_noop_third_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory retry path-equivalent reopen modify post-noop third save"
    zip_report, openpyxl_report = verify_generated_in_memory_retry_reopen_modify_post_noop_third_save(
        path,
        tool_report,
    )
    require("first:save_as(path-equivalent-source) rejected" in tool_report.get("mutations", []),
            f"{label}: tool did not report the path-equivalent rejected save stage")
    zip_report["path_equivalent_retry"] = "checked"
    return zip_report, openpyxl_report


def verify_generated_in_memory_reopen_modify_save(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names,
            "generated in-memory reopen/modify/save: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory reopen/modify/save: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Notes": "xl/worksheets/sheet2.xml",
    }, f"generated in-memory reopen/modify/save: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    require("seed-a1" not in data_xml,
            "generated in-memory reopen/modify/save: original A1 text remained")
    require("B1+1" not in data_xml,
            "generated in-memory reopen/modify/save: original C1 formula remained")
    require('r="A1"' in data_xml and "first-edit" in data_xml,
            "generated in-memory reopen/modify/save: missing first-stage A1 text")
    require('<c r="B1"><v>10</v></c>' in data_xml,
            "generated in-memory reopen/modify/save: missing second-stage B1 number")
    require('r="D1"' in data_xml and "second-edit" in data_xml,
            "generated in-memory reopen/modify/save: missing second-stage D1 text")
    require('r="A2"' in data_xml and "keep-row-two" in data_xml,
            "generated in-memory reopen/modify/save: missing preserved A2 text")
    require('r="A3"' in data_xml and "reopened-row" in data_xml,
            "generated in-memory reopen/modify/save: missing first-stage appended A3 text")
    require('<c r="B3"><v>4</v></c>' in data_xml,
            "generated in-memory reopen/modify/save: missing first-stage appended B3 number")
    formulas = worksheet_formula_cells(path, "Data")
    require(formulas.get("C1") == "B1+5",
            f"generated in-memory reopen/modify/save: C1 formula mismatch {formulas!r}")
    require(formulas.get("C3") == "B3*2",
            f"generated in-memory reopen/modify/save: C3 formula mismatch {formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["formulas"] = formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Notes"],
                f"generated in-memory reopen/modify/save: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        notes = workbook["Notes"]
        require(data["A1"].value == "first-edit",
                "generated in-memory reopen/modify/save: A1 mismatch")
        require(data["B1"].value == 10,
                "generated in-memory reopen/modify/save: B1 mismatch")
        require(data["C1"].value == "=B1+5",
                f"generated in-memory reopen/modify/save: C1 mismatch {data['C1'].value!r}")
        require(data["D1"].value == "second-edit",
                "generated in-memory reopen/modify/save: D1 mismatch")
        require(data["A2"].value == "keep-row-two",
                "generated in-memory reopen/modify/save: A2 mismatch")
        require(data["A3"].value == "reopened-row",
                "generated in-memory reopen/modify/save: A3 mismatch")
        require(data["B3"].value == 4,
                "generated in-memory reopen/modify/save: B3 mismatch")
        require(data["C3"].value == "=B3*2",
                f"generated in-memory reopen/modify/save: C3 mismatch {data['C3'].value!r}")
        require(notes["A1"].value == "preserved",
                "generated in-memory reopen/modify/save: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!C1": data["C1"].value,
            "Data!D1": data["D1"].value,
            "Data!A2": data["A2"].value,
            "Data!A3": data["A3"].value,
            "Data!B3": data["B3"].value,
            "Data!C3": data["C3"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_reopen_modify_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory reopen/modify/no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_reopen_modify_save(path)
    require("third:save_as(noop-output)" in tool_report.get("mutations", []),
            f"{label}: tool did not report the no-op save stage")
    zip_report["noop_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_reopen_modify_post_noop_third_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory reopen/modify post-noop third save"
    zip_report, openpyxl_report = verify_generated_in_memory_reopen_modify_noop_save(
        path,
        tool_report,
    )
    mutations = tool_report.get("mutations", [])
    require("fourth:worksheet(Data).set_cell(E1,text)" in mutations,
            f"{label}: tool did not report the post-noop Data edit")
    require("fifth:save_as(third-noop-output)" in mutations,
            f"{label}: tool did not report the final no-op save stage")

    sheet_entries = zip_report["sheet_entries"]
    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="E1"' in data_xml and "third-edit" in data_xml,
            f"{label}: missing post-noop Data!E1 text")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        data = workbook["Data"]
        require(data["E1"].value == "third-edit",
                f"{label}: Data!E1 mismatch")
        openpyxl_report["Data!E1"] = data["E1"].value
    finally:
        workbook.close()

    zip_report["post_noop_third_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_save(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names,
            "generated in-memory multi-sheet save: missing workbook.xml")
    require("xl/calcChain.xml" not in names,
            "generated in-memory multi-sheet save: unexpected calcChain.xml")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Summary": "xl/worksheets/sheet2.xml",
        "Notes": "xl/worksheets/sheet3.xml",
    }, f"generated in-memory multi-sheet save: unexpected sheet map {sheet_entries!r}")

    data_xml = read_zip_text(path, sheet_entries["Data"])
    summary_xml = read_zip_text(path, sheet_entries["Summary"])
    require("old-data" not in data_xml,
            "generated in-memory multi-sheet save: old Data!A1 text remained")
    require("old-summary" not in summary_xml,
            "generated in-memory multi-sheet save: old Summary!A1 text remained")
    require("Data!B1*2" not in summary_xml,
            "generated in-memory multi-sheet save: old Summary!B1 formula remained")
    require('r="A1"' in data_xml and "edited-data" in data_xml,
            "generated in-memory multi-sheet save: missing edited Data!A1 text")
    require('<c r="B1"><v>7</v></c>' in data_xml,
            "generated in-memory multi-sheet save: missing edited Data!B1 number")
    require('r="A2"' in data_xml and "keep-data-row" in data_xml,
            "generated in-memory multi-sheet save: missing preserved Data!A2")
    require('r="A3"' in data_xml and "multi-row" in data_xml,
            "generated in-memory multi-sheet save: missing appended Data!A3")
    require('<c r="B3"><v>3</v></c>' in data_xml,
            "generated in-memory multi-sheet save: missing appended Data!B3 number")
    require('r="A1"' in summary_xml and "edited-summary" in summary_xml,
            "generated in-memory multi-sheet save: missing edited Summary!A1 text")
    data_formulas = worksheet_formula_cells(path, "Data")
    summary_formulas = worksheet_formula_cells(path, "Summary")
    require(data_formulas.get("C3") == "B3+Data!B1",
            f"generated in-memory multi-sheet save: Data!C3 formula mismatch {data_formulas!r}")
    require(summary_formulas.get("B1") == "Data!B1+Data!B3",
            f"generated in-memory multi-sheet save: Summary!B1 formula mismatch {summary_formulas!r}")
    zip_report["sheet_entries"] = sheet_entries
    zip_report["data_formulas"] = data_formulas
    zip_report["summary_formulas"] = summary_formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Summary", "Notes"],
                f"generated in-memory multi-sheet save: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        summary = workbook["Summary"]
        notes = workbook["Notes"]
        require(data["A1"].value == "edited-data",
                "generated in-memory multi-sheet save: Data!A1 mismatch")
        require(data["B1"].value == 7,
                "generated in-memory multi-sheet save: Data!B1 mismatch")
        require(data["A2"].value == "keep-data-row",
                "generated in-memory multi-sheet save: Data!A2 mismatch")
        require(data["A3"].value == "multi-row",
                "generated in-memory multi-sheet save: Data!A3 mismatch")
        require(data["B3"].value == 3,
                "generated in-memory multi-sheet save: Data!B3 mismatch")
        require(data["C3"].value == "=B3+Data!B1",
                f"generated in-memory multi-sheet save: Data!C3 mismatch {data['C3'].value!r}")
        require(summary["A1"].value == "edited-summary",
                "generated in-memory multi-sheet save: Summary!A1 mismatch")
        require(summary["B1"].value == "=Data!B1+Data!B3",
                f"generated in-memory multi-sheet save: Summary!B1 mismatch {summary['B1'].value!r}")
        require(notes["A1"].value == "preserved",
                "generated in-memory multi-sheet save: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!A2": data["A2"].value,
            "Data!A3": data["A3"].value,
            "Data!B3": data["B3"].value,
            "Data!C3": data["C3"].value,
            "Summary!A1": summary["A1"].value,
            "Summary!B1": summary["B1"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_save(path)
    require("save_as(noop-output)" in tool_report.get("mutations", []),
            f"{label}: tool did not report the no-op save stage")
    zip_report["noop_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_in_memory_multi_sheet_retry_source_unchanged(
    source_path: Path,
    label: str,
) -> dict[str, Any]:
    source_sheet_entries = workbook_sheet_entry_map(source_path)
    require(source_sheet_entries == {
        "Data": "xl/worksheets/sheet1.xml",
        "Summary": "xl/worksheets/sheet2.xml",
        "Notes": "xl/worksheets/sheet3.xml",
    }, f"{label}: unexpected source sheet map {source_sheet_entries!r}")
    source_data_xml = read_zip_text(source_path, source_sheet_entries["Data"])
    source_summary_xml = read_zip_text(source_path, source_sheet_entries["Summary"])
    require("old-data" in source_data_xml and "edited-data" not in source_data_xml,
            f"{label}: source Data payload was overwritten")
    require("old-summary" in source_summary_xml and "edited-summary" not in source_summary_xml,
            f"{label}: source Summary payload was overwritten")
    require("Data!B1*2" in source_summary_xml and "Data!B1+Data!B3" not in source_summary_xml,
            f"{label}: source Summary formula was overwritten")
    return {
        "source_sheet_entries": source_sheet_entries,
        "source_payload": "checked",
    }


def verify_generated_in_memory_multi_sheet_retry_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry save"
    require(tool_report.get("status") == "expected_retry_observed",
            f"{label}: unexpected tool status {tool_report!r}")
    require(tool_report.get("error_message"),
            f"{label}: missing rejected save diagnostic")
    source_path = Path(tool_report["source"])
    source_report = verify_in_memory_multi_sheet_retry_source_unchanged(source_path, label)

    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_save(path)
    zip_report.update(source_report)
    zip_report["retry_error_message"] = tool_report["error_message"]
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_save(
        path,
        tool_report,
    )
    require("save_as(noop-output)" in tool_report.get("mutations", []),
            f"{label}: tool did not report the no-op save stage")
    zip_report["retry_noop_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_path_equivalent_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry path-equivalent no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_noop_save(
        path,
        tool_report,
    )
    require("save_as(path-equivalent-source) rejected" in tool_report.get("mutations", []),
            f"{label}: tool did not report the path-equivalent rejected save stage")
    zip_report["path_equivalent_retry"] = "checked"
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_reopen_modify_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry reopen modify save"
    require(tool_report.get("status") == "expected_retry_observed",
            f"{label}: unexpected tool status {tool_report!r}")
    require(tool_report.get("error_message"),
            f"{label}: missing rejected save diagnostic")
    source_report = verify_in_memory_multi_sheet_retry_source_unchanged(
        Path(tool_report["source"]),
        label,
    )

    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_save(path)
    sheet_entries = zip_report["sheet_entries"]
    data_xml = read_zip_text(path, sheet_entries["Data"])
    summary_xml = read_zip_text(path, sheet_entries["Summary"])
    require('r="D1"' in data_xml and "retry-reopened-data" in data_xml,
            f"{label}: missing second-stage Data!D1 text")
    data_formulas = worksheet_formula_cells(path, "Data")
    summary_formulas = worksheet_formula_cells(path, "Summary")
    require(data_formulas.get("C3") == "B3+Data!B1",
            f"{label}: Data!C3 formula mismatch {data_formulas!r}")
    require(summary_formulas.get("B1") == "Data!B1+Data!B3",
            f"{label}: Summary!B1 formula mismatch {summary_formulas!r}")
    require(summary_formulas.get("C1") == "Data!B1+10",
            f"{label}: Summary!C1 formula mismatch {summary_formulas!r}")
    require('r="C1"' in summary_xml,
            f"{label}: missing second-stage Summary!C1 cell")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Summary", "Notes"],
                f"{label}: unexpected sheetnames {workbook.sheetnames!r}")
        data = workbook["Data"]
        summary = workbook["Summary"]
        notes = workbook["Notes"]
        require(data["A1"].value == "edited-data", f"{label}: Data!A1 mismatch")
        require(data["B1"].value == 7, f"{label}: Data!B1 mismatch")
        require(data["A2"].value == "keep-data-row", f"{label}: Data!A2 mismatch")
        require(data["A3"].value == "multi-row", f"{label}: Data!A3 mismatch")
        require(data["B3"].value == 3, f"{label}: Data!B3 mismatch")
        require(data["C3"].value == "=B3+Data!B1",
                f"{label}: Data!C3 mismatch {data['C3'].value!r}")
        require(data["D1"].value == "retry-reopened-data", f"{label}: Data!D1 mismatch")
        require(summary["A1"].value == "edited-summary", f"{label}: Summary!A1 mismatch")
        require(summary["B1"].value == "=Data!B1+Data!B3",
                f"{label}: Summary!B1 mismatch {summary['B1'].value!r}")
        require(summary["C1"].value == "=Data!B1+10",
                f"{label}: Summary!C1 mismatch {summary['C1'].value!r}")
        require(notes["A1"].value == "preserved", f"{label}: Notes!A1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": data["A1"].value,
            "Data!B1": data["B1"].value,
            "Data!A2": data["A2"].value,
            "Data!A3": data["A3"].value,
            "Data!B3": data["B3"].value,
            "Data!C3": data["C3"].value,
            "Data!D1": data["D1"].value,
            "Summary!A1": summary["A1"].value,
            "Summary!B1": summary["B1"].value,
            "Summary!C1": summary["C1"].value,
            "Notes!A1": notes["A1"].value,
        }
    finally:
        workbook.close()

    zip_report.update(source_report)
    zip_report["data_formulas"] = data_formulas
    zip_report["summary_formulas"] = summary_formulas
    zip_report["retry_error_message"] = tool_report["error_message"]
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_reopen_modify_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry reopen modify no-op save"
    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_reopen_modify_save(
        path,
        tool_report,
    )
    require("third:save_as(noop-output)" in tool_report.get("mutations", []),
            f"{label}: tool did not report the no-op save stage")
    zip_report["noop_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_noop_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry path-equivalent reopen modify no-op save"
    zip_report, openpyxl_report = (
        verify_generated_in_memory_multi_sheet_retry_reopen_modify_noop_save(
            path,
            tool_report,
        )
    )
    require("first:save_as(path-equivalent-source) rejected" in tool_report.get("mutations", []),
            f"{label}: tool did not report the path-equivalent rejected save stage")
    zip_report["path_equivalent_retry"] = "checked"
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry reopen modify post-noop third save"
    zip_report, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_reopen_modify_noop_save(
        path,
        tool_report,
    )
    mutations = tool_report.get("mutations", [])
    require("fourth:worksheet(Data).set_cell(E1,text)" in mutations,
            f"{label}: tool did not report the post-noop Data edit")
    require("fourth:worksheet(Summary).set_cell(D1,formula)" in mutations,
            f"{label}: tool did not report the post-noop Summary edit")
    require("fifth:save_as(third-noop-output)" in mutations,
            f"{label}: tool did not report the final no-op save stage")

    sheet_entries = zip_report["sheet_entries"]
    data_xml = read_zip_text(path, sheet_entries["Data"])
    require('r="E1"' in data_xml and "retry-reopened-post-noop-data" in data_xml,
            f"{label}: missing post-noop Data!E1 text")
    summary_formulas = worksheet_formula_cells(path, "Summary")
    require(summary_formulas.get("D1") == "Data!B1+20",
            f"{label}: Summary!D1 formula mismatch {summary_formulas!r}")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        data = workbook["Data"]
        summary = workbook["Summary"]
        require(data["E1"].value == "retry-reopened-post-noop-data",
                f"{label}: Data!E1 mismatch")
        require(summary["D1"].value == "=Data!B1+20",
                f"{label}: Summary!D1 mismatch {summary['D1'].value!r}")
        openpyxl_report["Data!E1"] = data["E1"].value
        openpyxl_report["Summary!D1"] = summary["D1"].value
    finally:
        workbook.close()

    zip_report["summary_formulas"] = summary_formulas
    zip_report["post_noop_third_save"] = "byte-identical"
    return zip_report, openpyxl_report


def verify_generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    label = "generated in-memory multi-sheet retry path-equivalent reopen modify post-noop third save"
    zip_report, openpyxl_report = (
        verify_generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save(
            path,
            tool_report,
        )
    )
    require("first:save_as(path-equivalent-source) rejected" in tool_report.get("mutations", []),
            f"{label}: tool did not report the path-equivalent rejected save stage")
    zip_report["path_equivalent_retry"] = "checked"
    return zip_report, openpyxl_report


def verify_generated_source_formula_audit(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    formula_sheet_xml = read_zip_text(path, "xl/worksheets/sheet4.xml")

    require('name="RenamedData"' in workbook_xml,
            "generated source formula audit: missing renamed sheet catalog entry")
    require('name="Data"' not in workbook_xml,
            "generated source formula audit: old Data sheet catalog entry remained")
    require("Data!A1" in formula_sheet_xml,
            "generated source formula audit: original formula text was unexpectedly rewritten")
    require("RenamedData!A1" not in formula_sheet_xml,
            "generated source formula audit: non-materialized formula was silently repaired")

    require(tool_report.get("source_formula_audit_count") == 5,
            f"generated source formula audit: count mismatch {tool_report!r}")
    require(tool_report.get("source_formula_rename_risk_count") == 1,
            f"generated source formula audit: rename-risk count mismatch {tool_report!r}")
    require(tool_report.get("source_formula_external_count") == 1,
            f"generated source formula audit: external count mismatch {tool_report!r}")
    require(tool_report.get("source_formula_sheet_range_count") == 1,
            f"generated source formula audit: 3D count mismatch {tool_report!r}")
    require(tool_report.get("source_formula_matched_count") == 3,
            f"generated source formula audit: matched count mismatch {tool_report!r}")
    references = tool_report.get("source_formula_audit_references", [])
    require("Data!A1" in references,
            f"generated source formula audit: missing Data!A1 reference {references!r}")
    require("[Book.xlsx]Data!A1" in references,
            f"generated source formula audit: missing external reference {references!r}")
    require("Data:Formula!A1" in references,
            f"generated source formula audit: missing 3D reference {references!r}")
    zip_report["source_formula_audit"] = {
        "count": tool_report["source_formula_audit_count"],
        "rename_risk_count": tool_report["source_formula_rename_risk_count"],
        "external_count": tool_report["source_formula_external_count"],
        "sheet_range_count": tool_report["source_formula_sheet_range_count"],
        "matched_count": tool_report["source_formula_matched_count"],
        "references": references,
    }

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["RenamedData", "Other Sheet", "O'Brien", "Formula"],
                f"generated source formula audit: unexpected sheetnames {workbook.sheetnames!r}")
        formula_sheet = workbook["Formula"]
        require(formula_sheet["A1"].value is not None and "Data!A1" in formula_sheet["A1"].value,
                f"generated source formula audit: Formula!A1 mismatch {formula_sheet['A1'].value!r}")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Formula!A1": formula_sheet["A1"].value,
            "source_formula_audit_count": tool_report["source_formula_audit_count"],
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_formula_rename_rewrite(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    workbook_xml = read_zip_text(path, "xl/workbook.xml")

    require('name="RenamedData"' in workbook_xml,
            "generated formula rename rewrite: missing renamed sheet catalog entry")
    require('name="Data"' not in workbook_xml,
            "generated formula rename rewrite: old Data sheet catalog entry remained")
    require(
        R"<definedName name=\"ReportRange\">'RenamedData'!$A$1:$B$2</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename rewrite: ReportRange was not rewritten",
    )
    require(
        R"<definedName name=\"QuotedDataRef\">'RenamedData'!$A$1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename rewrite: QuotedDataRef was not rewritten",
    )
    require(
        R"<definedName name=\"ScopedOther\" localSheetId=\"2\">'Other Sheet'!$A$1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename rewrite: unrelated scoped definedName changed",
    )
    require(
        R"<definedName name=\"ExternalRef\">[Book.xlsx]Data!A1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename rewrite: external definedName reference changed",
    )
    require(
        R"<definedName name=\"ThreeDRef\">Data:Formula!A1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename rewrite: 3D definedName reference changed",
    )
    require(
        R"<definedName name=\"LiteralText\">\"Data!A1\"</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename rewrite: definedName string literal changed",
    )
    require("xl/calcChain.xml" not in zip_names(path),
            "generated formula rename rewrite: calcChain.xml should not be invented")

    formulas = worksheet_formula_cells(path, "Formula")
    expected_formulas = {
        "A1": "'RenamedData'!A1",
        "A2": "'RenamedData'!$A$1",
        "A3": "[Book.xlsx]Data!A1",
        "A4": "Data:Formula!A1",
        "A5": "'RenamedData'!A1+\"Data!A1\"",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated formula rename rewrite: Formula!{reference} mismatch "
            f"{formulas.get(reference)!r}",
        )
    unmaterialized_formulas = worksheet_formula_cells(path, "Unmaterialized")
    require(
        unmaterialized_formulas.get("A1") == "Data!A1",
        f"generated formula rename rewrite: non-materialized formula was rewritten "
        f"{unmaterialized_formulas!r}",
    )
    require(
        not worksheet_formula_cached_values(path, "Formula"),
        "generated formula rename rewrite: cached formula values remained on dirty Formula sheet",
    )

    defined_records = workbook_defined_name_records(path)
    zip_report["formula_cells"] = sorted_string_mapping(formulas)
    zip_report["unmaterialized_formula_cells"] = sorted_string_mapping(unmaterialized_formulas)
    zip_report["defined_names"] = summarize_defined_name_records(defined_records)
    zip_report["defined_name_records"] = defined_records
    zip_report["formula_rewrite"] = {
        "materialized_formula_cells_rewritten": 3,
        "external_references_preserved": True,
        "three_d_references_preserved": True,
        "string_literals_preserved": True,
        "non_materialized_formulas_rewritten": False,
        "calc_chain_invented": False,
    }
    zip_report["tool_formula_audit"] = {
        "count": tool_report["source_formula_audit_count"],
        "rename_risk_count": tool_report["source_formula_rename_risk_count"],
        "external_count": tool_report["source_formula_external_count"],
        "sheet_range_count": tool_report["source_formula_sheet_range_count"],
        "matched_count": tool_report["source_formula_matched_count"],
        "references": tool_report.get("source_formula_audit_references", []),
    }
    zip_report["tool_defined_name_audit"] = {
        "count": tool_report["defined_name_audit_count"],
        "rename_risk_count": tool_report["defined_name_audit_rename_risk_count"],
        "external_count": tool_report["defined_name_audit_external_count"],
        "sheet_range_count": tool_report["defined_name_audit_sheet_range_count"],
        "matched_count": tool_report["defined_name_audit_matched_count"],
        "references": tool_report.get("defined_name_audit_references", []),
    }
    require(tool_report.get("source_formula_rename_risk_count") == 0,
            f"generated formula rename rewrite: materialized rename risks remained {tool_report!r}")
    require(tool_report.get("defined_name_audit_rename_risk_count") == 0,
            f"generated formula rename rewrite: definedName rename risks remained {tool_report!r}")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(
            workbook.sheetnames == ["RenamedData", "Other Sheet", "Formula", "Unmaterialized"],
            f"generated formula rename rewrite: unexpected sheetnames {workbook.sheetnames!r}",
        )
        formula_sheet = workbook["Formula"]
        unmaterialized = workbook["Unmaterialized"]
        openpyxl_formula_report = openpyxl_formula_cells(formula_sheet, expected_formulas)
        require(
            unmaterialized["A1"].value == "=Data!A1",
            f"generated formula rename rewrite: openpyxl non-materialized formula mismatch "
            f"{unmaterialized['A1'].value!r}",
        )
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "Unmaterialized!A1": unmaterialized["A1"].value,
            "defined_name_count": len(defined_records),
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_formula_rename_escaped_sheet_name(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    workbook_xml = read_zip_text(path, "xl/workbook.xml")

    require(
        'name="Renamed &amp; O&apos;Brien"' in workbook_xml,
        "generated formula rename escaped sheet name: missing renamed sheet catalog entry",
    )
    require(
        'name="Data"' not in workbook_xml,
        "generated formula rename escaped sheet name: old Data sheet catalog entry remained",
    )
    require(
        '<definedName name="ReportRange">\'Renamed &amp; O\'\'Brien\'!$A$1:$B$2</definedName>'
        in workbook_xml,
        "generated formula rename escaped sheet name: ReportRange was not rewritten",
    )
    require(
        '<definedName name="QuotedDataRef">\'Renamed &amp; O\'\'Brien\'!$A$1</definedName>'
        in workbook_xml,
        "generated formula rename escaped sheet name: QuotedDataRef was not rewritten",
    )
    require(
        '<definedName name="ScopedOther" localSheetId="2">\'Other Sheet\'!$A$1</definedName>'
        in workbook_xml,
        "generated formula rename escaped sheet name: unrelated scoped definedName changed",
    )
    require(
        '<definedName name="ExternalRef">[Book.xlsx]Data!A1</definedName>' in workbook_xml,
        "generated formula rename escaped sheet name: external definedName reference changed",
    )
    require(
        '<definedName name="ThreeDRef">Data:Formula!A1</definedName>' in workbook_xml,
        "generated formula rename escaped sheet name: 3D definedName reference changed",
    )
    require(
        "xl/calcChain.xml" not in zip_names(path),
        "generated formula rename escaped sheet name: calcChain.xml should not be invented",
    )

    formulas = worksheet_formula_cells(path, "Formula")
    expected_formulas = {
        "A1": "'Renamed & O''Brien'!A1",
        "A2": "'Renamed & O''Brien'!$A$1",
        "A3": "[Book.xlsx]Data!A1",
        "A4": "Data:Formula!A1",
        "A5": "'Renamed & O''Brien'!A1+\"Data!A1\"",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated formula rename escaped sheet name: Formula!{reference} mismatch "
            f"{formulas.get(reference)!r}",
        )
    unmaterialized_formulas = worksheet_formula_cells(path, "Unmaterialized")
    require(
        unmaterialized_formulas.get("A1") == "Data!A1",
        f"generated formula rename escaped sheet name: non-materialized formula was rewritten "
        f"{unmaterialized_formulas!r}",
    )
    require(
        not worksheet_formula_cached_values(path, "Formula"),
        "generated formula rename escaped sheet name: cached formula values remained on dirty Formula sheet",
    )

    defined_records = workbook_defined_name_records(path)
    zip_report["formula_cells"] = sorted_string_mapping(formulas)
    zip_report["unmaterialized_formula_cells"] = sorted_string_mapping(unmaterialized_formulas)
    zip_report["defined_names"] = summarize_defined_name_records(defined_records)
    zip_report["defined_name_records"] = defined_records
    zip_report["escaped_formula_policy"] = {
        "catalog_renamed": True,
        "materialized_formula_cells_rewritten": 3,
        "defined_names_rewritten": 2,
        "calc_chain_invented": False,
    }
    zip_report["tool_formula_audit"] = {
        "count": tool_report["source_formula_audit_count"],
        "rename_risk_count": tool_report["source_formula_rename_risk_count"],
        "external_count": tool_report["source_formula_external_count"],
        "sheet_range_count": tool_report["source_formula_sheet_range_count"],
        "matched_count": tool_report["source_formula_matched_count"],
        "references": tool_report.get("source_formula_audit_references", []),
    }
    zip_report["tool_defined_name_audit"] = {
        "count": tool_report["defined_name_audit_count"],
        "rename_risk_count": tool_report["defined_name_audit_rename_risk_count"],
        "external_count": tool_report["defined_name_audit_external_count"],
        "sheet_range_count": tool_report["defined_name_audit_sheet_range_count"],
        "matched_count": tool_report["defined_name_audit_matched_count"],
        "references": tool_report.get("defined_name_audit_references", []),
    }
    require(tool_report.get("source_formula_rename_risk_count") == 0,
            f"generated formula rename escaped sheet name: materialized rename risks remained {tool_report!r}")
    require(tool_report.get("defined_name_audit_rename_risk_count") == 0,
            f"generated formula rename escaped sheet name: definedName rename risks remained {tool_report!r}")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(
            workbook.sheetnames == ["Renamed & O'Brien", "Other Sheet", "Formula", "Unmaterialized"],
            f"generated formula rename escaped sheet name: unexpected sheetnames {workbook.sheetnames!r}",
        )
        formula_sheet = workbook["Formula"]
        unmaterialized = workbook["Unmaterialized"]
        openpyxl_formula_report = openpyxl_formula_cells(formula_sheet, expected_formulas)
        require(
            unmaterialized["A1"].value == "=Data!A1",
            f"generated formula rename escaped sheet name: openpyxl non-materialized formula mismatch "
            f"{unmaterialized['A1'].value!r}",
        )
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "Unmaterialized!A1": unmaterialized["A1"].value,
            "defined_name_count": len(defined_records),
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_formula_rename_chain_rewrite(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    workbook_xml = read_zip_text(path, "xl/workbook.xml")

    require(
        'name="FinalData"' in workbook_xml,
        "generated formula rename chain: missing final sheet catalog entry",
    )
    require(
        'name="TemporaryData"' not in workbook_xml,
        "generated formula rename chain: temporary sheet catalog entry remained",
    )
    require(
        '<definedName name="ReportRange">\'FinalData\'!$A$1:$B$2</definedName>' in workbook_xml,
        "generated formula rename chain: ReportRange was not rewritten",
    )
    require(
        '<definedName name="TemporaryAlias">\'FinalData\'!$B$1</definedName>' in workbook_xml,
        "generated formula rename chain: TemporaryAlias was not rewritten",
    )
    require(
        '<definedName name="ExternalRef">[Book.xlsx]Data!A1</definedName>' in workbook_xml,
        "generated formula rename chain: external definedName reference changed",
    )
    require(
        '<definedName name="ThreeDRef">Data:Formula!A1</definedName>' in workbook_xml,
        "generated formula rename chain: 3D definedName reference changed",
    )
    require(
        "xl/calcChain.xml" not in zip_names(path),
        "generated formula rename chain: calcChain.xml should not be invented",
    )

    formulas = worksheet_formula_cells(path, "Formula")
    expected_formulas = {
        "A1": "'FinalData'!A1",
        "A2": "'FinalData'!B1",
        "A3": "[Book.xlsx]Data!A1",
        "A4": "Data:Formula!A1",
        "A5": "'FinalData'!A1+\"TemporaryData!B1\"",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated formula rename chain: Formula!{reference} mismatch "
            f"{formulas.get(reference)!r}",
        )
    unmaterialized_formulas = worksheet_formula_cells(path, "Unmaterialized")
    require(
        unmaterialized_formulas.get("A1") == "Data!A1+TemporaryData!B1",
        f"generated formula rename chain: non-materialized formula was rewritten "
        f"{unmaterialized_formulas!r}",
    )
    require(
        not worksheet_formula_cached_values(path, "Formula"),
        "generated formula rename chain: cached formula values remained on dirty Formula sheet",
    )

    defined_records = workbook_defined_name_records(path)
    zip_report["formula_cells"] = sorted_string_mapping(formulas)
    zip_report["unmaterialized_formula_cells"] = sorted_string_mapping(unmaterialized_formulas)
    zip_report["defined_names"] = summarize_defined_name_records(defined_records)
    zip_report["defined_name_records"] = defined_records
    zip_report["chain_formula_policy"] = {
        "catalog_renamed": True,
        "intermediate_catalog_renamed": True,
        "materialized_formula_cells_rewritten": 3,
        "defined_names_rewritten": 2,
        "calc_chain_invented": False,
    }
    zip_report["tool_formula_audit"] = {
        "count": tool_report["source_formula_audit_count"],
        "rename_risk_count": tool_report["source_formula_rename_risk_count"],
        "external_count": tool_report["source_formula_external_count"],
        "sheet_range_count": tool_report["source_formula_sheet_range_count"],
        "matched_count": tool_report["source_formula_matched_count"],
        "references": tool_report.get("source_formula_audit_references", []),
    }
    zip_report["tool_defined_name_audit"] = {
        "count": tool_report["defined_name_audit_count"],
        "rename_risk_count": tool_report["defined_name_audit_rename_risk_count"],
        "external_count": tool_report["defined_name_audit_external_count"],
        "sheet_range_count": tool_report["defined_name_audit_sheet_range_count"],
        "matched_count": tool_report["defined_name_audit_matched_count"],
        "references": tool_report.get("defined_name_audit_references", []),
    }
    require(tool_report.get("source_formula_rename_risk_count") == 0,
            f"generated formula rename chain: materialized rename risks remained {tool_report!r}")
    require(tool_report.get("defined_name_audit_rename_risk_count") == 0,
            f"generated formula rename chain: definedName rename risks remained {tool_report!r}")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(
            workbook.sheetnames == ["FinalData", "Formula", "Unmaterialized"],
            f"generated formula rename chain: unexpected sheetnames {workbook.sheetnames!r}",
        )
        formula_sheet = workbook["Formula"]
        unmaterialized = workbook["Unmaterialized"]
        openpyxl_formula_report = openpyxl_formula_cells(formula_sheet, expected_formulas)
        require(
            unmaterialized["A1"].value == "=Data!A1+TemporaryData!B1",
            f"generated formula rename chain: openpyxl non-materialized formula mismatch "
            f"{unmaterialized['A1'].value!r}",
        )
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "Unmaterialized!A1": unmaterialized["A1"].value,
            "defined_name_count": len(defined_records),
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_formula_rename_defined_names_only(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    workbook_xml = read_zip_text(path, "xl/workbook.xml")

    require('name="RenamedData"' in workbook_xml,
            "generated formula rename definedNames-only: missing renamed sheet catalog entry")
    require('name="Data"' not in workbook_xml,
            "generated formula rename definedNames-only: old Data sheet catalog entry remained")
    require(
        R"<definedName name=\"ReportRange\">'RenamedData'!$A$1:$B$2</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename definedNames-only: ReportRange was not rewritten",
    )
    require(
        R"<definedName name=\"QuotedDataRef\">'RenamedData'!$A$1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename definedNames-only: QuotedDataRef was not rewritten",
    )
    require(
        R"<definedName name=\"ScopedOther\" localSheetId=\"2\">'Other Sheet'!$A$1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename definedNames-only: unrelated scoped definedName changed",
    )
    require(
        R"<definedName name=\"ExternalRef\">[Book.xlsx]Data!A1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename definedNames-only: external definedName reference changed",
    )
    require(
        R"<definedName name=\"ThreeDRef\">Data:Formula!A1</definedName>"
        .replace('\\"', '"') in workbook_xml,
        "generated formula rename definedNames-only: 3D definedName reference changed",
    )
    require("xl/calcChain.xml" not in zip_names(path),
            "generated formula rename definedNames-only: calcChain.xml should not be invented")

    formulas = worksheet_formula_cells(path, "Formula")
    expected_formulas = {
        "A1": "Data!A1",
        "A2": "'Data'!$A$1",
        "A3": "[Book.xlsx]Data!A1",
        "A4": "Data:Formula!A1",
        "A5": "Data!A1+\"Data!A1\"",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated formula rename definedNames-only: Formula!{reference} mismatch "
            f"{formulas.get(reference)!r}",
        )
    unmaterialized_formulas = worksheet_formula_cells(path, "Unmaterialized")
    require(
        unmaterialized_formulas.get("A1") == "Data!A1",
        f"generated formula rename definedNames-only: non-materialized formula changed "
        f"{unmaterialized_formulas!r}",
    )

    defined_records = workbook_defined_name_records(path)
    zip_report["formula_cells"] = sorted_string_mapping(formulas)
    zip_report["unmaterialized_formula_cells"] = sorted_string_mapping(unmaterialized_formulas)
    zip_report["defined_names"] = summarize_defined_name_records(defined_records)
    zip_report["defined_name_records"] = defined_records
    zip_report["defined_names_only_policy"] = {
        "catalog_renamed": True,
        "materialized_formula_cells_rewritten": 0,
        "defined_names_rewritten": 2,
        "calc_chain_invented": False,
    }
    zip_report["tool_formula_audit"] = {
        "count": tool_report["source_formula_audit_count"],
        "rename_risk_count": tool_report["source_formula_rename_risk_count"],
        "external_count": tool_report["source_formula_external_count"],
        "sheet_range_count": tool_report["source_formula_sheet_range_count"],
        "matched_count": tool_report["source_formula_matched_count"],
        "references": tool_report.get("source_formula_audit_references", []),
    }
    zip_report["tool_defined_name_audit"] = {
        "count": tool_report["defined_name_audit_count"],
        "rename_risk_count": tool_report["defined_name_audit_rename_risk_count"],
        "external_count": tool_report["defined_name_audit_external_count"],
        "sheet_range_count": tool_report["defined_name_audit_sheet_range_count"],
        "matched_count": tool_report["defined_name_audit_matched_count"],
        "references": tool_report.get("defined_name_audit_references", []),
    }
    require(tool_report.get("source_formula_rename_risk_count") == 3,
            f"generated formula rename definedNames-only: expected materialized rename risks {tool_report!r}")
    require(tool_report.get("defined_name_audit_rename_risk_count") == 0,
            f"generated formula rename definedNames-only: definedName rename risks remained {tool_report!r}")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(
            workbook.sheetnames == ["RenamedData", "Other Sheet", "Formula", "Unmaterialized"],
            f"generated formula rename definedNames-only: unexpected sheetnames {workbook.sheetnames!r}",
        )
        formula_sheet = workbook["Formula"]
        unmaterialized = workbook["Unmaterialized"]
        openpyxl_formula_report = openpyxl_formula_cells(formula_sheet, expected_formulas)
        require(
            unmaterialized["A1"].value == "=Data!A1",
            f"generated formula rename definedNames-only: openpyxl non-materialized formula mismatch "
            f"{unmaterialized['A1'].value!r}",
        )
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "Unmaterialized!A1": unmaterialized["A1"].value,
            "defined_name_count": len(defined_records),
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_formula_rename_default_audit(
    path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    workbook_xml = read_zip_text(path, "xl/workbook.xml")

    require('name="RenamedData"' in workbook_xml,
            "generated formula rename default: missing renamed sheet catalog entry")
    require('name="Data"' not in workbook_xml,
            "generated formula rename default: old Data sheet catalog entry remained")
    expected_defined_names = [
        R"<definedName name=\"ReportRange\">Data!$A$1:$B$2</definedName>",
        R"<definedName name=\"QuotedDataRef\">'Data'!$A$1</definedName>",
        R"<definedName name=\"ScopedOther\" localSheetId=\"2\">'Other Sheet'!$A$1</definedName>",
        R"<definedName name=\"ExternalRef\">[Book.xlsx]Data!A1</definedName>",
        R"<definedName name=\"ThreeDRef\">Data:Formula!A1</definedName>",
        R"<definedName name=\"LiteralText\">\"Data!A1\"</definedName>",
    ]
    for expected in expected_defined_names:
        require(
            expected.replace('\\"', '"') in workbook_xml,
            f"generated formula rename default: definedName changed or missing {expected!r}",
        )
    require("'RenamedData'!$A$1" not in workbook_xml,
            "generated formula rename default: definedNames were rewritten by default")
    require("xl/calcChain.xml" not in zip_names(path),
            "generated formula rename default: calcChain.xml should not be invented")

    formulas = worksheet_formula_cells(path, "Formula")
    expected_formulas = {
        "A1": "Data!A1",
        "A2": "'Data'!$A$1",
        "A3": "[Book.xlsx]Data!A1",
        "A4": "Data:Formula!A1",
        "A5": "Data!A1+\"Data!A1\"",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated formula rename default: Formula!{reference} mismatch "
            f"{formulas.get(reference)!r}",
        )
    unmaterialized_formulas = worksheet_formula_cells(path, "Unmaterialized")
    require(
        unmaterialized_formulas.get("A1") == "Data!A1",
        f"generated formula rename default: non-materialized formula changed "
        f"{unmaterialized_formulas!r}",
    )

    defined_records = workbook_defined_name_records(path)
    zip_report["formula_cells"] = sorted_string_mapping(formulas)
    zip_report["unmaterialized_formula_cells"] = sorted_string_mapping(unmaterialized_formulas)
    zip_report["defined_names"] = summarize_defined_name_records(defined_records)
    zip_report["defined_name_records"] = defined_records
    zip_report["default_formula_policy"] = {
        "catalog_renamed": True,
        "materialized_formula_cells_rewritten": 0,
        "defined_names_rewritten": 0,
        "calc_chain_invented": False,
    }
    zip_report["tool_formula_audit"] = {
        "count": tool_report["source_formula_audit_count"],
        "rename_risk_count": tool_report["source_formula_rename_risk_count"],
        "external_count": tool_report["source_formula_external_count"],
        "sheet_range_count": tool_report["source_formula_sheet_range_count"],
        "matched_count": tool_report["source_formula_matched_count"],
        "references": tool_report.get("source_formula_audit_references", []),
    }
    zip_report["tool_defined_name_audit"] = {
        "count": tool_report["defined_name_audit_count"],
        "rename_risk_count": tool_report["defined_name_audit_rename_risk_count"],
        "external_count": tool_report["defined_name_audit_external_count"],
        "sheet_range_count": tool_report["defined_name_audit_sheet_range_count"],
        "matched_count": tool_report["defined_name_audit_matched_count"],
        "references": tool_report.get("defined_name_audit_references", []),
    }
    require(tool_report.get("source_formula_rename_risk_count") == 3,
            f"generated formula rename default: expected materialized rename risks {tool_report!r}")
    require(tool_report.get("defined_name_audit_rename_risk_count") == 2,
            f"generated formula rename default: expected definedName rename risks {tool_report!r}")

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(
            workbook.sheetnames == ["RenamedData", "Other Sheet", "Formula", "Unmaterialized"],
            f"generated formula rename default: unexpected sheetnames {workbook.sheetnames!r}",
        )
        formula_sheet = workbook["Formula"]
        unmaterialized = workbook["Unmaterialized"]
        openpyxl_formula_report = openpyxl_formula_cells(formula_sheet, expected_formulas)
        require(
            unmaterialized["A1"].value == "=Data!A1",
            f"generated formula rename default: openpyxl non-materialized formula mismatch "
            f"{unmaterialized['A1'].value!r}",
        )
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "Unmaterialized!A1": unmaterialized["A1"].value,
            "defined_name_count": len(defined_records),
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_shared_formula_materialization(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated shared formula: missing workbook.xml")
    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('t="shared"' not in worksheet_xml,
            "generated shared formula: stale shared formula metadata remained in output")
    expected_fragments = [
        R'<c r="C1"><f>A1+B1</f></c>',
        R'<c r="C2"><f>A2+B2</f></c>',
        R'<c r="C3"><f>A3+B3</f></c>',
        R'<c r="D2"><f>SUM(A2:B2)+$A2+A$1+$A$1</f></c>',
        R'<c r="D3"><f>SUM(A3:B3)+$A3+A$1+$A$1</f></c>',
        R'<c r="E4" t="inlineStr"><is><t>shared-formula-qa-edit</t></is></c>',
    ]
    for fragment in expected_fragments:
        require(fragment in worksheet_xml,
                f"generated shared formula: missing worksheet XML fragment {fragment!r}")
    for stale_value in ["9", "14", "15", "23"]:
        require(f"<v>{stale_value}</v>" not in worksheet_xml,
                f"generated shared formula: stale cached formula value {stale_value} remained")
    expected_formula_cells = {
        "C1": "A1+B1",
        "C2": "A2+B2",
        "C3": "A3+B3",
        "D1": "SUM(A1:B1)+$A1+A$1+$A$1",
        "D2": "SUM(A2:B2)+$A2+A$1+$A$1",
        "D3": "SUM(A3:B3)+$A3+A$1+$A$1",
    }
    attach_shared_formula_report(
        zip_report,
        path,
        "SharedFormula",
        expected_formula_cells,
        stale_cached_values_removed=True,
        excel_ui_smoke="excel_com_supported",
    )
    zip_report["formulas"] = sorted(expected_formula_cells)
    zip_report["stale_cached_values"] = "absent"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["SharedFormula", "Untouched"],
                f"generated shared formula: unexpected sheetnames {workbook.sheetnames!r}")
        worksheet = workbook["SharedFormula"]
        untouched = workbook["Untouched"]
        require(worksheet["A1"].value == 1, "generated shared formula: A1 mismatch")
        require(worksheet["B3"].value == 8, "generated shared formula: B3 mismatch")
        openpyxl_formula_report = openpyxl_formula_cells(worksheet, expected_formula_cells)
        require(worksheet["E4"].value == "shared-formula-qa-edit",
                "generated shared formula: E4 edit mismatch")
        require(untouched["A1"].value == "keep-shared-formula-qa",
                "generated shared formula: untouched sheet mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "SharedFormula!E4": worksheet["E4"].value,
            "Untouched!A1": untouched["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_shared_formula_boundary_materialization(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated shared formula boundaries: missing workbook.xml")
    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('t="shared"' not in worksheet_xml,
            "generated shared formula boundaries: stale shared metadata remained in output")

    formulas = worksheet_formula_cells(path, "SharedBoundaries")
    expected_formulas = {
        "C1": (
            "A1+SharedBoundaries!A1+'Other Sheet'!A1+SUM(A1:B1)+LOG10(A1)"
            '+A1foo+_A1+A1_+R1C1+Table1[A1]+SUM(A:A)+SUM(1:1)&"A1"'
            "+[Book.xlsx]Sheet1!A1"
        ),
        "C2": (
            "A2+SharedBoundaries!A2+'Other Sheet'!A2+SUM(A2:B2)+LOG10(A2)"
            '+A1foo+_A1+A1_+R1C1+Table1[A1]+SUM(A:A)+SUM(2:2)&"A1"'
            "+[Book.xlsx]Sheet1!A2"
        ),
        "D1": "C1+D$1+$C1+$C$1",
        "E2": "D2+E$1+$C2+$C$1",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated shared formula boundaries: {reference} formula mismatch "
            f"{formulas.get(reference)!r}",
        )
    require("F4" not in formulas, "generated shared formula boundaries: F4 should be text, not formula")

    attach_shared_formula_report(
        zip_report,
        path,
        "SharedBoundaries",
        expected_formulas,
        stale_cached_values_removed=True,
        excel_ui_smoke="not_run_synthetic_parser_boundary",
    )
    zip_report["formulas"] = expected_formulas

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["SharedBoundaries", "Other Sheet", "Untouched"],
                f"generated shared formula boundaries: unexpected sheetnames {workbook.sheetnames!r}")
        worksheet = workbook["SharedBoundaries"]
        untouched = workbook["Untouched"]
        openpyxl_formula_report = openpyxl_formula_cells(worksheet, expected_formulas)
        require(worksheet["F4"].value == "shared-formula-boundary-edit",
                "generated shared formula boundaries: F4 edit mismatch")
        require(untouched["A1"].value == "keep-shared-formula-boundary-qa",
                "generated shared formula boundaries: untouched sheet mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "SharedBoundaries!F4": worksheet["F4"].value,
            "Untouched!A1": untouched["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_shared_formula_office_like_materialization(
    path: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/workbook.xml" in names, "generated office-like shared formula: missing workbook.xml")
    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('t="shared"' not in worksheet_xml,
            "generated office-like shared formula: stale shared metadata remained in output")

    formulas = worksheet_formula_cells(path, "OfficeLikeShared")
    expected_formulas = {
        "C1": "A1+B1",
        "D1": "B1+C1",
        "C2": "A2+B2",
        "D2": "B2+C2",
        "C3": "A3+B3",
        "D3": "B3+C3",
        "E1": "A1*2",
        "F2": "SUM($A2:B2)+C$1",
        "G2": "SUM($A2:C2)+D$1",
        "F3": "SUM($A3:B3)+C$1",
        "G3": "SUM($A3:C3)+D$1",
    }
    for reference, expected in expected_formulas.items():
        require(
            formulas.get(reference) == expected,
            f"generated office-like shared formula: {reference} formula mismatch "
            f"{formulas.get(reference)!r}",
        )
    require("H6" not in formulas, "generated office-like shared formula: H6 should be text")

    for stale_value in range(9901, 9912):
        require(f"<v>{stale_value}</v>" not in worksheet_xml,
                f"generated office-like shared formula: stale cached value {stale_value} remained")
    require(
        R'<c r="H6" t="inlineStr"><is><t>office-like-shared-formula-edit</t></is></c>'
        in worksheet_xml,
        "generated office-like shared formula: missing H6 edit",
    )

    attach_shared_formula_report(
        zip_report,
        path,
        "OfficeLikeShared",
        expected_formulas,
        stale_cached_values_removed=True,
        excel_ui_smoke="excel_com_supported",
    )
    zip_report["formulas"] = expected_formulas
    zip_report["stale_cached_values"] = "absent"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["OfficeLikeShared", "Untouched"],
                f"generated office-like shared formula: unexpected sheetnames {workbook.sheetnames!r}")
        worksheet = workbook["OfficeLikeShared"]
        untouched = workbook["Untouched"]
        openpyxl_formula_report = openpyxl_formula_cells(worksheet, expected_formulas)
        require(worksheet["H6"].value == "office-like-shared-formula-edit",
                "generated office-like shared formula: H6 edit mismatch")
        require(untouched["A1"].value == "keep-office-like-shared-formula-qa",
                "generated office-like shared formula: untouched sheet mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "formula_cells": openpyxl_formula_report,
            "OfficeLikeShared!H6": worksheet["H6"].value,
            "Untouched!A1": untouched["A1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_style_passthrough(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/styles.xml" in names, "generated style passthrough: missing styles.xml")
    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<c r="A1" s="1"><v>9.5</v></c>' in worksheet_xml,
            "generated style passthrough: missing styled numeric cell")
    require('<c r="B1" t="inlineStr"><is><t>explicit default</t></is></c>' in worksheet_xml,
            "generated style passthrough: missing explicit default text cell")
    require('s="0"' not in worksheet_xml, "generated style passthrough: explicit default serialized as s=0")
    zip_report["style_cell"] = "A1"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        worksheet = workbook["Data"]
        require(worksheet["A1"].value == 9.5, "generated style passthrough: A1 mismatch")
        require(worksheet["B1"].value == "explicit default", "generated style passthrough: B1 mismatch")
        require(worksheet["A1"].number_format == "0.00",
                f"generated style passthrough: unexpected A1 number format {worksheet['A1'].number_format!r}")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": worksheet["A1"].value,
            "Data!A1.number_format": worksheet["A1"].number_format,
            "Data!B1": worksheet["B1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_image_replace(path: Path, replacement_image: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    require("xl/media/image1.png" in names, "generated image replace: missing image1.png")
    require("xl/drawings/drawing1.xml" in names, "generated image replace: missing drawing1.xml")
    output_bytes = read_zip_bytes(path, "xl/media/image1.png")
    expected_bytes = replacement_image.read_bytes()
    require(output_bytes == expected_bytes, "generated image replace: media bytes mismatch")
    zip_report["image_part"] = "xl/media/image1.png"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["Data", "Pictures"],
                f"generated image replace: unexpected sheetnames {workbook.sheetnames!r}")
        pictures = workbook["Pictures"]
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Pictures!A1": pictures["A1"].value,
            "Pictures.image_count": openpyxl_image_count(pictures),
        }
        require(pictures["A1"].value == "image-sheet", "generated image replace: Pictures!A1 mismatch")
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_fixture_image_replace(
    source_path: Path,
    output_path: Path,
    sheet_name: str,
    image_part_name: str,
    replacement_image: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    source_names = zip_names(source_path)
    output_names = zip_names(output_path)
    require(image_part_name in source_names,
            f"fixture image replace: missing source image part {image_part_name}")
    require(image_part_name in output_names,
            f"fixture image replace: missing output image part {image_part_name}")
    require(source_names == output_names,
            "fixture image replace: output ZIP entry set changed unexpectedly")

    expected_bytes = replacement_image.read_bytes()
    actual_bytes = read_zip_bytes(output_path, image_part_name)
    require(actual_bytes == expected_bytes,
            f"fixture image replace: {image_part_name} bytes mismatch")
    for entry_name in source_names:
        if entry_name == image_part_name:
            continue
        require(
            read_zip_bytes(source_path, entry_name) == read_zip_bytes(output_path, entry_name),
            f"fixture image replace: preserved entry changed unexpectedly: {entry_name}",
        )

    zip_report["image_part"] = image_part_name
    zip_report["preserved_entries"] = len(source_names) - 1

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
    try:
        require(sheet_name in workbook.sheetnames,
                f"fixture image replace: sheet missing from {workbook.sheetnames!r}")
        worksheet = workbook[sheet_name]
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            f"{sheet_name}.image_count": openpyxl_image_count(worksheet),
            "workbook.image_count": openpyxl_total_image_count(workbook),
        }
        require(openpyxl_total_image_count(workbook) >= 1,
                "fixture image replace: openpyxl did not load any images")
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_public_e2e(path: Path, replacement_image: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    names = zip_names(path)
    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    sheet1_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    sheet2_xml = read_zip_text(path, "xl/worksheets/sheet2.xml")
    sheet3_xml = read_zip_text(path, "xl/worksheets/sheet3.xml")
    require('name="EditedData"' in workbook_xml, "generated public e2e: missing renamed EditedData")
    require('name="Data"' not in workbook_xml, "generated public e2e: old Data name still present")
    sheet_entries = workbook_sheet_entry_map(path)
    require(sheet_entries == {
        "EditedData": "xl/worksheets/sheet1.xml",
        "ReplaceMe": "xl/worksheets/sheet2.xml",
        "Pictures": "xl/worksheets/sheet3.xml",
    }, f"generated public e2e: unexpected workbook sheet map {sheet_entries!r}")
    require("materialized-edit" in sheet1_xml, "generated public e2e: missing materialized edit")
    require('<c r="B2"><v>42</v></c>' in sheet1_xml, "generated public e2e: missing B2 edit")
    require("sheetdata-final" in sheet2_xml, "generated public e2e: missing replaced sheet text")
    require('<c r="B1"><v>7</v></c>' in sheet2_xml, "generated public e2e: missing replaced sheet number")
    require("xl/worksheets/_rels/sheet3.xml.rels" in names,
            "generated public e2e: missing Pictures worksheet relationships")
    require("xl/drawings/drawing1.xml" in names,
            "generated public e2e: missing drawing1.xml")
    require("xl/drawings/_rels/drawing1.xml.rels" in names,
            "generated public e2e: missing drawing1.xml.rels")
    require('<drawing r:id="rId1"/>' in sheet3_xml,
            "generated public e2e: Pictures worksheet missing drawing relationship")
    sheet_rels = relationship_entries(path, "xl/worksheets/_rels/sheet3.xml.rels")
    require(
        any(
            relationship["id"] == "rId1"
            and relationship["type"].endswith("/drawing")
            and relationship["target"] == "../drawings/drawing1.xml"
            for relationship in sheet_rels
        ),
        f"generated public e2e: Pictures worksheet drawing rel missing {sheet_rels!r}",
    )
    drawing_rels = relationship_entries(path, "xl/drawings/_rels/drawing1.xml.rels")
    require(
        any(
            relationship["type"].endswith("/image")
            and relationship["target"] == "../media/image1.png"
            for relationship in drawing_rels
        ),
        f"generated public e2e: drawing image rel missing {drawing_rels!r}",
    )
    expected_image_bytes = replacement_image.read_bytes()
    actual_image_bytes = read_zip_bytes(path, "xl/media/image1.png")
    require(actual_image_bytes == expected_image_bytes,
            "generated public e2e: image bytes mismatch")
    zip_report["sheets"] = list(sheet_entries)
    zip_report["sheet_entries"] = sheet_entries
    zip_report["image_part"] = "xl/media/image1.png"
    zip_report["image_bytes"] = len(actual_image_bytes)
    zip_report["worksheet_drawing_rels"] = len(sheet_rels)
    zip_report["drawing_rels"] = len(drawing_rels)

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == ["EditedData", "ReplaceMe", "Pictures"],
                f"generated public e2e: unexpected sheetnames {workbook.sheetnames!r}")
        edited = workbook["EditedData"]
        replaced = workbook["ReplaceMe"]
        pictures = workbook["Pictures"]
        require(edited["A1"].value == "materialized-edit", "generated public e2e: A1 mismatch")
        require(edited["B2"].value == 42, "generated public e2e: B2 mismatch")
        require(replaced["A1"].value == "sheetdata-final", "generated public e2e: ReplaceMe!A1 mismatch")
        require(replaced["B1"].value == 7, "generated public e2e: ReplaceMe!B1 mismatch")
        require(pictures["A1"].value == "image-sheet", "generated public e2e: Pictures!A1 mismatch")
        require(openpyxl_image_count(pictures) >= 1,
                "generated public e2e: openpyxl did not load the replaced image")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "EditedData!A1": edited["A1"].value,
            "EditedData!B2": edited["B2"].value,
            "ReplaceMe!A1": replaced["A1"].value,
            "ReplaceMe!B1": replaced["B1"].value,
            "Pictures!A1": pictures["A1"].value,
            "Pictures.image_count": openpyxl_image_count(pictures),
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_generated_non_default_style_rejection(
    source_path: Path,
    output_path: Path,
    tool_report: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    require(tool_report.get("status") == "expected_rejection_observed",
            f"generated style rejection: unexpected tool status {tool_report.get('status')!r}")
    error_message = tool_report.get("error_message", "")
    require("does not support non-default StyleId" in error_message,
            f"generated style rejection: unexpected error message {error_message!r}")

    source_names = zip_names(source_path)
    output_names = zip_names(output_path)
    require(source_names == output_names, "generated style rejection: ZIP entries changed after rejection")
    for name in source_names:
        require(read_zip_bytes(source_path, name) == read_zip_bytes(output_path, name),
                f"generated style rejection: entry changed after rejection: {name}")
    zip_report["entries_compared"] = len(source_names)

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
    try:
        worksheet = workbook["Data"]
        require(worksheet["A1"].value == "placeholder-a1", "generated style rejection: A1 mismatch")
        require(worksheet["B1"].value == 1, "generated style rejection: B1 mismatch")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            "Data!A1": worksheet["A1"].value,
            "Data!B1": worksheet["B1"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def workbook_sheetnames(path: Path) -> list[str]:
    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def verify_fixture_rename_materialized(
    source_path: Path,
    output_path: Path,
    original_sheet_name: str,
    renamed_sheet_name: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    source_names = zip_names(source_path)
    output_names = zip_names(output_path)
    require("xl/workbook.xml" in output_names, "fixture rename/materialized: missing workbook.xml")
    if "xl/sharedStrings.xml" in source_names:
        require("xl/sharedStrings.xml" in output_names,
                "fixture rename/materialized: sharedStrings part should be preserved when source has one")
        require(read_zip_bytes(source_path, "xl/sharedStrings.xml")
                == read_zip_bytes(output_path, "xl/sharedStrings.xml"),
                "fixture rename/materialized: sharedStrings bytes changed unexpectedly")
        zip_report["shared_strings"] = "preserved"
    else:
        require("xl/sharedStrings.xml" not in output_names,
                "fixture rename/materialized: output should not invent sharedStrings.xml")
        zip_report["shared_strings"] = "absent"

    if "xl/styles.xml" in source_names:
        require("xl/styles.xml" in output_names, "fixture rename/materialized: styles.xml should remain present")
        zip_report["styles"] = "present"
    else:
        zip_report["styles"] = "absent"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
    try:
        require(renamed_sheet_name in workbook.sheetnames,
                f"fixture rename/materialized: renamed sheet missing from {workbook.sheetnames!r}")
        require(original_sheet_name not in workbook.sheetnames,
                f"fixture rename/materialized: old sheet still present in {workbook.sheetnames!r}")
        worksheet = workbook[renamed_sheet_name]
        require(worksheet["A1"].value == "fixture-materialized-edit",
                "fixture rename/materialized: A1 mismatch")
        require(worksheet["B2"].value == 42,
                f"fixture rename/materialized: B2 mismatch {worksheet['B2'].value!r}")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            f"{renamed_sheet_name}!A1": worksheet["A1"].value,
            f"{renamed_sheet_name}!B2": worksheet["B2"].value,
        }
        zip_report["sheetnames"] = workbook.sheetnames
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_fixture_materialized_only(
    source_path: Path,
    output_path: Path,
    sheet_name: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    zip_report: dict[str, Any] = {}
    source_names = zip_names(source_path)
    output_names = zip_names(output_path)
    require("xl/workbook.xml" in output_names or "workbook.xml" in output_names,
            "fixture materialized-only: missing workbook part")

    if "xl/sharedStrings.xml" in source_names:
        require("xl/sharedStrings.xml" in output_names,
                "fixture materialized-only: sharedStrings part should be preserved when source has one")
        require(read_zip_bytes(source_path, "xl/sharedStrings.xml")
                == read_zip_bytes(output_path, "xl/sharedStrings.xml"),
                "fixture materialized-only: sharedStrings bytes changed unexpectedly")
        zip_report["shared_strings"] = "preserved"
    else:
        require("xl/sharedStrings.xml" not in output_names,
                "fixture materialized-only: output should not invent sharedStrings.xml")
        zip_report["shared_strings"] = "absent"

    if "xl/styles.xml" in source_names:
        require("xl/styles.xml" in output_names,
                "fixture materialized-only: styles.xml should remain present when source has one")
        zip_report["styles"] = "present"
    else:
        zip_report["styles"] = "absent"

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
    try:
        require(sheet_name in workbook.sheetnames,
                f"fixture materialized-only: sheet {sheet_name!r} missing from {workbook.sheetnames!r}")
        worksheet = workbook[sheet_name]
        require(worksheet["A1"].value == "fixture-materialized-edit",
                f"fixture materialized-only: A1 mismatch {worksheet['A1'].value!r}")
        require(worksheet["B2"].value == 42,
                f"fixture materialized-only: B2 mismatch {worksheet['B2'].value!r}")
        openpyxl_report = {
            "sheetnames": workbook.sheetnames,
            f"{sheet_name}!A1": worksheet["A1"].value,
            f"{sheet_name}!B2": worksheet["B2"].value,
        }
    finally:
        workbook.close()

    return zip_report, openpyxl_report


def verify_defined_names_preserved(source_path: Path, output_path: Path) -> dict[str, Any]:
    source_records = workbook_defined_name_records(source_path)
    output_records = workbook_defined_name_records(output_path)
    require(source_records, "definedName fixture: source workbook has no definedName records")
    require(
        canonical_defined_name_records(source_records)
        == canonical_defined_name_records(output_records),
        "definedName fixture: definedName records changed after materialized edit",
    )
    return {
        "source": summarize_defined_name_records(source_records),
        "output": summarize_defined_name_records(output_records),
        "records_preserved": len(output_records),
    }


def create_xlsxwriter_reference(
    scenario: str,
    reference_path: Path,
    replacement_image: Path | None,
) -> dict[str, Any]:
    xlsxwriter = load_xlsxwriter()
    if xlsxwriter is None:
        return {"status": "skipped", "reason": "xlsxwriter not installed"}
    if scenario in {
        "generated_source_formula_audit",
        "generated_formula_rename_rewrite",
        "generated_formula_rename_escaped_sheet_name",
        "generated_formula_rename_chain_rewrite",
        "generated_formula_rename_defined_names_only",
        "generated_formula_rename_default_audit",
    }:
        return {"status": "skipped", "reason": f"no xlsxwriter reference for {scenario}"}

    reference_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(reference_path)
    try:
        if scenario == "generated_rename_materialized":
            edited = workbook.add_worksheet("EditedData")
            untouched = workbook.add_worksheet("Untouched")
            edited.write("A1", "materialized-edit")
            edited.write_number("B2", 42)
            untouched.write("A1", "keep-me")
        elif scenario == "generated_in_memory_insert_formula":
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write_row("A1", ["item", "value", "double"])
            data.write("A2", "inserted-row")
            data.write_number("B2", 5)
            data.write_formula("C2", "=B2*2")
            data.write("A3", "source-row")
            data.write_number("B3", 3)
            data.write_formula("C3", "=B3*2")
            notes.write("A1", "preserved")
        elif scenario == "generated_in_memory_delete_column_formula":
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write_number("A1", 7)
            data.write_formula("B1", "=A1+C1")
            data.write("C1", "tail")
            notes.write("A1", "preserved")
        elif scenario == "generated_in_memory_insert_column_formula":
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write("A1", "item")
            data.write("B1", "inserted-col")
            data.write_number("C1", 2)
            data.write_formula("D1", "=C1*2")
            notes.write("A1", "preserved")
        elif scenario == "generated_in_memory_delete_row_formula":
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write_number("A1", 4)
            data.write_formula("B1", "=A1+A2")
            data.write_number("A2", 6)
            notes.write("A1", "preserved")
        elif scenario == "generated_in_memory_stationary_formula_shift":
            expected_formula = (
                "SUM($A$4,C$4,Data!$A$4,$E$1,$E1,Data!$E$1,#REF!,#REF!,"
                "Data!#REF!,#REF!,Data!#REF!)"
            )
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write("A1", "item")
            data.write("B1", "label")
            data.write_formula("C1", f"={expected_formula}")
            data.write_number("E1", 4)
            data.write("F1", "survive-column")
            data.write("A2", "keep-row-two")
            data.write("A4", "row-target")
            data.write("C4", "c3-target")
            data.write("A5", "row-four")
            data.write("A6", "row-five")
            notes.write("A1", "preserved")
        elif scenario == "generated_in_memory_clear_erase":
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write("A1", "keep-a1")
            data.write_blank("B1", None)
            data.write("D1", "new-d1")
            data.write_number("A2", 8)
            notes.write("A1", "preserved")
        elif scenario == "generated_in_memory_append_row_formula":
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write_row("A1", ["item", "value", "double"])
            data.write("A2", "source-row")
            data.write_number("B2", 10)
            data.write("A3", "appended-row")
            data.write_number("B3", 4)
            data.write_formula("C3", "=B3*2")
            notes.write("A1", "preserved")
        elif scenario in {
            "generated_in_memory_overwrite_formula_text",
            "generated_in_memory_retry_noop_save",
            "generated_in_memory_retry_path_equivalent_noop_save",
        }:
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write("A1", "new-text")
            data.write_number("B1", 5)
            data.write_formula("C1", "=B1+10")
            data.write("A2", "keep-row-two")
            notes.write("A1", "preserved")
        elif scenario in {
            "generated_in_memory_retry_reopen_modify_noop_save",
            "generated_in_memory_retry_path_equivalent_reopen_modify_noop_save",
            "generated_in_memory_retry_path_equivalent_reopen_modify_post_noop_third_save",
            "generated_in_memory_retry_reopen_modify_post_noop_third_save",
            "generated_in_memory_reopen_modify_save",
            "generated_in_memory_reopen_modify_noop_save",
            "generated_in_memory_reopen_modify_post_noop_third_save",
        }:
            data = workbook.add_worksheet("Data")
            notes = workbook.add_worksheet("Notes")
            data.write("A1", "first-edit")
            data.write_number("B1", 10)
            data.write_formula("C1", "=B1+5")
            data.write("D1", "second-edit")
            data.write("A2", "keep-row-two")
            data.write("A3", "reopened-row")
            data.write_number("B3", 4)
            data.write_formula("C3", "=B3*2")
            if scenario in {
                "generated_in_memory_retry_path_equivalent_reopen_modify_post_noop_third_save",
                "generated_in_memory_retry_reopen_modify_post_noop_third_save",
                "generated_in_memory_reopen_modify_post_noop_third_save",
            }:
                data.write("E1", "third-edit")
            notes.write("A1", "preserved")
        elif scenario in {
            "generated_in_memory_multi_sheet_save",
            "generated_in_memory_multi_sheet_noop_save",
            "generated_in_memory_multi_sheet_retry_save",
            "generated_in_memory_multi_sheet_retry_noop_save",
            "generated_in_memory_multi_sheet_retry_path_equivalent_noop_save",
            "generated_in_memory_multi_sheet_retry_reopen_modify_save",
            "generated_in_memory_multi_sheet_retry_reopen_modify_noop_save",
            "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_noop_save",
            "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save",
            "generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save",
        }:
            data = workbook.add_worksheet("Data")
            summary = workbook.add_worksheet("Summary")
            notes = workbook.add_worksheet("Notes")
            data.write("A1", "edited-data")
            data.write_number("B1", 7)
            data.write("A2", "keep-data-row")
            data.write("A3", "multi-row")
            data.write_number("B3", 3)
            data.write_formula("C3", "=B3+Data!B1")
            summary.write("A1", "edited-summary")
            summary.write_formula("B1", "=Data!B1+Data!B3")
            if scenario in {
                "generated_in_memory_multi_sheet_retry_reopen_modify_save",
                "generated_in_memory_multi_sheet_retry_reopen_modify_noop_save",
                "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_noop_save",
                "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save",
                "generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save",
            }:
                data.write("D1", "retry-reopened-data")
                summary.write_formula("C1", "=Data!B1+10")
            if scenario in {
                "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save",
                "generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save",
            }:
                data.write("E1", "retry-reopened-post-noop-data")
                summary.write_formula("D1", "=Data!B1+20")
            notes.write("A1", "preserved")
        elif scenario == "generated_shared_formula_materialization":
            shared = workbook.add_worksheet("SharedFormula")
            untouched = workbook.add_worksheet("Untouched")
            for row_index, (left, right) in enumerate([(1, 2), (4, 5), (7, 8)]):
                excel_row = row_index + 1
                shared.write_number(row_index, 0, left)
                shared.write_number(row_index, 1, right)
                shared.write_formula(row_index, 2, f"=A{excel_row}+B{excel_row}")
                shared.write_formula(
                    row_index, 3,
                    f"=SUM(A{excel_row}:B{excel_row})+$A{excel_row}+A$1+$A$1")
            shared.write("E4", "shared-formula-qa-edit")
            untouched.write("A1", "keep-shared-formula-qa")
        elif scenario == "generated_shared_formula_office_like_materialization":
            shared = workbook.add_worksheet("OfficeLikeShared")
            untouched = workbook.add_worksheet("Untouched")
            values = [(1, 2), (10, 20), (100, 200)]
            for row_index, (left, right) in enumerate(values):
                excel_row = row_index + 1
                shared.write_number(row_index, 0, left)
                shared.write_number(row_index, 1, right)
                shared.write_formula(row_index, 2, f"=A{excel_row}+B{excel_row}")
                shared.write_formula(row_index, 3, f"=B{excel_row}+C{excel_row}")
            shared.write_formula(0, 4, "=A1*2")
            shared.write("E2", "between-shared-groups")
            shared.write_formula(1, 5, "=SUM($A2:B2)+C$1")
            shared.write_formula(1, 6, "=SUM($A2:C2)+D$1")
            shared.write_formula(2, 5, "=SUM($A3:B3)+C$1")
            shared.write_formula(2, 6, "=SUM($A3:C3)+D$1")
            shared.write("H6", "office-like-shared-formula-edit")
            untouched.write("A1", "keep-office-like-shared-formula-qa")
        elif scenario == "generated_style_passthrough":
            data = workbook.add_worksheet("Data")
            style = workbook.add_format({"num_format": "0.00"})
            data.write_number("A1", 9.5, style)
            data.write("B1", "explicit default")
        elif scenario == "generated_image_replace":
            require(replacement_image is not None, "xlsxwriter reference: replacement image is required")
            data = workbook.add_worksheet("Data")
            pictures = workbook.add_worksheet("Pictures")
            data.write("A1", "placeholder-a1")
            pictures.write("A1", "image-sheet")
            pictures.insert_image("A2", str(replacement_image))
        elif scenario == "generated_public_e2e":
            require(replacement_image is not None, "xlsxwriter reference: replacement image is required")
            edited = workbook.add_worksheet("EditedData")
            replace_me = workbook.add_worksheet("ReplaceMe")
            pictures = workbook.add_worksheet("Pictures")
            edited.write("A1", "materialized-edit")
            edited.write_number("B2", 42)
            replace_me.write("A1", "sheetdata-final")
            replace_me.write_number("B1", 7)
            pictures.insert_image("A1", str(replacement_image))
        else:
            return {"status": "skipped", "reason": f"no xlsxwriter reference for {scenario}"}
    finally:
        workbook.close()

    if not reference_path.exists():
        return {"status": "skipped", "reason": f"no xlsxwriter reference for {scenario}"}

    openpyxl = load_openpyxl()
    reference = openpyxl.load_workbook(reference_path, read_only=False, data_only=False)
    try:
        info = {"status": "created", "sheetnames": reference.sheetnames}
        if scenario == "generated_style_passthrough":
            info["Data!A1.number_format"] = reference["Data"]["A1"].number_format
        return info
    finally:
        reference.close()


def run_generated_case(
    qa_exe: Path,
    work_dir: Path,
    scenario: str,
) -> ScenarioResult:
    case_dir = work_dir / GENERATED_CASE_DIRECTORY_ALIASES.get(scenario, scenario)
    tool_report = run_tool(qa_exe, scenario, case_dir)
    output_path = Path(tool_report["output"])
    replacement_image = Path(tool_report["replacement_image"]) if tool_report.get("replacement_image") else None

    if scenario == "generated_rename_materialized":
        zip_xml, openpyxl_report = verify_generated_rename_materialized(output_path)
    elif scenario == "generated_in_memory_insert_formula":
        zip_xml, openpyxl_report = verify_generated_in_memory_insert_formula(output_path)
    elif scenario == "generated_in_memory_delete_column_formula":
        zip_xml, openpyxl_report = verify_generated_in_memory_delete_column_formula(output_path)
    elif scenario == "generated_in_memory_insert_column_formula":
        zip_xml, openpyxl_report = verify_generated_in_memory_insert_column_formula(output_path)
    elif scenario == "generated_in_memory_delete_row_formula":
        zip_xml, openpyxl_report = verify_generated_in_memory_delete_row_formula(output_path)
    elif scenario == "generated_in_memory_stationary_formula_shift":
        zip_xml, openpyxl_report = verify_generated_in_memory_stationary_formula_shift(output_path)
    elif scenario == "generated_in_memory_clear_erase":
        zip_xml, openpyxl_report = verify_generated_in_memory_clear_erase(output_path)
    elif scenario == "generated_in_memory_append_row_formula":
        zip_xml, openpyxl_report = verify_generated_in_memory_append_row_formula(output_path)
    elif scenario == "generated_in_memory_overwrite_formula_text":
        zip_xml, openpyxl_report = verify_generated_in_memory_overwrite_formula_text(output_path)
    elif scenario == "generated_in_memory_retry_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_retry_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_retry_path_equivalent_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_retry_path_equivalent_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_retry_reopen_modify_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_retry_reopen_modify_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_retry_path_equivalent_reopen_modify_noop_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_retry_path_equivalent_reopen_modify_noop_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_in_memory_retry_path_equivalent_reopen_modify_post_noop_third_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_retry_path_equivalent_reopen_modify_post_noop_third_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_in_memory_retry_reopen_modify_post_noop_third_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_retry_reopen_modify_post_noop_third_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_in_memory_reopen_modify_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_reopen_modify_save(output_path)
    elif scenario == "generated_in_memory_reopen_modify_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_reopen_modify_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_reopen_modify_post_noop_third_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_reopen_modify_post_noop_third_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_multi_sheet_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_multi_sheet_save(output_path)
    elif scenario == "generated_in_memory_multi_sheet_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_multi_sheet_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_path_equivalent_noop_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_multi_sheet_retry_path_equivalent_noop_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_reopen_modify_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_reopen_modify_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_reopen_modify_noop_save":
        zip_xml, openpyxl_report = verify_generated_in_memory_multi_sheet_retry_reopen_modify_noop_save(
            output_path,
            tool_report,
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_noop_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_noop_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_multi_sheet_retry_path_equivalent_reopen_modify_post_noop_third_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save":
        zip_xml, openpyxl_report = (
            verify_generated_in_memory_multi_sheet_retry_reopen_modify_post_noop_third_save(
                output_path,
                tool_report,
            )
        )
    elif scenario == "generated_source_formula_audit":
        zip_xml, openpyxl_report = verify_generated_source_formula_audit(
            output_path,
            tool_report,
        )
    elif scenario == "generated_formula_rename_rewrite":
        zip_xml, openpyxl_report = verify_generated_formula_rename_rewrite(
            output_path,
            tool_report,
        )
    elif scenario == "generated_formula_rename_escaped_sheet_name":
        zip_xml, openpyxl_report = verify_generated_formula_rename_escaped_sheet_name(
            output_path,
            tool_report,
        )
    elif scenario == "generated_formula_rename_chain_rewrite":
        zip_xml, openpyxl_report = verify_generated_formula_rename_chain_rewrite(
            output_path,
            tool_report,
        )
    elif scenario == "generated_formula_rename_defined_names_only":
        zip_xml, openpyxl_report = verify_generated_formula_rename_defined_names_only(
            output_path,
            tool_report,
        )
    elif scenario == "generated_formula_rename_default_audit":
        zip_xml, openpyxl_report = verify_generated_formula_rename_default_audit(
            output_path,
            tool_report,
        )
    elif scenario == "generated_shared_formula_materialization":
        zip_xml, openpyxl_report = verify_generated_shared_formula_materialization(output_path)
    elif scenario == "generated_shared_formula_boundary_materialization":
        zip_xml, openpyxl_report = verify_generated_shared_formula_boundary_materialization(output_path)
    elif scenario == "generated_shared_formula_office_like_materialization":
        zip_xml, openpyxl_report = verify_generated_shared_formula_office_like_materialization(output_path)
    elif scenario == "generated_style_passthrough":
        zip_xml, openpyxl_report = verify_generated_style_passthrough(output_path)
    elif scenario == "generated_image_replace":
        require(replacement_image is not None, "generated image replace: missing replacement image in tool report")
        zip_xml, openpyxl_report = verify_generated_image_replace(output_path, replacement_image)
    elif scenario == "generated_public_e2e":
        require(replacement_image is not None, "generated public e2e: missing replacement image in tool report")
        zip_xml, openpyxl_report = verify_generated_public_e2e(output_path, replacement_image)
    elif scenario == "generated_non_default_style_rejection":
        source_path = Path(tool_report["source"])
        zip_xml, openpyxl_report = verify_generated_non_default_style_rejection(
            source_path, output_path, tool_report
        )
    else:  # pragma: no cover - guarded by argparse
        raise RuntimeError(f"unsupported generated scenario: {scenario}")

    reference_path = case_dir / "xlsxwriter-reference.xlsx"
    xlsxwriter_reference = create_xlsxwriter_reference(scenario, reference_path, replacement_image)
    return ScenarioResult(
        name=scenario,
        report=tool_report,
        zip_xml=zip_xml,
        openpyxl=openpyxl_report,
        xlsxwriter_reference=xlsxwriter_reference,
    )


def run_fixture_case(
    qa_exe: Path,
    work_dir: Path,
    fixture_path: Path,
    group_name: str,
    *,
    fixture_root: Path | None = None,
) -> ScenarioResult:
    case_slug = (
        fixture_case_slug(fixture_root, fixture_path)
        if fixture_root is not None
        else fixture_path.stem
    )
    case_dir = work_dir / group_name / case_slug
    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(fixture_path, read_only=False, data_only=False)
    try:
        source_sheet_name = workbook.sheetnames[0]
        rename_to = choose_unique_sheet_name(list(workbook.sheetnames))
    finally:
        workbook.close()

    source_names = zip_names(fixture_path)
    tool_scenario = (
        "fixture_rename_materialized"
        if "xl/workbook.xml" in source_names
        else "fixture_materialized_only"
    )
    case_dir.mkdir(parents=True, exist_ok=True)
    tool_source_path = case_dir / "source.xlsx"
    shutil.copyfile(fixture_path, tool_source_path)

    tool_kwargs: dict[str, Any] = {
        "source": tool_source_path,
        "rename_to": rename_to,
    }
    if tool_scenario == "fixture_materialized_only":
        tool_kwargs["sheet_name"] = source_sheet_name

    tool_report = run_tool(
        qa_exe,
        tool_scenario,
        case_dir,
        **tool_kwargs,
    )
    output_path = Path(tool_report["output"])
    actual_source_sheet_name = tool_report.get("source_sheet_name") or source_sheet_name
    actual_renamed_sheet_name = tool_report.get("renamed_sheet_name") or rename_to
    if tool_scenario == "fixture_rename_materialized":
        zip_xml, openpyxl_report = verify_fixture_rename_materialized(
            fixture_path,
            output_path,
            actual_source_sheet_name,
            actual_renamed_sheet_name,
        )
    else:
        zip_xml, openpyxl_report = verify_fixture_materialized_only(
            fixture_path,
            output_path,
            actual_source_sheet_name,
        )
    if fixture_root is not None:
        try:
            case_name = f"{group_name}:{fixture_path.relative_to(fixture_root).as_posix()}"
        except ValueError:
            case_name = f"{group_name}:{fixture_path.name}"
    else:
        case_name = f"{group_name}:{fixture_path.name}"
    return ScenarioResult(
        name=case_name,
        report=tool_report,
        zip_xml=zip_xml,
        openpyxl=openpyxl_report,
        xlsxwriter_reference={"status": "skipped", "reason": "fixture scenario"},
    )


def run_fixture_image_replace_case(
    qa_exe: Path,
    work_dir: Path,
    image_info: ImageFixtureInfo,
    group_name: str,
    *,
    fixture_root: Path,
) -> ScenarioResult:
    fixture_path = image_info.fixture_path
    case_slug = fixture_case_slug(fixture_root, fixture_path) + "__image"
    case_dir = work_dir / group_name / case_slug
    replacement_image = (
        repository_asset("docs/assets/donation/zhifubao.jpg")
        if image_info.image_part_name.lower().endswith((".jpg", ".jpeg"))
        else repository_asset("docs/assets/donation/weixin.png")
    )
    if not replacement_image.exists():
        raise RuntimeError(f"replacement image asset not found: {replacement_image}")

    case_dir.mkdir(parents=True, exist_ok=True)
    tool_source_path = case_dir / "source.xlsx"
    shutil.copyfile(fixture_path, tool_source_path)
    tool_report = run_tool(
        qa_exe,
        "fixture_image_replace",
        case_dir,
        source=tool_source_path,
        sheet_name=image_info.sheet_name,
        image_part_name=image_info.image_part_name,
    )
    output_path = Path(tool_report["output"])
    actual_sheet_name = tool_report.get("source_sheet_name") or image_info.sheet_name
    actual_image_part_name = tool_report.get("image_part_name") or image_info.image_part_name
    zip_xml, openpyxl_report = verify_fixture_image_replace(
        fixture_path,
        output_path,
        actual_sheet_name,
        actual_image_part_name,
        replacement_image,
    )

    try:
        case_name = f"{group_name}:{fixture_path.relative_to(fixture_root).as_posix()}:{actual_image_part_name}"
    except ValueError:
        case_name = f"{group_name}:{fixture_path.name}:{actual_image_part_name}"
    return ScenarioResult(
        name=case_name,
        report=tool_report,
        zip_xml=zip_xml,
        openpyxl=openpyxl_report,
        xlsxwriter_reference={"status": "skipped", "reason": "fixture image scenario"},
    )


def run_formula_fixture_case(
    qa_exe: Path,
    work_dir: Path,
    formula_info: FormulaWorksheetInfo,
    group_name: str,
    *,
    fixture_root: Path,
) -> ScenarioResult:
    fixture_path = formula_info.fixture_path
    case_slug = formula_fixture_case_slug(fixture_root, fixture_path, formula_info.sheet_name)
    case_dir = work_dir / group_name / case_slug
    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(fixture_path, read_only=False, data_only=False)
    try:
        rename_to = choose_unique_sheet_name(list(workbook.sheetnames))
    finally:
        workbook.close()

    source_names = zip_names(fixture_path)
    tool_scenario = (
        "fixture_rename_materialized"
        if "xl/workbook.xml" in source_names
        else "fixture_materialized_only"
    )
    case_dir.mkdir(parents=True, exist_ok=True)
    tool_source_path = case_dir / "source.xlsx"
    shutil.copyfile(fixture_path, tool_source_path)

    tool_report = run_tool(
        qa_exe,
        tool_scenario,
        case_dir,
        source=tool_source_path,
        sheet_name=formula_info.sheet_name,
        rename_to=rename_to,
    )
    output_path = Path(tool_report["output"])
    actual_source_sheet_name = tool_report.get("source_sheet_name") or formula_info.sheet_name
    actual_renamed_sheet_name = tool_report.get("renamed_sheet_name") or rename_to
    if tool_scenario == "fixture_rename_materialized":
        zip_xml, openpyxl_report = verify_fixture_rename_materialized(
            fixture_path,
            output_path,
            actual_source_sheet_name,
            actual_renamed_sheet_name,
        )
    else:
        zip_xml, openpyxl_report = verify_fixture_materialized_only(
            fixture_path,
            output_path,
            actual_source_sheet_name,
        )
    zip_xml["formula_scan"] = formula_info.as_report(fixture_root)
    output_formula_summary = formula_summary_for_sheet(
        output_path,
        actual_renamed_sheet_name if tool_scenario == "fixture_rename_materialized"
        else actual_source_sheet_name,
    )
    if formula_info.shared_formula_count > 0:
        require(
            output_formula_summary["shared_formula_count"] == 0,
            "formula fixture smoke: stale shared formula metadata remained in dirty output",
        )
    zip_xml["formula_output"] = output_formula_summary

    try:
        case_name = (
            f"{group_name}:{fixture_path.relative_to(fixture_root).as_posix()}"
            f":{formula_info.sheet_name}"
        )
    except ValueError:
        case_name = f"{group_name}:{fixture_path.name}:{formula_info.sheet_name}"
    return ScenarioResult(
        name=case_name,
        report=tool_report,
        zip_xml=zip_xml,
        openpyxl=openpyxl_report,
        xlsxwriter_reference={"status": "skipped", "reason": "formula fixture scenario"},
    )


def run_source_formula_fixture_audit_case(
    qa_exe: Path,
    work_dir: Path,
    formula_info: FormulaWorksheetInfo,
    group_name: str,
    *,
    fixture_root: Path,
) -> ScenarioResult:
    fixture_path = formula_info.fixture_path
    case_slug = formula_fixture_case_slug(
        fixture_root, fixture_path, formula_info.sheet_name
    ) + "__source_audit"
    case_dir = work_dir / group_name / case_slug
    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(fixture_path, read_only=False, data_only=False)
    try:
        rename_to = choose_unique_sheet_name(list(workbook.sheetnames))
    finally:
        workbook.close()

    case_dir.mkdir(parents=True, exist_ok=True)
    tool_source_path = case_dir / "source.xlsx"
    shutil.copyfile(fixture_path, tool_source_path)
    tool_report = run_tool(
        qa_exe,
        "fixture_source_formula_audit",
        case_dir,
        source=tool_source_path,
        sheet_name=formula_info.sheet_name,
        rename_to=rename_to,
    )
    require(
        "source_formula_audit_count" in tool_report,
        "source formula fixture audit: tool report missing audit count",
    )
    zip_xml = {
        "formula_scan": formula_info.as_report(fixture_root),
        "source_formula_audit": {
            "count": tool_report["source_formula_audit_count"],
            "rename_risk_count": tool_report["source_formula_rename_risk_count"],
            "external_count": tool_report["source_formula_external_count"],
            "sheet_range_count": tool_report["source_formula_sheet_range_count"],
            "matched_count": tool_report["source_formula_matched_count"],
            "references": tool_report.get("source_formula_audit_references", []),
        },
    }
    openpyxl_report = {
        "sheet_name": formula_info.sheet_name,
        "formula_count": formula_info.formula_count,
        "shared_formula_count": formula_info.shared_formula_count,
        "source_formula_audit_count": tool_report["source_formula_audit_count"],
        "rename_to": rename_to,
    }

    try:
        case_name = (
            f"{group_name}:{fixture_path.relative_to(fixture_root).as_posix()}"
            f":{formula_info.sheet_name}"
        )
    except ValueError:
        case_name = f"{group_name}:{fixture_path.name}:{formula_info.sheet_name}"
    return ScenarioResult(
        name=case_name,
        report=tool_report,
        zip_xml=zip_xml,
        openpyxl=openpyxl_report,
        xlsxwriter_reference={"status": "skipped", "reason": "source formula fixture audit"},
    )


def run_defined_name_fixture_case(
    qa_exe: Path,
    work_dir: Path,
    defined_name_info: DefinedNameFixtureInfo,
    group_name: str,
    *,
    fixture_root: Path,
) -> ScenarioResult:
    fixture_path = defined_name_info.fixture_path
    case_slug = fixture_case_slug(fixture_root, fixture_path) + "__defined"
    case_dir = work_dir / group_name / case_slug
    case_dir.mkdir(parents=True, exist_ok=True)
    tool_source_path = case_dir / "source.xlsx"
    shutil.copyfile(fixture_path, tool_source_path)

    tool_report = run_tool(
        qa_exe,
        "fixture_materialized_only",
        case_dir,
        source=tool_source_path,
    )
    output_path = Path(tool_report["output"])
    actual_source_sheet_name = tool_report.get("source_sheet_name") or defined_name_info.sheet_name
    zip_xml, openpyxl_report = verify_fixture_materialized_only(
        fixture_path,
        output_path,
        actual_source_sheet_name,
    )
    zip_xml["defined_name_scan"] = defined_name_info.as_report(fixture_root)
    zip_xml["defined_name_preservation"] = verify_defined_names_preserved(
        fixture_path,
        output_path,
    )
    zip_xml["defined_name_audit"] = {
        "count": tool_report.get("defined_name_audit_count", 0),
        "rename_risk_count": tool_report.get("defined_name_audit_rename_risk_count", 0),
        "external_count": tool_report.get("defined_name_audit_external_count", 0),
        "sheet_range_count": tool_report.get("defined_name_audit_sheet_range_count", 0),
        "matched_count": tool_report.get("defined_name_audit_matched_count", 0),
        "references": tool_report.get("defined_name_audit_references", []),
    }
    openpyxl_report["defined_name_count"] = zip_xml["defined_name_preservation"]["output"][
        "defined_name_count"
    ]
    openpyxl_report["defined_name_audit_count"] = zip_xml["defined_name_audit"]["count"]

    try:
        case_name = f"{group_name}:{fixture_path.relative_to(fixture_root).as_posix()}"
    except ValueError:
        case_name = f"{group_name}:{fixture_path.name}"
    return ScenarioResult(
        name=case_name,
        report=tool_report,
        zip_xml=zip_xml,
        openpyxl=openpyxl_report,
        xlsxwriter_reference={"status": "skipped", "reason": "definedName fixture scenario"},
    )


def run_self_test() -> int:
    openpyxl = load_openpyxl()
    temp_dir = Path(tempfile.mkdtemp(prefix="fastxlsx-workbook-editor-qa-selftest-"))
    try:
        workbook_path = temp_dir / "selftest.xlsx"
        workbook = openpyxl.Workbook()
        try:
            worksheet = workbook.active
            worksheet.title = "SelfTest"
            worksheet["A1"] = "ok"
            workbook.save(workbook_path)
        finally:
            workbook.close()

        names = zip_names(workbook_path)
        require("xl/workbook.xml" in names, "self-test: workbook.xml missing")
        require("SelfTest" in read_zip_text(workbook_path, "xl/workbook.xml"),
                "self-test: worksheet title missing from workbook.xml")

        reopened = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            require(reopened["SelfTest"]["A1"].value == "ok", "self-test: openpyxl readback mismatch")
        finally:
            reopened.close()

        xlsxwriter = load_xlsxwriter()
        xlsxwriter_status = "not_installed"
        if xlsxwriter is not None:
            reference_path = temp_dir / "xlsxwriter-selftest.xlsx"
            reference = xlsxwriter.Workbook(reference_path)
            try:
                sheet = reference.add_worksheet("Ref")
                sheet.write("A1", "ok")
            finally:
                reference.close()
            require(reference_path.exists(), "self-test: xlsxwriter did not produce workbook")
            xlsxwriter_status = "created"

        qa_build = temp_dir / "qa-build"
        stale_qa_exe = qa_build / DEFAULT_QA_EXE_RELATIVE_CANDIDATES[0]
        fresh_qa_exe = qa_build / DEFAULT_QA_EXE_RELATIVE_CANDIDATES[1]
        stale_qa_exe.parent.mkdir(parents=True)
        fresh_qa_exe.parent.mkdir(parents=True)
        stale_qa_exe.write_text("stale", encoding="utf-8")
        fresh_qa_exe.write_text("fresh", encoding="utf-8")
        os.utime(stale_qa_exe, ns=(1_000_000_000, 1_000_000_000))
        os.utime(fresh_qa_exe, ns=(2_000_000_000, 2_000_000_000))
        require(
            resolve_default_qa_exe(qa_build) == fresh_qa_exe,
            "self-test: default QA exe resolution should prefer the newest candidate",
        )

        fallback_build = temp_dir / "fallback-build"
        older_discovered_exe = fallback_build / "older/tools/fastxlsx_workbook_editor_qa_tool.exe"
        newer_discovered_exe = fallback_build / "custom/tools/fastxlsx_workbook_editor_qa_tool.exe"
        older_discovered_exe.parent.mkdir(parents=True)
        newer_discovered_exe.parent.mkdir(parents=True)
        older_discovered_exe.write_text("older", encoding="utf-8")
        newer_discovered_exe.write_text("newer", encoding="utf-8")
        os.utime(older_discovered_exe, ns=(3_000_000_000, 3_000_000_000))
        os.utime(newer_discovered_exe, ns=(4_000_000_000, 4_000_000_000))
        require(
            resolve_default_qa_exe(fallback_build) == newer_discovered_exe,
            "self-test: fallback QA exe discovery should prefer the newest match",
        )

        nested_dir = temp_dir / "fixtures" / "nested"
        nested_dir.mkdir(parents=True)
        discovered_workbook = nested_dir / "fixture.xlsx"
        shutil.copyfile(workbook_path, discovered_workbook)
        (nested_dir / "~$ignored.xlsx").write_bytes(b"ignored")
        discovered = discover_external_fixtures(temp_dir / "fixtures", ["**/*.xlsx"], 0)
        require(discovered == [discovered_workbook.resolve()],
                f"self-test: fixture discovery mismatch {discovered!r}")
        unicode_slug = fixture_case_slug(
            temp_dir / "fixtures", nested_dir / "9_unicode_Λ_😇.xlsx"
        )
        require(unicode_slug.isascii(), f"self-test: non-ASCII fixture slug {unicode_slug!r}")
        long_fixture_slug = fixture_case_slug(
            temp_dir / "fixtures",
            nested_dir / ("x" * 120 + ".xlsx"),
        )
        require(
            len(long_fixture_slug) <= 24,
            f"self-test: fixture slug too long {long_fixture_slug!r}",
        )
        long_formula_slug = formula_fixture_case_slug(
            temp_dir / "fixtures",
            nested_dir / ("x" * 120 + ".xlsx"),
            "VeryLongWorksheetNameForPathBudget",
        )
        require(
            len(long_formula_slug) <= 48,
            f"self-test: formula fixture slug too long {long_formula_slug!r}",
        )

        image_fixture = workspace_asset("xlnt/tests/data/14_images.xlsx")
        require(image_fixture.exists(), f"self-test: image fixture missing {image_fixture}")
        image_infos = discover_image_fixture_infos(image_fixture.parent, ["14_images.xlsx"], 0)
        require(len(image_infos) == 1, f"self-test: image fixture discovery mismatch {image_infos!r}")
        require(image_infos[0].sheet_name == "1",
                f"self-test: image fixture sheet mismatch {image_infos[0].sheet_name!r}")
        require(image_infos[0].image_part_name.lower().endswith(".jpg"),
                f"self-test: image fixture part mismatch {image_infos[0].image_part_name!r}")

        formula_path = temp_dir / "formula.xlsx"
        formula_workbook = openpyxl.Workbook()
        try:
            formula_sheet = formula_workbook.active
            formula_sheet.title = "FormulaSheet"
            formula_sheet["A1"] = 1
            formula_sheet["B1"] = "=A1+1"
            formula_workbook.save(formula_path)
        finally:
            formula_workbook.close()
        formula_infos = formula_worksheet_infos(formula_path)
        require(len(formula_infos) == 1, f"self-test: formula scan mismatch {formula_infos!r}")
        require(formula_infos[0].sheet_name == "FormulaSheet",
                f"self-test: formula sheet mismatch {formula_infos[0].sheet_name!r}")
        require(formula_infos[0].formula_count == 1,
                f"self-test: formula count mismatch {formula_infos[0].formula_count!r}")

        shared_summary = worksheet_formula_summary(
            b"""<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">"""
            b"""<sheetData><row r="1"><c r="A1"><f t="shared" si="1">B1+C1</f></c>"""
            b"""<c r="A2"><f t="shared" si="1"/></c><c r="B1"><f>A1+1</f></c>"""
            b"""</row></sheetData></worksheet>"""
        )
        require(shared_summary["formula_count"] == 3,
                f"self-test: shared summary formula count mismatch {shared_summary!r}")
        require(shared_summary["shared_formula_count"] == 2,
                f"self-test: shared formula count mismatch {shared_summary!r}")
        require(shared_summary["shared_definition_count"] == 1,
                f"self-test: shared definition count mismatch {shared_summary!r}")
        require(shared_summary["metadata_only_shared_count"] == 1,
                f"self-test: shared follower count mismatch {shared_summary!r}")

        defined_source_path = temp_dir / "defined-source.xlsx"
        defined_path = temp_dir / "defined-names.xlsx"
        defined_workbook = openpyxl.Workbook()
        try:
            defined_sheet = defined_workbook.active
            defined_sheet.title = "Data"
            defined_sheet["A1"] = "named"
            other_sheet = defined_workbook.create_sheet("Other Sheet")
            other_sheet["A1"] = "other"
            defined_workbook.save(defined_source_path)
        finally:
            defined_workbook.close()
        defined_names_xml = (
            "<definedNames>"
            '<definedName name="ReportRange">Data!$A$1:$B$2</definedName>'
            '<definedName name="ScopedRange" localSheetId="1">\'Other Sheet\'!$A$1</definedName>'
            '<definedName name="ExternalRef">[Book.xlsx]Data!A1</definedName>'
            '<definedName name="ThreeDRef">Data:\'Other Sheet\'!A1</definedName>'
            '<definedName name="LiteralText">"Data!A1"</definedName>'
            "</definedNames>"
        )
        with zipfile.ZipFile(defined_source_path, "r") as source_archive:
            with zipfile.ZipFile(defined_path, "w", zipfile.ZIP_DEFLATED) as target_archive:
                for info in source_archive.infolist():
                    payload = source_archive.read(info.filename)
                    if info.filename == "xl/workbook.xml":
                        workbook_xml = payload.decode("utf-8")
                        if "<definedNames/>" in workbook_xml:
                            workbook_xml = workbook_xml.replace("<definedNames/>", defined_names_xml, 1)
                        elif "<definedNames />" in workbook_xml:
                            workbook_xml = workbook_xml.replace("<definedNames />", defined_names_xml, 1)
                        elif "<calcPr" in workbook_xml:
                            workbook_xml = workbook_xml.replace("<calcPr", defined_names_xml + "<calcPr", 1)
                        else:
                            require("</workbook>" in workbook_xml,
                                    "self-test: workbook close tag missing")
                            workbook_xml = workbook_xml.replace(
                                "</workbook>", defined_names_xml + "</workbook>", 1
                            )
                        payload = workbook_xml.encode("utf-8")
                    target_archive.writestr(info, payload)
        defined_records = workbook_defined_name_records(defined_path)
        require(len(defined_records) == 5,
                f"self-test: definedName record count mismatch {defined_records!r}")
        defined_summary = summarize_defined_name_records(defined_records)
        require(defined_summary["local_sheet_scoped_count"] == 1,
                f"self-test: definedName local scope mismatch {defined_summary!r}")
        require(defined_summary["external_reference_count"] == 1,
                f"self-test: definedName external reference mismatch {defined_summary!r}")
        require(defined_summary["three_d_reference_count"] == 1,
                f"self-test: definedName 3D reference mismatch {defined_summary!r}")
        defined_infos = discover_defined_name_fixture_infos(temp_dir, ["defined-names.xlsx"], 0)
        require(len(defined_infos) == 1,
                f"self-test: definedName fixture discovery mismatch {defined_infos!r}")
        require(defined_infos[0].sheet_name == "Data",
                f"self-test: definedName fixture sheet mismatch {defined_infos[0].sheet_name!r}")

        print(
            json.dumps(
                {
                    "status": "ok",
                    "workbook": str(workbook_path),
                    "xlsxwriter": xlsxwriter_status,
                    "qa_exe_resolution": "newest",
                    "fixture_discovery": len(discovered),
                    "formula_scan": len(formula_infos),
                    "defined_name_scan": len(defined_infos),
                },
                indent=2,
            )
        )
        return 0
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def run_excel_verification(report_path: Path, work_dir: Path, excel_script: Path) -> dict[str, Any]:
    office_report_path = work_dir / "office-report.json"
    completed = subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(excel_script),
            "-ReportPath",
            str(report_path),
            "-OfficeReportPath",
            str(office_report_path),
        ],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    if completed.returncode != 0:
        return {
            "status": "failed",
            "report": str(office_report_path),
            "stdout": completed.stdout,
            "stderr": completed.stderr,
        }

    if office_report_path.exists():
        with office_report_path.open("r", encoding="utf-8-sig") as stream:
            office_report = json.load(stream)
        office_report["stdout"] = completed.stdout
        return office_report

    return {
        "status": "failed",
        "report": str(office_report_path),
        "stdout": completed.stdout,
        "stderr": "Excel verification completed but did not write an office report.",
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--qa-exe",
        type=Path,
        default=resolve_default_qa_exe(),
        help="Path to fastxlsx_workbook_editor_qa_tool.exe.",
    )
    parser.add_argument(
        "--fixture-root",
        type=Path,
        help="Root directory containing external .xlsx fixtures, for example xlnt tests/data.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/workbook-editor"),
        help="Directory for QA artifacts and reports.",
    )
    parser.add_argument(
        "--report",
        type=Path,
        help="Optional aggregate JSON report path. Defaults to <work-dir>/report.json.",
    )
    parser.add_argument(
        "--scenario",
        action="append",
        choices=GENERATED_SCENARIOS + FIXTURE_SCENARIOS,
        help="Scenario group to run. Can be passed multiple times.",
    )
    parser.add_argument(
        "--fixture-glob",
        action="append",
        default=None,
        help=(
            "Glob pattern under --fixture-root for external fixture scenarios. "
            "Defaults to **/*.xlsx. Can be passed multiple times."
        ),
    )
    parser.add_argument(
        "--fixture-limit",
        type=int,
        default=0,
        help="Maximum external fixtures, formula worksheets, or definedName workbooks to run; 0 means no limit.",
    )
    parser.add_argument(
        "--formula-shared-only",
        action="store_true",
        help=(
            "For external_formula_fixture_materialized_smoke, only run worksheets "
            "whose XML contains shared formula metadata."
        ),
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Run a lightweight helper self-test and exit.",
    )
    parser.add_argument(
        "--excel-verify",
        action="store_true",
        help="Open successful output workbooks with local Excel COM and write an office report.",
    )
    parser.add_argument(
        "--excel-script",
        type=Path,
        default=resolve_default_excel_script(),
        help="Path to verify_workbook_editor_qa_excel.ps1.",
    )
    args = parser.parse_args()

    if args.self_test:
        return run_self_test()

    require(args.qa_exe is not None, "qa tool path is required; build with FASTXLSX_BUILD_QA_TOOLS=ON")
    qa_exe = args.qa_exe.resolve()
    require(qa_exe.exists(), f"qa tool not found: {qa_exe}")

    work_dir: Path = args.work_dir
    work_dir.mkdir(parents=True, exist_ok=True)

    selected = list(args.scenario) if args.scenario else list(GENERATED_SCENARIOS)
    if args.fixture_root is not None and not args.scenario:
        selected.extend([
            "xlnt_fixture_rename_smoke",
            "xlnt_fixture_string_smoke",
            IMAGE_FIXTURE_SCENARIO,
        ])

    results: list[ScenarioResult] = []
    for scenario in selected:
        if scenario in GENERATED_SCENARIOS:
            try:
                results.append(run_generated_case(qa_exe, work_dir, scenario))
            except Exception as exc:
                results.append(
                    ScenarioResult(
                        name=scenario,
                        report={"scenario": scenario, "status": "failed"},
                        zip_xml={},
                        openpyxl={},
                        xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                        error=str(exc),
                    )
                )
            continue

        require(args.fixture_root is not None,
                f"{scenario}: --fixture-root is required for fixture scenarios")
        fixture_root: Path = args.fixture_root.resolve()
        require(fixture_root.exists(), f"fixture root does not exist: {fixture_root}")

        if scenario == EXTERNAL_FIXTURE_SCENARIO:
            fixture_globs = args.fixture_glob if args.fixture_glob else ["**/*.xlsx"]
            fixtures = discover_external_fixtures(fixture_root, fixture_globs, args.fixture_limit)
            require(fixtures, f"{scenario}: no .xlsx fixtures found under {fixture_root}")
            for fixture_path in fixtures:
                try:
                    results.append(
                        run_fixture_case(
                            qa_exe,
                            work_dir,
                            fixture_path,
                            scenario,
                            fixture_root=fixture_root,
                        )
                    )
                except Exception as exc:
                    try:
                        case_name = f"{scenario}:{fixture_path.relative_to(fixture_root).as_posix()}"
                    except ValueError:
                        case_name = f"{scenario}:{fixture_path.name}"
                    results.append(
                        ScenarioResult(
                            name=case_name,
                            report={
                                "scenario": scenario,
                                "status": "failed",
                                "fixture": str(fixture_path),
                            },
                            zip_xml={},
                            openpyxl={},
                            xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                            error=str(exc),
                        )
                    )
            continue

        if scenario == FORMULA_FIXTURE_SCENARIO:
            fixture_globs = args.fixture_glob if args.fixture_glob else ["**/*.xlsx"]
            formula_infos = discover_formula_fixture_infos(
                fixture_root,
                fixture_globs,
                args.fixture_limit,
                shared_only=args.formula_shared_only,
            )
            require(
                formula_infos,
                f"{scenario}: no formula-bearing .xlsx worksheets found under {fixture_root}",
            )
            for formula_info in formula_infos:
                try:
                    results.append(
                        run_formula_fixture_case(
                            qa_exe,
                            work_dir,
                            formula_info,
                            scenario,
                            fixture_root=fixture_root,
                        )
                    )
                except Exception as exc:
                    try:
                        fixture_name = formula_info.fixture_path.relative_to(fixture_root).as_posix()
                    except ValueError:
                        fixture_name = formula_info.fixture_path.name
                    results.append(
                        ScenarioResult(
                            name=f"{scenario}:{fixture_name}:{formula_info.sheet_name}",
                            report={
                                "scenario": scenario,
                                "status": "failed",
                                "fixture": str(formula_info.fixture_path),
                                "sheet_name": formula_info.sheet_name,
                            },
                            zip_xml={"formula_scan": formula_info.as_report(fixture_root)},
                            openpyxl={},
                            xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                            error=str(exc),
                        )
                    )
            continue

        if scenario == SOURCE_FORMULA_FIXTURE_SCENARIO:
            fixture_globs = args.fixture_glob if args.fixture_glob else ["**/*.xlsx"]
            formula_infos = discover_formula_fixture_infos(
                fixture_root,
                fixture_globs,
                args.fixture_limit,
                shared_only=False,
            )
            require(
                formula_infos,
                f"{scenario}: no formula-bearing .xlsx worksheets found under {fixture_root}",
            )
            for formula_info in formula_infos:
                try:
                    results.append(
                        run_source_formula_fixture_audit_case(
                            qa_exe,
                            work_dir,
                            formula_info,
                            scenario,
                            fixture_root=fixture_root,
                        )
                    )
                except Exception as exc:
                    try:
                        fixture_name = formula_info.fixture_path.relative_to(fixture_root).as_posix()
                    except ValueError:
                        fixture_name = formula_info.fixture_path.name
                    results.append(
                        ScenarioResult(
                            name=f"{scenario}:{fixture_name}:{formula_info.sheet_name}",
                            report={
                                "scenario": scenario,
                                "status": "failed",
                                "fixture": str(formula_info.fixture_path),
                                "sheet_name": formula_info.sheet_name,
                            },
                            zip_xml={"formula_scan": formula_info.as_report(fixture_root)},
                            openpyxl={},
                            xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                            error=str(exc),
                        )
                    )
            continue

        if scenario == DEFINED_NAME_FIXTURE_SCENARIO:
            fixture_globs = args.fixture_glob if args.fixture_glob else ["**/*.xlsx"]
            defined_name_infos = discover_defined_name_fixture_infos(
                fixture_root,
                fixture_globs,
                args.fixture_limit,
            )
            require(
                defined_name_infos,
                f"{scenario}: no definedName-bearing .xlsx workbooks found under {fixture_root}",
            )
            for defined_name_info in defined_name_infos:
                try:
                    results.append(
                        run_defined_name_fixture_case(
                            qa_exe,
                            work_dir,
                            defined_name_info,
                            scenario,
                            fixture_root=fixture_root,
                        )
                    )
                except Exception as exc:
                    try:
                        fixture_name = defined_name_info.fixture_path.relative_to(fixture_root).as_posix()
                    except ValueError:
                        fixture_name = defined_name_info.fixture_path.name
                    results.append(
                        ScenarioResult(
                            name=f"{scenario}:{fixture_name}",
                            report={
                                "scenario": scenario,
                                "status": "failed",
                                "fixture": str(defined_name_info.fixture_path),
                                "sheet_name": defined_name_info.sheet_name,
                            },
                            zip_xml={"defined_name_scan": defined_name_info.as_report(fixture_root)},
                            openpyxl={},
                            xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                            error=str(exc),
                        )
                    )
            continue

        if scenario == IMAGE_FIXTURE_SCENARIO:
            fixture_globs = args.fixture_glob if args.fixture_glob else ["**/*.xlsx"]
            image_infos = discover_image_fixture_infos(
                fixture_root,
                fixture_globs,
                args.fixture_limit,
            )
            require(
                image_infos,
                f"{scenario}: no image-bearing .xlsx worksheets found under {fixture_root}",
            )
            for image_info in image_infos:
                try:
                    results.append(
                        run_fixture_image_replace_case(
                            qa_exe,
                            work_dir,
                            image_info,
                            scenario,
                            fixture_root=fixture_root,
                        )
                    )
                except Exception as exc:
                    try:
                        fixture_name = image_info.fixture_path.relative_to(fixture_root).as_posix()
                    except ValueError:
                        fixture_name = image_info.fixture_path.name
                    results.append(
                        ScenarioResult(
                            name=f"{scenario}:{fixture_name}:{image_info.image_part_name}",
                            report={
                                "scenario": scenario,
                                "status": "failed",
                                "fixture": str(image_info.fixture_path),
                                "sheet_name": image_info.sheet_name,
                                "image_part_name": image_info.image_part_name,
                            },
                            zip_xml={"image_scan": image_info.as_report(fixture_root)},
                            openpyxl={},
                            xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                            error=str(exc),
                        )
                    )
            continue

        fixture_names = (
            DEFAULT_XLNT_RENAME_FIXTURES
            if scenario == "xlnt_fixture_rename_smoke"
            else DEFAULT_XLNT_STRING_FIXTURES
        )
        for fixture_name in fixture_names:
            fixture_path = fixture_root / fixture_name
            require(fixture_path.exists(), f"fixture not found: {fixture_path}")
            try:
                results.append(run_fixture_case(qa_exe, work_dir, fixture_path, scenario))
            except Exception as exc:
                results.append(
                    ScenarioResult(
                        name=f"{scenario}:{fixture_path.name}",
                        report={"scenario": scenario, "status": "failed", "fixture": str(fixture_path)},
                        zip_xml={},
                        openpyxl={},
                        xlsxwriter_reference={"status": "skipped", "reason": "case failed"},
                        error=str(exc),
                    )
                )

    failed_cases = [result.name for result in results if result.error]
    aggregate = {
        "qa_exe": str(qa_exe),
        "work_dir": str(work_dir.resolve()),
        "status": "ok" if not failed_cases else "partial_failure",
        "failed_cases": failed_cases,
        "cases": [
            {
                "name": result.name,
                "tool_report": result.report,
                "zip_xml": result.zip_xml,
                "openpyxl": result.openpyxl,
                "xlsxwriter_reference": result.xlsxwriter_reference,
                "error": result.error,
            }
            for result in results
        ],
    }
    report_path = args.report if args.report is not None else work_dir / "report.json"
    with report_path.open("w", encoding="utf-8") as stream:
        json.dump(aggregate, stream, indent=2, ensure_ascii=False)
        stream.write("\n")

    excel_failed = False
    if args.excel_verify:
        excel_script: Path = args.excel_script.resolve()
        require(excel_script.exists(), f"Excel verification script not found: {excel_script}")
        excel_report = run_excel_verification(report_path, work_dir, excel_script)
        aggregate["excel"] = excel_report
        excel_failed = excel_report.get("status") != "ok"
        with report_path.open("w", encoding="utf-8") as stream:
            json.dump(aggregate, stream, indent=2, ensure_ascii=False)
            stream.write("\n")

    final_status = "ok" if not failed_cases and not excel_failed else "partial_failure"
    print(
        json.dumps(
            {
                "status": final_status,
                "report": str(report_path),
                "case_count": len(results),
                "failed_cases": failed_cases,
                "excel": aggregate.get("excel", {"status": "not_run"}),
            },
            indent=2,
        )
    )
    return 0 if not failed_cases and not excel_failed else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"run_workbook_editor_qa.py failed: {exc}", file=sys.stderr)
        raise
