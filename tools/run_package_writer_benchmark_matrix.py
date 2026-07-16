#!/usr/bin/env python3
"""Run an isolated repeated FastXLSX package-writer backend/buffer matrix."""

from __future__ import annotations

import argparse
import hashlib
import json
import platform
import statistics
import subprocess
import sys
import tempfile
import zipfile
import zlib
from pathlib import Path
from typing import Any


BENCHMARK_SCHEMA_VERSION = "2"
MATRIX_SCHEMA_VERSION = "2"
DEFAULT_BUFFER_KIB = [256, 512, 1024, 2048, 4096]
DEFAULT_DEFLATE_OUTPUT_KIB = [32, 128, 512, 1024]
BACKENDS = ["minizip", "direct-zlib-raw", "direct-zlib-one-pass"]
METRICS = [
    "write_ms",
    "pipeline_total_us",
    "pipeline_total_process_cpu_us",
    "output_bytes",
    "peak_memory_mb",
    "direct_deflate_total_us",
    "direct_deflate_total_process_cpu_us",
    "direct_deflate_input_bytes",
    "direct_deflate_input_read_calls",
    "direct_deflate_input_read_us",
    "direct_deflate_input_crc32_us",
    "direct_deflate_calls",
    "direct_deflate_us",
    "direct_deflate_max_us",
    "direct_deflate_output_bytes",
    "direct_deflate_output_write_calls",
    "direct_deflate_output_write_us",
    "direct_deflate_output_write_max_us",
    "direct_deflate_input_buffer_bytes",
    "direct_deflate_output_buffer_peak_bytes",
    "package_writer_total_us",
    "package_writer_total_process_cpu_us",
    "package_writer_open_us",
    "package_writer_close_us",
    "target_entry_total_us",
    "target_entry_total_process_cpu_us",
    "target_entry_open_us",
    "target_entry_close_us",
    "target_entry_close_process_cpu_us",
    "target_entry_direct_zlib_output_bytes",
    "target_entry_direct_zlib_output_buffer_peak_bytes",
    "target_entry_direct_zlib_deflate_calls",
    "target_entry_direct_zlib_deflate_us",
    "target_entry_direct_zlib_engine_process_cpu_us",
    "target_entry_direct_zlib_deflate_max_us",
    "target_entry_direct_zlib_crc32_us",
    "target_entry_input_read_calls",
    "target_entry_input_read_us",
    "target_entry_input_read_wait_us",
    "target_entry_writer_write_calls",
    "target_entry_writer_write_us",
    "target_entry_writer_write_process_cpu_us",
    "target_entry_writer_write_input_peak_bytes",
    "target_entry_writer_write_max_us",
    "target_entry_deflate_writer_process_cpu_us",
    "target_entry_staged_crc_validation_us",
    "target_entry_prefetched_staged_file_chunk_count",
    "target_entry_prefetched_staged_input_bytes",
    "target_entry_prefetch_peak_buffer_bytes",
]


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def default_benchmark_exe() -> Path:
    suffix = ".exe" if platform.system() == "Windows" else ""
    return (
        Path("build/windows-nmake-release-benchmark/benchmarks")
        / f"fastxlsx_bench_package_writer{suffix}"
    )


def validate_buffer_kib(value: int) -> int:
    if value < 64 or value > 4096 or value & (value - 1):
        raise ValueError("file buffer must be a power of two from 64 KiB to 4096 KiB")
    return value


def validate_deflate_output_kib(value: int) -> int:
    if value < 16 or value > 4096 or value & (value - 1):
        raise ValueError(
            "direct DEFLATE output buffer must be a power of two from 16 KiB to 4096 KiB"
        )
    return value


def validate_backend(value: str) -> str:
    if value not in BACKENDS:
        raise ValueError(f"unsupported package writer benchmark backend: {value}")
    return value


def validate_compression_level(value: int) -> int:
    if value < 1 or value > 9:
        raise ValueError("compression level must be between 1 and 9")
    return value


def indexed_run_name(base_name: str, run_kind: str, run_index: int, run_count: int) -> str:
    require(run_index >= 1, "run index must be positive")
    require(run_count >= 1, "run count must be positive")
    width = max(2, len(str(run_count)))
    return f"{base_name}-{run_kind}-{run_index:0{width}d}"


def metric_summary(values: list[int | float]) -> dict[str, int | float]:
    return {
        "min": min(values),
        "median": statistics.median(values),
        "max": max(values),
    }


def measured_run_summary(
    runs: list[dict[str, Any]],
) -> tuple[int, dict[str, dict[str, int | float]]]:
    require(bool(runs), "measured runs must not be empty")
    elapsed = [run["result"]["pipeline_total_us"] for run in runs]
    elapsed_median = statistics.median(elapsed)
    representative_index = min(
        range(len(runs)),
        key=lambda index: (abs(elapsed[index] - elapsed_median), index),
    )
    report: dict[str, dict[str, int | float]] = {}
    for metric in METRICS:
        values = [run["result"][metric] for run in runs]
        report[metric] = metric_summary(values)
    return representative_index, report


def run_process(command: list[str]) -> subprocess.CompletedProcess[str]:
    creationflags = 0x00000080 if platform.system() == "Windows" else 0
    completed = subprocess.run(
        command,
        check=False,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        creationflags=creationflags,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "benchmark command failed\n"
            + " ".join(command)
            + f"\nstdout:\n{completed.stdout}\nstderr:\n{completed.stderr}"
        )
    return completed


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as stream:
        value = json.load(stream)
    require(isinstance(value, dict), f"expected JSON object: {path}")
    return value


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(report, stream, ensure_ascii=False, indent=2)
        stream.write("\n")


def hash_and_crc32(path: Path) -> tuple[str, int, int]:
    digest = hashlib.sha256()
    crc32 = 0
    size = 0
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
            crc32 = zlib.crc32(chunk, crc32)
            size += len(chunk)
    return digest.hexdigest(), crc32 & 0xFFFFFFFF, size


def prepare_payload(
    bench_exe: Path,
    output_dir: Path,
    rows: int,
    cols: int,
    pattern: str,
) -> dict[str, Any]:
    payload = output_dir / "package-writer-payload.xml"
    result_path = output_dir / "payload-preparation.json"
    command = [
        str(bench_exe),
        "--prepare-only",
        "--rows",
        str(rows),
        "--cols",
        str(cols),
        "--pattern",
        pattern,
        "--payload",
        str(payload),
        "--result",
        str(result_path),
    ]
    completed = run_process(command)
    result = read_json(result_path)
    require(
        result.get("package_writer_benchmark_schema_version")
        == BENCHMARK_SCHEMA_VERSION,
        "preparation benchmark schema mismatch",
    )
    require(result.get("mode") == "prepare", "preparation mode mismatch")
    require(result.get("rows") == rows and result.get("cols") == cols, "preparation shape mismatch")
    require(result.get("pattern") == pattern, "preparation pattern mismatch")
    sha256, crc32, size = hash_and_crc32(payload)
    require(result.get("payload_bytes") == size, "preparation payload size mismatch")
    require(result.get("payload_crc32") == crc32, "preparation payload CRC mismatch")
    return {
        "command": command,
        "stdout": completed.stdout,
        "stderr": completed.stderr,
        "payload": str(payload),
        "payload_bytes": size,
        "payload_crc32": crc32,
        "payload_sha256": sha256,
        "payload_generation_ms": result["payload_generation_ms"],
    }


def inspect_archive(
    path: Path,
    payload_size: int,
    payload_crc32: int,
) -> dict[str, Any]:
    expected_names = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
    }
    with zipfile.ZipFile(path, "r") as archive:
        require(set(archive.namelist()) == expected_names, "output package entry set mismatch")
        worksheet = archive.getinfo("xl/worksheets/sheet1.xml")
        require(worksheet.compress_type == zipfile.ZIP_DEFLATED, "worksheet is not DEFLATE")
        require(worksheet.file_size == payload_size, "worksheet logical size mismatch")
        require(worksheet.CRC == payload_crc32, "worksheet CRC mismatch")
        bad_entry = archive.testzip()
        require(bad_entry is None, f"output ZIP CRC failure: {bad_entry}")
        return {
            "entry_count": len(archive.infolist()),
            "worksheet_compression_method": "deflate",
            "worksheet_uncompressed_bytes": worksheet.file_size,
            "worksheet_compressed_bytes": worksheet.compress_size,
            "worksheet_crc32": worksheet.CRC,
        }


def validate_result(
    result: dict[str, Any],
    output: Path,
    rows: int,
    cols: int,
    pattern: str,
    backend: str,
    compression_level: int,
    buffer_kib: int,
    deflate_output_kib: int,
    payload_size: int,
    payload_crc32: int,
) -> dict[str, Any]:
    require(
        result.get("package_writer_benchmark_schema_version")
        == BENCHMARK_SCHEMA_VERSION,
        "benchmark schema mismatch",
    )
    require(result.get("mode") == "timed-write", "timed mode mismatch")
    require(result.get("rows") == rows and result.get("cols") == cols, "timed shape mismatch")
    require(result.get("pattern") == pattern, "timed pattern mismatch")
    require(result.get("backend") == backend, "backend mismatch")
    require(result.get("compression_level") == compression_level, "compression level mismatch")
    buffer_bytes = buffer_kib * 1024
    require(result.get("file_io_buffer_bytes") == buffer_bytes, "file buffer mismatch")
    expected_deflate_output_bytes = (
        deflate_output_kib * 1024 if backend != "minizip" else 0
    )
    require(
        result.get("direct_deflate_output_buffer_bytes") == expected_deflate_output_bytes,
        "direct DEFLATE output buffer mismatch",
    )
    require(result.get("payload_bytes") == payload_size, "timed payload size mismatch")
    require(result.get("payload_crc32") == payload_crc32, "timed payload CRC mismatch")
    require(result.get("pipeline_total_us", 0) > 0, "pipeline timing is missing")
    require(
        result.get("temporary_compressed_file_removed") is True,
        "temporary compressed output was not removed",
    )

    direct_helper_metrics = [
        "direct_deflate_total_us",
        "direct_deflate_total_process_cpu_us",
        "direct_deflate_input_bytes",
        "direct_deflate_input_read_calls",
        "direct_deflate_input_read_us",
        "direct_deflate_input_crc32_us",
        "direct_deflate_input_crc32",
        "direct_deflate_calls",
        "direct_deflate_us",
        "direct_deflate_max_us",
        "direct_deflate_output_bytes",
        "direct_deflate_output_write_calls",
        "direct_deflate_output_write_us",
        "direct_deflate_output_write_max_us",
        "direct_deflate_input_buffer_bytes",
        "direct_deflate_output_buffer_peak_bytes",
    ]
    one_pass_metrics = [
        "target_entry_direct_zlib_output_bytes",
        "target_entry_direct_zlib_output_buffer_bytes",
        "target_entry_direct_zlib_output_buffer_peak_bytes",
        "target_entry_direct_zlib_deflate_calls",
        "target_entry_direct_zlib_deflate_us",
        "target_entry_direct_zlib_engine_process_cpu_us",
        "target_entry_direct_zlib_deflate_max_us",
        "target_entry_direct_zlib_crc32_us",
    ]

    if backend == "direct-zlib-raw":
        require(
            result.get("direct_deflate_input_bytes") == payload_size,
            "direct DEFLATE input byte mismatch",
        )
        require(
            result.get("direct_deflate_input_crc32") == payload_crc32,
            "direct DEFLATE input CRC mismatch",
        )
        expected_direct_reads = (payload_size + buffer_bytes - 1) // buffer_bytes
        require(
            result.get("direct_deflate_input_read_calls") == expected_direct_reads,
            "direct DEFLATE input read count mismatch",
        )
        require(result.get("direct_deflate_calls", 0) > 0, "direct DEFLATE calls missing")
        require(
            result.get("direct_deflate_output_write_calls", 0) > 0,
            "direct DEFLATE output writes missing",
        )
        require(
            0 < result.get("direct_deflate_output_buffer_peak_bytes", 0)
            <= expected_deflate_output_bytes,
            "direct DEFLATE output buffer peak mismatch",
        )
        compressed_input_bytes = result.get("direct_deflate_output_bytes", 0)
        require(compressed_input_bytes > 0, "direct DEFLATE produced no bytes")
        require(
            result.get("target_entry_raw_compressed_copy") is True,
            "direct backend did not use raw compressed package copy",
        )
        require(
            result.get("target_entry_reused_staged_crc32") is False,
            "raw package copy unexpectedly reported staged CRC reuse",
        )
        require(
            result.get("target_entry_direct_zlib_raw") is False,
            "two-pass backend unexpectedly reported one-pass direct zlib",
        )
        require(
            all(result.get(metric) == 0 for metric in one_pass_metrics),
            "two-pass backend reported one-pass direct zlib telemetry",
        )
        exact_writer_calls = True
    elif backend == "direct-zlib-one-pass":
        require(
            all(result.get(metric) == 0 for metric in direct_helper_metrics),
            "one-pass backend reported two-pass helper telemetry",
        )
        require(
            result.get("target_entry_raw_compressed_copy") is False,
            "one-pass backend unexpectedly used a precompressed source",
        )
        require(
            result.get("target_entry_direct_zlib_raw") is True,
            "one-pass backend did not activate direct zlib raw output",
        )
        require(
            result.get("target_entry_reused_staged_crc32") is True,
            "one-pass backend did not retain staged CRC metadata",
        )
        require(
            result.get("target_entry_direct_zlib_output_buffer_bytes")
            == expected_deflate_output_bytes,
            "one-pass output buffer mismatch",
        )
        require(
            0 < result.get("target_entry_direct_zlib_output_buffer_peak_bytes", 0)
            <= expected_deflate_output_bytes,
            "one-pass output buffer peak mismatch",
        )
        require(
            result.get("target_entry_direct_zlib_deflate_calls", 0) > 0,
            "one-pass DEFLATE calls missing",
        )
        compressed_input_bytes = payload_size
        exact_writer_calls = False
    else:
        require(
            all(result.get(metric) == 0 for metric in direct_helper_metrics),
            "minizip backend reported direct DEFLATE telemetry",
        )
        require(
            all(result.get(metric) == 0 for metric in one_pass_metrics),
            "minizip backend reported one-pass direct zlib telemetry",
        )
        compressed_input_bytes = payload_size
        require(
            result.get("target_entry_raw_compressed_copy") is False,
            "minizip backend unexpectedly used raw compressed package copy",
        )
        require(result.get("target_entry_reused_staged_crc32") is True, "staged CRC was not reused")
        require(
            result.get("target_entry_direct_zlib_raw") is False,
            "minizip backend unexpectedly reported one-pass direct zlib",
        )
        exact_writer_calls = True

    require(
        result.get("target_entry_input_bytes") == compressed_input_bytes,
        "entry input byte mismatch",
    )
    expected_calls = (compressed_input_bytes + buffer_bytes - 1) // buffer_bytes
    require(result.get("target_entry_input_read_calls") == expected_calls, "input read count mismatch")
    if exact_writer_calls:
        require(result.get("target_entry_writer_write_calls") == expected_calls, "writer call count mismatch")
        require(
            result.get("target_entry_writer_write_input_peak_bytes")
            == min(compressed_input_bytes, buffer_bytes),
            "writer input peak mismatch",
        )
    else:
        require(result.get("target_entry_writer_write_calls", 0) > 0, "one-pass writer calls missing")
        require(
            result.get("target_entry_writer_write_input_peak_bytes")
            == result.get("target_entry_direct_zlib_output_buffer_peak_bytes"),
            "one-pass writer/output peak mismatch",
        )
    if (
        backend != "direct-zlib-raw"
        and platform.system() == "Windows"
        and compressed_input_bytes >= 4 * 1024 * 1024
    ):
        require(result.get("target_entry_staged_file_read_prefetch") is True, "prefetch was not active")
        require(
            result.get("target_entry_prefetched_staged_file_chunk_count") == 1,
            "prefetched chunk count mismatch",
        )
        require(
            result.get("target_entry_prefetched_staged_input_bytes") == compressed_input_bytes,
            "prefetched byte count mismatch",
        )
        require(
            result.get("target_entry_prefetch_peak_buffer_bytes") == 2 * buffer_bytes,
            "prefetch buffer peak mismatch",
        )
    else:
        require(
            result.get("target_entry_staged_file_read_prefetch") is False,
            "prefetch unexpectedly active below threshold",
        )
    require(output.exists() and output.stat().st_size == result.get("output_bytes"), "output size mismatch")
    archive = inspect_archive(output, payload_size, payload_crc32)
    if backend == "direct-zlib-raw":
        require(
            archive["worksheet_compressed_bytes"]
            == result.get("direct_deflate_output_bytes"),
            "raw package compressed byte count mismatch",
        )
    elif backend == "direct-zlib-one-pass":
        require(
            archive["worksheet_compressed_bytes"]
            == result.get("target_entry_direct_zlib_output_bytes"),
            "one-pass compressed byte count mismatch",
        )
    return archive


def run_once(
    bench_exe: Path,
    output_dir: Path,
    base_name: str,
    run_kind: str,
    run_index: int,
    run_count: int,
    rows: int,
    cols: int,
    pattern: str,
    backend: str,
    compression_level: int,
    buffer_kib: int,
    deflate_output_kib: int,
    payload: Path,
    payload_size: int,
    payload_crc32: int,
) -> dict[str, Any]:
    name = indexed_run_name(base_name, run_kind, run_index, run_count)
    output = output_dir / f"{name}.xlsx"
    result_path = output_dir / f"{name}.json"
    command = [
        str(bench_exe),
        "--rows",
        str(rows),
        "--cols",
        str(cols),
        "--pattern",
        pattern,
        "--backend",
        backend,
        "--compression-level",
        str(compression_level),
        "--file-buffer-kib",
        str(buffer_kib),
        "--deflate-output-kib",
        str(deflate_output_kib if deflate_output_kib > 0 else 32),
        "--payload-size",
        str(payload_size),
        "--payload-crc32",
        str(payload_crc32),
        "--payload",
        str(payload),
        "--output",
        str(output),
        "--result",
        str(result_path),
    ]
    completed = run_process(command)
    result = read_json(result_path)
    archive = validate_result(
        result,
        output,
        rows,
        cols,
        pattern,
        backend,
        compression_level,
        buffer_kib,
        deflate_output_kib,
        payload_size,
        payload_crc32,
    )
    return {
        "name": name,
        "command": command,
        "stdout": completed.stdout,
        "stderr": completed.stderr,
        "result": result,
        "archive": archive,
    }


def verify_workbook_with_openpyxl(path: Path, rows: int, cols: int) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for --verify-openpyxl") from exc
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(workbook.sheetnames == ["Data"], "openpyxl sheet catalog mismatch")
        worksheet = workbook["Data"]
        require(worksheet.max_row == rows, "openpyxl max row mismatch")
        require(worksheet.max_column == cols, "openpyxl max column mismatch")
        require(worksheet["A1"].value == 1, "openpyxl A1 value mismatch")
        return {
            "status": "opened",
            "openpyxl_version": openpyxl.__version__,
            "sheet_count": 1,
            "sheet1_max_row": worksheet.max_row,
            "sheet1_max_column": worksheet.max_column,
            "sheet1_first_cell": worksheet["A1"].value,
        }
    finally:
        workbook.close()


def run_case(
    bench_exe: Path,
    output_dir: Path,
    rows: int,
    cols: int,
    pattern: str,
    backend: str,
    compression_level: int,
    buffer_kib: int,
    deflate_output_kib: int,
    payload: Path,
    payload_size: int,
    payload_crc32: int,
    warmup_runs: int,
    measured_runs: int,
    verify_openpyxl: bool,
) -> dict[str, Any]:
    output_suffix = (
        f"-deflate-output-{deflate_output_kib}kib"
        if backend != "minizip"
        else ""
    )
    base_name = (
        f"backend-{backend}-buffer-{buffer_kib}kib-level-{compression_level}"
        f"{output_suffix}"
    )
    warmups = [
        run_once(
            bench_exe,
            output_dir,
            base_name,
            "warmup",
            index,
            warmup_runs,
            rows,
            cols,
            pattern,
            backend,
            compression_level,
            buffer_kib,
            deflate_output_kib,
            payload,
            payload_size,
            payload_crc32,
        )
        for index in range(1, warmup_runs + 1)
    ]
    measured = [
        run_once(
            bench_exe,
            output_dir,
            base_name,
            "measured",
            index,
            measured_runs,
            rows,
            cols,
            pattern,
            backend,
            compression_level,
            buffer_kib,
            deflate_output_kib,
            payload,
            payload_size,
            payload_crc32,
        )
        for index in range(1, measured_runs + 1)
    ]
    representative_index, statistics_report = measured_run_summary(measured)
    representative = measured[representative_index]
    openpyxl_report = (
        verify_workbook_with_openpyxl(
            Path(representative["result"]["output"]), rows, cols
        )
        if verify_openpyxl
        else {"status": "not_requested"}
    )
    return {
        "name": base_name,
        "backend": backend,
        "compression_level": compression_level,
        "file_buffer_kib": buffer_kib,
        "direct_deflate_output_buffer_kib": (
            deflate_output_kib if backend != "minizip" else 0
        ),
        "warmup_runs": warmup_runs,
        "measured_runs": measured_runs,
        "representative_run": representative_index + 1,
        "statistics": statistics_report,
        "result": representative["result"],
        "archive": representative["archive"],
        "openpyxl": openpyxl_report,
        "warmups": warmups,
        "runs": measured,
    }


def run_self_test() -> None:
    for value in DEFAULT_BUFFER_KIB:
        require(validate_buffer_kib(value) == value, "valid buffer rejected")
    for invalid in [0, 63, 96, 8192]:
        try:
            validate_buffer_kib(invalid)
        except ValueError:
            pass
        else:
            raise AssertionError(f"invalid buffer accepted: {invalid}")
    for value in DEFAULT_DEFLATE_OUTPUT_KIB:
        require(validate_deflate_output_kib(value) == value, "valid DEFLATE buffer rejected")
    for invalid in [0, 15, 24, 8192]:
        try:
            validate_deflate_output_kib(invalid)
        except ValueError:
            pass
        else:
            raise AssertionError(f"invalid DEFLATE buffer accepted: {invalid}")
    for backend in BACKENDS:
        require(validate_backend(backend) == backend, "valid backend rejected")
    reports = [
        {"result": {metric: index + 1 for metric in METRICS}}
        for index in [2, 0, 1]
    ]
    representative, summary = measured_run_summary(reports)
    require(representative == 2, "representative selection mismatch")
    require(summary["pipeline_total_us"] == {"min": 1, "median": 2, "max": 3},
        "statistics mismatch")
    require(indexed_run_name("case", "run", 1, 3) == "case-run-01", "run name mismatch")
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        payload = root / "payload.bin"
        payload.write_bytes(b"package-writer-self-test")
        sha256, crc32, size = hash_and_crc32(payload)
        require(len(sha256) == 64 and size == 24, "hash metadata mismatch")
        archive_path = root / "archive.xlsx"
        with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", "types")
            archive.writestr("_rels/.rels", "rels")
            archive.writestr("xl/workbook.xml", "workbook")
            archive.writestr("xl/_rels/workbook.xml.rels", "workbook-rels")
            archive.writestr("xl/worksheets/sheet1.xml", payload.read_bytes())
        inspected = inspect_archive(archive_path, size, crc32)
        require(inspected["worksheet_uncompressed_bytes"] == size, "archive inspection mismatch")
    print("OK: run_package_writer_benchmark_matrix.py self-test passed")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bench-exe", type=Path, default=default_benchmark_exe())
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("build/qa/package-writer-benchmark-matrix"),
    )
    parser.add_argument("--rows", type=int, default=100000)
    parser.add_argument("--cols", type=int, default=10)
    parser.add_argument(
        "--pattern",
        choices=["numeric", "mixed-inline", "formula"],
        default="numeric",
    )
    parser.add_argument("--compression-level", type=int, action="append")
    parser.add_argument("--file-buffer-kib", type=int, action="append")
    parser.add_argument("--backend", choices=BACKENDS, action="append")
    parser.add_argument("--deflate-output-kib", type=int, action="append")
    parser.add_argument("--warmup-runs", type=int, default=2)
    parser.add_argument("--measured-runs", type=int, default=5)
    parser.add_argument("--verify-openpyxl", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        return 0

    require(args.rows > 0 and args.rows <= 1048576, "rows out of range")
    require(args.cols > 0 and args.cols <= 16384, "cols out of range")
    require(args.warmup_runs >= 0, "warmup runs must be non-negative")
    require(args.measured_runs > 0, "measured runs must be positive")
    compression_levels = args.compression_level or [1]
    buffer_values = args.file_buffer_kib or DEFAULT_BUFFER_KIB
    backends = args.backend or ["minizip"]
    deflate_output_values = args.deflate_output_kib or DEFAULT_DEFLATE_OUTPUT_KIB
    compression_levels = [validate_compression_level(value) for value in compression_levels]
    buffer_values = [validate_buffer_kib(value) for value in buffer_values]
    backends = [validate_backend(value) for value in backends]
    deflate_output_values = [
        validate_deflate_output_kib(value) for value in deflate_output_values
    ]
    require(len(set(compression_levels)) == len(compression_levels), "duplicate compression level")
    require(len(set(buffer_values)) == len(buffer_values), "duplicate file buffer")
    require(len(set(backends)) == len(backends), "duplicate backend")
    require(
        len(set(deflate_output_values)) == len(deflate_output_values),
        "duplicate direct DEFLATE output buffer",
    )

    bench_exe = args.bench_exe.resolve()
    require(bench_exe.is_file(), f"benchmark executable does not exist: {bench_exe}")
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    preparation = prepare_payload(
        bench_exe, output_dir, args.rows, args.cols, args.pattern
    )
    payload = Path(preparation["payload"])

    cases: list[dict[str, Any]] = []
    for compression_level in compression_levels:
        for buffer_kib in buffer_values:
            for backend in backends:
                backend_output_values = (
                    deflate_output_values if backend != "minizip" else [0]
                )
                for deflate_output_kib in backend_output_values:
                    case = run_case(
                        bench_exe,
                        output_dir,
                        args.rows,
                        args.cols,
                        args.pattern,
                        backend,
                        compression_level,
                        buffer_kib,
                        deflate_output_kib,
                        payload,
                        preparation["payload_bytes"],
                        preparation["payload_crc32"],
                        args.warmup_runs,
                        args.measured_runs,
                        args.verify_openpyxl,
                    )
                    cases.append(case)
                    statistics_report = case["statistics"]
                    deflate_us = (
                        statistics_report["target_entry_direct_zlib_deflate_us"]["median"]
                        if backend == "direct-zlib-one-pass"
                        else statistics_report["direct_deflate_us"]["median"]
                    )
                    print(
                        f"{case['name']}: pipeline="
                        f"{statistics_report['pipeline_total_us']['median']} us, "
                        f"deflate={deflate_us} us, "
                        f"writer={statistics_report['target_entry_writer_write_us']['median']} us, "
                        f"peak={statistics_report['peak_memory_mb']['median']} MB"
                    )

    report = {
        "package_writer_matrix_schema_version": MATRIX_SCHEMA_VERSION,
        "benchmark_executable": str(bench_exe),
        "output_dir": str(output_dir),
        "rows": args.rows,
        "cols": args.cols,
        "source_cells": args.rows * args.cols,
        "pattern": args.pattern,
        "compression_levels": compression_levels,
        "file_buffer_kib_values": buffer_values,
        "backends": backends,
        "direct_deflate_output_buffer_kib_values": deflate_output_values,
        "warmup_runs_per_case": args.warmup_runs,
        "measured_runs_per_case": args.measured_runs,
        "representative_result_policy":
            "Measured run nearest pipeline_total_us median; earliest run breaks ties.",
        "process_priority": "Windows High priority" if platform.system() == "Windows" else "inherited",
        "preparation": preparation,
        "cases": cases,
        "scope":
            "Isolated staged worksheet payload -> production minizip-ng DEFLATE or "
            "benchmark-only direct zlib raw-DEFLATE plus production raw-copy package writer. "
            "The one-pass direct zlib backend streams bounded raw DEFLATE output directly "
            "into the production minizip raw-entry writer. "
            "Payload generation and openpyxl validation are outside timed benchmark processes; "
            "Office remains not_run.",
    }
    report_path = output_dir / "package-writer-benchmark-matrix-report.json"
    write_report(report_path, report)
    print(f"Wrote {report_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (AssertionError, RuntimeError, ValueError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
