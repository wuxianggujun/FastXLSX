#!/usr/bin/env python3
"""Run opt-in C++ XLSX writer reference benchmark adapters.

The adapters are standalone executables for third-party XLSX libraries such as
OpenXLSX and xlnt. They are benchmark/reference tools only and must not be
linked into the FastXLSX runtime library.
"""

from __future__ import annotations

import argparse
import json
import math
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


REPORT_SCHEMA_VERSION = "1"
DEFAULT_CASES = [
    "numeric:mixed",
    "mixed:mixed",
    "strings:repeated",
    "strings:unique",
]
BYTES_PER_MIB = 1024.0 * 1024.0


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


@dataclass(frozen=True)
class Adapter:
    name: str
    path: Path


@dataclass(frozen=True)
class BenchmarkCase:
    scenario: str
    string_pattern: str

    @property
    def name(self) -> str:
        if self.scenario == "numeric":
            return "numeric"
        return f"{self.scenario}-{self.string_pattern}"


def parse_adapter(text: str) -> Adapter:
    if "=" not in text:
        raise ValueError("adapter must use name=path format")
    name, path = text.split("=", 1)
    name = name.strip().lower()
    if name not in {"openxlsx", "xlnt"}:
        raise ValueError(f"unsupported reference adapter: {name}")
    if not path.strip():
        raise ValueError("adapter path cannot be empty")
    return Adapter(name=name, path=Path(path))


def parse_case(text: str) -> BenchmarkCase:
    parts = text.split(":")
    if len(parts) != 2:
        raise ValueError(f"case must use scenario:string_pattern format: {text}")
    scenario, string_pattern = parts
    if scenario not in {"numeric", "mixed", "strings"}:
        raise ValueError(f"unsupported scenario in case {text!r}")
    if string_pattern not in {"mixed", "repeated", "unique"}:
        raise ValueError(f"unsupported string pattern in case {text!r}")
    if scenario == "numeric" and string_pattern != "mixed":
        raise ValueError("numeric cases should use the mixed placeholder string pattern")
    return BenchmarkCase(scenario=scenario, string_pattern=string_pattern)


def expected_string_ratio(case: BenchmarkCase, mixed_string_ratio: float) -> float:
    if case.scenario == "numeric":
        return 0.0
    if case.scenario == "strings":
        return 1.0
    return mixed_string_ratio


def should_write_string(case: BenchmarkCase, mixed_string_ratio: float, row: int, col: int) -> bool:
    ratio = expected_string_ratio(case, mixed_string_ratio)
    if case.scenario == "numeric" or ratio <= 0.0:
        return False
    if case.scenario == "strings" or ratio >= 1.0:
        return True
    raw_bucket = 1.0 / ratio
    if not math.isfinite(raw_bucket) or raw_bucket >= 2**32 - 1:
        return False
    bucket = max(int(raw_bucket), 1)
    return ((row + col) % bucket) == 0


def string_value_is_repeated(case: BenchmarkCase, row: int, col: int) -> bool:
    if case.string_pattern == "repeated":
        return True
    if case.string_pattern == "unique":
        return False
    return (row + col) % 5 == 0


def make_string_value(case: BenchmarkCase, sheet: int, row: int, col: int) -> str:
    if string_value_is_repeated(case, row, col):
        return "repeat"
    return f"s{sheet}-r{row}-c{col}"


def make_number_value(sheet: int, row: int, col: int) -> int:
    return sheet * 1_000_000 + row * 1_000 + col


def make_cell_value(
    case: BenchmarkCase,
    mixed_string_ratio: float,
    sheet: int,
    row: int,
    col: int,
) -> int | str:
    if should_write_string(case, mixed_string_ratio, row, col):
        return make_string_value(case, sheet, row, col)
    return make_number_value(sheet, row, col)


def verify_workbook_with_openpyxl(
    path: Path,
    case: BenchmarkCase,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(len(workbook.sheetnames) == sheets, "sheet count mismatch")
        require("Sheet1" in workbook.sheetnames, "missing Sheet1")
        worksheet = workbook["Sheet1"]
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()
        observed_rows = 0
        observed_cols = 0
        first_value = None
        last_value = None
        for row_values in worksheet.iter_rows(values_only=True):
            observed_rows += 1
            observed_cols = max(observed_cols, len(row_values))
            if observed_rows == 1 and row_values:
                first_value = row_values[0]
            if len(row_values) >= cols:
                last_value = row_values[cols - 1]
            elif row_values:
                last_value = row_values[-1]
            else:
                last_value = None
        require(observed_rows == rows, f"row count mismatch: {observed_rows}")
        require(observed_cols == cols, f"column count mismatch: {observed_cols}")
        first_expected = make_cell_value(case, mixed_string_ratio, 1, 1, 1)
        last_expected = make_cell_value(case, mixed_string_ratio, 1, rows, cols)
        require(first_value == first_expected,
            f"A1 mismatch: expected {first_expected!r}, got {first_value!r}")
        require(last_value == last_expected,
            f"last cell mismatch: expected {last_expected!r}, got {last_value!r}")
        return {
            "status": "opened",
            "sheet_count": len(workbook.sheetnames),
            "sheet1_max_row": observed_rows,
            "sheet1_max_column": observed_cols,
            "sheet1_first_cell": first_value,
            "sheet1_last_cell": last_value,
        }
    finally:
        workbook.close()


def run_case(
    adapter: Adapter,
    case: BenchmarkCase,
    output_dir: Path,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
    verify_openpyxl: bool,
    strict_missing: bool,
    timeout_seconds: float | None,
) -> dict[str, Any]:
    case_name = f"{adapter.name}-{case.name}"
    adapter_path = adapter.path.resolve()
    if not adapter_path.exists():
        result = {
            "name": case_name,
            "library": adapter.name,
            "case": case.name,
            "status": "skipped",
            "skip_reason": f"adapter executable not found: {adapter_path}",
            "rows": rows,
            "cols": cols,
            "sheets": sheets,
            "cells": rows * cols * sheets,
        }
        if strict_missing:
            raise AssertionError(result["skip_reason"])
        return result

    output_path = output_dir / f"cpp-reference-bench-{case_name}.xlsx"
    result_path = output_dir / f"cpp-reference-bench-{case_name}.json"
    command = [
        str(adapter_path),
        "--scenario",
        case.scenario,
        "--rows",
        str(rows),
        "--cols",
        str(cols),
        "--sheets",
        str(sheets),
        "--string-pattern",
        case.string_pattern,
        "--string-ratio",
        str(mixed_string_ratio),
        "--output",
        str(output_path),
        "--result",
        str(result_path),
    ]
    try:
        completed = subprocess.run(
            command,
            text=True,
            capture_output=True,
            check=False,
            timeout=timeout_seconds,
        )
    except subprocess.TimeoutExpired as exc:
        for artifact_path in (output_path, result_path):
            try:
                artifact_path.unlink()
            except FileNotFoundError:
                pass
        return {
            "name": case_name,
            "library": adapter.name,
            "library_mode": "workbook-api",
            "case": case.name,
            "scenario": case.scenario,
            "string_pattern": case.string_pattern,
            "status": "timeout",
            "timeout_seconds": timeout_seconds,
            "rows": rows,
            "cols": cols,
            "sheets": sheets,
            "cells": rows * cols * sheets,
            "string_ratio": expected_string_ratio(case, mixed_string_ratio),
            "output": str(output_path),
            "result_json": str(result_path),
            "command": command,
            "stdout": (exc.stdout or "").strip() if isinstance(exc.stdout, str) else "",
            "stderr": (exc.stderr or "").strip() if isinstance(exc.stderr, str) else "",
            "openpyxl": {"status": "not_run"},
        }
    require(completed.returncode == 0,
        f"{case_name} failed with {completed.returncode}: {completed.stderr.strip()}")
    require(output_path.exists(), f"{case_name} output workbook was not created")
    require(result_path.exists(), f"{case_name} result JSON was not created")

    result = json.loads(result_path.read_text(encoding="utf-8"))
    output_bytes = output_path.stat().st_size
    elapsed_ms = int(result["elapsed_ms"])
    cells = rows * cols * sheets
    verify_result = (
        verify_workbook_with_openpyxl(output_path, case, rows, cols, sheets, mixed_string_ratio)
        if verify_openpyxl
        else {"status": "not_requested"}
    )
    return {
        "name": case_name,
        "library": result.get("library", adapter.name),
        "library_mode": result.get("library_mode", "workbook-api"),
        "case": case.name,
        "scenario": case.scenario,
        "string_pattern": case.string_pattern,
        "status": "completed",
        "rows": rows,
        "cols": cols,
        "sheets": sheets,
        "cells": cells,
        "string_ratio": expected_string_ratio(case, mixed_string_ratio),
        "elapsed_ms": elapsed_ms,
        "cells_per_second": cells / (elapsed_ms / 1000.0) if elapsed_ms > 0 else 0.0,
        "million_cells_per_second": cells / (elapsed_ms / 1000.0) / 1_000_000.0
            if elapsed_ms > 0 else 0.0,
        "peak_memory_mb": result.get("peak_memory_mb"),
        "output_bytes": output_bytes,
        "output_mib": output_bytes / BYTES_PER_MIB,
        "office_open": result.get("office_open", "not_run"),
        "openpyxl": verify_result,
        "output": str(output_path),
        "result_json": str(result_path),
        "command": command,
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
    }


def build_report(
    output_dir: Path,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
    cases: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "cpp_reference_writer_benchmark_schema_version": REPORT_SCHEMA_VERSION,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "output_dir": str(output_dir),
        "rows": rows,
        "cols": cols,
        "sheets": sheets,
        "cells_per_case": rows * cols * sheets,
        "mixed_string_ratio": mixed_string_ratio,
        "cases": cases,
        "notes": [
            "OpenXLSX and xlnt adapters are opt-in benchmark/reference tools only.",
            "Adapters use public workbook APIs and are not linked into the FastXLSX runtime library.",
            "Office validation is separate; adapter result JSON keeps office_open=not_run.",
        ],
    }


def fmt_int(value: int | None) -> str:
    return "-" if value is None else f"{value:,}"


def fmt_float(value: float | None, digits: int = 2) -> str:
    return "-" if value is None else f"{value:.{digits}f}"


def render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# C++ XLSX Reference Writer Benchmark Summary",
        "",
        f"- Generated UTC: `{report['generated_at_utc']}`",
        f"- Rows x cols x sheets: `{report['rows']} x {report['cols']} x {report['sheets']}`",
        f"- Cells per case: `{report['cells_per_case']:,}`",
        "",
        "| library | mode | case | status | elapsed ms | M cells/s | peak MB | output MiB | openpyxl |",
        "| --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for case in report["cases"]:
        openpyxl_status = "-"
        if isinstance(case.get("openpyxl"), dict):
            openpyxl_status = str(case["openpyxl"].get("status", "-"))
        lines.append(
            "| "
            f"{case.get('library', '-')} | "
            f"{case.get('library_mode', '-')} | "
            f"{case.get('case', '-')} | "
            f"{case.get('status', '-')} | "
            f"{fmt_int(case.get('elapsed_ms'))} | "
            f"{fmt_float(case.get('million_cells_per_second'))} | "
            f"{fmt_float(case.get('peak_memory_mb'))} | "
            f"{fmt_float(case.get('output_mib'))} | "
            f"{openpyxl_status} |"
        )

    skipped = [case for case in report["cases"] if case.get("status") == "skipped"]
    if skipped:
        lines.extend(["", "## Skipped"])
        for case in skipped:
            lines.append(f"- {case['name']}: {case.get('skip_reason', 'skipped')}")

    timed_out = [case for case in report["cases"] if case.get("status") == "timeout"]
    if timed_out:
        lines.extend(["", "## Timed out"])
        for case in timed_out:
            lines.append(f"- {case['name']}: exceeded {case.get('timeout_seconds')} seconds")

    lines.extend(["", "## Notes"])
    lines.extend(f"- {note}" for note in report["notes"])
    lines.append("")
    return "\n".join(lines)


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def expect_value_error(text: str, parser: Any) -> None:
    try:
        parser(text)
    except ValueError:
        return
    raise AssertionError(f"expected parser to reject {text!r}")


def run_self_test() -> None:
    adapter = parse_adapter("openxlsx=build/ref/openxlsx.exe")
    require(adapter.name == "openxlsx", "adapter name mismatch")
    require(adapter.path == Path("build/ref/openxlsx.exe"), "adapter path mismatch")
    expect_value_error("openxlsx", parse_adapter)
    expect_value_error("other=tool.exe", parse_adapter)

    numeric = parse_case("numeric:mixed")
    mixed = parse_case("mixed:mixed")
    repeated = parse_case("strings:repeated")
    unique = parse_case("strings:unique")
    require(numeric.name == "numeric", "numeric case name mismatch")
    require(mixed.name == "mixed-mixed", "mixed case name mismatch")
    require(repeated.name == "strings-repeated", "repeated case name mismatch")
    require(unique.name == "strings-unique", "unique case name mismatch")
    expect_value_error("numeric:repeated", parse_case)
    expect_value_error("strings", parse_case)
    expect_value_error("other:mixed", parse_case)

    require(make_cell_value(numeric, 0.25, 1, 2, 3) == 1002003,
        "numeric cell value mismatch")
    require(make_cell_value(unique, 0.25, 1, 2, 3) == "s1-r2-c3",
        "unique cell value mismatch")

    report = build_report(Path("out"), 2, 3, 1, 0.25, [{
        "name": "openxlsx-numeric",
        "library": "openxlsx",
        "library_mode": "workbook-api",
        "case": "numeric",
        "status": "completed",
        "elapsed_ms": 10,
        "million_cells_per_second": 0.0006,
        "peak_memory_mb": 20.0,
        "output_mib": 0.01,
        "openpyxl": {"status": "opened"},
    }, {
        "name": "xlnt-numeric",
        "library": "xlnt",
        "library_mode": "workbook-api",
        "case": "numeric",
        "status": "skipped",
        "skip_reason": "missing",
    }])
    markdown = render_markdown(report)
    require("C++ XLSX Reference Writer Benchmark Summary" in markdown, "markdown title mismatch")
    require("openxlsx" in markdown, "markdown completed case mismatch")
    require("missing" in markdown, "markdown skipped reason mismatch")
    timeout_report = build_report(Path("out"), 2, 3, 1, 0.25, [{
        "name": "openxlsx-strings-unique",
        "library": "openxlsx",
        "library_mode": "workbook-api",
        "case": "strings-unique",
        "status": "timeout",
        "timeout_seconds": 1.0,
        "openpyxl": {"status": "not_run"},
    }])
    timeout_markdown = render_markdown(timeout_report)
    require("Timed out" in timeout_markdown, "markdown timeout section mismatch")
    with tempfile.TemporaryDirectory() as temp:
        markdown_path = Path(temp) / "summary.md"
        write_text(markdown_path, markdown)
        require(markdown_path.read_text(encoding="utf-8") == markdown,
            "markdown write mismatch")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--self-test", action="store_true",
        help="Run lightweight internal checks without invoking adapters.")
    parser.add_argument("--adapter", action="append", dest="adapters",
        help="Reference adapter in name=path format. Names: openxlsx, xlnt.")
    parser.add_argument("--case", action="append", dest="cases",
        help="Case in scenario:string_pattern format. May be passed multiple times.")
    parser.add_argument("--output-dir", type=Path,
        default=Path("build/qa/cpp-reference-writer-benchmarks"),
        help="Directory for generated workbooks and reports.")
    parser.add_argument("--output-markdown", type=Path,
        help="Write a Markdown summary to this path.")
    parser.add_argument("--rows", type=int, default=1000, help="Rows per worksheet.")
    parser.add_argument("--cols", type=int, default=10, help="Columns per worksheet.")
    parser.add_argument("--sheets", type=int, default=1, help="Worksheet count.")
    parser.add_argument("--mixed-string-ratio", type=float, default=0.25,
        help="String ratio used by mixed scenario cases.")
    parser.add_argument("--verify-openpyxl", action="store_true",
        help="Open generated workbooks with openpyxl and verify Sheet1 first/last cells.")
    parser.add_argument("--strict-missing", action="store_true",
        help="Fail if a requested adapter executable is missing.")
    parser.add_argument("--timeout-seconds", type=float,
        help="Per adapter-case timeout. Timed-out child processes are killed and reported.")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        print("OK: run_cpp_reference_writer_benchmarks.py self-test passed")
        return 0

    require(args.adapters, "at least one --adapter name=path is required")
    require(args.rows > 0 and args.cols > 0 and args.sheets > 0,
        "rows, cols, and sheets must be positive")
    require(0.0 <= args.mixed_string_ratio <= 1.0,
        "mixed string ratio must be between 0 and 1")
    if args.timeout_seconds is not None:
        require(args.timeout_seconds > 0, "timeout seconds must be positive")

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    adapters = [parse_adapter(value) for value in args.adapters]
    cases = [parse_case(value) for value in (args.cases or DEFAULT_CASES)]

    results = []
    for adapter in adapters:
        for case in cases:
            results.append(run_case(
                adapter=adapter,
                case=case,
                output_dir=output_dir,
                rows=args.rows,
                cols=args.cols,
                sheets=args.sheets,
                mixed_string_ratio=args.mixed_string_ratio,
                verify_openpyxl=args.verify_openpyxl,
                strict_missing=args.strict_missing,
                timeout_seconds=args.timeout_seconds,
            ))

    report = build_report(
        output_dir=output_dir,
        rows=args.rows,
        cols=args.cols,
        sheets=args.sheets,
        mixed_string_ratio=args.mixed_string_ratio,
        cases=results,
    )
    report_path = output_dir / "cpp-reference-writer-benchmark-report.json"
    report_path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    markdown = render_markdown(report)
    markdown_path = args.output_markdown.resolve() if args.output_markdown else (
        output_dir / "cpp-reference-writer-benchmark-summary.md")
    write_text(markdown_path, markdown)
    print(markdown)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI should emit one concise failure line.
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
