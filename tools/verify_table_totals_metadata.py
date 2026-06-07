#!/usr/bin/env python3
"""Local table totals-row metadata QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It verifies FastXLSX table XML
semantics directly, then uses Python XLSX libraries only as local QA/reference
tools. These libraries are not FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


NAMESPACES = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def relationship_count(xml: str) -> int:
    root = ElementTree.fromstring(xml)
    return len(root.findall("rel:Relationship", NAMESPACES))


def verify_fastxlsx_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet3.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/tables/table1.xml",
        "xl/tables/table2.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")
    require("xl/worksheets/_rels/sheet3.xml.rels" not in names, "plain sheet should not have rels")
    require("xl/styles.xml" not in names, "table metadata should not create styles.xml")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require(
        content_types.count(
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"'
        )
        == 2,
        "table content type override count mismatch",
    )

    first_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(first_rels) == 2, "first sheet relationship count mismatch")
    require('Id="rId1"' in first_rels and 'TargetMode="External"' in first_rels,
            "first sheet hyperlink rId mismatch")
    require('Id="rId2"' in first_rels and 'Target="../tables/table1.xml"' in first_rels,
            "first sheet table rId mismatch")

    second_rels = read_zip_text(path, "xl/worksheets/_rels/sheet2.xml.rels")
    require(relationship_count(second_rels) == 1, "second sheet relationship count mismatch")
    require('Id="rId1"' in second_rels and 'Target="../tables/table2.xml"' in second_rels,
            "second sheet table rId mismatch")

    first_table_xml = read_zip_text(path, "xl/tables/table1.xml")
    require('name="InventoryTable"' in first_table_xml, "missing InventoryTable")
    require('ref="A1:C3"' in first_table_xml, "InventoryTable ref mismatch")
    require('totalsRowShown="0"' in first_table_xml, "InventoryTable should hide totals row")
    require('totalsRowCount=' not in first_table_xml, "InventoryTable should not write totalsRowCount")
    require('<autoFilter ref="A1:C3"/>' in first_table_xml, "InventoryTable autoFilter mismatch")

    second_table_xml = read_zip_text(path, "xl/tables/table2.xml")
    require('name="TotalsTable"' in second_table_xml, "missing TotalsTable")
    require('ref="A1:B3"' in second_table_xml, "TotalsTable ref mismatch")
    require('totalsRowCount="1"' in second_table_xml, "TotalsTable should expose one totals row")
    require('totalsRowShown="0"' not in second_table_xml,
            "TotalsTable should not also write hidden totals metadata")
    require('<autoFilter ref="A1:B2"/>' in second_table_xml,
            "TotalsTable autoFilter should exclude totals row")
    require('totalsRowFunction="sum"' in second_table_xml,
            "FastXLSX should emit the requested totals function metadata")
    require("totalsRowLabel" not in second_table_xml,
            "FastXLSX should not generate totals labels")
    require("totalsRowFormula" not in second_table_xml,
            "FastXLSX should not generate totals formula text")
    require("calculatedColumnFormula" not in second_table_xml,
            "FastXLSX should not generate calculated column formulas")
    require("<tableStyleInfo" not in second_table_xml,
            "empty style name should still omit tableStyleInfo")

    root = ElementTree.fromstring(second_table_xml)
    columns = root.findall("main:tableColumns/main:tableColumn", NAMESPACES)
    require([column.attrib.get("name") for column in columns] == ["Metric", "Value"],
            "TotalsTable parsed column names mismatch")
    require([column.attrib.get("totalsRowFunction") for column in columns] == [None, "sum"],
            "TotalsTable parsed totals functions mismatch")

    return {
        "verified_entries": required,
        "inventory_table": {
            "ref": "A1:C3",
            "totals": "hidden",
        },
        "totals_table": {
            "ref": root.attrib.get("ref"),
            "totalsRowCount": root.attrib.get("totalsRowCount"),
            "column_names": [column.attrib.get("name") for column in columns],
            "totals_functions": [column.attrib.get("totalsRowFunction") for column in columns],
        },
    }


def verify_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("Inventory" in workbook.sheetnames, "missing Inventory sheet")
        require("Totals" in workbook.sheetnames, "missing Totals sheet")
        require("Plain" in workbook.sheetnames, "missing Plain sheet")

        inventory = workbook["Inventory"]
        totals = workbook["Totals"]
        plain = workbook["Plain"]
        require("InventoryTable" in inventory.tables, "openpyxl missing InventoryTable")
        require("TotalsTable" in totals.tables, "openpyxl missing TotalsTable")
        require(len(plain.tables) == 0, "Plain sheet should not expose tables")

        inventory_table = inventory.tables["InventoryTable"]
        totals_table = totals.tables["TotalsTable"]
        require(inventory_table.ref == "A1:C3", "openpyxl InventoryTable ref mismatch")
        require(totals_table.ref == "A1:B3", "openpyxl TotalsTable ref mismatch")
        require(getattr(totals_table, "totalsRowCount", None) == 1,
                "openpyxl TotalsTable totalsRowCount mismatch")
        require([column.name for column in totals_table.tableColumns] == ["Metric", "Value"],
                "openpyxl TotalsTable column names mismatch")
        require([getattr(column, "totalsRowFunction", None) for column in totals_table.tableColumns]
                == [None, "sum"],
                "openpyxl TotalsTable totals functions mismatch")

        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "inventory_ref": inventory_table.ref,
            "totals_ref": totals_table.ref,
            "totalsRowCount": getattr(totals_table, "totalsRowCount", None),
            "totals_functions": [
                getattr(column, "totalsRowFunction", None)
                for column in totals_table.tableColumns
            ],
        }
    finally:
        workbook.close()


def create_xlsxwriter_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("Totals")
    worksheet.write_row("A1", ["Metric", "Value"])
    worksheet.write_row("A2", ["Rows", 2])
    worksheet.write_row("A3", ["Total", 2])
    worksheet.add_table(
        "A1:B3",
        {
            "name": "TotalsTable",
            "total_row": True,
            "columns": [
                {"header": "Metric"},
                {"header": "Value", "total_function": "sum"},
            ],
        },
    )
    workbook.close()
    reference_xml = read_zip_text(path, "xl/tables/table1.xml")
    require('totalsRowCount="1"' in reference_xml,
            "XlsxWriter reference did not emit totalsRowCount")
    require('<autoFilter ref="A1:B2"/>' in reference_xml,
            "XlsxWriter reference autoFilter did not exclude totals row")
    require('totalsRowFunction="sum"' in reference_xml,
            "XlsxWriter reference did not emit totalsRowFunction")
    return {"status": "created", "path": str(path)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-tables.xlsx"),
        help="FastXLSX table workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/table-totals"),
        help="Directory for reference files and the JSON report.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    package_report = verify_fastxlsx_package(input_path)
    openpyxl_report = verify_with_openpyxl(input_path)
    xlsxwriter_report = create_xlsxwriter_reference(work_dir / "reference-xlsxwriter-table-totals.xlsx")

    report = {
        "fastxlsx_input": str(input_path),
        "fastxlsx_package": package_report,
        "xlsx_libraries": {
            "openpyxl": openpyxl_report,
            "xlsxwriter": xlsxwriter_report,
        },
    }
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - local QA helper
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
