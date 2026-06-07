#!/usr/bin/env python3
"""Local Phase 3 worksheet metadata QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It verifies FastXLSX OpenXML
structure directly, then uses openpyxl only as a local QA/reference reader.
Python XLSX libraries are not FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def count(text: str, fragment: str) -> int:
    return text.count(fragment)


def verify_fastxlsx_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/core.xml",
        "docProps/app.xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")

    forbidden = [
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/sharedStrings.xml",
        "xl/calcChain.xml",
        "xl/styles.xml",
        "xl/drawings/drawing1.xml",
        "xl/tables/table1.xml",
    ]
    for name in forbidden:
        require(name not in names, f"unexpected package entry: {name}")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require("drawing" not in content_types, "phase3 metadata should not create drawing content types")
    require("spreadsheetml.table+xml" not in content_types,
            "phase3 metadata should not create table content type overrides")
    require("spreadsheetml.styles+xml" not in content_types,
            "phase3 metadata should not create styles content type overrides")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require(count(workbook_rels, "<Relationship ") == 1,
            "phase3 metadata should not add workbook relationships")

    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    require('name="Metadata"' in workbook_xml, "missing Metadata sheet")
    require('<calcPr calcId="124519" fullCalcOnLoad="1"/>' in workbook_xml,
            "phase3 formula metadata should request full recalculation on load")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("xmlns:r=" not in worksheet_xml,
            "phase3 metadata worksheet should not declare relationship namespace")
    require('<dimension ref="A1:D4"/>' in worksheet_xml,
            "phase3 metadata worksheet dimension mismatch")
    require(
        '<sheetViews><sheetView workbookViewId="0"><pane xSplit="3" ySplit="2" '
        'topLeftCell="D3" activePane="bottomRight" state="frozen"/></sheetView></sheetViews>'
        in worksheet_xml,
        "last freeze pane XML mismatch",
    )
    require('topLeftCell="A2"' not in worksheet_xml,
            "obsolete freeze pane setting should not be serialized")
    require(count(worksheet_xml, "<col ") == 2,
            "phase3 metadata column width count mismatch")
    require(
        '<cols><col min="1" max="1" width="12.25" customWidth="1"/>'
        '<col min="3" max="4" width="8.75" customWidth="1"/></cols>'
        in worksheet_xml,
        "phase3 metadata column width XML mismatch",
    )
    require('<row r="2" ht="19.25" customHeight="1">' in worksheet_xml,
            "phase3 metadata row height XML mismatch")
    require('<c r="B2"><f>A2*2</f></c>' in worksheet_xml,
            "plain formula XML mismatch")
    require('<c r="C2"><f>IF(A2&gt;0,"&lt;yes&gt;","&amp;no")</f></c>' in worksheet_xml,
            "escaped formula XML mismatch")
    require('<autoFilter ref="B2:D4"/>' in worksheet_xml,
            "last autoFilter range mismatch")
    require('<autoFilter ref="A1:B2"/>' not in worksheet_xml,
            "obsolete autoFilter range should not be serialized")
    require(
        '<mergeCells count="2"><mergeCell ref="A3:B3"/><mergeCell ref="C4:D4"/></mergeCells>'
        in worksheet_xml,
        "phase3 metadata mergeCells XML mismatch",
    )
    require(
        '</sheetData><autoFilter ref="B2:D4"/><mergeCells count="2">'
        in worksheet_xml,
        "phase3 metadata suffix ordering mismatch",
    )

    return {
        "verified_entries": required,
        "forbidden_entries_absent": forbidden,
        "worksheet_dimension": "A1:D4",
        "auto_filter": "B2:D4",
        "merged_ranges": ["A3:B3", "C4:D4"],
    }


def verify_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for this local QA helper") from exc

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("Metadata" in workbook.sheetnames, "missing Metadata worksheet")
        worksheet = workbook["Metadata"]
        require(worksheet.max_row == 4, f"worksheet row count mismatch: {worksheet.max_row}")
        require(worksheet.max_column == 4,
                f"worksheet column count mismatch: {worksheet.max_column}")
        require(worksheet["A2"].value == 42, f"A2 value mismatch: {worksheet['A2'].value!r}")
        require(worksheet["B2"].value == "=A2*2",
                f"B2 formula mismatch: {worksheet['B2'].value!r}")
        require(worksheet["C2"].value == '=IF(A2>0,"<yes>","&no")',
                f"C2 formula mismatch: {worksheet['C2'].value!r}")
        require(worksheet["D2"].value is True,
                f"D2 boolean mismatch: {worksheet['D2'].value!r}")
        require(worksheet["B4"].value == 7,
                f"B4 value mismatch: {worksheet['B4'].value!r}")
        require(abs(float(worksheet.row_dimensions[2].height) - 19.25) < 0.001,
                "row 2 height mismatch")
        require(abs(float(worksheet.column_dimensions["A"].width) - 12.25) < 0.001,
                "column A width mismatch")
        require(abs(float(worksheet.column_dimensions["C"].width) - 8.75) < 0.001,
                "column C:D width record mismatch")
        require(worksheet.freeze_panes == "D3",
                f"freeze panes mismatch: {worksheet.freeze_panes!r}")
        require(worksheet.auto_filter.ref == "B2:D4",
                f"autoFilter mismatch: {worksheet.auto_filter.ref!r}")
        merged_ranges = sorted(str(item) for item in worksheet.merged_cells.ranges)
        require(merged_ranges == ["A3:B3", "C4:D4"],
                f"merged range mismatch: {merged_ranges!r}")
        require(getattr(workbook.calculation, "calcId", None) == 124519,
                "calcId mismatch")
        require(getattr(workbook.calculation, "fullCalcOnLoad", None) is True,
                "fullCalcOnLoad mismatch")
        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "freeze_panes": worksheet.freeze_panes,
            "auto_filter": worksheet.auto_filter.ref,
            "merged_ranges": merged_ranges,
            "row2_height": worksheet.row_dimensions[2].height,
            "column_a_width": worksheet.column_dimensions["A"].width,
            "column_c_width": worksheet.column_dimensions["C"].width,
            "calcId": workbook.calculation.calcId,
            "fullCalcOnLoad": workbook.calculation.fullCalcOnLoad,
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-phase3-metadata.xlsx"),
        help="FastXLSX Phase 3 metadata workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/phase3-metadata"),
        help="Directory for the JSON report.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    package_report = verify_fastxlsx_package(input_path)
    openpyxl_report = verify_with_openpyxl(input_path)

    report = {
        "fastxlsx_input": str(input_path),
        "fastxlsx_package": package_report,
        "xlsx_libraries": {
            "openpyxl": openpyxl_report,
        },
        "comparison_scope": (
            "Phase 3 metadata structure and reader-visible workbook semantics only; "
            "formula calculation, cached values, calcChain, styles, and full Phase 3 "
            "coverage are intentionally not claimed."
        ),
    }
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - local QA helper
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
