#!/usr/bin/env python3
"""Local hyperlink QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It uses Python XLSX libraries only
as local QA/reference tools and never as FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


EXTERNAL_LINKS = {
    "Links": {
        "A1": {
            "target": "https://openai.com/",
            "location": None,
            "display": None,
            "tooltip": None,
            "value": "OpenAI",
        },
        "B2": {
            "target": "https://example.com/path?a=1&b=2",
            "location": None,
            "display": None,
            "tooltip": None,
            "value": "Docs & <API>",
        },
    },
    "MoreLinks": {
        "A1": {
            "target": "mailto:test@example.com",
            "location": None,
            "display": None,
            "tooltip": None,
            "value": "Second sheet",
        },
    },
    "Plain": {},
}

INTERNAL_LINKS = {
    "Internal": {
        "A1": {
            "target": None,
            "location": "'Target & <Sheet>'!A1",
            "display": None,
            "tooltip": None,
            "value": "Jump to target",
        },
        "A2": {
            "target": None,
            "location": "'Target & <Sheet>'!B2:\"quoted\"",
            "display": None,
            "tooltip": None,
            "value": "Second jump",
        },
    },
    "Target & <Sheet>": {},
    "Mixed": {
        "A1": {
            "target": "https://example.com/",
            "location": None,
            "display": None,
            "tooltip": None,
            "value": "External",
        },
        "B1": {
            "target": None,
            "location": "'Target & <Sheet>'!A1",
            "display": None,
            "tooltip": None,
            "value": "Internal",
        },
        "A2": {
            "target": "https://example.com/more",
            "location": None,
            "display": None,
            "tooltip": None,
            "value": "External 2",
        },
    },
    "Plain": {},
}

DISPLAY_TOOLTIP_LINKS = {
    "ExternalAttrs": {
        "A1": {
            "target": "https://example.com/both",
            "location": None,
            "display": "Open & <Docs> \"Q\" 'A'",
            "tooltip": "External tip & <more> \"Q\" 'A'",
            "value": "External both",
        },
        "B1": {
            "target": "https://example.com/display",
            "location": None,
            "display": "Display only",
            "tooltip": None,
            "value": "External display",
        },
        "C1": {
            "target": "https://example.com/tooltip",
            "location": None,
            "display": None,
            "tooltip": "Tooltip only",
            "value": "External tooltip",
        },
    },
    "InternalAttrs": {
        "A1": {
            "target": None,
            "location": "Target!A1",
            "display": "Jump & <Target> \"Q\" 'A'",
            "tooltip": "Internal tip & <more> \"Q\" 'A'",
            "value": "Internal both",
        },
        "B1": {
            "target": None,
            "location": "Target!A2",
            "display": "Internal display only",
            "tooltip": None,
            "value": "Internal display",
        },
        "C1": {
            "target": None,
            "location": "Target!A3",
            "display": None,
            "tooltip": "Internal tooltip only",
            "value": "Internal tooltip",
        },
        "D1": {
            "target": None,
            "location": "Target!D4",
            "display": None,
            "tooltip": None,
            "value": "Internal empty options",
        },
    },
    "Target": {},
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def count_relationships(xml: str) -> int:
    return xml.count("<Relationship ")


def require_no_common_side_effects(path: Path) -> None:
    names = zip_names(path)
    require("[Content_Types].xml" in names, "missing content types part")
    require("xl/_rels/workbook.xml.rels" in names, "missing workbook relationships")
    content_types = read_zip_text(path, "[Content_Types].xml")
    require("hyperlink" not in content_types, "hyperlinks should not add content types")
    require("styles.xml" not in content_types, "hyperlinks should not add styles content type")
    require("xl/styles.xml" not in names, "hyperlinks should not create styles.xml")
    require("xl/metadata.xml" not in names, "hyperlinks should not create metadata.xml")
    require("xl/calcChain.xml" not in names, "hyperlinks should not create calcChain.xml")


def verify_external_package(path: Path) -> dict[str, Any]:
    require_no_common_side_effects(path)
    names = zip_names(path)
    require("xl/worksheets/_rels/sheet1.xml.rels" in names, "missing first external sheet rels")
    require("xl/worksheets/_rels/sheet2.xml.rels" in names, "missing second external sheet rels")
    require("xl/worksheets/_rels/sheet3.xml.rels" not in names, "plain sheet should not have rels")

    sheet1 = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("xmlns:r=" in sheet1, "external hyperlink sheet should declare relationships namespace")
    require(
        '<hyperlinks><hyperlink ref="A1" r:id="rId1"/><hyperlink ref="B2" r:id="rId2"/></hyperlinks>'
        in sheet1,
        "external hyperlink XML mismatch",
    )
    require("Docs &amp; &lt;API&gt;" in sheet1, "external hyperlink should preserve escaped cell text")

    rels1 = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(count_relationships(rels1) == 2, "first external sheet relationship count mismatch")
    require(
        'Id="rId1"' in rels1
        and 'Target="https://openai.com/"' in rels1
        and 'TargetMode="External"' in rels1,
        "first external hyperlink relationship mismatch",
    )
    require(
        'Id="rId2"' in rels1
        and 'Target="https://example.com/path?a=1&amp;b=2"' in rels1
        and 'TargetMode="External"' in rels1,
        "second external hyperlink relationship mismatch",
    )

    sheet2 = read_zip_text(path, "xl/worksheets/sheet2.xml")
    rels2 = read_zip_text(path, "xl/worksheets/_rels/sheet2.xml.rels")
    require('<hyperlink ref="A1" r:id="rId1"/>' in sheet2, "second external sheet rId mismatch")
    require(
        'Id="rId1"' in rels2 and 'Target="mailto:test@example.com"' in rels2,
        "second external sheet relationship mismatch",
    )
    sheet3 = read_zip_text(path, "xl/worksheets/sheet3.xml")
    require("<hyperlinks>" not in sheet3 and "xmlns:r=" not in sheet3, "plain external sheet polluted")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require(count_relationships(workbook_rels) == 3, "external hyperlinks polluted workbook rels")
    return {"relationship_model": "worksheet-owner-local external rIds", "worksheets": 3}


def verify_internal_package(path: Path) -> dict[str, Any]:
    require_no_common_side_effects(path)
    names = zip_names(path)
    require("xl/worksheets/_rels/sheet1.xml.rels" not in names, "internal-only sheet should not have rels")
    require("xl/worksheets/_rels/sheet2.xml.rels" not in names, "target sheet should not have rels")
    require("xl/worksheets/_rels/sheet3.xml.rels" in names, "mixed sheet should have external rels")
    require("xl/worksheets/_rels/sheet4.xml.rels" not in names, "plain internal sheet should not have rels")

    internal = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("xmlns:r=" not in internal and "r:id=" not in internal, "internal-only sheet used relationships")
    require(
        '<hyperlink ref="A1" location="&apos;Target &amp; &lt;Sheet&gt;&apos;!A1"/>'
        in internal,
        "first internal hyperlink location mismatch",
    )
    require(
        '<hyperlink ref="A2" location="&apos;Target &amp; &lt;Sheet&gt;&apos;!B2:&quot;quoted&quot;"/>'
        in internal,
        "second internal hyperlink location escape mismatch",
    )

    mixed = read_zip_text(path, "xl/worksheets/sheet3.xml")
    require(
        '<hyperlinks><hyperlink ref="A1" r:id="rId1"/><hyperlink ref="A2" r:id="rId2"/>'
        '<hyperlink ref="B1" location="&apos;Target &amp; &lt;Sheet&gt;&apos;!A1"/></hyperlinks>'
        in mixed,
        "mixed external/internal hyperlink XML mismatch",
    )
    mixed_rels = read_zip_text(path, "xl/worksheets/_rels/sheet3.xml.rels")
    require(count_relationships(mixed_rels) == 2, "mixed sheet relationship count mismatch")
    require('Target="https://example.com/"' in mixed_rels, "mixed first external target mismatch")
    require('Target="https://example.com/more"' in mixed_rels, "mixed second external target mismatch")
    require("Target &amp; &lt;Sheet&gt;" not in mixed_rels, "internal target leaked into rels")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require(count_relationships(workbook_rels) == 4, "internal hyperlinks polluted workbook rels")
    return {"relationship_model": "internal locations do not consume rIds", "worksheets": 4}


def verify_display_tooltip_package(path: Path) -> dict[str, Any]:
    require_no_common_side_effects(path)
    names = zip_names(path)
    require("xl/worksheets/_rels/sheet1.xml.rels" in names, "external attrs sheet should have rels")
    require("xl/worksheets/_rels/sheet2.xml.rels" not in names, "internal attrs sheet should not have rels")
    require("xl/worksheets/_rels/sheet3.xml.rels" not in names, "target sheet should not have rels")

    external = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require(
        'display="Open &amp; &lt;Docs&gt; &quot;Q&quot; &apos;A&apos;"' in external
        and 'tooltip="External tip &amp; &lt;more&gt; &quot;Q&quot; &apos;A&apos;"' in external,
        "external display/tooltip escape mismatch",
    )
    require('display="Display only"' in external, "external display-only attribute missing")
    require('tooltip="Tooltip only"' in external, "external tooltip-only attribute missing")

    external_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(count_relationships(external_rels) == 3, "external attrs relationship count mismatch")
    require("display=" not in external_rels and "tooltip=" not in external_rels,
            "display/tooltip leaked into relationships")

    internal = read_zip_text(path, "xl/worksheets/sheet2.xml")
    require("xmlns:r=" not in internal and "r:id=" not in internal, "internal attrs sheet used rels")
    require(
        'location="Target!A1" display="Jump &amp; &lt;Target&gt; &quot;Q&quot; &apos;A&apos;" '
        'tooltip="Internal tip &amp; &lt;more&gt; &quot;Q&quot; &apos;A&apos;"'
        in internal,
        "internal display/tooltip escape mismatch",
    )
    require('location="Target!D4"/>' in internal, "empty internal hyperlink options missing")
    require('display=""' not in internal and 'tooltip=""' not in internal,
            "empty display/tooltip attributes should be omitted")
    return {"attribute_model": "display/tooltip stay on worksheet hyperlink elements", "worksheets": 3}


def import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for this local QA helper") from exc
    return openpyxl


def verify_openpyxl_links(path: Path, expected: dict[str, dict[str, dict[str, str | None]]]) -> dict[str, Any]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require(workbook.sheetnames == list(expected.keys()),
                f"sheet names mismatch: expected {list(expected.keys())!r}, got {workbook.sheetnames!r}")
        observed: dict[str, dict[str, Any]] = {}
        for sheet_name, expected_cells in expected.items():
            worksheet = workbook[sheet_name]
            observed[sheet_name] = {}
            actual_cells: dict[str, Any] = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    hyperlink = cell.hyperlink
                    if hyperlink is None:
                        continue
                    actual_cells[cell.coordinate] = {
                        "target": hyperlink.target,
                        "location": hyperlink.location,
                        "display": hyperlink.display,
                        "tooltip": hyperlink.tooltip,
                        "value": cell.value,
                    }
            require(set(actual_cells) == set(expected_cells),
                    f"{sheet_name} hyperlink cells mismatch: {set(actual_cells)!r}")
            for coordinate, expected_link in expected_cells.items():
                actual = actual_cells[coordinate]
                for key, expected_value in expected_link.items():
                    require(
                        actual[key] == expected_value,
                        f"{sheet_name}!{coordinate} {key} mismatch: "
                        f"expected {expected_value!r}, got {actual[key]!r}",
                    )
                observed[sheet_name][coordinate] = actual
        return observed
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--external-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-external-hyperlinks.xlsx"),
        help="FastXLSX external hyperlink workbook to verify.",
    )
    parser.add_argument(
        "--internal-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-internal-hyperlinks.xlsx"),
        help="FastXLSX internal hyperlink workbook to verify.",
    )
    parser.add_argument(
        "--display-tooltip-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-hyperlink-display-tooltips.xlsx"),
        help="FastXLSX hyperlink display/tooltip workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/hyperlinks"),
        help="Directory for the JSON report.",
    )
    args = parser.parse_args()

    external_input = args.external_input.resolve()
    internal_input = args.internal_input.resolve()
    display_tooltip_input = args.display_tooltip_input.resolve()
    work_dir = args.work_dir.resolve()
    for path in [external_input, internal_input, display_tooltip_input]:
        require(path.exists(), f"input workbook does not exist: {path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    report = {
        "external_hyperlinks": {
            "fastxlsx_input": str(external_input),
            "package": verify_external_package(external_input),
            "openpyxl": verify_openpyxl_links(external_input, EXTERNAL_LINKS),
        },
        "internal_hyperlinks": {
            "fastxlsx_input": str(internal_input),
            "package": verify_internal_package(internal_input),
            "openpyxl": verify_openpyxl_links(internal_input, INTERNAL_LINKS),
        },
        "display_tooltip_hyperlinks": {
            "fastxlsx_input": str(display_tooltip_input),
            "package": verify_display_tooltip_package(display_tooltip_input),
            "openpyxl": verify_openpyxl_links(display_tooltip_input, DISPLAY_TOOLTIP_LINKS),
        },
        "notes": [
            "This helper verifies package XML and openpyxl semantics only.",
            "Excel COM visual checks remain separate local QA.",
        ],
    }
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - local QA helper
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
