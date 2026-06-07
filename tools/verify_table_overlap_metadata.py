#!/usr/bin/env python3
"""Local table range-overlap QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It verifies FastXLSX package XML
semantics directly, then uses openpyxl as a local QA reader. openpyxl is not a
FastXLSX runtime dependency.
"""

from __future__ import annotations

import argparse
import json
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


NAMESPACES = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def parse_table(path: Path, name: str) -> dict[str, str | None]:
    root = ElementTree.fromstring(read_zip_text(path, name))
    return {
        "name": root.attrib.get("name"),
        "displayName": root.attrib.get("displayName"),
        "ref": root.attrib.get("ref"),
    }


def verify_fastxlsx_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/tables/table1.xml",
        "xl/tables/table2.xml",
        "xl/tables/table3.xml",
        "xl/tables/table4.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")
    require("xl/tables/table5.xml" not in names, "rejected overlap should not create table5.xml")
    require("xl/styles.xml" not in names, "table overlap metadata should not create styles.xml")

    first_sheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    second_sheet_xml = read_zip_text(path, "xl/worksheets/sheet2.xml")
    require('<tableParts count="3">' in first_sheet_xml, "first worksheet table count mismatch")
    require('<tableParts count="1">' in second_sheet_xml, "second worksheet table count mismatch")

    expected_tables = {
        "xl/tables/table1.xml": ("FirstTable", "A1:B2"),
        "xl/tables/table2.xml": ("AdjacentColumnsTable", "C1:D2"),
        "xl/tables/table3.xml": ("AdjacentRowsTable", "A3:B4"),
        "xl/tables/table4.xml": ("OtherSheetTable", "A1:B2"),
    }
    parsed_tables = {}
    for entry, (expected_name, expected_ref) in expected_tables.items():
        table = parse_table(path, entry)
        require(table["name"] == expected_name, f"{entry} table name mismatch")
        require(table["displayName"] == expected_name, f"{entry} displayName mismatch")
        require(table["ref"] == expected_ref, f"{entry} ref mismatch")
        parsed_tables[entry] = table

    return {
        "verified_entries": required,
        "tables": parsed_tables,
    }


def verify_with_openpyxl(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("Tables" in workbook.sheetnames, "missing Tables sheet")
        require("OtherTables" in workbook.sheetnames, "missing OtherTables sheet")

        tables_sheet = workbook["Tables"]
        other_sheet = workbook["OtherTables"]
        expected_first = {
            "FirstTable": "A1:B2",
            "AdjacentColumnsTable": "C1:D2",
            "AdjacentRowsTable": "A3:B4",
        }
        require(set(tables_sheet.tables.keys()) == set(expected_first.keys()),
                "openpyxl first worksheet table names mismatch")
        for table_name, table_ref in expected_first.items():
            require(tables_sheet.tables[table_name].ref == table_ref,
                    f"openpyxl {table_name} ref mismatch")
        require(set(other_sheet.tables.keys()) == {"OtherSheetTable"},
                "openpyxl second worksheet table names mismatch")
        require(other_sheet.tables["OtherSheetTable"].ref == "A1:B2",
                "openpyxl cross-worksheet table ref mismatch")

        return {
            "status": "opened",
            "sheetnames": workbook.sheetnames,
            "tables": {
                "Tables": {
                    name: tables_sheet.tables[name].ref
                    for name in tables_sheet.tables.keys()
                },
                "OtherTables": {
                    name: other_sheet.tables[name].ref
                    for name in other_sheet.tables.keys()
                },
            },
        }
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-table-range-overlap.xlsx"),
        help="FastXLSX table range-overlap workbook to verify.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")

    report = {
        "fastxlsx_input": str(input_path),
        "fastxlsx_package": verify_fastxlsx_package(input_path),
        "xlsx_libraries": {
            "openpyxl": verify_with_openpyxl(input_path),
        },
    }
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
