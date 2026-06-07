#!/usr/bin/env python3
"""Local data-validation prompt/error QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It uses Python XLSX libraries only
as local QA/reference tools and never as FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


EXPECTED_RULES: dict[str, dict[str, str | bool | None]] = {
    "A2:A10": {
        "type": "whole",
        "operator": "between",
        "formula1": "1",
        "formula2": "10",
        "showInputMessage": True,
        "showErrorMessage": True,
        "errorStyle": "warning",
        "errorTitle": 'Error "Title" & <bad>',
        "error": "Bad 'value' & <cell>",
        "promptTitle": 'Input <Title> & "Quote"',
        "prompt": "Enter 'whole' & <value>",
    },
    "B2:B10": {
        "type": "list",
        "operator": None,
        "formula1": '"A,B,C"',
        "formula2": None,
        "showInputMessage": True,
        "showErrorMessage": False,
        "errorStyle": None,
        "errorTitle": None,
        "error": None,
        "promptTitle": "Choice",
        "prompt": "Pick A, B, or C",
    },
    "C2:C10": {
        "type": "decimal",
        "operator": "greaterThan",
        "formula1": "0",
        "formula2": None,
        "showInputMessage": False,
        "showErrorMessage": True,
        "errorStyle": "information",
        "errorTitle": "Decimal",
        "error": "Use a positive decimal",
        "promptTitle": None,
        "prompt": None,
    },
    "D2:D10": {
        "type": "custom",
        "operator": None,
        "formula1": "LEN(D2)>0",
        "formula2": None,
        "showInputMessage": False,
        "showErrorMessage": False,
        "errorStyle": "stop",
        "errorTitle": None,
        "error": None,
        "promptTitle": None,
        "prompt": None,
    },
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def verify_fastxlsx_package(path: Path) -> list[str]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/core.xml",
        "docProps/app.xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")

    forbidden = [
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/metadata.xml",
        "xl/styles.xml",
        "xl/calcChain.xml",
    ]
    for name in forbidden:
        require(name not in names, f"unexpected package entry: {name}")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require("dataValidation" not in content_types, "unexpected dataValidation content type")
    require("styles" not in content_types, "unexpected styles content type")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require(workbook_rels.count("<Relationship ") == 1, "unexpected workbook relationship count")

    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    require("<calcPr" not in workbook_xml, "data validation prompts should not request calculation")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("xmlns:r=" not in worksheet_xml, "validation-only worksheet should not declare relationships")
    require('<dimension ref="A1:D2"/>' in worksheet_xml, "worksheet dimension mismatch")
    require("<dataValidations count=\"4\">" in worksheet_xml, "dataValidations count mismatch")
    require(worksheet_xml.count("<dataValidation ") == 4, "dataValidation item count mismatch")

    expected_fragments = [
        (
            '<dataValidation type="whole" allowBlank="1" showInputMessage="1" '
            'showErrorMessage="1" errorStyle="warning" '
            'errorTitle="Error &quot;Title&quot; &amp; &lt;bad&gt;" '
            'error="Bad &apos;value&apos; &amp; &lt;cell&gt;" '
            'promptTitle="Input &lt;Title&gt; &amp; &quot;Quote&quot;" '
            'prompt="Enter &apos;whole&apos; &amp; &lt;value&gt;" operator="between" '
            'sqref="A2:A10"><formula1>1</formula1><formula2>10</formula2></dataValidation>'
        ),
        (
            '<dataValidation type="list" showInputMessage="1" promptTitle="Choice" '
            'prompt="Pick A, B, or C" sqref="B2:B10"><formula1>"A,B,C"</formula1></dataValidation>'
        ),
        (
            '<dataValidation type="decimal" showErrorMessage="1" errorStyle="information" '
            'errorTitle="Decimal" error="Use a positive decimal" operator="greaterThan" '
            'sqref="C2:C10"><formula1>0</formula1></dataValidation>'
        ),
        (
            '<dataValidation type="custom" errorStyle="stop" sqref="D2:D10">'
            "<formula1>LEN(D2)&gt;0</formula1></dataValidation>"
        ),
    ]
    for fragment in expected_fragments:
        require(fragment in worksheet_xml, f"missing dataValidation XML fragment: {fragment}")

    for forbidden_fragment in [
        'showInputMessage="0"',
        'showErrorMessage="0"',
        'promptTitle=""',
        'prompt=""',
        'errorTitle=""',
        'error=""',
    ]:
        require(forbidden_fragment not in worksheet_xml, f"unexpected empty/false attribute: {forbidden_fragment}")

    return required


def import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for this local QA helper") from exc
    return openpyxl


def normalize_bool(value: Any) -> bool | None:
    if value is None:
        return None
    return bool(value)


def verify_rules_with_openpyxl(path: Path) -> dict[str, dict[str, Any]]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        require("ValidationPrompt" in workbook.sheetnames, "missing ValidationPrompt worksheet")
        worksheet = workbook["ValidationPrompt"]
        validations = list(worksheet.data_validations.dataValidation)
        require(len(validations) == 4, f"expected 4 validations, got {len(validations)}")

        observed: dict[str, dict[str, Any]] = {}
        for validation in validations:
            sqref = str(validation.sqref)
            require(sqref in EXPECTED_RULES, f"unexpected validation range: {sqref}")
            expected = EXPECTED_RULES[sqref]
            actual = {
                "type": validation.type,
                "operator": validation.operator,
                "formula1": validation.formula1,
                "formula2": validation.formula2,
                "showInputMessage": normalize_bool(validation.showInputMessage),
                "showErrorMessage": normalize_bool(validation.showErrorMessage),
                "errorStyle": validation.errorStyle,
                "errorTitle": validation.errorTitle,
                "error": validation.error,
                "promptTitle": validation.promptTitle,
                "prompt": validation.prompt,
            }
            for key, expected_value in expected.items():
                require(
                    actual[key] == expected_value,
                    f"{sqref} {key} mismatch: expected {expected_value!r}, got {actual[key]!r}",
                )
            observed[sqref] = actual
        return observed
    finally:
        workbook.close()


def create_openpyxl_reference(path: Path) -> None:
    openpyxl = import_openpyxl()
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ValidationPrompt"
    for col, value in enumerate(["Whole", "List", "Decimal", "Custom"], start=1):
        worksheet.cell(row=1, column=col, value=value)

    whole = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
    whole.showInputMessage = True
    whole.showErrorMessage = True
    whole.errorStyle = "warning"
    whole.errorTitle = 'Error "Title" & <bad>'
    whole.error = "Bad 'value' & <cell>"
    whole.promptTitle = 'Input <Title> & "Quote"'
    whole.prompt = "Enter 'whole' & <value>"
    worksheet.add_data_validation(whole)
    whole.add("A2:A10")

    list_rule = DataValidation(type="list", formula1='"A,B,C"')
    list_rule.showInputMessage = True
    list_rule.promptTitle = "Choice"
    list_rule.prompt = "Pick A, B, or C"
    worksheet.add_data_validation(list_rule)
    list_rule.add("B2:B10")

    decimal = DataValidation(type="decimal", operator="greaterThan", formula1="0")
    decimal.showErrorMessage = True
    decimal.errorStyle = "information"
    decimal.errorTitle = "Decimal"
    decimal.error = "Use a positive decimal"
    worksheet.add_data_validation(decimal)
    decimal.add("C2:C10")

    custom = DataValidation(type="custom", formula1="LEN(D2)>0")
    custom.errorStyle = "stop"
    worksheet.add_data_validation(custom)
    custom.add("D2:D10")

    workbook.save(path)


def create_xlsxwriter_reference(path: Path) -> str | None:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return None

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("ValidationPrompt")
    worksheet.write_row(0, 0, ["Whole", "List", "Decimal", "Custom"])
    worksheet.data_validation(
        "A2:A10",
        {
            "validate": "integer",
            "criteria": "between",
            "minimum": 1,
            "maximum": 10,
            "ignore_blank": True,
            "input_title": 'Input <Title> & "Quote"',
            "input_message": "Enter 'whole' & <value>",
            "error_title": 'Error "Title" & <bad>',
            "error_message": "Bad 'value' & <cell>",
            "error_type": "warning",
        },
    )
    worksheet.data_validation(
        "B2:B10",
        {
            "validate": "list",
            "source": ["A", "B", "C"],
            "input_title": "Choice",
            "input_message": "Pick A, B, or C",
        },
    )
    worksheet.data_validation(
        "C2:C10",
        {
            "validate": "decimal",
            "criteria": ">",
            "value": 0,
            "error_title": "Decimal",
            "error_message": "Use a positive decimal",
            "error_type": "information",
        },
    )
    workbook.close()
    return "created"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-data-validation-prompts.xlsx"),
        help="FastXLSX data validation prompt/error workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/data-validation-prompts"),
        help="Directory for reference files and the JSON report.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    verified_entries = verify_fastxlsx_package(input_path)
    fastxlsx_rules = verify_rules_with_openpyxl(input_path)

    openpyxl_reference = work_dir / "reference-openpyxl-data-validation-prompts.xlsx"
    create_openpyxl_reference(openpyxl_reference)
    openpyxl_reference_rules = verify_rules_with_openpyxl(openpyxl_reference)

    xlsxwriter_reference = work_dir / "reference-xlsxwriter-data-validation-prompts.xlsx"
    xlsxwriter_status = create_xlsxwriter_reference(xlsxwriter_reference)

    report = {
        "fastxlsx_input": str(input_path),
        "verified_fastxlsx_entries": verified_entries,
        "fastxlsx_rules": fastxlsx_rules,
        "references": {
            "openpyxl": {
                "status": "created",
                "path": str(openpyxl_reference),
                "rules": openpyxl_reference_rules,
            },
            "xlsxwriter": {
                "status": "created" if xlsxwriter_status else "skipped",
                "path": str(xlsxwriter_reference) if xlsxwriter_status else None,
                "reason": None if xlsxwriter_status else "Python module xlsxwriter is not installed",
            },
        },
    }
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - local QA helper
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
