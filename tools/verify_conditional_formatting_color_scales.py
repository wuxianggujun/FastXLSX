#!/usr/bin/env python3
"""Local QA for FastXLSX conditional-formatting color scales.

This helper is intentionally local QA, not a runtime dependency and not a
default CI gate. It checks generated OpenXML first, then uses openpyxl as a
reader-visible semantic check and XlsxWriter as an optional reference writer.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


EXPECTED_BASIC_SQREF = "A2:A10"
EXPECTED_MULTI_RANGE_SQREF = "A2:A3 C2:C3 E2:E3"
EXPECTED_LOW_COLOR = "FFFF0000"
EXPECTED_HIGH_COLOR = "FF00B050"


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        return archive.read(name).decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def worksheet_relationship_entries(names: set[str]) -> list[str]:
    return sorted(
        name
        for name in names
        if name.startswith("xl/worksheets/_rels/") and name.endswith(".xml.rels")
    )


def verify_no_package_side_effects(path: Path, worksheet_rels_allowed: bool = False) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/core.xml",
        "docProps/app.xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")

    forbidden = ["xl/styles.xml", "xl/metadata.xml", "xl/calcChain.xml"]
    rel_entries = worksheet_relationship_entries(names)
    if not worksheet_rels_allowed:
        require(not rel_entries, f"unexpected worksheet relationship entries: {rel_entries}")
    for name in forbidden:
        require(name not in names, f"unexpected package entry: {name}")

    content_types = read_zip_text(path, "[Content_Types].xml")
    for fragment in ["conditionalFormatting", "styles", "metadata"]:
        require(fragment not in content_types, f"unexpected content type fragment: {fragment}")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require("styles" not in workbook_rels, "conditional formatting should not add styles relationship")
    require("conditionalFormatting" not in workbook_rels,
            "conditional formatting should not add workbook relationship")

    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    require("<calcPr" not in workbook_xml, "conditional formatting should not request recalculation")

    return {
        "required_entries": required,
        "forbidden_entries_absent": forbidden,
        "worksheet_relationship_entries": rel_entries,
    }


def verify_basic_package(path: Path) -> dict[str, Any]:
    side_effects = verify_no_package_side_effects(path)
    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("xmlns:r=" not in worksheet_xml,
            "conditional-formatting-only worksheet should not declare relationship namespace")
    require('<dimension ref="A1:A10"/>' in worksheet_xml, "basic worksheet dimension mismatch")
    fragment = (
        f'<conditionalFormatting sqref="{EXPECTED_BASIC_SQREF}">'
        '<cfRule type="colorScale" priority="1"><colorScale>'
        '<cfvo type="min"/><cfvo type="max"/>'
        f'<color rgb="{EXPECTED_LOW_COLOR}"/><color rgb="{EXPECTED_HIGH_COLOR}"/>'
        '</colorScale></cfRule></conditionalFormatting>'
    )
    require(fragment in worksheet_xml, "basic two-color scale XML fragment mismatch")
    require(worksheet_xml.count("<conditionalFormatting ") == 1,
            "basic conditionalFormatting count mismatch")
    require(worksheet_xml.count("<cfRule ") == 1, "basic cfRule count mismatch")
    require(worksheet_xml.count("<cfvo ") == 2, "basic cfvo count mismatch")
    require(worksheet_xml.count("<color rgb=") == 2, "basic color count mismatch")
    for forbidden_fragment in ["dxfId=", "<dataBar", "<iconSet", "<formula>", "<dxfs"]:
        require(forbidden_fragment not in worksheet_xml,
                f"unexpected unsupported conditional-formatting fragment: {forbidden_fragment}")
    return {"side_effects": side_effects, "sqref": EXPECTED_BASIC_SQREF}


def verify_metadata_order_package(path: Path) -> dict[str, Any]:
    verify_no_package_side_effects(path, worksheet_rels_allowed=True)
    names = zip_names(path)
    require("xl/worksheets/_rels/sheet1.xml.rels" in names,
            "metadata-order workbook should keep relationship-backed objects")
    require("xl/tables/table1.xml" in names, "metadata-order workbook should keep table part")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    expected_order = (
        '</sheetData><mergeCells count="1"><mergeCell ref="A4:B4"/></mergeCells>'
        f'<conditionalFormatting sqref="{EXPECTED_BASIC_SQREF}">'
        '<cfRule type="colorScale" priority="1"><colorScale>'
        '<cfvo type="min"/><cfvo type="max"/>'
        f'<color rgb="{EXPECTED_LOW_COLOR}"/><color rgb="{EXPECTED_HIGH_COLOR}"/>'
        '</colorScale></cfRule></conditionalFormatting>'
        '<dataValidations count="1">'
    )
    require(expected_order in worksheet_xml,
            "conditional formatting should be between mergeCells and dataValidations")
    require(
        '<hyperlinks><hyperlink ref="B2" r:id="rId1"/></hyperlinks>'
        '<tableParts count="1"><tablePart r:id="rId2"/></tableParts>' in worksheet_xml,
        "conditional formatting should not consume worksheet-local relationship ids",
    )

    worksheet_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(worksheet_rels.count("<Relationship ") == 2,
            "metadata-order worksheet relationship count mismatch")
    require('Id="rId1"' in worksheet_rels and "relationships/hyperlink" in worksheet_rels,
            "hyperlink relationship should remain rId1")
    require('Id="rId2"' in worksheet_rels and "relationships/table" in worksheet_rels,
            "table relationship should remain rId2")
    return {"suffix_order": "mergeCells -> conditionalFormatting -> dataValidations"}


def verify_multi_range_package(path: Path) -> dict[str, Any]:
    side_effects = verify_no_package_side_effects(path)
    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("xmlns:r=" not in worksheet_xml,
            "multi-range conditional formatting should not declare relationship namespace")
    require('<dimension ref="A1:E3"/>' in worksheet_xml, "multi-range worksheet dimension mismatch")
    fragment = (
        f'<conditionalFormatting sqref="{EXPECTED_MULTI_RANGE_SQREF}">'
        '<cfRule type="colorScale" priority="1"><colorScale>'
        '<cfvo type="min"/><cfvo type="max"/>'
        f'<color rgb="{EXPECTED_LOW_COLOR}"/><color rgb="{EXPECTED_HIGH_COLOR}"/>'
        '</colorScale></cfRule></conditionalFormatting>'
    )
    require(fragment in worksheet_xml, "multi-range color scale XML fragment mismatch")
    require(worksheet_xml.count("<conditionalFormatting ") == 1,
            "multi-range conditionalFormatting count mismatch")
    return {"side_effects": side_effects, "sqref": EXPECTED_MULTI_RANGE_SQREF}


def verify_priorities_package(path: Path) -> dict[str, Any]:
    verify_no_package_side_effects(path)
    first_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    second_xml = read_zip_text(path, "xl/worksheets/sheet2.xml")
    third_xml = read_zip_text(path, "xl/worksheets/sheet3.xml")

    require(first_xml.count("<conditionalFormatting ") == 2,
            "first worksheet conditional formatting count mismatch")
    require('priority="1"' in first_xml and 'priority="2"' in first_xml,
            "first worksheet priorities mismatch")
    require(
        '<cfvo type="num" val="0"/><cfvo type="percentile" val="90"/>'
        '<color rgb="FFFFEB84"/><color rgb="FF5A8AD6"/>' in first_xml,
        "numeric/percentile endpoint XML mismatch",
    )
    require(second_xml.count("<conditionalFormatting ") == 1,
            "second worksheet conditional formatting count mismatch")
    require('priority="1"' in second_xml and 'priority="2"' not in second_xml,
            "second worksheet priority should reset")
    require("<conditionalFormatting" not in third_xml, "plain worksheet should not contain rules")
    return {"first_sheet_priorities": [1, 2], "second_sheet_priorities": [1]}


def verify_openpyxl_basic(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        worksheet = workbook["ColorScale"]
        rules = list(worksheet.conditional_formatting)
        require(len(rules) == 1, f"expected 1 conditional formatting range, got {len(rules)}")
        cf_range = rules[0]
        sqref = str(getattr(cf_range, "sqref", cf_range))
        require(sqref == EXPECTED_BASIC_SQREF,
                f"conditional formatting sqref mismatch: {sqref}")
        rule = worksheet.conditional_formatting[cf_range][0]
        require(rule.type == "colorScale", f"rule type mismatch: {rule.type!r}")
        require(rule.priority == 1, f"rule priority mismatch: {rule.priority!r}")
        require(len(rule.colorScale.cfvo) == 2, "openpyxl cfvo count mismatch")
        require(rule.colorScale.cfvo[0].type == "min", "first cfvo type mismatch")
        require(rule.colorScale.cfvo[1].type == "max", "second cfvo type mismatch")
        colors = [color.rgb for color in rule.colorScale.color]
        require(colors == [EXPECTED_LOW_COLOR, EXPECTED_HIGH_COLOR],
                f"openpyxl color mismatch: {colors!r}")
        return {
            "status": "opened",
            "sqref": sqref,
            "priority": rule.priority,
            "colors": colors,
        }
    finally:
        workbook.close()


def verify_openpyxl_multi_range(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        worksheet = workbook["ColorScaleRanges"]
        ranges = [
            str(getattr(cf_range, "sqref", cf_range))
            for cf_range in worksheet.conditional_formatting
        ]
        require(ranges == [EXPECTED_MULTI_RANGE_SQREF],
                f"multi-range conditional formatting sqref mismatch: {ranges!r}")
        cf_range = next(iter(worksheet.conditional_formatting))
        rule = worksheet.conditional_formatting[cf_range][0]
        require(rule.type == "colorScale", f"multi-range rule type mismatch: {rule.type!r}")
        require(rule.priority == 1, f"multi-range rule priority mismatch: {rule.priority!r}")
        colors = [color.rgb for color in rule.colorScale.color]
        require(colors == [EXPECTED_LOW_COLOR, EXPECTED_HIGH_COLOR],
                f"multi-range color mismatch: {colors!r}")
        return {"status": "opened", "sqref": ranges[0], "colors": colors}
    finally:
        workbook.close()


def create_xlsxwriter_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(str(path))
    try:
        worksheet = workbook.add_worksheet("ColorScale")
        worksheet.write(0, 0, "Score")
        for index, value in enumerate(range(1, 10), start=1):
            worksheet.write_number(index, 0, value)
        worksheet.conditional_format(
            "A2:A10",
            {
                "type": "2_color_scale",
                "min_type": "min",
                "max_type": "max",
                "min_color": "#FF0000",
                "max_color": "#00B050",
            },
        )
    finally:
        workbook.close()

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require("<conditionalFormatting" in worksheet_xml,
            "XlsxWriter reference missing conditionalFormatting")
    return {"status": "created", "path": str(path)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path(
            "build/windows-nmake-release/tests/"
            "fastxlsx-streaming-conditional-formatting-two-color-scale.xlsx"
        ),
        help="FastXLSX two-color scale workbook to verify.",
    )
    parser.add_argument(
        "--metadata-order-input",
        type=Path,
        default=Path(
            "build/windows-nmake-release/tests/"
            "fastxlsx-streaming-conditional-formatting-metadata-order.xlsx"
        ),
        help="FastXLSX conditional formatting + relationship-backed metadata workbook.",
    )
    parser.add_argument(
        "--multi-range-input",
        type=Path,
        default=Path(
            "build/windows-nmake-release/tests/"
            "fastxlsx-streaming-conditional-formatting-multi-range.xlsx"
        ),
        help="FastXLSX multi-range conditional-formatting color scale workbook.",
    )
    parser.add_argument(
        "--priorities-input",
        type=Path,
        default=Path(
            "build/windows-nmake-release/tests/"
            "fastxlsx-streaming-conditional-formatting-priorities.xlsx"
        ),
        help="FastXLSX workbook that verifies per-worksheet priorities.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/conditional-formatting-color-scales"),
        help="Directory for local QA reports and reference workbooks.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    metadata_order_path = args.metadata_order_input.resolve()
    multi_range_path = args.multi_range_input.resolve()
    priorities_path = args.priorities_input.resolve()
    work_dir = args.work_dir.resolve()
    for path in [input_path, metadata_order_path, multi_range_path, priorities_path]:
        require(path.exists(), f"input workbook does not exist: {path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    report = {
        "fastxlsx_input": str(input_path),
        "fastxlsx_package": verify_basic_package(input_path),
        "metadata_order_input": str(metadata_order_path),
        "metadata_order": verify_metadata_order_package(metadata_order_path),
        "multi_range_input": str(multi_range_path),
        "multi_range": verify_multi_range_package(multi_range_path),
        "priorities_input": str(priorities_path),
        "priorities": verify_priorities_package(priorities_path),
        "xlsx_libraries": {
            "openpyxl": verify_openpyxl_basic(input_path),
            "openpyxl_multi_range": verify_openpyxl_multi_range(multi_range_path),
            "xlsxwriter": create_xlsxwriter_reference(
                work_dir / "reference-xlsxwriter-conditional-formatting-color-scale.xlsx"
            ),
        },
    }

    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # noqa: BLE001
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
