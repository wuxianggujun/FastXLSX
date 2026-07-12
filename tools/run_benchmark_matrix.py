#!/usr/bin/env python3
"""Run an opt-in FastXLSX streaming writer benchmark matrix.

This helper is intentionally outside CTest and CI. It wraps the existing
fastxlsx_bench_streaming_writer executable, collects its schema-v5 JSON results,
and optionally verifies generated workbooks with openpyxl as local QA only.
"""

from __future__ import annotations

import argparse
import json
import math
import platform
import statistics
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any


DEFAULT_CASES = [
    "numeric:inline:mixed",
    "mixed:inline:mixed",
    "mixed:shared:mixed",
    "strings:inline:repeated",
    "strings:shared:repeated",
    "strings:inline:unique",
    "strings:shared:unique",
]

BENCHMARK_SCHEMA_VERSION = "5"
DEFAULT_ZIP_COMPRESSION_LEVEL = -1
MIN_ZIP_COMPRESSION_LEVEL = 0
MAX_ZIP_COMPRESSION_LEVEL = 9


def default_benchmark_exe() -> Path:
    suffix = ".exe" if platform.system() == "Windows" else ""
    return Path("build/windows-nmake-release-benchmark/benchmarks") / f"fastxlsx_bench_streaming_writer{suffix}"


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


@dataclass(frozen=True)
class MatrixCase:
    scenario: str
    string_strategy: str
    string_pattern: str

    @property
    def name(self) -> str:
        if self.scenario == "numeric":
            return "numeric-inline"
        return f"{self.scenario}-{self.string_pattern}-{self.string_strategy}"


def parse_case(text: str) -> MatrixCase:
    parts = text.split(":")
    if len(parts) != 3:
        raise ValueError(f"case must use scenario:string_strategy:string_pattern format: {text}")
    scenario, string_strategy, string_pattern = parts
    if scenario not in {"numeric", "mixed", "strings"}:
        raise ValueError(f"unsupported scenario in case {text!r}")
    if string_strategy not in {"inline", "shared"}:
        raise ValueError(f"unsupported string strategy in case {text!r}")
    if string_pattern not in {"mixed", "repeated", "unique"}:
        raise ValueError(f"unsupported string pattern in case {text!r}")
    if scenario == "numeric" and string_strategy != "inline":
        raise ValueError("numeric cases should use inline strategy to avoid duplicate no-string runs")
    return MatrixCase(scenario=scenario, string_strategy=string_strategy, string_pattern=string_pattern)


def parse_compression_level(text: str) -> int:
    try:
        level = int(text, 10)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("compression level must be -1 or between 0 and 9") from exc
    if level != DEFAULT_ZIP_COMPRESSION_LEVEL and not (
        MIN_ZIP_COMPRESSION_LEVEL <= level <= MAX_ZIP_COMPRESSION_LEVEL
    ):
        raise argparse.ArgumentTypeError("compression level must be -1 or between 0 and 9")
    return level


def compression_level_label(level: int) -> str:
    if level == DEFAULT_ZIP_COMPRESSION_LEVEL:
        return "default"
    return f"level-{level}"


def case_run_name(case: MatrixCase, compression_level: int | None) -> str:
    if compression_level is None:
        return case.name
    return f"{case.name}-compression-{compression_level_label(compression_level)}"


def indexed_run_name(base_name: str, run_kind: str, run_index: int, run_count: int) -> str:
    require(run_index >= 1, "run index must be positive")
    require(run_count >= 1, "run count must be positive")
    width = max(2, len(str(run_count)))
    return f"{base_name}-{run_kind}-{run_index:0{width}d}"


def metric_summary(values: list[int | float]) -> dict[str, int | float]:
    require(bool(values), "metric summary requires at least one value")
    return {
        "min": min(values),
        "median": statistics.median(values),
        "max": max(values),
    }


def measured_run_summary(reports: list[dict[str, Any]]) -> tuple[int, dict[str, Any]]:
    require(bool(reports), "measured run summary requires at least one report")
    elapsed_values = [int(report["result"]["elapsed_ms"]) for report in reports]
    elapsed_median = statistics.median(elapsed_values)
    representative_index = min(
        range(len(reports)),
        key=lambda index: (abs(elapsed_values[index] - elapsed_median), index),
    )
    statistics_report = {
        "elapsed_ms": metric_summary(elapsed_values),
        "generation_ms": metric_summary(
            [int(report["result"]["generation_ms"]) for report in reports]
        ),
        "package_close_ms": metric_summary(
            [int(report["result"]["package_close_ms"]) for report in reports]
        ),
        "peak_memory_mb": metric_summary(
            [float(report["result"]["peak_memory_mb"]) for report in reports]
        ),
        "output_bytes": metric_summary(
            [int(report["result"]["output_bytes"]) for report in reports]
        ),
        "temporary_worksheet_part_footprint_bytes": metric_summary([
            int(report["result"]["temporary_worksheet_part_footprint_bytes"])
            for report in reports
        ]),
        "worksheet_body_buffer_peak_bytes": metric_summary([
            int(report["result"]["worksheet_body_buffer_peak_bytes"])
            for report in reports
        ]),
        "worksheet_body_flush_count": metric_summary([
            int(report["result"]["worksheet_body_flush_count"])
            for report in reports
        ]),
    }
    return representative_index, statistics_report


def expected_string_ratio(case: MatrixCase, mixed_string_ratio: float) -> float:
    if case.scenario == "numeric":
        return 0.0
    if case.scenario == "strings":
        return 1.0
    return mixed_string_ratio


def should_write_string(case: MatrixCase, mixed_string_ratio: float, row: int, col: int) -> bool:
    ratio = expected_string_ratio(case, mixed_string_ratio)
    if case.scenario == "numeric" or ratio <= 0.0:
        return False
    if case.scenario == "strings" or ratio >= 1.0:
        return True
    raw_bucket = 1.0 / ratio
    if not math.isfinite(raw_bucket) or raw_bucket >= 2**32 - 1:
        return False
    bucket = max(int(raw_bucket), 1)
    return ((row + col) % bucket) == 0


def make_string_value(case: MatrixCase, sheet: int, row: int, col: int) -> str:
    if string_value_is_repeated(case, row, col):
        return "repeat"
    return f"s{sheet}-r{row}-c{col}"


def string_value_is_repeated(case: MatrixCase, row: int, col: int) -> bool:
    if case.string_pattern == "repeated":
        return True
    if case.string_pattern == "unique":
        return False
    return (row + col) % 5 == 0


def make_number_value(sheet: int, row: int, col: int) -> float:
    return float(sheet) * 1_000_000.0 + float(row) * 1_000.0 + float(col)


def expected_cell_value(case: MatrixCase, mixed_string_ratio: float, row: int, col: int) -> int | str:
    if should_write_string(case, mixed_string_ratio, row, col):
        return make_string_value(case, 1, row, col)
    value = make_number_value(1, row, col)
    return int(value) if value.is_integer() else value


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def expected_string_distribution(
    case: MatrixCase, rows: int, cols: int, sheets: int, mixed_string_ratio: float
) -> dict[str, float | int]:
    per_sheet_string_cells = 0
    per_sheet_repeated_string_cells = 0
    per_sheet_non_repeated_unique_string_cells = 0
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            if not should_write_string(case, mixed_string_ratio, row, col):
                continue
            per_sheet_string_cells += 1
            if string_value_is_repeated(case, row, col):
                per_sheet_repeated_string_cells += 1
            else:
                per_sheet_non_repeated_unique_string_cells += 1

    string_cells = per_sheet_string_cells * sheets
    repeated_string_cells = per_sheet_repeated_string_cells * sheets
    unique_string_values = per_sheet_non_repeated_unique_string_cells * sheets
    if repeated_string_cells > 0:
        unique_string_values += 1
    duplicate_string_cells = string_cells - unique_string_values
    string_dedup_ratio = duplicate_string_cells / string_cells if string_cells else 0.0
    return {
        "string_cells": string_cells,
        "unique_string_values": unique_string_values,
        "duplicate_string_cells": duplicate_string_cells,
        "string_dedup_ratio": string_dedup_ratio,
    }


def verify_result_json(path: Path, case: MatrixCase, rows: int, cols: int, sheets: int,
    mixed_string_ratio: float, output_path: Path, compression_level: int | None) -> dict[str, Any]:
    data = load_json(path)
    require(data.get("benchmark_schema_version") == BENCHMARK_SCHEMA_VERSION,
        "benchmark schema version mismatch")
    require(data.get("scenario") == case.scenario, f"{case.name} scenario mismatch")
    require(data.get("rows") == rows, f"{case.name} rows mismatch")
    require(data.get("cols") == cols, f"{case.name} cols mismatch")
    require(data.get("sheets") == sheets, f"{case.name} sheets mismatch")
    require(data.get("cells") == rows * cols * sheets, f"{case.name} cell count mismatch")
    require(data.get("string_pattern") == case.string_pattern, f"{case.name} string pattern mismatch")
    require(data.get("string_strategy") == case.string_strategy, f"{case.name} string strategy mismatch")
    if compression_level is not None:
        require(data.get("compression_level") == compression_level,
            f"{case.name} compression level mismatch")
    require(
        abs(float(data.get("string_ratio")) - expected_string_ratio(case, mixed_string_ratio)) < 1e-9,
        f"{case.name} string ratio mismatch",
    )
    expected_strings = expected_string_distribution(case, rows, cols, sheets, mixed_string_ratio)
    require(int(data.get("string_cells")) == expected_strings["string_cells"],
        f"{case.name} string cell count mismatch")
    require(int(data.get("unique_string_values")) == expected_strings["unique_string_values"],
        f"{case.name} unique string value count mismatch")
    require(int(data.get("duplicate_string_cells")) == expected_strings["duplicate_string_cells"],
        f"{case.name} duplicate string cell count mismatch")
    require(
        abs(float(data.get("string_dedup_ratio")) - float(expected_strings["string_dedup_ratio"])) < 1e-6,
        f"{case.name} string dedup ratio mismatch",
    )
    require(data.get("package_entry_source_mode") == "worksheet-file-backed-chunked",
        f"{case.name} package entry source mode mismatch")
    require(data.get("temporary_worksheet_part_footprint") == "worksheet-body-file-bytes",
        f"{case.name} temporary footprint label mismatch")
    require(int(data.get("temporary_worksheet_part_footprint_bytes")) > 0,
        f"{case.name} temporary footprint bytes should be positive")
    require(int(data.get("elapsed_ms")) >= 0, f"{case.name} elapsed_ms mismatch")
    require(int(data.get("generation_ms")) >= 0, f"{case.name} generation_ms mismatch")
    require(int(data.get("package_close_ms")) >= 0, f"{case.name} package_close_ms mismatch")
    require(int(data.get("worksheet_body_buffer_limit_bytes")) == 256 * 1024,
        f"{case.name} body buffer limit mismatch")
    require(0 < int(data.get("worksheet_body_buffer_peak_bytes")) <= 256 * 1024,
        f"{case.name} body buffer peak mismatch")
    require(int(data.get("worksheet_body_flush_count")) > 0,
        f"{case.name} body flush count mismatch")
    require(int(data.get("active_worksheet_temporary_files_after_close")) == 0,
        f"{case.name} close should release worksheet temporary files")
    require(int(data.get("output_bytes")) == output_path.stat().st_size,
        f"{case.name} output size mismatch")
    require(data.get("office_open") == "not_run", f"{case.name} office_open should remain not_run")
    return data


def verify_workbook_with_openpyxl(path: Path, case: MatrixCase, rows: int, cols: int, sheets: int,
    mixed_string_ratio: float) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(len(workbook.sheetnames) == sheets, f"{case.name} sheet count mismatch")
        require("Sheet1" in workbook.sheetnames, f"{case.name} missing Sheet1")
        worksheet = workbook["Sheet1"]
        require(worksheet.max_row == rows, f"{case.name} row count mismatch: {worksheet.max_row}")
        require(worksheet.max_column == cols, f"{case.name} column count mismatch: {worksheet.max_column}")
        first_expected = expected_cell_value(case, mixed_string_ratio, 1, 1)
        last_expected = expected_cell_value(case, mixed_string_ratio, rows, cols)
        first_value = worksheet.cell(row=1, column=1).value
        last_value = worksheet.cell(row=rows, column=cols).value
        require(first_value == first_expected,
            f"{case.name} A1 mismatch: expected {first_expected!r}, got {first_value!r}")
        require(last_value == last_expected,
            f"{case.name} last cell mismatch: expected {last_expected!r}, got {last_value!r}")
        return {
            "status": "opened",
            "sheet_count": len(workbook.sheetnames),
            "sheet1_max_row": worksheet.max_row,
            "sheet1_max_column": worksheet.max_column,
            "sheet1_first_cell": first_value,
            "sheet1_last_cell": last_value,
        }
    finally:
        workbook.close()


def build_matrix_report(bench_exe: Path, output_dir: Path, rows: int, cols: int, sheets: int,
    mixed_string_ratio: float, compression_levels: list[int | None],
    warmup_runs: int, measured_runs: int, reports: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "benchmark_matrix_schema_version": "1",
        "benchmark_executable": str(bench_exe),
        "output_dir": str(output_dir),
        "rows": rows,
        "cols": cols,
        "sheets": sheets,
        "cells_per_case": rows * cols * sheets,
        "mixed_string_ratio": mixed_string_ratio,
        "compression_levels": compression_levels,
        "warmup_runs_per_case": warmup_runs,
        "measured_runs_per_case": measured_runs,
        "representative_result_policy": "measured run nearest elapsed_ms median; earliest run breaks ties",
        "cases": reports,
        "comparison_scope": (
            "Manual opt-in repeated benchmark matrix. Each case retains every measured "
            "schema-v5 JSON result and reports min/median/max statistics. Optional openpyxl "
            "validation checks the representative workbook only; Office validation is a "
            "separate local step and benchmark office_open fields remain not_run."
        ),
    }


def run_single_case(bench_exe: Path, output_dir: Path, case: MatrixCase, rows: int, cols: int,
    sheets: int, mixed_string_ratio: float, compression_level: int | None,
    run_name: str) -> dict[str, Any]:
    output_path = output_dir / f"fastxlsx-bench-{run_name}.xlsx"
    result_path = output_dir / f"fastxlsx-bench-{run_name}.json"
    command = [
        str(bench_exe),
        "--scenario",
        case.scenario,
        "--rows",
        str(rows),
        "--cols",
        str(cols),
        "--sheets",
        str(sheets),
        "--string-pattern",
        case.string_pattern,
        "--string-strategy",
        case.string_strategy,
        "--string-ratio",
        str(mixed_string_ratio),
        "--output",
        str(output_path),
        "--result",
        str(result_path),
    ]
    if compression_level is not None:
        command.extend(["--compression-level", str(compression_level)])
    completed = subprocess.run(command, check=False, text=True, capture_output=True)
    require(completed.returncode == 0,
        f"{case.name} benchmark failed with {completed.returncode}: {completed.stderr.strip()}")
    require(output_path.exists(), f"{case.name} output workbook was not created")
    require(result_path.exists(), f"{case.name} result JSON was not created")

    result_json = verify_result_json(
        result_path, case, rows, cols, sheets, mixed_string_ratio, output_path,
        compression_level)
    return {
        "name": run_name,
        "scenario": case.scenario,
        "string_strategy": case.string_strategy,
        "string_pattern": case.string_pattern,
        "compression_level": compression_level,
        "command": command,
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
        "output": str(output_path),
        "result_json": str(result_path),
        "expected": {
            "sheet1_first_cell": expected_cell_value(case, mixed_string_ratio, 1, 1),
            "sheet1_last_cell": expected_cell_value(case, mixed_string_ratio, rows, cols),
            "string_distribution": expected_string_distribution(
                case, rows, cols, sheets, mixed_string_ratio),
        },
        "result": result_json,
        "openpyxl": {"status": "not_requested"},
    }


def run_case(bench_exe: Path, output_dir: Path, case: MatrixCase, rows: int, cols: int,
    sheets: int, mixed_string_ratio: float, verify_openpyxl: bool,
    compression_level: int | None, warmup_runs: int, measured_runs: int) -> dict[str, Any]:
    base_name = case_run_name(case, compression_level)
    warmup_dir = output_dir / "warmup"
    warmup_dir.mkdir(parents=True, exist_ok=True)
    for run_index in range(1, warmup_runs + 1):
        run_single_case(
            bench_exe, warmup_dir, case, rows, cols, sheets, mixed_string_ratio,
            compression_level,
            indexed_run_name(base_name, "warmup", run_index, warmup_runs),
        )

    measured_reports = [
        run_single_case(
            bench_exe, output_dir, case, rows, cols, sheets, mixed_string_ratio,
            compression_level,
            indexed_run_name(base_name, "run", run_index, measured_runs),
        )
        for run_index in range(1, measured_runs + 1)
    ]
    representative_index, statistics_report = measured_run_summary(measured_reports)
    representative = measured_reports[representative_index]
    representative_output = Path(representative["output"])
    openpyxl_report = verify_workbook_with_openpyxl(
        representative_output, case, rows, cols, sheets, mixed_string_ratio
    ) if verify_openpyxl else {"status": "not_requested"}

    return {
        "name": base_name,
        "scenario": case.scenario,
        "string_strategy": case.string_strategy,
        "string_pattern": case.string_pattern,
        "compression_level": compression_level,
        "warmup_runs": warmup_runs,
        "measured_runs": measured_runs,
        "representative_run": representative_index + 1,
        "statistics": statistics_report,
        "command": representative["command"],
        "stdout": representative["stdout"],
        "stderr": representative["stderr"],
        "output": representative["output"],
        "result_json": representative["result_json"],
        "expected": representative["expected"],
        "result": representative["result"],
        "openpyxl": openpyxl_report,
        "runs": measured_reports,
    }


def expect_value_error(text: str) -> None:
    try:
        parse_case(text)
    except ValueError:
        return
    raise AssertionError(f"expected parse_case({text!r}) to fail")


def run_self_test() -> None:
    numeric = parse_case("numeric:inline:mixed")
    mixed = parse_case("mixed:shared:mixed")
    repeated = parse_case("strings:shared:repeated")
    unique = parse_case("strings:inline:unique")
    require(numeric.name == "numeric-inline", "numeric case name mismatch")
    require(mixed.name == "mixed-mixed-shared", "mixed case name mismatch")
    require(repeated.name == "strings-repeated-shared", "repeated case name mismatch")
    require(unique.name == "strings-unique-inline", "unique case name mismatch")
    require(parse_compression_level("-1") == -1, "compression default parse mismatch")
    require(parse_compression_level("0") == 0, "compression level 0 parse mismatch")
    require(parse_compression_level("9") == 9, "compression level 9 parse mismatch")
    require(case_run_name(repeated, None) == "strings-repeated-shared",
        "default case run name mismatch")
    require(case_run_name(repeated, -1) == "strings-repeated-shared-compression-default",
        "compression default case run name mismatch")
    require(case_run_name(repeated, 3) == "strings-repeated-shared-compression-level-3",
        "compression level case run name mismatch")

    expect_value_error("numeric:shared:mixed")
    expect_value_error("strings:inline")
    expect_value_error("other:inline:mixed")
    expect_value_error("strings:other:mixed")
    expect_value_error("strings:inline:other")

    repeated_distribution = expected_string_distribution(repeated, rows=2, cols=3, sheets=2,
        mixed_string_ratio=0.25)
    require(repeated_distribution["string_cells"] == 12, "repeated string cell count mismatch")
    require(repeated_distribution["unique_string_values"] == 1,
        "repeated unique string value count mismatch")
    require(repeated_distribution["duplicate_string_cells"] == 11,
        "repeated duplicate string count mismatch")
    require(abs(float(repeated_distribution["string_dedup_ratio"]) - (11.0 / 12.0)) < 1e-9,
        "repeated dedup ratio mismatch")

    unique_distribution = expected_string_distribution(unique, rows=2, cols=3, sheets=2,
        mixed_string_ratio=0.25)
    require(unique_distribution["string_cells"] == 12, "unique string cell count mismatch")
    require(unique_distribution["unique_string_values"] == 12,
        "unique value count mismatch")
    require(unique_distribution["duplicate_string_cells"] == 0,
        "unique duplicate string count mismatch")
    require(unique_distribution["string_dedup_ratio"] == 0.0, "unique dedup ratio mismatch")

    numeric_distribution = expected_string_distribution(numeric, rows=5, cols=4, sheets=3,
        mixed_string_ratio=1.0)
    require(numeric_distribution["string_cells"] == 0, "numeric string cell count mismatch")
    require(expected_cell_value(numeric, 0.25, 2, 3) == 1002003,
        "numeric expected cell mismatch")
    require(expected_cell_value(repeated, 0.25, 2, 3) == "repeat",
        "repeated expected cell mismatch")
    require(expected_cell_value(unique, 0.25, 2, 3) == "s1-r2-c3",
        "unique expected cell mismatch")
    require(expected_cell_value(parse_case("mixed:inline:mixed"), 1.0, 4, 1) == "repeat",
        "mixed repeated expected cell mismatch")

    measured_reports = [
        {"result": {
            "elapsed_ms": elapsed_ms,
            "generation_ms": elapsed_ms // 2,
            "package_close_ms": elapsed_ms - elapsed_ms // 2,
            "peak_memory_mb": peak_memory_mb,
            "output_bytes": 100,
            "temporary_worksheet_part_footprint_bytes": 200,
            "worksheet_body_buffer_peak_bytes": 128,
            "worksheet_body_flush_count": 2,
        }}
        for elapsed_ms, peak_memory_mb in [(30, 7.0), (10, 5.0), (20, 6.0)]
    ]
    representative_index, statistics_report = measured_run_summary(measured_reports)
    require(representative_index == 2, "median representative run mismatch")
    require(statistics_report["elapsed_ms"] == {"min": 10, "median": 20, "max": 30},
        "elapsed statistics mismatch")
    require(statistics_report["peak_memory_mb"] == {"min": 5.0, "median": 6.0, "max": 7.0},
        "memory statistics mismatch")
    require(indexed_run_name("numeric-inline", "run", 2, 3) == "numeric-inline-run-02",
        "indexed run name mismatch")

    report = build_matrix_report(Path("bench"), Path("out"), 2, 3, 2, 0.25, [-1, 3], 1, 3, [{
        "name": repeated.name,
        "expected": {
            "sheet1_first_cell": "repeat",
            "sheet1_last_cell": "repeat",
            "string_distribution": repeated_distribution,
        },
    }])
    require(report["benchmark_matrix_schema_version"] == "1", "matrix schema mismatch")
    require(report["cells_per_case"] == 12, "matrix cell count mismatch")
    require(report["compression_levels"] == [-1, 3], "matrix compression levels mismatch")
    require(report["warmup_runs_per_case"] == 1, "matrix warmup count mismatch")
    require(report["measured_runs_per_case"] == 3, "matrix measured count mismatch")
    require(report["cases"][0]["name"] == "strings-repeated-shared", "matrix case mismatch")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--self-test", action="store_true",
        help="Run lightweight internal checks without invoking the benchmark executable.")
    parser.add_argument("--bench-exe", type=Path, default=default_benchmark_exe(),
        help="Path to fastxlsx_bench_streaming_writer.")
    parser.add_argument("--output-dir", type=Path, default=Path("build/qa/benchmark-matrix"),
        help="Directory for generated workbooks, per-case JSON, and the matrix report.")
    parser.add_argument("--rows", type=int, default=1000, help="Rows per worksheet.")
    parser.add_argument("--cols", type=int, default=10, help="Columns per worksheet.")
    parser.add_argument("--sheets", type=int, default=1, help="Worksheet count.")
    parser.add_argument("--mixed-string-ratio", type=float, default=0.25,
        help="String ratio passed to mixed scenario cases.")
    parser.add_argument("--compression-level", action="append", dest="compression_levels",
        type=parse_compression_level,
        help="ZIP compression level to pass to the benchmark. May be passed multiple times.")
    parser.add_argument("--warmup-runs", type=int, default=1,
        help="Untimed process runs per case before measured runs. Default: 1.")
    parser.add_argument("--measured-runs", "--repeat", type=int, default=3,
        help="Measured process runs retained per case. Default: 3.")
    parser.add_argument("--case", action="append", dest="cases",
        help="Case in scenario:string_strategy:string_pattern format. May be passed multiple times.")
    parser.add_argument("--verify-openpyxl", action="store_true",
        help="Open generated workbooks with openpyxl and verify Sheet1 first/last cells.")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        print("OK: run_benchmark_matrix.py self-test passed")
        return 0

    bench_exe = args.bench_exe.resolve()
    output_dir = args.output_dir.resolve()
    require(bench_exe.exists(), f"benchmark executable does not exist: {bench_exe}")
    require(args.rows > 0 and args.cols > 0 and args.sheets > 0, "rows, cols, and sheets must be positive")
    require(0.0 <= args.mixed_string_ratio <= 1.0, "mixed string ratio must be between 0 and 1")
    require(args.warmup_runs >= 0, "warmup runs must not be negative")
    require(args.measured_runs > 0, "measured runs must be positive")
    output_dir.mkdir(parents=True, exist_ok=True)

    raw_cases = args.cases if args.cases else DEFAULT_CASES
    cases = [parse_case(raw_case) for raw_case in raw_cases]
    compression_levels = args.compression_levels if args.compression_levels else [None]
    names = [case_run_name(case, compression_level)
        for compression_level in compression_levels
        for case in cases]
    require(len(names) == len(set(names)), "benchmark case names must be unique")

    reports = [
        run_case(bench_exe, output_dir, case, args.rows, args.cols, args.sheets,
            args.mixed_string_ratio, args.verify_openpyxl, compression_level,
            args.warmup_runs, args.measured_runs)
        for compression_level in compression_levels
        for case in cases
    ]
    report = build_matrix_report(
        bench_exe, output_dir, args.rows, args.cols, args.sheets, args.mixed_string_ratio,
        compression_levels, args.warmup_runs, args.measured_runs, reports)
    report_path = output_dir / "benchmark-matrix-report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI should emit one concise failure line.
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
