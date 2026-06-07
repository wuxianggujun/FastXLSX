#!/usr/bin/env python3
"""Local QA for the sharedStrings artifact-absence edge.

This helper is intentionally outside CTest. It verifies that enabling
StringStrategy::SharedString does not create sharedStrings package artifacts
when no string cells were written. Python XLSX libraries are used only as local
QA/reference tools and never as FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


EXPECTED_VALUES: dict[str, int | bool | str] = {
    "A1": 42,
    "B1": True,
    "C1": "=A1+1",
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for this local QA helper") from exc
    return openpyxl


def verify_fastxlsx_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/core.xml",
        "docProps/app.xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")
    forbidden_entries = [
        "xl/sharedStrings.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/styles.xml",
        "xl/calcChain.xml",
    ]
    for name in forbidden_entries:
        require(name not in names, f"unexpected package entry: {name}")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require("/xl/sharedStrings.xml" not in content_types, "unexpected sharedStrings content type override")
    require(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
        not in content_types,
        "unexpected sharedStrings MIME type",
    )

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require("relationships/sharedStrings" not in workbook_rels, "unexpected sharedStrings relationship")
    require('Target="sharedStrings.xml"' not in workbook_rels, "unexpected sharedStrings relationship target")

    workbook_xml = read_zip_text(path, "xl/workbook.xml")
    require('name="NoStrings"' in workbook_xml, "missing NoStrings worksheet in workbook.xml")
    require(
        '<calcPr calcId="124519" fullCalcOnLoad="1"/>' in workbook_xml,
        "formula workbook recalculation metadata is missing",
    )

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<dimension ref="A1:C1"/>' in worksheet_xml, "sharedStrings absence worksheet dimension mismatch")
    require(' t="s"' not in worksheet_xml, "worksheet unexpectedly references shared string indexes")
    require("inlineStr" not in worksheet_xml, "worksheet unexpectedly used inlineStr")
    expected_fragments = [
        '<c r="A1"><v>42</v></c>',
        '<c r="B1" t="b"><v>1</v></c>',
        '<c r="C1"><f>A1+1</f></c>',
    ]
    for fragment in expected_fragments:
        require(fragment in worksheet_xml, f"missing worksheet fragment: {fragment}")

    return {
        "verified_entries": required,
        "forbidden_entries_absent": forbidden_entries,
        "shared_strings_artifacts_absent": True,
        "absence_checks": [
            "sharedStrings content type override",
            "sharedStrings MIME type",
            "workbook sharedStrings relationship",
            "worksheet t=\"s\" references",
            "worksheet inlineStr cells",
        ],
    }


def verify_values_with_openpyxl(path: Path, sheet_name: str = "NoStrings") -> dict[str, str]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(sheet_name in workbook.sheetnames, f"missing worksheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        observed: dict[str, str] = {}
        for cell, expected in EXPECTED_VALUES.items():
            value = worksheet[cell].value
            require(value == expected, f"{cell} value mismatch: expected {expected!r}, got {value!r}")
            observed[cell] = str(value)
        require(worksheet.max_row == 1, f"worksheet row count mismatch: {worksheet.max_row}")
        require(worksheet.max_column == 3, f"worksheet column count mismatch: {worksheet.max_column}")
        return observed
    finally:
        workbook.close()


def create_openpyxl_reference(path: Path) -> None:
    openpyxl = import_openpyxl()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "NoStrings"
    worksheet["A1"] = 42
    worksheet["B1"] = True
    worksheet["C1"] = "=A1+1"
    workbook.save(path)


def create_xlsxwriter_reference(path: Path) -> str | None:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return None

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("NoStrings")
    worksheet.write_number(0, 0, 42)
    worksheet.write_boolean(0, 1, True)
    worksheet.write_formula(0, 2, "=A1+1")
    workbook.close()
    return "created"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-shared-strings-empty-table.xlsx"),
        help="FastXLSX workbook for the no-string-cell sharedStrings edge.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/shared-strings-absence"),
        help="Directory for reference files and the JSON report.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    package_report = verify_fastxlsx_package(input_path)
    fastxlsx_values = verify_values_with_openpyxl(input_path)

    openpyxl_reference = work_dir / "reference-openpyxl-shared-strings-absence.xlsx"
    create_openpyxl_reference(openpyxl_reference)
    openpyxl_reference_values = verify_values_with_openpyxl(openpyxl_reference)

    xlsxwriter_reference = work_dir / "reference-xlsxwriter-shared-strings-absence.xlsx"
    xlsxwriter_status = create_xlsxwriter_reference(xlsxwriter_reference)
    xlsxwriter_values: dict[str, str] | None = None
    if xlsxwriter_status is not None:
        xlsxwriter_values = verify_values_with_openpyxl(xlsxwriter_reference)

    report = {
        "fastxlsx_input": str(input_path),
        "package": package_report,
        "fastxlsx_values": fastxlsx_values,
        "references": {
            "openpyxl": {
                "status": "created",
                "path": str(openpyxl_reference),
                "values": openpyxl_reference_values,
            },
            "xlsxwriter": {
                "status": "created" if xlsxwriter_status else "skipped",
                "path": str(xlsxwriter_reference) if xlsxwriter_status else None,
                "reason": None if xlsxwriter_status else "Python module xlsxwriter is not installed",
                "values": xlsxwriter_values,
            },
        },
        "comparison_scope": (
            "FastXLSX package absence checks plus semantic value/formula checks; "
            "reference XML byte identity is intentionally not required."
        ),
    }
    report_path = work_dir / "shared-strings-absence-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK: verified FastXLSX sharedStrings absence workbook: {input_path}")
    print(f"OK: created openpyxl reference: {openpyxl_reference}")
    if xlsxwriter_status is None:
        print("SKIP: XlsxWriter reference (Python module xlsxwriter is not installed)")
    else:
        print(f"OK: created XlsxWriter reference: {xlsxwriter_reference}")
    print(f"OK: wrote report: {report_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI should emit one concise failure line.
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
