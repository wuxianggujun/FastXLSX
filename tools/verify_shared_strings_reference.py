#!/usr/bin/env python3
"""Local sharedStrings reference QA for FastXLSX generated workbooks.

This helper is intentionally outside CTest. It uses Python XLSX libraries only
as local QA/reference tools and never as FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


EXPECTED_CELLS: dict[str, str | None] = {
    "A1": "repeat",
    "B1": "space ",
    "C1": "escaped & <tag>",
    "A2": "repeat",
    "B2": "space ",
    "A3": "",
    "B3": " leading",
    "C3": "\tindent",
    "D3": "repeat",
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def count(text: str, fragment: str) -> int:
    return text.count(fragment)


def verify_fastxlsx_package(path: Path) -> list[str]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/core.xml",
        "docProps/app.xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/sharedStrings.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require("/xl/sharedStrings.xml" in content_types, "missing sharedStrings content type override")
    require(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
        in content_types,
        "sharedStrings content type mismatch",
    )

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require("relationships/sharedStrings" in workbook_rels, "missing workbook sharedStrings relationship")
    require('Target="sharedStrings.xml"' in workbook_rels, "sharedStrings relationship target mismatch")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<dimension ref="A1:D3"/>' in worksheet_xml, "sharedStrings worksheet dimension mismatch")
    require(count(worksheet_xml, ' t="s"') == 9, "shared string cell reference count mismatch")
    require("inlineStr" not in worksheet_xml, "sharedStrings sample unexpectedly used inlineStr")
    expected_refs = {
        "A1": 0,
        "B1": 1,
        "C1": 2,
        "A2": 0,
        "B2": 1,
        "A3": 3,
        "B3": 4,
        "C3": 5,
        "D3": 0,
    }
    for cell, index in expected_refs.items():
        fragment = f'<c r="{cell}" t="s"><v>{index}</v></c>'
        require(fragment in worksheet_xml, f"{cell} shared string index mismatch")

    shared_strings_xml = read_zip_text(path, "xl/sharedStrings.xml")
    require("<sst " in shared_strings_xml, "missing sharedStrings root")
    require('count="9"' in shared_strings_xml, "sharedStrings count mismatch")
    require('uniqueCount="6"' in shared_strings_xml, "sharedStrings uniqueCount mismatch")
    require(count(shared_strings_xml, "<si>") == 6, "unique shared string entry count mismatch")
    expected_items = [
        "<si><t>repeat</t></si>",
        '<si><t xml:space="preserve">space </t></si>',
        "<si><t>escaped &amp; &lt;tag&gt;</t></si>",
        "<si><t></t></si>",
        '<si><t xml:space="preserve"> leading</t></si>',
        '<si><t xml:space="preserve">\tindent</t></si>',
    ]
    for fragment in expected_items:
        require(fragment in shared_strings_xml, f"missing shared string item: {fragment}")

    return required


def import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for this local QA helper") from exc
    return openpyxl


def verify_values_with_openpyxl(path: Path, sheet_name: str = "Shared") -> dict[str, str | None]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(sheet_name in workbook.sheetnames, f"missing worksheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        observed: dict[str, str | None] = {}
        for cell, expected in EXPECTED_CELLS.items():
            value = worksheet[cell].value
            if value is None and expected == "":
                normalized = ""
            else:
                normalized = value
            require(normalized == expected, f"{cell} value mismatch: expected {expected!r}, got {value!r}")
            observed[cell] = None if value is None else str(value)
        require(worksheet.max_row == 3, f"worksheet row count mismatch: {worksheet.max_row}")
        require(worksheet.max_column == 4, f"worksheet column count mismatch: {worksheet.max_column}")
        return observed
    finally:
        workbook.close()


def create_openpyxl_reference(path: Path) -> None:
    openpyxl = import_openpyxl()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Shared"
    for cell, value in EXPECTED_CELLS.items():
        worksheet[cell] = value
    workbook.save(path)


def create_xlsxwriter_reference(path: Path) -> str | None:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return None

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("Shared")
    for cell, value in EXPECTED_CELLS.items():
        row = int(cell[1:]) - 1
        col = ord(cell[0]) - ord("A")
        worksheet.write_string(row, col, value or "")
    workbook.close()
    return "created"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-shared-strings.xlsx"),
        help="FastXLSX sharedStrings workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/shared-strings-reference"),
        help="Directory for reference files and the JSON report.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    verified_entries = verify_fastxlsx_package(input_path)
    fastxlsx_values = verify_values_with_openpyxl(input_path)

    openpyxl_reference = work_dir / "reference-openpyxl-shared-strings.xlsx"
    create_openpyxl_reference(openpyxl_reference)
    openpyxl_reference_values = verify_values_with_openpyxl(openpyxl_reference)

    xlsxwriter_reference = work_dir / "reference-xlsxwriter-shared-strings.xlsx"
    xlsxwriter_status = create_xlsxwriter_reference(xlsxwriter_reference)
    xlsxwriter_values: dict[str, str | None] | None = None
    if xlsxwriter_status is not None:
        xlsxwriter_values = verify_values_with_openpyxl(xlsxwriter_reference)

    report = {
        "fastxlsx_input": str(input_path),
        "verified_fastxlsx_entries": verified_entries,
        "fastxlsx_values": fastxlsx_values,
        "references": {
            "openpyxl": {
                "status": "created",
                "path": str(openpyxl_reference),
                "values": openpyxl_reference_values,
            },
            "xlsxwriter": {
                "status": "created" if xlsxwriter_status else "skipped",
                "path": str(xlsxwriter_reference) if xlsxwriter_status else None,
                "reason": None if xlsxwriter_status else "Python module xlsxwriter is not installed",
                "values": xlsxwriter_values,
            },
        },
        "comparison_scope": (
            "Semantic value and key OpenXML package checks only; XML byte-level "
            "identity is intentionally not required."
        ),
    }
    report_path = work_dir / "shared-strings-reference-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK: verified FastXLSX sharedStrings workbook: {input_path}")
    print(f"OK: created openpyxl reference: {openpyxl_reference}")
    if xlsxwriter_status is None:
        print("SKIP: XlsxWriter reference (Python module xlsxwriter is not installed)")
    else:
        print(f"OK: created XlsxWriter reference: {xlsxwriter_reference}")
    print(f"OK: wrote report: {report_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI should emit one concise failure line.
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
