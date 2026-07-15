#!/usr/bin/env python3
"""Run an opt-in repeated FastXLSX public Patch benchmark matrix."""

from __future__ import annotations

import argparse
import json
import platform
import statistics
import subprocess
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as element_tree
from dataclasses import dataclass
from pathlib import Path
from typing import Any


BENCHMARK_SCHEMA_VERSION = "9"
DEFAULT_CASES = [
    "noop-copy:0",
    "document-properties:1",
    "patch-replace:1000",
    "patch-upsert:1000",
]
METRICS = [
    "total_editor_ms",
    "open_ms",
    "mutation_ms",
    "save_ms",
    "peak_memory_mb",
    "output_bytes",
    "copied_uncompressed_bytes",
    "copied_source_compressed_bytes",
    "copied_output_compressed_bytes",
    "raw_compressed_copy_bytes",
    "rewritten_uncompressed_bytes",
    "rewritten_compressed_bytes",
    "single_pass_transform_ms",
    "single_pass_transform_us",
    "single_pass_source_scan_action_us",
    "single_pass_source_parsed_event_count",
    "single_pass_source_callback_event_count",
    "single_pass_source_coalesced_input_event_count",
    "single_pass_source_coalesced_output_event_count",
    "single_pass_source_simple_inline_string_fast_path_count",
    "single_pass_source_simple_inline_string_fast_path_bytes",
    "single_pass_source_simple_inline_string_fallback_count",
    "single_pass_transform_action_callback_count",
    "single_pass_relationship_scan_us",
    "single_pass_relationship_scan_input_call_count",
    "single_pass_relationship_scan_input_bytes",
    "single_pass_relationship_scan_boundary_carry_count",
    "single_pass_relationship_scan_slow_path_tag_count",
    "single_pass_temporary_write_us",
    "single_pass_output_append_call_count",
    "single_pass_output_flush_count",
    "single_pass_output_peak_buffer_bytes",
    "package_writer_total_us",
    "package_writer_total_process_cpu_us",
    "target_worksheet_entry_total_us",
    "target_worksheet_entry_total_process_cpu_us",
    "target_worksheet_entry_input_read_us",
    "target_worksheet_entry_writer_write_us",
    "target_worksheet_entry_writer_write_process_cpu_us",
    "target_worksheet_entry_staged_crc_validation_us",
    "target_worksheet_entry_close_us",
    "target_worksheet_entry_close_process_cpu_us",
    "target_worksheet_entry_deflate_writer_process_cpu_us",
]


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def default_benchmark_exe() -> Path:
    suffix = ".exe" if platform.system() == "Windows" else ""
    return (
        Path("build/windows-nmake-release-benchmark/benchmarks")
        / f"fastxlsx_bench_workbook_editor{suffix}"
    )


@dataclass(frozen=True)
class PatchCase:
    scenario: str
    edits: int

    @property
    def name(self) -> str:
        return self.scenario


def parse_case(text: str) -> PatchCase:
    scenario, separator, edits_text = text.partition(":")
    if not separator or not edits_text:
        raise ValueError(f"case must use scenario:edits format: {text}")
    if scenario not in {
        "noop-copy",
        "document-properties",
        "patch-replace",
        "patch-upsert",
    }:
        raise ValueError(f"unsupported Patch scenario: {scenario}")
    try:
        edits = int(edits_text, 10)
    except ValueError as exc:
        raise ValueError(f"case edits must be a non-negative integer: {text}") from exc
    if edits < 0:
        raise ValueError(f"case edits must be a non-negative integer: {text}")
    if scenario == "noop-copy" and edits != 0:
        raise ValueError("noop-copy requires zero edits")
    if scenario != "noop-copy" and edits == 0:
        raise ValueError(f"{scenario} requires at least one edit")
    return PatchCase(scenario=scenario, edits=edits)


def indexed_run_name(base_name: str, run_kind: str, run_index: int, run_count: int) -> str:
    require(run_index >= 1, "run index must be positive")
    require(run_count >= 1, "run count must be positive")
    width = max(2, len(str(run_count)))
    return f"{base_name}-{run_kind}-{run_index:0{width}d}"


def compression_case_name(case: PatchCase, compression_level: int) -> str:
    level = "default" if compression_level == -1 else str(compression_level)
    return f"{case.name}-compression-{level}"


def metric_summary(values: list[int | float]) -> dict[str, int | float]:
    require(bool(values), "metric summary requires at least one value")
    return {
        "min": min(values),
        "median": statistics.median(values),
        "max": max(values),
    }


def measured_run_summary(reports: list[dict[str, Any]]) -> tuple[int, dict[str, Any]]:
    require(bool(reports), "measured run summary requires at least one report")
    elapsed = [int(report["result"]["total_editor_ms"]) for report in reports]
    elapsed_median = statistics.median(elapsed)
    representative_index = min(
        range(len(reports)),
        key=lambda index: (abs(elapsed[index] - elapsed_median), index),
    )
    statistics_report: dict[str, Any] = {}
    for metric in METRICS:
        values = []
        for report in reports:
            source = report["result"]
            if metric in report["byte_accounting"]:
                source = report["byte_accounting"]
            values.append(source[metric])
        statistics_report[metric] = metric_summary(values)
    return representative_index, statistics_report


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def account_plan_bytes(
    source_path: Path, output_path: Path, result: dict[str, Any]
) -> dict[str, int]:
    copied_names = list(result["copied_entry_names"])
    rewritten_names = list(result["rewritten_entry_names"])
    omitted_names = list(result["omitted_entry_names"])
    require(
        int(result["output_plan_entry_count"])
        == len(copied_names) + len(rewritten_names) + len(omitted_names),
        "output plan entry count mismatch",
    )

    with zipfile.ZipFile(source_path) as source, zipfile.ZipFile(output_path) as output:
        source_info = {entry.filename: entry for entry in source.infolist()}
        output_info = {entry.filename: entry for entry in output.infolist()}

        raw_copy_names = set(result.get("raw_compressed_copy_entry_names", []))
        require(
            len(raw_copy_names) == int(result.get("raw_compressed_copy_entry_count", 0)),
            "raw compressed-copy entry count mismatch",
        )
        require(raw_copy_names.issubset(set(copied_names)),
            "raw compressed-copy entries must be copy-original entries")

        copied_uncompressed = 0
        copied_source_compressed = 0
        copied_output_compressed = 0
        raw_compressed_copy_bytes = 0
        for name in copied_names:
            require(name in source_info, f"copied source entry missing: {name}")
            require(name in output_info, f"copied output entry missing: {name}")
            require(
                source_info[name].file_size == output_info[name].file_size
                and source_info[name].CRC == output_info[name].CRC,
                f"copy-original entry changed: {name}",
            )
            copied_uncompressed += source_info[name].file_size
            copied_source_compressed += source_info[name].compress_size
            copied_output_compressed += output_info[name].compress_size
            if name in raw_copy_names:
                require(
                    raw_entry_payload(source_path, source_info[name])
                    == raw_entry_payload(output_path, output_info[name]),
                    f"raw compressed-copy entry bytes changed: {name}",
                )
                raw_compressed_copy_bytes += source_info[name].compress_size

        rewritten_uncompressed = 0
        rewritten_compressed = 0
        for name in rewritten_names:
            require(name in output_info, f"rewritten output entry missing: {name}")
            rewritten_uncompressed += output_info[name].file_size
            rewritten_compressed += output_info[name].compress_size

        for name in omitted_names:
            require(name not in output_info, f"omitted entry still exists: {name}")

    require(raw_compressed_copy_bytes == int(result.get("raw_compressed_copy_bytes", 0)),
        "raw compressed-copy byte count mismatch")
    return {
        "copied_uncompressed_bytes": copied_uncompressed,
        "copied_source_compressed_bytes": copied_source_compressed,
        "copied_output_compressed_bytes": copied_output_compressed,
        "raw_compressed_copy_bytes": raw_compressed_copy_bytes,
        "rewritten_uncompressed_bytes": rewritten_uncompressed,
        "rewritten_compressed_bytes": rewritten_compressed,
    }


def raw_entry_payload(path: Path, entry: zipfile.ZipInfo) -> bytes:
    with path.open("rb") as stream:
        stream.seek(entry.header_offset)
        header = stream.read(30)
        require(len(header) == 30 and header[:4] == b"PK\x03\x04",
            f"invalid local ZIP header for raw payload: {entry.filename}")
        name_size = int.from_bytes(header[26:28], "little")
        extra_size = int.from_bytes(header[28:30], "little")
        stream.seek(name_size + extra_size, 1)
        payload = stream.read(entry.compress_size)
    require(len(payload) == entry.compress_size,
        f"truncated raw ZIP payload: {entry.filename}")
    return payload


def inspect_archive(path: Path, compression_level: int, purpose: str) -> dict[str, Any]:
    method_names = {
        zipfile.ZIP_STORED: "stored",
        zipfile.ZIP_DEFLATED: "deflate",
    }
    with zipfile.ZipFile(path) as archive:
        entries = archive.infolist()
        worksheet = archive.getinfo("xl/worksheets/sheet1.xml")
    expected_method = zipfile.ZIP_STORED if compression_level == 0 else zipfile.ZIP_DEFLATED
    require(
        worksheet.compress_type == expected_method,
        f"{purpose} worksheet ZIP compression method does not match requested level",
    )
    method_counts: dict[str, int] = {}
    for entry in entries:
        name = method_names.get(entry.compress_type, f"method-{entry.compress_type}")
        method_counts[name] = method_counts.get(name, 0) + 1
    return {
        "entry_count": len(entries),
        "compression_method_counts": method_counts,
        "worksheet_compression_method": method_names.get(
            worksheet.compress_type, f"method-{worksheet.compress_type}"
        ),
        "worksheet_uncompressed_bytes": worksheet.file_size,
        "worksheet_compressed_bytes": worksheet.compress_size,
    }


def expected_inserted(case: PatchCase) -> int:
    return case.edits - case.edits // 2 if case.scenario == "patch-upsert" else 0


def excel_column_name(column: int) -> str:
    name = ""
    while column > 0:
        column, remainder = divmod(column - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name


def expected_unedited_source_value(row: int, col: int, source_pattern: str) -> Any:
    if source_pattern in {"mixed-inline", "mixed-shared"} and (row + col) % 3 == 0:
        return f"group-{(row + col) % 32}"
    if source_pattern == "formula" and col % 4 == 0:
        return f"={excel_column_name(col - 1)}{row}*2"
    return row * 1000 + col


def expected_source_tail(
    case: PatchCase, rows: int, cols: int, source_pattern: str
) -> Any:
    source_cells = rows * cols
    existing_edits = case.edits
    if case.scenario == "patch-upsert":
        existing_edits = case.edits // 2
    if case.scenario in {"patch-replace", "patch-upsert"} and existing_edits >= source_cells:
        return 900000000 + source_cells - 1
    return expected_unedited_source_value(rows, cols, source_pattern)


def verify_result(
    path: Path,
    case: PatchCase,
    rows: int,
    cols: int,
    source_compression_level: int,
    output_compression_level: int,
    source_path: Path,
    output_path: Path,
    expect_reused_source: bool,
    source_pattern: str,
    source_external_hyperlink: bool,
) -> tuple[dict[str, Any], dict[str, int], dict[str, Any]]:
    result = load_json(path)
    require(
        result.get("workbook_editor_benchmark_schema_version") == BENCHMARK_SCHEMA_VERSION,
        "workbook editor benchmark schema mismatch",
    )
    require(result.get("scenario") == case.scenario, f"{case.name} scenario mismatch")
    require(result.get("rows") == rows, f"{case.name} rows mismatch")
    require(result.get("cols") == cols, f"{case.name} cols mismatch")
    require(result.get("source_cells") == rows * cols, f"{case.name} source cell mismatch")
    require(result.get("requested_edits") == case.edits, f"{case.name} edit count mismatch")
    require(
        result.get("inserted_coordinates") == expected_inserted(case),
        f"{case.name} inserted count mismatch",
    )
    if case.scenario == "patch-upsert":
        require(result.get("single_pass_worksheet_transform") is True,
            "patch-upsert should use the single-pass worksheet transform")
        require(
            int(result.get("single_pass_inserted_cell_count")) == expected_inserted(case),
            "patch-upsert single-pass inserted count mismatch",
        )
        require(int(result.get("single_pass_scanned_source_cell_count")) == rows * cols,
            "patch-upsert single-pass source scan count mismatch")
        require(int(result.get("single_pass_staged_output_bytes")) > 0,
            "patch-upsert single-pass staged output bytes should be positive")
        require(int(result.get("single_pass_transform_us")) > 0,
            "patch-upsert single-pass transform microseconds should be positive")
        parsed_events = int(result.get("single_pass_source_parsed_event_count"))
        callback_events = int(result.get("single_pass_source_callback_event_count"))
        coalesced_input_events = int(
            result.get("single_pass_source_coalesced_input_event_count")
        )
        coalesced_output_events = int(
            result.get("single_pass_source_coalesced_output_event_count")
        )
        action_callbacks = int(result.get("single_pass_transform_action_callback_count"))
        require(parsed_events > callback_events > action_callbacks > 0,
            "patch-upsert should reduce parser events before transform actions")
        require(coalesced_input_events > coalesced_output_events > 0,
            "patch-upsert should report value-event coalescing")
        inline_fast_path_count = int(
            result.get("single_pass_source_simple_inline_string_fast_path_count")
        )
        inline_fast_path_bytes = int(
            result.get("single_pass_source_simple_inline_string_fast_path_bytes")
        )
        inline_fallback_count = int(
            result.get("single_pass_source_simple_inline_string_fallback_count")
        )
        if source_pattern == "mixed-inline":
            require(inline_fast_path_count > 0,
                "mixed-inline patch-upsert should use the simple inline-string fast path")
            require(inline_fast_path_bytes > inline_fast_path_count,
                "mixed-inline fast-path bytes should exceed its payload count")
            require(inline_fallback_count < inline_fast_path_count,
                "mixed-inline fast-path fallbacks should stay below completed payloads")
        else:
            require(inline_fast_path_count == 0 and inline_fast_path_bytes == 0,
                "non-inline source should not report simple inline-string fast-path traffic")
        require(int(result.get("single_pass_output_append_call_count")) > 0,
            "patch-upsert should report output append calls")
        require(int(result.get("single_pass_output_flush_count")) > 0,
            "patch-upsert should report bounded output flushes")
        require(
            int(result.get("single_pass_output_flush_count"))
            < int(result.get("single_pass_output_append_call_count")),
            "patch-upsert output batching should coalesce append calls",
        )
        require(
            0 < int(result.get("single_pass_output_peak_buffer_bytes")) <= 256 * 1024,
            "patch-upsert output buffer peak should stay within 256 KiB",
        )
        require(
            int(result.get("single_pass_relationship_scan_input_call_count"))
            <= int(result.get("single_pass_output_flush_count")),
            "patch-upsert relationship scanner batches should fit output flushes",
        )
        require(
            0 < int(result.get("single_pass_relationship_scan_input_bytes"))
            < int(result.get("single_pass_staged_output_bytes")),
            "patch-upsert relationship scanner should skip cell XML",
        )
        require(
            int(result.get("single_pass_relationship_scan_boundary_carry_count"))
            <= int(result.get("single_pass_relationship_scan_input_call_count")),
            "patch-upsert relationship scanner boundary carries should fit input calls",
        )
        require(
            int(result.get("single_pass_relationship_scan_slow_path_tag_count"))
            < int(result.get("single_pass_scanned_source_cell_count")),
            "patch-upsert relationship scanner should limit attribute slow paths",
        )
        measured_sink_us = (
            int(result.get("single_pass_relationship_scan_us"))
            + int(result.get("single_pass_temporary_write_us"))
        )
        require(
            measured_sink_us <= int(result.get("single_pass_transform_us")),
            "patch-upsert measured sink time should not exceed transform time",
        )
    require(
        result.get("source_fixture_mode")
        == ("reused-existing-source" if expect_reused_source else "generated-source"),
        f"{case.name} source fixture mode mismatch",
    )
    require(result.get("source_pattern") == source_pattern,
        f"{case.name} source pattern mismatch")
    require(result.get("source_external_hyperlink") is source_external_hyperlink,
        f"{case.name} source hyperlink mode mismatch")
    require(
        result.get("source_compression_level") == source_compression_level,
        f"{case.name} source compression mismatch",
    )
    require(
        result.get("output_compression_level") == output_compression_level,
        f"{case.name} output compression mismatch",
    )
    require(result.get("source_bytes") == source_path.stat().st_size, "source size mismatch")
    require(result.get("output_bytes") == output_path.stat().st_size, "output size mismatch")
    require(result.get("office_open") == "not_run", "Office state must remain not_run")
    require(float(result.get("peak_memory_mb")) >= 0.0, "peak memory must be non-negative")
    require(int(result.get("total_editor_ms")) >= 0, "total editor time must be non-negative")
    require(int(result.get("package_writer_total_us")) > 0,
        "package writer telemetry should report positive total time")
    require(int(result.get("package_writer_total_process_cpu_us")) >= 0,
        "package writer process CPU time must be non-negative")
    if output_compression_level != 0:
        require(result.get("target_worksheet_entry_telemetry") is True,
            "minizip Patch output should report target worksheet entry telemetry")
        require(
            int(result.get("target_worksheet_entry_requested_compression_level"))
            == output_compression_level,
            "target worksheet compression level mismatch",
        )
        require(int(result.get("target_worksheet_entry_input_bytes")) > 0,
            "target worksheet entry input bytes should be positive")
        require(int(result.get("target_worksheet_entry_input_read_calls")) > 0,
            "target worksheet entry should report source read calls")
        require(int(result.get("target_worksheet_entry_writer_write_calls")) > 0,
            "target worksheet entry should report writer calls")
        require(int(result.get("target_worksheet_entry_total_us")) > 0,
            "target worksheet entry should report positive total time")
        writer_process_cpu_us = int(
            result.get("target_worksheet_entry_writer_write_process_cpu_us")
        )
        close_process_cpu_us = int(
            result.get("target_worksheet_entry_close_process_cpu_us")
        )
        deflate_writer_process_cpu_us = int(
            result.get("target_worksheet_entry_deflate_writer_process_cpu_us")
        )
        require(
            int(result.get("target_worksheet_entry_total_process_cpu_us")) >= 0
            and writer_process_cpu_us >= 0
            and close_process_cpu_us >= 0,
            "target worksheet process CPU timings must be non-negative",
        )
        if result.get("target_worksheet_entry_raw_compressed_copy") is True:
            require(deflate_writer_process_cpu_us == 0,
                "raw-copy target should not report DEFLATE process CPU")
        else:
            require(
                deflate_writer_process_cpu_us
                == writer_process_cpu_us + close_process_cpu_us,
                "rewritten target DEFLATE process CPU accounting mismatch",
            )
            if int(result.get("target_worksheet_entry_uncompressed_bytes")) >= 1024 * 1024:
                require(deflate_writer_process_cpu_us > 0,
                    "large rewritten target should report positive DEFLATE process CPU")
        if case.scenario == "patch-upsert":
            require(result.get("target_worksheet_entry_reused_staged_crc32") is True,
                "patch-upsert should reuse staged CRC32 metadata")
            require(
                int(result.get("target_worksheet_entry_reused_staged_file_chunk_count")) > 0,
                "patch-upsert should report reused staged file chunks",
            )
            require(int(result.get("target_worksheet_entry_staged_crc_validation_us")) >= 0,
                "patch-upsert staged CRC validation time must be non-negative")
    require(not result.get("materialized_worksheet"), "Patch case unexpectedly materialized sheet")

    copied_names = list(result.get("copied_entry_names", []))
    rewritten_names = list(result.get("rewritten_entry_names", []))
    require(
        len(copied_names) == int(result.get("copied_entry_count")),
        "copied entry count mismatch",
    )
    require(
        len(rewritten_names) == int(result.get("rewritten_entry_count")),
        "rewritten entry count mismatch",
    )
    if case.scenario == "noop-copy":
        require(not rewritten_names, "no-op copy should not rewrite package entries")
        require(copied_names, "no-op copy should copy source entries")
    elif case.scenario == "document-properties":
        require(
            {"docProps/core.xml", "docProps/app.xml"}.issubset(rewritten_names),
            "document-properties should rewrite core/app parts",
        )
    else:
        require(
            "xl/worksheets/sheet1.xml" in rewritten_names,
            "targeted Patch should rewrite the source worksheet entry",
        )

    accounting = account_plan_bytes(source_path, output_path, result)
    output_archive = inspect_archive(output_path, output_compression_level, "output")
    return result, accounting, output_archive


def run_process(command: list[str]) -> subprocess.CompletedProcess[str]:
    completed = subprocess.run(command, capture_output=True, text=True, check=False)
    if completed.returncode != 0:
        raise RuntimeError(
            f"benchmark command failed ({completed.returncode}): {' '.join(command)}\n"
            f"stdout:\n{completed.stdout}\nstderr:\n{completed.stderr}"
        )
    return completed


def run_single_case(
    bench_exe: Path,
    output_dir: Path,
    source_path: Path,
    case: PatchCase,
    rows: int,
    cols: int,
    source_compression_level: int,
    output_compression_level: int,
    source_pattern: str,
    source_external_hyperlink: bool,
    run_name: str,
    reuse_source: bool,
) -> dict[str, Any]:
    output_path = output_dir / f"{run_name}.xlsx"
    result_path = output_dir / f"{run_name}.json"
    command = [
        str(bench_exe),
        "--scenario",
        case.scenario,
        "--rows",
        str(rows),
        "--cols",
        str(cols),
        "--edits",
        str(case.edits),
        "--source-compression-level",
        str(source_compression_level),
        "--output-compression-level",
        str(output_compression_level),
        "--source-pattern",
        source_pattern,
        "--source",
        str(source_path),
        "--output",
        str(output_path),
        "--result",
        str(result_path),
    ]
    if source_external_hyperlink:
        command.append("--source-external-hyperlink")
    if reuse_source:
        command.append("--reuse-source")
    completed = run_process(command)
    result, accounting, output_archive = verify_result(
        result_path,
        case,
        rows,
        cols,
        source_compression_level,
        output_compression_level,
        source_path,
        output_path,
        reuse_source,
        source_pattern,
        source_external_hyperlink,
    )
    return {
        "name": run_name,
        "command": command,
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
        "source": str(source_path),
        "output": str(output_path),
        "result_json": str(result_path),
        "result": result,
        "byte_accounting": accounting,
        "output_archive": output_archive,
    }


def verify_external_hyperlink_archive(path: Path) -> str:
    relationship_part = "xl/worksheets/_rels/sheet1.xml.rels"
    worksheet_part = "xl/worksheets/sheet1.xml"
    expected_target = "https://example.com/fastxlsx-benchmark"
    with zipfile.ZipFile(path) as archive:
        relationship_root = element_tree.fromstring(archive.read(relationship_part))
        relationship_id = None
        for relationship in relationship_root:
            if (
                relationship.attrib.get("Type", "").endswith("/hyperlink")
                and relationship.attrib.get("Target") == expected_target
                and relationship.attrib.get("TargetMode") == "External"
            ):
                relationship_id = relationship.attrib.get("Id")
                break
        require(relationship_id is not None, "source hyperlink relationship was removed")

        relationship_marker = f'r:id="{relationship_id}"'.encode("utf-8")
        reference_marker = b'ref="A1"'
        carry = b""
        hyperlink_tag_found = False
        with archive.open(worksheet_part) as worksheet_xml:
            while chunk := worksheet_xml.read(64 * 1024):
                window = carry + chunk
                search_offset = 0
                while True:
                    tag_begin = window.find(b"<hyperlink", search_offset)
                    if tag_begin < 0:
                        break
                    tag_end = window.find(b">", tag_begin)
                    if tag_end < 0:
                        break
                    tag = window[tag_begin : tag_end + 1]
                    if relationship_marker in tag and reference_marker in tag:
                        hyperlink_tag_found = True
                        break
                    search_offset = tag_end + 1
                if hyperlink_tag_found:
                    break
                carry = window[-4096:]
        require(hyperlink_tag_found, "source hyperlink worksheet binding was removed")
    return expected_target


def verify_workbook_with_openpyxl(
    path: Path,
    case: PatchCase,
    rows: int,
    cols: int,
    source_pattern: str,
    source_external_hyperlink: bool,
) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for --verify-openpyxl") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(workbook.sheetnames == ["Data"], f"{case.name} sheet names mismatch")
        worksheet = workbook["Data"]
        expected_rows = rows + expected_inserted(case)
        require(worksheet.max_row == expected_rows, f"{case.name} max row mismatch")
        require(worksheet.max_column == cols, f"{case.name} max column mismatch")
        expected_first = 900000000 if case.scenario in {"patch-replace", "patch-upsert"} else 1001
        require(worksheet["A1"].value == expected_first, f"{case.name} first cell mismatch")
        require(
            worksheet.cell(rows, cols).value
            == expected_source_tail(case, rows, cols, source_pattern),
            f"{case.name} source tail cell mismatch",
        )
        hyperlink_target = (
            verify_external_hyperlink_archive(path)
            if source_external_hyperlink
            else None
        )
        if case.scenario == "patch-upsert":
            require(
                worksheet.cell(expected_rows, 1).value
                == 900000000 + case.edits - 1,
                "patch-upsert inserted tail mismatch",
            )
        if case.scenario == "document-properties":
            require(
                workbook.properties.title == "FastXLSX Patch Benchmark",
                "document-properties title mismatch",
            )
        return {
            "status": "opened",
            "openpyxl_version": openpyxl.__version__,
            "sheet_count": len(workbook.sheetnames),
            "sheet1_max_row": worksheet.max_row,
            "sheet1_max_column": worksheet.max_column,
            "sheet1_first_cell": worksheet["A1"].value,
            "sheet1_source_tail_cell": worksheet.cell(rows, cols).value,
            "document_title": workbook.properties.title,
            "source_hyperlink_target": hyperlink_target,
        }
    finally:
        workbook.close()


def run_case(
    bench_exe: Path,
    output_dir: Path,
    source_path: Path,
    case: PatchCase,
    rows: int,
    cols: int,
    source_compression_level: int,
    output_compression_level: int,
    source_pattern: str,
    source_external_hyperlink: bool,
    warmup_runs: int,
    measured_runs: int,
    verify_openpyxl: bool,
) -> dict[str, Any]:
    base_name = compression_case_name(case, output_compression_level)
    for run_index in range(1, warmup_runs + 1):
        run_single_case(
            bench_exe,
            output_dir,
            source_path,
            case,
            rows,
            cols,
            source_compression_level,
            output_compression_level,
            source_pattern,
            source_external_hyperlink,
            indexed_run_name(base_name, "warmup", run_index, warmup_runs),
            True,
        )

    measured = [
        run_single_case(
            bench_exe,
            output_dir,
            source_path,
            case,
            rows,
            cols,
            source_compression_level,
            output_compression_level,
            source_pattern,
            source_external_hyperlink,
            indexed_run_name(base_name, "run", run_index, measured_runs),
            True,
        )
        for run_index in range(1, measured_runs + 1)
    ]
    representative_index, statistics_report = measured_run_summary(measured)
    representative = measured[representative_index]
    openpyxl_report = (
        verify_workbook_with_openpyxl(
            Path(representative["output"]),
            case,
            rows,
            cols,
            source_pattern,
            source_external_hyperlink,
        )
        if verify_openpyxl
        else {"status": "not_requested"}
    )
    return {
        "name": base_name,
        "scenario": case.scenario,
        "output_compression_level": output_compression_level,
        "requested_edits": case.edits,
        "warmup_runs": warmup_runs,
        "measured_runs": measured_runs,
        "representative_run": representative_index + 1,
        "statistics": statistics_report,
        "result": representative["result"],
        "byte_accounting": representative["byte_accounting"],
        "output_archive": representative["output_archive"],
        "openpyxl": openpyxl_report,
        "runs": measured,
    }


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(report, stream, ensure_ascii=False, indent=2)
        stream.write("\n")


def expect_value_error(text: str) -> None:
    try:
        parse_case(text)
    except ValueError:
        return
    raise AssertionError(f"expected ValueError for {text!r}")


def run_self_test() -> None:
    require(parse_case("noop-copy:0") == PatchCase("noop-copy", 0), "no-op parse mismatch")
    require(
        parse_case("patch-replace:1000") == PatchCase("patch-replace", 1000),
        "replace parse mismatch",
    )
    for invalid in ["noop-copy:1", "patch-replace:0", "unknown:1", "patch-replace:x"]:
        expect_value_error(invalid)

    reports = [
        {
            "result": {
                "total_editor_ms": elapsed,
                "open_ms": elapsed // 10,
                "mutation_ms": 1,
                "save_ms": elapsed - elapsed // 10 - 1,
                "peak_memory_mb": 10.0 + index,
                "output_bytes": 100 + index,
                "single_pass_transform_ms": 1,
                "single_pass_transform_us": 1000,
                "single_pass_source_scan_action_us": 700,
                "single_pass_source_parsed_event_count": 500,
                "single_pass_source_callback_event_count": 300,
                "single_pass_source_coalesced_input_event_count": 300,
                "single_pass_source_coalesced_output_event_count": 100,
                "single_pass_source_simple_inline_string_fast_path_count": 0,
                "single_pass_source_simple_inline_string_fast_path_bytes": 0,
                "single_pass_source_simple_inline_string_fallback_count": 0,
                "single_pass_transform_action_callback_count": 200,
                "single_pass_relationship_scan_us": 200,
                "single_pass_relationship_scan_input_call_count": 1,
                "single_pass_relationship_scan_input_bytes": 100000,
                "single_pass_relationship_scan_boundary_carry_count": 1,
                "single_pass_relationship_scan_slow_path_tag_count": 2,
                "single_pass_temporary_write_us": 100,
                "single_pass_output_append_call_count": 100,
                "single_pass_output_flush_count": 2,
                "single_pass_output_peak_buffer_bytes": 262144,
                "package_writer_total_us": 1000,
                "package_writer_total_process_cpu_us": 900,
                "target_worksheet_entry_total_us": 500,
                "target_worksheet_entry_total_process_cpu_us": 450,
                "target_worksheet_entry_input_read_us": 50,
                "target_worksheet_entry_writer_write_us": 400,
                "target_worksheet_entry_writer_write_process_cpu_us": 350,
                "target_worksheet_entry_staged_crc_validation_us": 1,
                "target_worksheet_entry_close_us": 50,
                "target_worksheet_entry_close_process_cpu_us": 40,
                "target_worksheet_entry_deflate_writer_process_cpu_us": 390,
            },
            "byte_accounting": {
                "copied_uncompressed_bytes": 1000,
                "copied_source_compressed_bytes": 500,
                "copied_output_compressed_bytes": 450,
                "raw_compressed_copy_bytes": 0,
                "rewritten_uncompressed_bytes": 100 + index,
                "rewritten_compressed_bytes": 50 + index,
            },
        }
        for index, elapsed in enumerate([30, 10, 20])
    ]
    representative, statistics_report = measured_run_summary(reports)
    require(representative == 2, "median representative mismatch")
    require(statistics_report["total_editor_ms"] == {"min": 10, "median": 20, "max": 30},
        "elapsed statistics mismatch")
    require(indexed_run_name("case", "run", 1, 3) == "case-run-01", "run name mismatch")
    require(
        compression_case_name(PatchCase("patch-upsert", 1000), 6)
        == "patch-upsert-compression-6",
        "compression case name mismatch",
    )

    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        source = root / "source.xlsx"
        output = root / "output.xlsx"
        with zipfile.ZipFile(source, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("copied.xml", "same")
            archive.writestr("xl/worksheets/sheet1.xml", "sheet")
            archive.writestr("rewritten.xml", "before")
            archive.writestr("omitted.xml", "gone")
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("copied.xml", "same")
            archive.writestr("xl/worksheets/sheet1.xml", "sheet")
            archive.writestr("rewritten.xml", "after")
        accounting = account_plan_bytes(
            source,
            output,
            {
                "output_plan_entry_count": 4,
                "copied_entry_names": ["copied.xml", "xl/worksheets/sheet1.xml"],
                "rewritten_entry_names": ["rewritten.xml"],
                "omitted_entry_names": ["omitted.xml"],
            },
        )
        require(accounting["copied_uncompressed_bytes"] == 9, "copied byte accounting mismatch")
        require(accounting["rewritten_uncompressed_bytes"] == 5, "rewrite byte accounting mismatch")
        source_report = inspect_archive(source, 6, "source")
        require(source_report["worksheet_compression_method"] == "deflate",
            "source compression inspection mismatch")

    print("OK: run_patch_benchmark_matrix.py self-test passed")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bench-exe", type=Path, default=default_benchmark_exe())
    parser.add_argument("--output-dir", type=Path, default=Path("build/qa/patch-benchmark-matrix"))
    parser.add_argument("--rows", type=int, default=100000)
    parser.add_argument("--cols", type=int, default=10)
    parser.add_argument("--source-compression-level", type=int, default=6)
    parser.add_argument(
        "--output-compression-level",
        action="append",
        dest="output_compression_levels",
        type=int,
        help="Output ZIP compression level. May be passed multiple times. Default: 6.",
    )
    parser.add_argument(
        "--source-pattern",
        choices=["numeric", "mixed-inline", "mixed-shared", "formula"],
        default="numeric",
    )
    parser.add_argument("--source-external-hyperlink", action="store_true")
    parser.add_argument("--case", action="append", dest="cases")
    parser.add_argument("--warmup-runs", type=int, default=1)
    parser.add_argument("--measured-runs", "--repeat", type=int, default=3)
    parser.add_argument("--verify-openpyxl", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        return 0

    require(args.rows > 0, "--rows must be positive")
    require(args.cols > 0, "--cols must be positive")
    require(0 <= args.source_compression_level <= 9,
        "source compression level must be 0..9")
    output_compression_levels = args.output_compression_levels or [6]
    require(
        all(-1 <= level <= 9 for level in output_compression_levels),
        "output compression level must be -1..9",
    )
    require(
        len(output_compression_levels) == len(set(output_compression_levels)),
        "output compression levels must be unique",
    )
    require(args.warmup_runs >= 0, "--warmup-runs must be non-negative")
    require(args.measured_runs > 0, "--measured-runs must be positive")

    bench_exe = args.bench_exe.resolve()
    require(bench_exe.is_file(), f"benchmark executable not found: {bench_exe}")
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    cases = [parse_case(text) for text in (args.cases or DEFAULT_CASES)]
    require(len({case.name for case in cases}) == len(cases), "case scenarios must be unique")
    case_names = [
        compression_case_name(case, compression_level)
        for compression_level in output_compression_levels
        for case in cases
    ]
    require(len(case_names) == len(set(case_names)), "matrix case names must be unique")

    source_path = output_dir / "patch-matrix-source.xlsx"
    preparation = run_single_case(
        bench_exe,
        output_dir,
        source_path,
        PatchCase("noop-copy", 0),
        args.rows,
        args.cols,
        args.source_compression_level,
        output_compression_levels[0],
        args.source_pattern,
        args.source_external_hyperlink,
        "source-preparation",
        False,
    )
    source_archive = inspect_archive(source_path, args.source_compression_level, "source")

    case_reports = []
    for output_compression_level in output_compression_levels:
        for case in cases:
            report = run_case(
                bench_exe,
                output_dir,
                source_path,
                case,
                args.rows,
                args.cols,
                args.source_compression_level,
                output_compression_level,
                args.source_pattern,
                args.source_external_hyperlink,
                args.warmup_runs,
                args.measured_runs,
                args.verify_openpyxl,
            )
            case_reports.append(report)
            median = report["statistics"]["total_editor_ms"]["median"]
            peak = report["statistics"]["peak_memory_mb"]["median"]
            rewritten = report["statistics"]["rewritten_uncompressed_bytes"]["median"]
            print(
                f"{report['name']}: median={median} ms, peak={peak} MB, "
                f"rewritten={rewritten} logical bytes"
            )

    matrix_report = {
        "patch_benchmark_matrix_schema_version": "3",
        "benchmark_executable": str(bench_exe),
        "output_dir": str(output_dir),
        "rows": args.rows,
        "cols": args.cols,
        "source_cells": args.rows * args.cols,
        "source_compression_level": args.source_compression_level,
        "output_compression_levels": output_compression_levels,
        "source_pattern": args.source_pattern,
        "source_external_hyperlink": args.source_external_hyperlink,
        "source": str(source_path),
        "source_bytes": source_path.stat().st_size,
        "source_archive": source_archive,
        "source_preparation": preparation,
        "warmup_runs_per_case": args.warmup_runs,
        "measured_runs_per_case": args.measured_runs,
        "representative_result_policy":
            "measured run nearest total_editor_ms median; earliest run breaks ties",
        "byte_accounting":
            "ZIP central-directory logical file_size and compressed compress_size; copied "
            "entries report source and output compressed bytes separately and additionally "
            "require equal source/output CRC and logical file_size",
        "cases": case_reports,
        "comparison_scope":
            "Manual opt-in public WorkbookEditor Patch matrix. Source generation is a separate "
            "process; every warm-up/measured run reuses the same source so benchmark process "
            "PeakWorkingSet excludes WorkbookWriter fixture construction. openpyxl validates "
            "only representative outputs after benchmark timing; Office remains not_run.",
    }
    report_path = output_dir / "patch-benchmark-matrix-report.json"
    write_report(report_path, matrix_report)
    print(f"Wrote {report_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (AssertionError, RuntimeError, ValueError) as error:
        print(f"Patch benchmark matrix failed: {error}", file=sys.stderr)
        raise SystemExit(1)
