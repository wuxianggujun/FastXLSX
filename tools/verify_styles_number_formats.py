#!/usr/bin/env python3
"""Local QA for FastXLSX streaming styles.

This helper is intentionally local QA, not a runtime dependency and not a
default CI gate. It checks the generated OpenXML package, then uses openpyxl as
a reader-visible semantic check and XlsxWriter as an optional reference writer.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        return archive.read(name).decode("utf-8")


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def count(text: str, fragment: str) -> int:
    return text.count(fragment)


def verify_styles_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/styles.xml",
    ]
    for name in required:
        require(name in names, f"styles sample missing package entry: {name}")

    require("xl/worksheets/_rels/sheet1.xml.rels" not in names,
            "styles sample should not create worksheet relationships")
    require("xl/sharedStrings.xml" not in names,
            "styles number-format sample should not create sharedStrings.xml")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require(
        '<Override PartName="/xl/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        in content_types,
        "styles content type override missing",
    )

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require(count(workbook_rels, "<Relationship ") == 2,
            "styles workbook relationship count mismatch")
    require(
        'Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"' in workbook_rels,
        "styles workbook relationship mismatch",
    )

    styles_xml = read_zip_text(path, "xl/styles.xml")
    require('<numFmts count="2">' in styles_xml, "custom numFmt count mismatch")
    require('<numFmt numFmtId="164" formatCode="$#,##0.00"/>' in styles_xml,
            "currency number format mismatch")
    require(
        '<numFmt numFmtId="165" formatCode="0.00 &quot;kg &amp; &lt;unit&gt;&quot;"/>'
        in styles_xml,
        "escaped number format mismatch",
    )
    require('<fonts count="1">' in styles_xml, "default fonts missing")
    require('<fills count="2">' in styles_xml, "default fills missing")
    require('<borders count="1">' in styles_xml, "default borders missing")
    require('<cellXfs count="3">' in styles_xml, "cellXfs count mismatch")
    require('numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" '
            'applyNumberFormat="1"' in styles_xml,
            "first style xf mismatch")
    require('numFmtId="165" fontId="0" fillId="0" borderId="0" xfId="0" '
            'applyNumberFormat="1"' in styles_xml,
            "second style xf mismatch")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<dimension ref="A1:F2"/>' in worksheet_xml, "worksheet dimension mismatch")
    require('<c r="A2" s="1"><v>1234.5</v></c>' in worksheet_xml,
            "styled currency cell mismatch")
    require('<c r="B2" s="2"><v>7.25</v></c>' in worksheet_xml,
            "styled escaped-format cell mismatch")
    require('<c r="C2"><v>9</v></c>' in worksheet_xml,
            "default numeric cell should omit style attribute")
    require('<c r="D2" s="2" t="inlineStr"><is><t>styled text</t></is></c>'
            in worksheet_xml,
            "styled inlineStr cell mismatch")
    require('<c r="E2" s="1" t="b"><v>1</v></c>' in worksheet_xml,
            "styled boolean cell mismatch")
    require('<c r="F2" s="1"><f>A2*2</f></c>' in worksheet_xml,
            "styled formula cell mismatch")
    require('s="0"' not in worksheet_xml, "default style should not be serialized as s=\"0\"")

    return {
        "verified_entries": required,
        "style_ids": {"currency": 1, "escaped_number_format": 2},
        "custom_number_format_ids": [164, 165],
    }


def verify_shared_styles_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    for name in ["xl/sharedStrings.xml", "xl/styles.xml", "xl/worksheets/sheet1.xml"]:
        require(name in names, f"shared styles sample missing package entry: {name}")
    require("xl/worksheets/_rels/sheet1.xml.rels" not in names,
            "shared styles sample should not create worksheet relationships")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require('Id="rId2" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
            'Target="sharedStrings.xml"' in workbook_rels,
            "sharedStrings relationship should precede styles")
    require('Id="rId3" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
            'Target="styles.xml"' in workbook_rels,
            "styles relationship should follow sharedStrings")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<c r="A1" s="1" t="s"><v>0</v></c>' in worksheet_xml,
            "styled shared string cell mismatch")
    require('<c r="B1" t="s"><v>1</v></c>' in worksheet_xml,
            "plain shared string cell mismatch")

    return {
        "relationship_model": "workbook-local sheet, sharedStrings, then styles ids",
    }


def verify_alignment_styles_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/styles.xml",
    ]
    for name in required:
        require(name in names, f"alignment styles sample missing package entry: {name}")

    require("xl/worksheets/_rels/sheet1.xml.rels" not in names,
            "alignment styles sample should not create worksheet relationships")
    require("xl/sharedStrings.xml" not in names,
            "alignment styles sample should not create sharedStrings.xml")

    workbook_rels = read_zip_text(path, "xl/_rels/workbook.xml.rels")
    require(count(workbook_rels, "<Relationship ") == 2,
            "alignment styles workbook relationship count mismatch")
    require('Target="styles.xml"' in workbook_rels,
            "alignment styles workbook relationship missing")

    styles_xml = read_zip_text(path, "xl/styles.xml")
    require('<numFmts count="1">' in styles_xml,
            "alignment-only style should not create a custom numFmt entry")
    require('<numFmt numFmtId="164" formatCode="0.0"/>' in styles_xml,
            "alignment combined style should reuse number format id")
    require('<fonts count="1">' in styles_xml, "alignment slice should keep default fonts")
    require('<fills count="2">' in styles_xml, "alignment slice should keep default fills")
    require('<borders count="1">' in styles_xml, "alignment slice should keep default borders")
    require('<cellXfs count="10">' in styles_xml, "alignment cellXfs count mismatch")
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment wrapText="1"/></xf>' in styles_xml,
        "wrap-text alignment xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment horizontal="left"/></xf>' in styles_xml,
        "left alignment xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment horizontal="center"/></xf>' in styles_xml,
        "center alignment xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment horizontal="right"/></xf>' in styles_xml,
        "right alignment xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment vertical="top"/></xf>' in styles_xml,
        "top alignment xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment vertical="center"/></xf>' in styles_xml,
        "middle alignment xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyAlignment="1"><alignment vertical="bottom"/></xf>' in styles_xml,
        "bottom alignment xf mismatch",
    )
    require(
        '<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyNumberFormat="1"/>' in styles_xml,
        "number-only xf mismatch",
    )
    require(
        '<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" '
        'applyNumberFormat="1" applyAlignment="1">'
        '<alignment wrapText="1" horizontal="center" vertical="center"/></xf>'
        in styles_xml,
        "number plus combined alignment xf mismatch",
    )
    require(count(styles_xml, 'applyAlignment="1"') == 8,
            "alignment sample should write eight alignment xfs")

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<dimension ref="A1:J2"/>' in worksheet_xml,
            "alignment worksheet dimension mismatch")
    require('<c r="A2" s="1" t="inlineStr"><is><t>line 1\nline 2</t></is></c>'
            in worksheet_xml,
            "wrapped string cell mismatch")
    require('<c r="B2" s="2" t="inlineStr"><is><t>left</t></is></c>'
            in worksheet_xml,
            "left aligned string cell mismatch")
    require('<c r="C2" s="3" t="inlineStr"><is><t>center</t></is></c>'
            in worksheet_xml,
            "center aligned string cell mismatch")
    require('<c r="D2" s="4" t="inlineStr"><is><t>right</t></is></c>'
            in worksheet_xml,
            "right aligned string cell mismatch")
    require('<c r="E2" s="5" t="inlineStr"><is><t>top</t></is></c>'
            in worksheet_xml,
            "top aligned string cell mismatch")
    require('<c r="F2" s="6" t="inlineStr"><is><t>middle</t></is></c>'
            in worksheet_xml,
            "middle aligned string cell mismatch")
    require('<c r="G2" s="7" t="inlineStr"><is><t>bottom</t></is></c>'
            in worksheet_xml,
            "bottom aligned string cell mismatch")
    require('<c r="H2" s="8"><v>12.5</v></c>' in worksheet_xml,
            "number-only styled cell mismatch")
    require('<c r="I2" s="9"><v>42.5</v></c>' in worksheet_xml,
            "number plus alignment styled cell mismatch")
    require('<c r="J2" t="inlineStr"><is><t>plain</t></is></c>' in worksheet_xml,
            "default inline string cell mismatch")
    require('s="0"' not in worksheet_xml,
            "default style should not be serialized as s=\"0\"")

    return {
        "style_ids": {
            "wrap_text": 1,
            "horizontal_left": 2,
            "horizontal_center": 3,
            "horizontal_right": 4,
            "vertical_top": 5,
            "vertical_center": 6,
            "vertical_bottom": 7,
            "number_format": 8,
            "number_format_combined_alignment": 9,
        },
        "custom_number_format_ids": [164],
        "alignment_records": 8,
    }


def verify_font_styles_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/styles.xml",
    ]
    for name in required:
        require(name in names, f"font styles sample missing package entry: {name}")

    require("xl/worksheets/_rels/sheet1.xml.rels" not in names,
            "font styles sample should not create worksheet relationships")
    require("xl/sharedStrings.xml" not in names,
            "font styles sample should not create sharedStrings.xml")

    styles_xml = read_zip_text(path, "xl/styles.xml")
    require('<numFmts count="1">' in styles_xml,
            "font sample should create exactly one custom numFmt")
    require('<numFmt numFmtId="164" formatCode="0.0"/>' in styles_xml,
            "font sample custom number format mismatch")
    require('<fonts count="4">' in styles_xml, "font collection count mismatch")
    require(
        '<font><b/><sz val="11"/><color theme="1"/><name val="Calibri"/>'
        '<family val="2"/><scheme val="minor"/></font>' in styles_xml,
        "bold font XML mismatch",
    )
    require(
        '<font><i/><sz val="11"/><color theme="1"/><name val="Calibri"/>'
        '<family val="2"/><scheme val="minor"/></font>' in styles_xml,
        "italic font XML mismatch",
    )
    require(
        '<font><b/><i/><sz val="11"/><color theme="1"/><name val="Calibri"/>'
        '<family val="2"/><scheme val="minor"/></font>' in styles_xml,
        "bold italic font XML mismatch",
    )
    require('<cellXfs count="5">' in styles_xml, "font cellXfs count mismatch")
    require(
        '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
        in styles_xml,
        "bold xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
        in styles_xml,
        "italic xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="3" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
        in styles_xml,
        "bold italic xf mismatch",
    )
    require(
        '<xf numFmtId="164" fontId="1" fillId="0" borderId="0" xfId="0" '
        'applyNumberFormat="1" applyFont="1"/>' in styles_xml,
        "number plus bold xf mismatch",
    )

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<dimension ref="A1:E2"/>' in worksheet_xml, "font worksheet dimension mismatch")
    require('<c r="A2" s="1" t="inlineStr"><is><t>bold</t></is></c>' in worksheet_xml,
            "bold styled text cell mismatch")
    require('<c r="B2" s="2" t="inlineStr"><is><t>italic</t></is></c>' in worksheet_xml,
            "italic styled text cell mismatch")
    require('<c r="C2" s="3" t="b"><v>1</v></c>' in worksheet_xml,
            "bold italic styled boolean cell mismatch")
    require('<c r="D2" s="4"><v>12.5</v></c>' in worksheet_xml,
            "number plus bold styled number cell mismatch")
    require('<c r="E2" t="inlineStr"><is><t>plain</t></is></c>' in worksheet_xml,
            "plain font sample cell mismatch")
    require('s="0"' not in worksheet_xml,
            "default style should not be serialized as s=\"0\"")

    return {
        "style_ids": {"bold": 1, "italic": 2, "bold_italic": 3, "number_bold": 4},
        "custom_number_format_ids": [164],
        "custom_font_ids": [1, 2, 3],
    }


def verify_fill_styles_package(path: Path) -> dict[str, Any]:
    names = zip_names(path)
    required = [
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/styles.xml",
    ]
    for name in required:
        require(name in names, f"fill styles sample missing package entry: {name}")

    require("xl/worksheets/_rels/sheet1.xml.rels" not in names,
            "fill styles sample should not create worksheet relationships")
    require("xl/sharedStrings.xml" not in names,
            "fill styles sample should not create sharedStrings.xml")

    styles_xml = read_zip_text(path, "xl/styles.xml")
    require('<numFmts count="1">' in styles_xml,
            "fill sample should create exactly one custom numFmt")
    require('<numFmt numFmtId="164" formatCode="0.0"/>' in styles_xml,
            "fill sample custom number format mismatch")
    require('<fonts count="2">' in styles_xml,
            "fill sample should create default font plus bold font")
    require('<fills count="4">' in styles_xml, "fill collection count mismatch")
    require(
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>' in styles_xml,
        "default fill records mismatch",
    )
    require(
        '<fill><patternFill patternType="solid"><fgColor rgb="FFFFEB84"/>'
        '<bgColor indexed="64"/></patternFill></fill>' in styles_xml,
        "yellow solid fill XML mismatch",
    )
    require(
        '<fill><patternFill patternType="solid"><fgColor rgb="FF5A8AD6"/>'
        '<bgColor indexed="64"/></patternFill></fill>' in styles_xml,
        "blue solid fill XML mismatch",
    )
    require('<cellXfs count="5">' in styles_xml, "fill cellXfs count mismatch")
    require(
        '<xf numFmtId="0" fontId="0" fillId="2" borderId="0" xfId="0" applyFill="1"/>'
        in styles_xml,
        "yellow fill xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="0" fillId="3" borderId="0" xfId="0" applyFill="1"/>'
        in styles_xml,
        "blue fill xf mismatch",
    )
    require(
        '<xf numFmtId="164" fontId="0" fillId="2" borderId="0" xfId="0" '
        'applyNumberFormat="1" applyFill="1"/>' in styles_xml,
        "number plus fill xf mismatch",
    )
    require(
        '<xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" '
        'applyFont="1" applyFill="1"/>' in styles_xml,
        "font plus fill xf mismatch",
    )

    worksheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require('<dimension ref="A1:E2"/>' in worksheet_xml, "fill worksheet dimension mismatch")
    require('<c r="A2" s="1" t="inlineStr"><is><t>yellow</t></is></c>' in worksheet_xml,
            "yellow styled text cell mismatch")
    require('<c r="B2" s="2" t="inlineStr"><is><t>blue</t></is></c>' in worksheet_xml,
            "blue styled text cell mismatch")
    require('<c r="C2" s="3"><v>12.5</v></c>' in worksheet_xml,
            "number plus fill styled number cell mismatch")
    require('<c r="D2" s="4" t="inlineStr"><is><t>bold yellow</t></is></c>'
            in worksheet_xml,
            "font plus fill styled text cell mismatch")
    require('<c r="E2" t="inlineStr"><is><t>plain</t></is></c>' in worksheet_xml,
            "plain fill sample cell mismatch")
    require('s="0"' not in worksheet_xml,
            "default style should not be serialized as s=\"0\"")

    return {
        "style_ids": {"yellow": 1, "blue": 2, "number_yellow": 3, "bold_yellow": 4},
        "custom_number_format_ids": [164],
        "custom_font_ids": [1],
        "custom_fill_ids": [2, 3],
    }


def verify_with_openpyxl(
    path: Path, shared_path: Path, alignment_path: Path, font_path: Path, fill_path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = workbook["Styles"]
        require(sheet["A2"].number_format == "$#,##0.00",
                f"A2 number format mismatch: {sheet['A2'].number_format!r}")
        require(sheet["B2"].number_format == '0.00 "kg & <unit>"',
                f"B2 number format mismatch: {sheet['B2'].number_format!r}")
        require(sheet["C2"].number_format == "General",
                f"C2 default number format mismatch: {sheet['C2'].number_format!r}")
        require(sheet["D2"].number_format == '0.00 "kg & <unit>"',
                f"D2 string style number format mismatch: {sheet['D2'].number_format!r}")
        require(sheet["F2"].data_type == "f", "F2 should remain a formula cell")
        require(sheet["F2"].number_format == "$#,##0.00",
                f"F2 formula style mismatch: {sheet['F2'].number_format!r}")
    finally:
        workbook.close()

    shared_workbook = openpyxl.load_workbook(shared_path, read_only=False, data_only=False)
    try:
        shared = shared_workbook["StyledShared"]
        require(shared["A1"].value == "styled shared", "styled shared string value mismatch")
        require(shared["A1"].number_format == "@",
                f"styled shared string format mismatch: {shared['A1'].number_format!r}")
        require(shared["B1"].number_format == "General",
                f"plain shared string format mismatch: {shared['B1'].number_format!r}")
    finally:
        shared_workbook.close()

    alignment_workbook = openpyxl.load_workbook(alignment_path, read_only=False, data_only=False)
    try:
        alignment = alignment_workbook["Alignment"]
        require(alignment["A2"].value == "line 1\nline 2",
                "wrapped string value mismatch")
        require(alignment["A2"].alignment.wrap_text is True,
                f"A2 wrap_text mismatch: {alignment['A2'].alignment.wrap_text!r}")
        require(alignment["B2"].alignment.horizontal == "left",
                f"B2 horizontal mismatch: {alignment['B2'].alignment.horizontal!r}")
        require(alignment["C2"].alignment.horizontal == "center",
                f"C2 horizontal mismatch: {alignment['C2'].alignment.horizontal!r}")
        require(alignment["D2"].alignment.horizontal == "right",
                f"D2 horizontal mismatch: {alignment['D2'].alignment.horizontal!r}")
        require(alignment["E2"].alignment.vertical == "top",
                f"E2 vertical mismatch: {alignment['E2'].alignment.vertical!r}")
        require(alignment["F2"].alignment.vertical == "center",
                f"F2 vertical mismatch: {alignment['F2'].alignment.vertical!r}")
        require(alignment["G2"].alignment.vertical == "bottom",
                f"G2 vertical mismatch: {alignment['G2'].alignment.vertical!r}")
        require(alignment["H2"].number_format == "0.0",
                f"H2 number format mismatch: {alignment['H2'].number_format!r}")
        require(alignment["H2"].alignment.wrap_text is not True,
                f"H2 should not wrap text: {alignment['H2'].alignment.wrap_text!r}")
        require(alignment["I2"].number_format == "0.0",
                f"I2 number format mismatch: {alignment['I2'].number_format!r}")
        require(alignment["I2"].alignment.wrap_text is True,
                f"I2 wrap_text mismatch: {alignment['I2'].alignment.wrap_text!r}")
        require(alignment["I2"].alignment.horizontal == "center",
                f"I2 horizontal mismatch: {alignment['I2'].alignment.horizontal!r}")
        require(alignment["I2"].alignment.vertical == "center",
                f"I2 vertical mismatch: {alignment['I2'].alignment.vertical!r}")
        require(alignment["J2"].alignment.wrap_text is not True,
                f"J2 should keep default wrap_text: {alignment['J2'].alignment.wrap_text!r}")
        require(alignment["J2"].alignment.horizontal is None,
                f"J2 should keep default horizontal alignment: {alignment['J2'].alignment.horizontal!r}")
        require(alignment["J2"].alignment.vertical is None,
                f"J2 should keep default vertical alignment: {alignment['J2'].alignment.vertical!r}")
    finally:
        alignment_workbook.close()

    font_workbook = openpyxl.load_workbook(font_path, read_only=False, data_only=False)
    try:
        fonts = font_workbook["Fonts"]
        require(fonts["A2"].font.bold is True,
                f"A2 bold mismatch: {fonts['A2'].font.bold!r}")
        require(fonts["A2"].font.italic is False,
                f"A2 italic mismatch: {fonts['A2'].font.italic!r}")
        require(fonts["B2"].font.bold is False,
                f"B2 bold mismatch: {fonts['B2'].font.bold!r}")
        require(fonts["B2"].font.italic is True,
                f"B2 italic mismatch: {fonts['B2'].font.italic!r}")
        require(fonts["C2"].font.bold is True and fonts["C2"].font.italic is True,
                "C2 bold italic mismatch")
        require(fonts["D2"].font.bold is True,
                f"D2 bold mismatch: {fonts['D2'].font.bold!r}")
        require(fonts["D2"].number_format == "0.0",
                f"D2 number format mismatch: {fonts['D2'].number_format!r}")
        require(fonts["E2"].font.bold is False and fonts["E2"].font.italic is False,
                "E2 should keep default font flags")
    finally:
        font_workbook.close()

    fill_workbook = openpyxl.load_workbook(fill_path, read_only=False, data_only=False)
    try:
        fills = fill_workbook["Fills"]
        require(fills["A2"].fill.fill_type == "solid",
                f"A2 fill type mismatch: {fills['A2'].fill.fill_type!r}")
        require(fills["A2"].fill.fgColor.rgb == "FFFFEB84",
                f"A2 fill color mismatch: {fills['A2'].fill.fgColor.rgb!r}")
        require(fills["B2"].fill.fill_type == "solid",
                f"B2 fill type mismatch: {fills['B2'].fill.fill_type!r}")
        require(fills["B2"].fill.fgColor.rgb == "FF5A8AD6",
                f"B2 fill color mismatch: {fills['B2'].fill.fgColor.rgb!r}")
        require(fills["C2"].number_format == "0.0",
                f"C2 number format mismatch: {fills['C2'].number_format!r}")
        require(fills["C2"].fill.fgColor.rgb == "FFFFEB84",
                f"C2 fill color mismatch: {fills['C2'].fill.fgColor.rgb!r}")
        require(fills["D2"].font.bold is True,
                f"D2 bold mismatch: {fills['D2'].font.bold!r}")
        require(fills["D2"].fill.fgColor.rgb == "FFFFEB84",
                f"D2 fill color mismatch: {fills['D2'].fill.fgColor.rgb!r}")
        require(fills["E2"].fill.fill_type != "solid",
                f"E2 should keep default non-solid fill: {fills['E2'].fill.fill_type!r}")
    finally:
        fill_workbook.close()

    return {
        "status": "opened",
        "styles": {
            "A2": "$#,##0.00",
            "B2": '0.00 "kg & <unit>"',
            "C2": "General",
            "F2": "$#,##0.00",
        },
        "shared_styles": {"A1": "@", "B1": "General"},
        "alignment_styles": {
            "A2_wrap_text": True,
            "B2_horizontal": "left",
            "C2_horizontal": "center",
            "D2_horizontal": "right",
            "E2_vertical": "top",
            "F2_vertical": "center",
            "G2_vertical": "bottom",
            "H2_number_format": "0.0",
            "I2_number_format": "0.0",
            "I2_wrap_text": True,
            "I2_horizontal": "center",
            "I2_vertical": "center",
        },
        "font_styles": {
            "A2_bold": True,
            "B2_italic": True,
            "C2_bold_italic": True,
            "D2_number_format": "0.0",
        },
        "fill_styles": {
            "A2_fill": "FFFFEB84",
            "B2_fill": "FF5A8AD6",
            "C2_number_format": "0.0",
            "D2_bold": True,
        },
    }


def create_xlsxwriter_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(path)
    try:
        sheet = workbook.add_worksheet("Styles")
        currency = workbook.add_format({"num_format": "$#,##0.00"})
        escaped = workbook.add_format({"num_format": '0.00 "kg & <unit>"'})
        sheet.write_row(0, 0, ["Currency", "Escaped", "Default", "Text", "Bool", "Formula"])
        sheet.write_number(1, 0, 1234.5, currency)
        sheet.write_number(1, 1, 7.25, escaped)
        sheet.write_number(1, 2, 9.0)
        sheet.write_string(1, 3, "styled text", escaped)
        sheet.write_boolean(1, 4, True, currency)
        sheet.write_formula(1, 5, "=A2*2", currency)
    finally:
        workbook.close()

    return {"status": "created", "path": str(path)}


def create_xlsxwriter_alignment_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(path)
    try:
        sheet = workbook.add_worksheet("Alignment")
        wrap_text = workbook.add_format({"text_wrap": True})
        left = workbook.add_format({"align": "left"})
        center = workbook.add_format({"align": "center"})
        right = workbook.add_format({"align": "right"})
        top = workbook.add_format({"valign": "top"})
        middle = workbook.add_format({"valign": "vcenter"})
        bottom = workbook.add_format({"valign": "bottom"})
        number = workbook.add_format({"num_format": "0.0"})
        number_combined = workbook.add_format({
            "num_format": "0.0",
            "text_wrap": True,
            "align": "center",
            "valign": "vcenter",
        })
        sheet.write_row(0, 0, [
            "Wrapped",
            "Left",
            "Center",
            "Right",
            "Top",
            "Middle",
            "Bottom",
            "Number",
            "Combined",
            "Default",
        ])
        sheet.write_string(1, 0, "line 1\nline 2", wrap_text)
        sheet.write_string(1, 1, "left", left)
        sheet.write_string(1, 2, "center", center)
        sheet.write_string(1, 3, "right", right)
        sheet.write_string(1, 4, "top", top)
        sheet.write_string(1, 5, "middle", middle)
        sheet.write_string(1, 6, "bottom", bottom)
        sheet.write_number(1, 7, 12.5, number)
        sheet.write_number(1, 8, 42.5, number_combined)
        sheet.write_string(1, 9, "plain")
    finally:
        workbook.close()

    return {"status": "created", "path": str(path)}


def create_xlsxwriter_font_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(path)
    try:
        sheet = workbook.add_worksheet("Fonts")
        bold = workbook.add_format({"bold": True})
        italic = workbook.add_format({"italic": True})
        bold_italic = workbook.add_format({"bold": True, "italic": True})
        number_bold = workbook.add_format({"num_format": "0.0", "bold": True})
        sheet.write_row(0, 0, ["Bold", "Italic", "BoldItalic", "NumberBold", "Default"])
        sheet.write_string(1, 0, "bold", bold)
        sheet.write_string(1, 1, "italic", italic)
        sheet.write_boolean(1, 2, True, bold_italic)
        sheet.write_number(1, 3, 12.5, number_bold)
        sheet.write_string(1, 4, "plain")
    finally:
        workbook.close()

    return {"status": "created", "path": str(path)}


def create_xlsxwriter_fill_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(path)
    try:
        sheet = workbook.add_worksheet("Fills")
        yellow = workbook.add_format({"bg_color": "#FFEB84"})
        blue = workbook.add_format({"bg_color": "#5A8AD6"})
        number_yellow = workbook.add_format({"num_format": "0.0", "bg_color": "#FFEB84"})
        bold_yellow = workbook.add_format({"bold": True, "bg_color": "#FFEB84"})
        sheet.write_row(0, 0, ["Yellow", "Blue", "NumberYellow", "BoldYellow", "Default"])
        sheet.write_string(1, 0, "yellow", yellow)
        sheet.write_string(1, 1, "blue", blue)
        sheet.write_number(1, 2, 12.5, number_yellow)
        sheet.write_string(1, 3, "bold yellow", bold_yellow)
        sheet.write_string(1, 4, "plain")
    finally:
        workbook.close()

    return {"status": "created", "path": str(path)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-styles-number-formats.xlsx"),
        help="FastXLSX styles workbook to verify.",
    )
    parser.add_argument(
        "--shared-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-styles-shared-strings.xlsx"),
        help="FastXLSX sharedStrings + styles workbook to verify.",
    )
    parser.add_argument(
        "--alignment-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-styles-alignment.xlsx"),
        help="FastXLSX limited alignment styles workbook to verify.",
    )
    parser.add_argument(
        "--font-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-styles-fonts.xlsx"),
        help="FastXLSX bold/italic font styles workbook to verify.",
    )
    parser.add_argument(
        "--fill-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-styles-fills.xlsx"),
        help="FastXLSX solid fill styles workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/styles-number-formats"),
        help="Directory for local QA reports and reference workbooks.",
    )
    args = parser.parse_args()

    input_path = args.input.resolve()
    shared_input_path = args.shared_input.resolve()
    alignment_input_path = args.alignment_input.resolve()
    font_input_path = args.font_input.resolve()
    fill_input_path = args.fill_input.resolve()
    work_dir = args.work_dir.resolve()
    require(input_path.exists(), f"input workbook does not exist: {input_path}")
    require(shared_input_path.exists(), f"shared input workbook does not exist: {shared_input_path}")
    require(alignment_input_path.exists(),
            f"alignment input workbook does not exist: {alignment_input_path}")
    require(font_input_path.exists(), f"font input workbook does not exist: {font_input_path}")
    require(fill_input_path.exists(), f"fill input workbook does not exist: {fill_input_path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    report = {
        "fastxlsx_input": str(input_path),
        "fastxlsx_shared_input": str(shared_input_path),
        "fastxlsx_alignment_input": str(alignment_input_path),
        "fastxlsx_font_input": str(font_input_path),
        "fastxlsx_fill_input": str(fill_input_path),
        "fastxlsx_package": verify_styles_package(input_path),
        "fastxlsx_shared_package": verify_shared_styles_package(shared_input_path),
        "fastxlsx_alignment_package": verify_alignment_styles_package(alignment_input_path),
        "fastxlsx_font_package": verify_font_styles_package(font_input_path),
        "fastxlsx_fill_package": verify_fill_styles_package(fill_input_path),
        "xlsx_libraries": {
            "openpyxl": verify_with_openpyxl(
                input_path, shared_input_path, alignment_input_path, font_input_path,
                fill_input_path),
            "xlsxwriter": create_xlsxwriter_reference(
                work_dir / "reference-xlsxwriter-styles-number-formats.xlsx"),
            "xlsxwriter_alignment": create_xlsxwriter_alignment_reference(
                work_dir / "reference-xlsxwriter-styles-alignment.xlsx"),
            "xlsxwriter_fonts": create_xlsxwriter_font_reference(
                work_dir / "reference-xlsxwriter-styles-fonts.xlsx"),
            "xlsxwriter_fills": create_xlsxwriter_fill_reference(
                work_dir / "reference-xlsxwriter-styles-fills.xlsx"),
        },
    }

    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # noqa: BLE001
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
