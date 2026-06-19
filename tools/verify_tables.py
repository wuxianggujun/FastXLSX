#!/usr/bin/env python3
"""Local table QA for FastXLSX workbooks.

This helper is intentionally outside CTest. It verifies FastXLSX table package
XML directly, then uses Python XLSX libraries only as local QA/reference tools.
These libraries are not FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


NAMESPACES = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def zip_names(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as archive:
        return set(archive.namelist())


def read_zip_text(path: Path, name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open(name) as entry:
            return entry.read().decode("utf-8")


def relationship_count(xml: str) -> int:
    root = ElementTree.fromstring(xml)
    return len(root.findall("rel:Relationship", NAMESPACES))


def parse_table(path: Path, name: str) -> ElementTree.Element:
    return ElementTree.fromstring(read_zip_text(path, name))


def require_no_table_side_effects(path: Path) -> None:
    names = zip_names(path)
    require("[Content_Types].xml" in names, "missing content types part")
    require("xl/styles.xml" not in names, "table metadata should not create styles.xml")
    require("xl/metadata.xml" not in names, "table metadata should not create metadata.xml")
    require("xl/calcChain.xml" not in names, "table metadata should not create calcChain.xml")


def verify_totals_package(path: Path) -> dict[str, Any]:
    require_no_table_side_effects(path)
    names = zip_names(path)
    required = [
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet3.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/tables/table1.xml",
        "xl/tables/table2.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")
    require("xl/worksheets/_rels/sheet3.xml.rels" not in names, "plain sheet should not have rels")

    content_types = read_zip_text(path, "[Content_Types].xml")
    require(
        content_types.count(
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"'
        )
        == 2,
        "table content type override count mismatch",
    )

    first_sheet = read_zip_text(path, "xl/worksheets/sheet1.xml")
    require(
        "</sheetData><hyperlinks><hyperlink ref=\"A2\" r:id=\"rId1\"/></hyperlinks>"
        "<tableParts count=\"1\"><tablePart r:id=\"rId2\"/></tableParts></worksheet>"
        in first_sheet,
        "tableParts XML should follow hyperlinks and use the next worksheet rId",
    )
    first_rels = read_zip_text(path, "xl/worksheets/_rels/sheet1.xml.rels")
    require(relationship_count(first_rels) == 2, "first sheet relationship count mismatch")
    require('Id="rId1"' in first_rels and 'TargetMode="External"' in first_rels,
            "first sheet hyperlink rId mismatch")
    require('Id="rId2"' in first_rels and 'Target="../tables/table1.xml"' in first_rels,
            "first sheet table rId mismatch")

    second_sheet = read_zip_text(path, "xl/worksheets/sheet2.xml")
    require('<tableParts count="1"><tablePart r:id="rId1"/></tableParts>' in second_sheet,
            "second worksheet table rId should be owner-local")
    second_rels = read_zip_text(path, "xl/worksheets/_rels/sheet2.xml.rels")
    require(relationship_count(second_rels) == 1, "second sheet relationship count mismatch")
    require('Id="rId1"' in second_rels and 'Target="../tables/table2.xml"' in second_rels,
            "second sheet table rId mismatch")

    first_table_xml = read_zip_text(path, "xl/tables/table1.xml")
    require('name="InventoryTable"' in first_table_xml, "missing InventoryTable")
    require('ref="A1:C3"' in first_table_xml, "InventoryTable ref mismatch")
    require('totalsRowShown="0"' in first_table_xml, "InventoryTable should hide totals row")
    require('totalsRowCount=' not in first_table_xml, "InventoryTable should not write totalsRowCount")
    require('<autoFilter ref="A1:C3"/>' in first_table_xml, "InventoryTable autoFilter mismatch")
    require('name="Price &amp; &lt;Cost&gt;"' in first_table_xml,
            "InventoryTable column XML escaping mismatch")
    require(
        '<tableStyleInfo name="TableStyleMedium9" showFirstColumn="0" '
        'showLastColumn="0" showRowStripes="1" showColumnStripes="0"/>'
        in first_table_xml,
        "InventoryTable style info mismatch",
    )

    second_table_xml = read_zip_text(path, "xl/tables/table2.xml")
    require('name="TotalsTable"' in second_table_xml, "missing TotalsTable")
    require('ref="A1:B3"' in second_table_xml, "TotalsTable ref mismatch")
    require('totalsRowCount="1"' in second_table_xml, "TotalsTable should expose one totals row")
    require('totalsRowShown="0"' not in second_table_xml,
            "TotalsTable should not also write hidden totals metadata")
    require('<autoFilter ref="A1:B2"/>' in second_table_xml,
            "TotalsTable autoFilter should exclude totals row")
    require('totalsRowFunction="sum"' in second_table_xml,
            "FastXLSX should emit the requested totals function metadata")
    require('totalsRowLabel="Total"' in second_table_xml,
            "FastXLSX should emit the requested totals label metadata")
    require('totalsRowLabel=""' not in second_table_xml,
            "FastXLSX should omit empty totals label metadata")
    require("totalsRowFormula" not in second_table_xml,
            "FastXLSX should not generate totals formula text")
    require("calculatedColumnFormula" not in second_table_xml,
            "FastXLSX should not generate calculated column formulas")
    require("<tableStyleInfo" not in second_table_xml,
            "empty style name should still omit tableStyleInfo")

    root = ElementTree.fromstring(second_table_xml)
    columns = root.findall("main:tableColumns/main:tableColumn", NAMESPACES)
    require([column.attrib.get("name") for column in columns] == ["Metric", "Value"],
            "TotalsTable parsed column names mismatch")
    require([column.attrib.get("totalsRowLabel") for column in columns] == ["Total", None],
            "TotalsTable parsed totals labels mismatch")
    require([column.attrib.get("totalsRowFunction") for column in columns] == [None, "sum"],
            "TotalsTable parsed totals functions mismatch")

    third_sheet = read_zip_text(path, "xl/worksheets/sheet3.xml")
    require("<tableParts" not in third_sheet, "plain worksheet should not contain tableParts")
    require("xmlns:r=" not in third_sheet, "plain worksheet should not include relationship namespace")

    return {
        "relationship_model": "worksheet-owner-local table rIds",
        "tables": {
            "InventoryTable": {"ref": "A1:C3", "totals": "hidden"},
            "TotalsTable": {
                "ref": root.attrib.get("ref"),
                "totalsRowCount": root.attrib.get("totalsRowCount"),
                "totals_labels": [column.attrib.get("totalsRowLabel") for column in columns],
                "totals_functions": [column.attrib.get("totalsRowFunction") for column in columns],
            },
        },
    }


def verify_style_flags_package(path: Path) -> dict[str, Any]:
    require_no_table_side_effects(path)
    names = zip_names(path)
    require("xl/tables/table1.xml" in names, "style flags sample should create table1.xml")
    table_xml = read_zip_text(path, "xl/tables/table1.xml")
    expected = (
        '<tableStyleInfo name="TableStyleMedium4" showFirstColumn="1" '
        'showLastColumn="1" showRowStripes="0" showColumnStripes="1"/>'
    )
    require(expected in table_xml, "table style flags XML mismatch")
    root = parse_table(path, "xl/tables/table1.xml")
    style = root.find("main:tableStyleInfo", NAMESPACES)
    require(style is not None, "table style info element missing")
    return {
        "table": root.attrib.get("name"),
        "style_name": style.attrib.get("name"),
        "showFirstColumn": style.attrib.get("showFirstColumn"),
        "showLastColumn": style.attrib.get("showLastColumn"),
        "showRowStripes": style.attrib.get("showRowStripes"),
        "showColumnStripes": style.attrib.get("showColumnStripes"),
    }


def verify_column_escape_package(path: Path) -> dict[str, Any]:
    require_no_table_side_effects(path)
    names = zip_names(path)
    require("xl/tables/table1.xml" in names, "column escape sample should create table1.xml")
    table_xml = read_zip_text(path, "xl/tables/table1.xml")
    require(
        '<tableColumn id="1" name="Text &quot;quoted&quot;" '
        'totalsRowLabel="Total &quot;quoted&quot; &amp; &lt;done&gt;"/>'
        in table_xml,
        "table column double-quote attribute escape mismatch",
    )
    require(
        '<tableColumn id="2" name="Owner&apos;s Share" totalsRowFunction="sum"/>'
        in table_xml,
        "table column apostrophe attribute escape mismatch",
    )
    require(
        '<tableColumn id="3" name="A&amp;B&lt;Limit&gt;"/>'
        in table_xml,
        "table column ampersand and angle-bracket attribute escape mismatch",
    )
    require('totalsRowLabel=""' not in table_xml, "empty totals row labels should be omitted")
    require("totalsRowFormula" not in table_xml, "totals labels should not generate formulas")

    root = parse_table(path, "xl/tables/table1.xml")
    columns = root.findall("main:tableColumns/main:tableColumn", NAMESPACES)
    return {
        "table": root.attrib.get("name"),
        "ref": root.attrib.get("ref"),
        "columns": [column.attrib.get("name") for column in columns],
        "totals_labels": [column.attrib.get("totalsRowLabel") for column in columns],
        "totals_functions": [column.attrib.get("totalsRowFunction") for column in columns],
    }


def verify_overlap_package(path: Path) -> dict[str, Any]:
    require_no_table_side_effects(path)
    names = zip_names(path)
    required = [
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/tables/table1.xml",
        "xl/tables/table2.xml",
        "xl/tables/table3.xml",
        "xl/tables/table4.xml",
    ]
    for name in required:
        require(name in names, f"missing package entry: {name}")
    require("xl/tables/table5.xml" not in names, "rejected overlap should not create table5.xml")

    first_sheet_xml = read_zip_text(path, "xl/worksheets/sheet1.xml")
    second_sheet_xml = read_zip_text(path, "xl/worksheets/sheet2.xml")
    require('<tableParts count="3">' in first_sheet_xml, "first worksheet table count mismatch")
    require('<tableParts count="1">' in second_sheet_xml, "second worksheet table count mismatch")

    expected_tables = {
        "xl/tables/table1.xml": ("FirstTable", "A1:B2"),
        "xl/tables/table2.xml": ("AdjacentColumnsTable", "C1:D2"),
        "xl/tables/table3.xml": ("AdjacentRowsTable", "A3:B4"),
        "xl/tables/table4.xml": ("OtherSheetTable", "A1:B2"),
    }
    parsed_tables = {}
    for entry, (expected_name, expected_ref) in expected_tables.items():
        root = parse_table(path, entry)
        require(root.attrib.get("name") == expected_name, f"{entry} table name mismatch")
        require(root.attrib.get("displayName") == expected_name, f"{entry} displayName mismatch")
        require(root.attrib.get("ref") == expected_ref, f"{entry} ref mismatch")
        parsed_tables[entry] = {
            "name": root.attrib.get("name"),
            "displayName": root.attrib.get("displayName"),
            "ref": root.attrib.get("ref"),
        }
    return {"tables": parsed_tables, "rejected_overlap": "no table5.xml"}


def import_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for this local QA helper") from exc
    return openpyxl


def verify_openpyxl_totals(path: Path) -> dict[str, Any]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        inventory = workbook["Inventory"]
        totals = workbook["Totals"]
        plain = workbook["Plain"]
        require("InventoryTable" in inventory.tables, "openpyxl missing InventoryTable")
        require("TotalsTable" in totals.tables, "openpyxl missing TotalsTable")
        require(len(plain.tables) == 0, "Plain sheet should not expose tables")
        inventory_table = inventory.tables["InventoryTable"]
        totals_table = totals.tables["TotalsTable"]
        require(inventory_table.ref == "A1:C3", "openpyxl InventoryTable ref mismatch")
        require(totals_table.ref == "A1:B3", "openpyxl TotalsTable ref mismatch")
        require(getattr(totals_table, "totalsRowCount", None) == 1,
                "openpyxl TotalsTable totalsRowCount mismatch")
        require([column.name for column in totals_table.tableColumns] == ["Metric", "Value"],
                "openpyxl TotalsTable column names mismatch")
        require([getattr(column, "totalsRowLabel", None) for column in totals_table.tableColumns]
                == ["Total", None],
                "openpyxl TotalsTable totals labels mismatch")
        require([getattr(column, "totalsRowFunction", None) for column in totals_table.tableColumns]
                == [None, "sum"],
                "openpyxl TotalsTable totals functions mismatch")
        return {
            "InventoryTable": inventory_table.ref,
            "TotalsTable": {
                "ref": totals_table.ref,
                "totalsRowCount": getattr(totals_table, "totalsRowCount", None),
            },
            "Plain": {},
        }
    finally:
        workbook.close()


def verify_openpyxl_style_flags(path: Path) -> dict[str, Any]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        table = workbook["StyleFlags"].tables["StyleFlagTable"]
        style = table.tableStyleInfo
        require(table.ref == "A1:B2", "openpyxl StyleFlagTable ref mismatch")
        require(style is not None, "openpyxl StyleFlagTable style missing")
        require(style.name == "TableStyleMedium4", "openpyxl style name mismatch")
        require(style.showFirstColumn is True, "openpyxl showFirstColumn mismatch")
        require(style.showLastColumn is True, "openpyxl showLastColumn mismatch")
        require(style.showRowStripes is False, "openpyxl showRowStripes mismatch")
        require(style.showColumnStripes is True, "openpyxl showColumnStripes mismatch")
        return {
            "StyleFlagTable": {
                "ref": table.ref,
                "style": style.name,
                "showFirstColumn": style.showFirstColumn,
                "showLastColumn": style.showLastColumn,
                "showRowStripes": style.showRowStripes,
                "showColumnStripes": style.showColumnStripes,
            },
        }
    finally:
        workbook.close()


def verify_openpyxl_column_escape(path: Path) -> dict[str, Any]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        table = workbook["TableEscapes"].tables["EscapedColumnTable"]
        require(table.ref == "A1:C3", "openpyxl EscapedColumnTable ref mismatch")
        columns = table.tableColumns
        require([column.name for column in columns] == ['Text "quoted"', "Owner's Share", "A&B<Limit>"],
                "openpyxl escaped column names mismatch")
        require([getattr(column, "totalsRowLabel", None) for column in columns]
                == ['Total "quoted" & <done>', None, None],
                "openpyxl escaped totals labels mismatch")
        require([getattr(column, "totalsRowFunction", None) for column in columns]
                == [None, "sum", None],
                "openpyxl escaped totals functions mismatch")
        return {
            "EscapedColumnTable": {
                "ref": table.ref,
                "columns": [column.name for column in columns],
                "totals_labels": [getattr(column, "totalsRowLabel", None) for column in columns],
                "totals_functions": [getattr(column, "totalsRowFunction", None) for column in columns],
            },
        }
    finally:
        workbook.close()


def verify_openpyxl_overlap(path: Path) -> dict[str, Any]:
    openpyxl = import_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        tables_sheet = workbook["Tables"]
        other_sheet = workbook["OtherTables"]
        expected_first = {
            "FirstTable": "A1:B2",
            "AdjacentColumnsTable": "C1:D2",
            "AdjacentRowsTable": "A3:B4",
        }
        require(set(tables_sheet.tables.keys()) == set(expected_first.keys()),
                "openpyxl first worksheet table names mismatch")
        for table_name, table_ref in expected_first.items():
            require(tables_sheet.tables[table_name].ref == table_ref,
                    f"openpyxl {table_name} ref mismatch")
        require(set(other_sheet.tables.keys()) == {"OtherSheetTable"},
                "openpyxl second worksheet table names mismatch")
        require(other_sheet.tables["OtherSheetTable"].ref == "A1:B2",
                "openpyxl cross-worksheet table ref mismatch")
        return {
            "Tables": {name: tables_sheet.tables[name].ref for name in tables_sheet.tables.keys()},
            "OtherTables": {name: other_sheet.tables[name].ref for name in other_sheet.tables.keys()},
        }
    finally:
        workbook.close()


def create_xlsxwriter_reference(path: Path) -> dict[str, Any]:
    try:
        import xlsxwriter  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module xlsxwriter is not installed"}

    workbook = xlsxwriter.Workbook(str(path))
    worksheet = workbook.add_worksheet("Reference")
    worksheet.write_row("A1", ['Text "quoted"', "Owner's Share", "A&B<Limit>"])
    worksheet.write_row("A2", ["alpha", 42, "done"])
    worksheet.write_row("A3", ['Total "quoted" & <done>', 42, "Owner's Total"])
    worksheet.add_table(
        "A1:C3",
        {
            "name": "EscapedColumnTable",
            "total_row": True,
            "style": None,
            "columns": [
                {"header": 'Text "quoted"', "total_string": 'Total "quoted" & <done>'},
                {"header": "Owner's Share", "total_function": "sum"},
                {"header": "A&B<Limit>"},
            ],
        },
    )
    workbook.close()

    reference_xml = read_zip_text(path, "xl/tables/table1.xml")
    require('totalsRowCount="1"' in reference_xml,
            "XlsxWriter reference did not emit totalsRowCount")
    require('name="Text &quot;quoted&quot;"' in reference_xml,
            "XlsxWriter reference did not escape double quotes as expected")
    require('totalsRowFunction="sum"' in reference_xml,
            "XlsxWriter reference did not emit totalsRowFunction")
    require('totalsRowLabel="Total &quot;quoted&quot; &amp; &lt;done&gt;"' in reference_xml,
            "XlsxWriter reference did not emit escaped totalsRowLabel")
    return {"status": "created", "path": str(path)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--tables-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-tables.xlsx"),
        help="FastXLSX basic/totals table workbook to verify.",
    )
    parser.add_argument(
        "--style-flags-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-table-style-flags.xlsx"),
        help="FastXLSX table style flags workbook to verify.",
    )
    parser.add_argument(
        "--column-escape-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-table-column-escape.xlsx"),
        help="FastXLSX table column escape workbook to verify.",
    )
    parser.add_argument(
        "--overlap-input",
        type=Path,
        default=Path("build/windows-nmake-release/tests/fastxlsx-streaming-table-range-overlap.xlsx"),
        help="FastXLSX table range-overlap workbook to verify.",
    )
    parser.add_argument(
        "--work-dir",
        type=Path,
        default=Path("build/qa/tables"),
        help="Directory for reference files and the JSON report.",
    )
    args = parser.parse_args()

    tables_input = args.tables_input.resolve()
    style_flags_input = args.style_flags_input.resolve()
    column_escape_input = args.column_escape_input.resolve()
    overlap_input = args.overlap_input.resolve()
    work_dir = args.work_dir.resolve()
    for path in [tables_input, style_flags_input, column_escape_input, overlap_input]:
        require(path.exists(), f"input workbook does not exist: {path}")
    work_dir.mkdir(parents=True, exist_ok=True)

    report = {
        "tables": {
            "fastxlsx_input": str(tables_input),
            "package": verify_totals_package(tables_input),
            "openpyxl": verify_openpyxl_totals(tables_input),
        },
        "style_flags": {
            "fastxlsx_input": str(style_flags_input),
            "package": verify_style_flags_package(style_flags_input),
            "openpyxl": verify_openpyxl_style_flags(style_flags_input),
        },
        "column_escape": {
            "fastxlsx_input": str(column_escape_input),
            "package": verify_column_escape_package(column_escape_input),
            "openpyxl": verify_openpyxl_column_escape(column_escape_input),
        },
        "range_overlap": {
            "fastxlsx_input": str(overlap_input),
            "package": verify_overlap_package(overlap_input),
            "openpyxl": verify_openpyxl_overlap(overlap_input),
        },
        "xlsxwriter_reference": create_xlsxwriter_reference(
            work_dir / "reference-xlsxwriter-table-escape.xlsx"
        ),
        "notes": [
            "This helper verifies package XML and openpyxl semantics only.",
            "Excel COM visual checks remain separate local QA.",
        ],
    }
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover - local QA helper
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
