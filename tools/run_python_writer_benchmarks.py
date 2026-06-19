#!/usr/bin/env python3
"""Run opt-in Python XLSX writer comparison benchmarks.

This helper is intentionally outside CTest and CI. It benchmarks Python writer
libraries in a separate worker process so elapsed write time, output size, and
peak process working-set/RSS can be compared with FastXLSX's opt-in benchmark
matrix. The Python libraries are benchmark/reference tools only; they are not
FastXLSX runtime dependencies.
"""

from __future__ import annotations

import argparse
import ctypes
import json
import math
import os
import platform
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


REPORT_SCHEMA_VERSION = "1"
DEFAULT_CASES = [
    "numeric:mixed",
    "mixed:mixed",
    "strings:repeated",
    "strings:unique",
]
DEFAULT_LIBRARIES = [
    "xlsxwriter-constant-memory",
    "openpyxl-write-only",
]
BYTES_PER_MIB = 1024.0 * 1024.0


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def module_available(name: str) -> bool:
    try:
        __import__(name)
        return True
    except ModuleNotFoundError:
        return False


def module_version(name: str) -> str | None:
    try:
        module = __import__(name)
    except ModuleNotFoundError:
        return None
    return str(getattr(module, "__version__", "installed"))


@dataclass(frozen=True)
class PythonLibrary:
    name: str
    module_name: str
    mode: str

    @property
    def case_prefix(self) -> str:
        return f"{self.name}-{self.mode}"


@dataclass(frozen=True)
class BenchmarkCase:
    scenario: str
    string_pattern: str

    @property
    def name(self) -> str:
        if self.scenario == "numeric":
            return "numeric"
        return f"{self.scenario}-{self.string_pattern}"


def parse_library(text: str) -> PythonLibrary:
    normalized = text.lower()
    if normalized in {"xlsxwriter", "xlsxwriter-constant-memory"}:
        return PythonLibrary(
            name="xlsxwriter",
            module_name="xlsxwriter",
            mode="constant-memory",
        )
    if normalized in {"openpyxl", "openpyxl-write-only"}:
        return PythonLibrary(
            name="openpyxl",
            module_name="openpyxl",
            mode="write-only",
        )
    raise ValueError(f"unsupported Python writer library: {text}")


def parse_case(text: str) -> BenchmarkCase:
    parts = text.split(":")
    if len(parts) != 2:
        raise ValueError(f"case must use scenario:string_pattern format: {text}")
    scenario, string_pattern = parts
    if scenario not in {"numeric", "mixed", "strings"}:
        raise ValueError(f"unsupported scenario in case {text!r}")
    if string_pattern not in {"mixed", "repeated", "unique"}:
        raise ValueError(f"unsupported string pattern in case {text!r}")
    if scenario == "numeric" and string_pattern != "mixed":
        raise ValueError("numeric cases should use the mixed placeholder string pattern")
    return BenchmarkCase(scenario=scenario, string_pattern=string_pattern)


def expected_string_ratio(case: BenchmarkCase, mixed_string_ratio: float) -> float:
    if case.scenario == "numeric":
        return 0.0
    if case.scenario == "strings":
        return 1.0
    return mixed_string_ratio


def should_write_string(case: BenchmarkCase, mixed_string_ratio: float, row: int, col: int) -> bool:
    ratio = expected_string_ratio(case, mixed_string_ratio)
    if case.scenario == "numeric" or ratio <= 0.0:
        return False
    if case.scenario == "strings" or ratio >= 1.0:
        return True
    raw_bucket = 1.0 / ratio
    if not math.isfinite(raw_bucket) or raw_bucket >= 2**32 - 1:
        return False
    bucket = max(int(raw_bucket), 1)
    return ((row + col) % bucket) == 0


def string_value_is_repeated(case: BenchmarkCase, row: int, col: int) -> bool:
    if case.string_pattern == "repeated":
        return True
    if case.string_pattern == "unique":
        return False
    return (row + col) % 5 == 0


def make_string_value(case: BenchmarkCase, sheet: int, row: int, col: int) -> str:
    if string_value_is_repeated(case, row, col):
        return "repeat"
    return f"s{sheet}-r{row}-c{col}"


def make_number_value(sheet: int, row: int, col: int) -> float:
    return float(sheet) * 1_000_000.0 + float(row) * 1_000.0 + float(col)


def make_cell_value(
    case: BenchmarkCase,
    mixed_string_ratio: float,
    sheet: int,
    row: int,
    col: int,
) -> int | float | str:
    if should_write_string(case, mixed_string_ratio, row, col):
        return make_string_value(case, sheet, row, col)
    value = make_number_value(sheet, row, col)
    return int(value) if value.is_integer() else value


def expected_string_distribution(
    case: BenchmarkCase,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
) -> dict[str, float | int]:
    per_sheet_string_cells = 0
    per_sheet_repeated_string_cells = 0
    per_sheet_non_repeated_unique_string_cells = 0
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            if not should_write_string(case, mixed_string_ratio, row, col):
                continue
            per_sheet_string_cells += 1
            if string_value_is_repeated(case, row, col):
                per_sheet_repeated_string_cells += 1
            else:
                per_sheet_non_repeated_unique_string_cells += 1

    string_cells = per_sheet_string_cells * sheets
    repeated_string_cells = per_sheet_repeated_string_cells * sheets
    unique_string_values = per_sheet_non_repeated_unique_string_cells * sheets
    if repeated_string_cells > 0:
        unique_string_values += 1
    duplicate_string_cells = string_cells - unique_string_values
    string_dedup_ratio = duplicate_string_cells / string_cells if string_cells else 0.0
    return {
        "string_cells": string_cells,
        "unique_string_values": unique_string_values,
        "duplicate_string_cells": duplicate_string_cells,
        "string_dedup_ratio": string_dedup_ratio,
    }


def iter_rows(
    case: BenchmarkCase,
    rows: int,
    cols: int,
    sheet: int,
    mixed_string_ratio: float,
) -> Iterable[list[int | float | str]]:
    for row in range(1, rows + 1):
        yield [
            make_cell_value(case, mixed_string_ratio, sheet, row, col)
            for col in range(1, cols + 1)
        ]


def write_xlsxwriter_workbook(
    output: Path,
    case: BenchmarkCase,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
) -> None:
    import xlsxwriter  # type: ignore

    workbook = xlsxwriter.Workbook(str(output), {"constant_memory": True})
    try:
        for sheet in range(1, sheets + 1):
            worksheet = workbook.add_worksheet(f"Sheet{sheet}")
            for row_index, values in enumerate(
                iter_rows(case, rows, cols, sheet, mixed_string_ratio)
            ):
                worksheet.write_row(row_index, 0, values)
    finally:
        workbook.close()


def write_openpyxl_workbook(
    output: Path,
    case: BenchmarkCase,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
) -> None:
    import openpyxl  # type: ignore

    workbook = openpyxl.Workbook(write_only=True)
    for sheet in range(1, sheets + 1):
        worksheet = workbook.create_sheet(title=f"Sheet{sheet}")
        for values in iter_rows(case, rows, cols, sheet, mixed_string_ratio):
            worksheet.append(values)
    workbook.save(output)


def run_worker(args: argparse.Namespace) -> int:
    library = parse_library(args.library)
    case = parse_case(args.case)
    output = args.output.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    started = time.perf_counter()
    if library.name == "xlsxwriter":
        write_xlsxwriter_workbook(
            output, case, args.rows, args.cols, args.sheets, args.mixed_string_ratio)
    elif library.name == "openpyxl":
        write_openpyxl_workbook(
            output, case, args.rows, args.cols, args.sheets, args.mixed_string_ratio)
    else:
        raise AssertionError(f"unhandled library: {library.name}")
    elapsed_ms = int(round((time.perf_counter() - started) * 1000.0))

    result = {
        "worker_elapsed_ms": elapsed_ms,
        "module_version": module_version(library.module_name),
    }
    args.worker_result.write_text(
        json.dumps(result, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return 0


def windows_process_memory_bytes(pid: int) -> int | None:
    if platform.system() != "Windows":
        return None

    class ProcessMemoryCounters(ctypes.Structure):
        _fields_ = [
            ("cb", ctypes.c_ulong),
            ("PageFaultCount", ctypes.c_ulong),
            ("PeakWorkingSetSize", ctypes.c_size_t),
            ("WorkingSetSize", ctypes.c_size_t),
            ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
            ("QuotaPagedPoolUsage", ctypes.c_size_t),
            ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
            ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
            ("PagefileUsage", ctypes.c_size_t),
            ("PeakPagefileUsage", ctypes.c_size_t),
        ]

    process_query_information = 0x0400
    process_vm_read = 0x0010
    handle = ctypes.windll.kernel32.OpenProcess(
        process_query_information | process_vm_read,
        False,
        pid,
    )
    if not handle:
        return None
    try:
        counters = ProcessMemoryCounters()
        counters.cb = ctypes.sizeof(ProcessMemoryCounters)
        ok = ctypes.windll.psapi.GetProcessMemoryInfo(
            handle,
            ctypes.byref(counters),
            counters.cb,
        )
        if not ok:
            return None
        return int(max(counters.WorkingSetSize, counters.PeakWorkingSetSize))
    finally:
        ctypes.windll.kernel32.CloseHandle(handle)


def procfs_process_memory_bytes(pid: int) -> int | None:
    status_path = Path("/proc") / str(pid) / "status"
    if not status_path.exists():
        return None
    try:
        for line in status_path.read_text(encoding="utf-8", errors="replace").splitlines():
            if line.startswith("VmRSS:"):
                parts = line.split()
                if len(parts) >= 2:
                    return int(parts[1]) * 1024
    except OSError:
        return None
    return None


def process_memory_bytes(pid: int) -> int | None:
    return windows_process_memory_bytes(pid) or procfs_process_memory_bytes(pid)


def verify_workbook_with_openpyxl(
    path: Path,
    case: BenchmarkCase,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError:
        return {"status": "skipped", "reason": "Python module openpyxl is not installed"}

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        require(len(workbook.sheetnames) == sheets, "sheet count mismatch")
        require("Sheet1" in workbook.sheetnames, "missing Sheet1")
        worksheet = workbook["Sheet1"]
        observed_rows = 0
        observed_cols = 0
        first_value = None
        last_value = None
        for row_values in worksheet.iter_rows(values_only=True):
            observed_rows += 1
            observed_cols = max(observed_cols, len(row_values))
            if observed_rows == 1 and row_values:
                first_value = row_values[0]
            if len(row_values) >= cols:
                last_value = row_values[cols - 1]
            elif row_values:
                last_value = row_values[-1]
            else:
                last_value = None
        require(observed_rows == rows, f"row count mismatch: {observed_rows}")
        require(observed_cols == cols, f"column count mismatch: {observed_cols}")
        first_expected = make_cell_value(case, mixed_string_ratio, 1, 1, 1)
        last_expected = make_cell_value(case, mixed_string_ratio, 1, rows, cols)
        require(first_value == first_expected,
            f"A1 mismatch: expected {first_expected!r}, got {first_value!r}")
        require(last_value == last_expected,
            f"last cell mismatch: expected {last_expected!r}, got {last_value!r}")
        return {
            "status": "opened",
            "sheet_count": len(workbook.sheetnames),
            "sheet1_max_row": observed_rows,
            "sheet1_max_column": observed_cols,
            "sheet1_first_cell": first_value,
            "sheet1_last_cell": last_value,
        }
    finally:
        workbook.close()


def run_case(
    script_path: Path,
    library: PythonLibrary,
    case: BenchmarkCase,
    output_dir: Path,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
    verify_openpyxl: bool,
    strict_missing: bool,
    sample_interval_seconds: float,
) -> dict[str, Any]:
    case_name = f"{library.case_prefix}-{case.name}"
    output_path = output_dir / f"python-bench-{case_name}.xlsx"
    worker_result_path = output_dir / f"python-bench-{case_name}.worker.json"

    if not module_available(library.module_name):
        result = {
            "name": case_name,
            "library": library.name,
            "library_mode": library.mode,
            "module_name": library.module_name,
            "module_version": None,
            "case": case.name,
            "scenario": case.scenario,
            "string_pattern": case.string_pattern,
            "status": "skipped",
            "skip_reason": f"Python module {library.module_name} is not installed",
            "rows": rows,
            "cols": cols,
            "sheets": sheets,
            "cells": rows * cols * sheets,
        }
        if strict_missing:
            raise AssertionError(result["skip_reason"])
        return result

    command = [
        sys.executable,
        str(script_path),
        "--worker",
        "--library",
        f"{library.name}-{library.mode}",
        "--case",
        f"{case.scenario}:{case.string_pattern}",
        "--rows",
        str(rows),
        "--cols",
        str(cols),
        "--sheets",
        str(sheets),
        "--mixed-string-ratio",
        str(mixed_string_ratio),
        "--output",
        str(output_path),
        "--worker-result",
        str(worker_result_path),
    ]

    started = time.perf_counter()
    process = subprocess.Popen(command, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    peak_memory_bytes = 0
    while process.poll() is None:
        sample = process_memory_bytes(process.pid)
        if sample is not None:
            peak_memory_bytes = max(peak_memory_bytes, sample)
        time.sleep(sample_interval_seconds)
    stdout, stderr = process.communicate()
    sample = process_memory_bytes(process.pid)
    if sample is not None:
        peak_memory_bytes = max(peak_memory_bytes, sample)
    wall_elapsed_ms = int(round((time.perf_counter() - started) * 1000.0))

    require(process.returncode == 0,
        f"{case_name} failed with {process.returncode}: {stderr.strip()}")
    require(output_path.exists(), f"{case_name} output workbook was not created")
    require(worker_result_path.exists(), f"{case_name} worker result was not created")
    worker_result = json.loads(worker_result_path.read_text(encoding="utf-8"))
    elapsed_ms = int(worker_result["worker_elapsed_ms"])
    output_bytes = output_path.stat().st_size
    verify_result = (
        verify_workbook_with_openpyxl(output_path, case, rows, cols, sheets, mixed_string_ratio)
        if verify_openpyxl
        else {"status": "not_requested"}
    )

    return {
        "name": case_name,
        "library": library.name,
        "library_mode": library.mode,
        "module_name": library.module_name,
        "module_version": worker_result.get("module_version"),
        "case": case.name,
        "scenario": case.scenario,
        "string_pattern": case.string_pattern,
        "status": "completed",
        "rows": rows,
        "cols": cols,
        "sheets": sheets,
        "cells": rows * cols * sheets,
        "string_ratio": expected_string_ratio(case, mixed_string_ratio),
        "string_distribution": expected_string_distribution(
            case, rows, cols, sheets, mixed_string_ratio),
        "elapsed_ms": elapsed_ms,
        "wall_elapsed_ms": wall_elapsed_ms,
        "cells_per_second": (rows * cols * sheets) / (elapsed_ms / 1000.0)
            if elapsed_ms > 0 else 0.0,
        "million_cells_per_second": (rows * cols * sheets) / (elapsed_ms / 1000.0) / 1_000_000.0
            if elapsed_ms > 0 else 0.0,
        "peak_memory_mb": peak_memory_bytes / BYTES_PER_MIB if peak_memory_bytes else None,
        "output_bytes": output_bytes,
        "output_mib": output_bytes / BYTES_PER_MIB,
        "output": str(output_path),
        "worker_result": str(worker_result_path),
        "command": command,
        "stdout": stdout.strip(),
        "stderr": stderr.strip(),
        "openpyxl": verify_result,
    }


def build_report(
    output_dir: Path,
    rows: int,
    cols: int,
    sheets: int,
    mixed_string_ratio: float,
    cases: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "python_writer_benchmark_schema_version": REPORT_SCHEMA_VERSION,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "output_dir": str(output_dir),
        "rows": rows,
        "cols": cols,
        "sheets": sheets,
        "cells_per_case": rows * cols * sheets,
        "mixed_string_ratio": mixed_string_ratio,
        "cases": cases,
        "notes": [
            "Python writer libraries are opt-in benchmark/reference tools, not FastXLSX runtime dependencies.",
            "Elapsed time is measured inside the worker around workbook generation after imports.",
            "Peak memory is sampled from the worker process working set/RSS when the platform exposes it.",
            "OpenXLSX/xlnt should be benchmarked by separate C++ adapters, not linked into FastXLSX runtime.",
        ],
    }


def fmt_int(value: int | None) -> str:
    return "-" if value is None else f"{value:,}"


def fmt_float(value: float | None, digits: int = 2) -> str:
    return "-" if value is None else f"{value:.{digits}f}"


def render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Python XLSX Writer Benchmark Summary",
        "",
        f"- Generated UTC: `{report['generated_at_utc']}`",
        f"- Rows x cols x sheets: `{report['rows']} x {report['cols']} x {report['sheets']}`",
        f"- Cells per case: `{report['cells_per_case']:,}`",
        "",
        "| library | mode | case | status | elapsed ms | M cells/s | peak MB | output MiB | openpyxl |",
        "| --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for case in report["cases"]:
        openpyxl_status = "-"
        if isinstance(case.get("openpyxl"), dict):
            openpyxl_status = str(case["openpyxl"].get("status", "-"))
        lines.append(
            "| "
            f"{case.get('library', '-')} | "
            f"{case.get('library_mode', '-')} | "
            f"{case.get('case', '-')} | "
            f"{case.get('status', '-')} | "
            f"{fmt_int(case.get('elapsed_ms'))} | "
            f"{fmt_float(case.get('million_cells_per_second'))} | "
            f"{fmt_float(case.get('peak_memory_mb'))} | "
            f"{fmt_float(case.get('output_mib'))} | "
            f"{openpyxl_status} |"
        )

    skipped = [case for case in report["cases"] if case.get("status") == "skipped"]
    if skipped:
        lines.extend(["", "## Skipped"])
        for case in skipped:
            lines.append(f"- {case['name']}: {case.get('skip_reason', 'skipped')}")

    lines.extend(["", "## Notes"])
    lines.extend(f"- {note}" for note in report["notes"])
    lines.append("")
    return "\n".join(lines)


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def expect_value_error(text: str, parser: Any) -> None:
    try:
        parser(text)
    except ValueError:
        return
    raise AssertionError(f"expected parser to reject {text!r}")


def run_self_test() -> None:
    numeric = parse_case("numeric:mixed")
    mixed = parse_case("mixed:mixed")
    repeated = parse_case("strings:repeated")
    unique = parse_case("strings:unique")
    require(numeric.name == "numeric", "numeric case name mismatch")
    require(mixed.name == "mixed-mixed", "mixed case name mismatch")
    require(repeated.name == "strings-repeated", "repeated case name mismatch")
    require(unique.name == "strings-unique", "unique case name mismatch")
    expect_value_error("numeric:repeated", parse_case)
    expect_value_error("strings", parse_case)
    expect_value_error("other:mixed", parse_case)
    expect_value_error("strings:other", parse_case)

    xlsxwriter = parse_library("xlsxwriter")
    openpyxl = parse_library("openpyxl-write-only")
    require(xlsxwriter.case_prefix == "xlsxwriter-constant-memory",
        "xlsxwriter prefix mismatch")
    require(openpyxl.case_prefix == "openpyxl-write-only", "openpyxl prefix mismatch")
    expect_value_error("other", parse_library)

    distribution = expected_string_distribution(repeated, rows=2, cols=3, sheets=2,
        mixed_string_ratio=0.25)
    require(distribution["string_cells"] == 12, "repeated string cell count mismatch")
    require(distribution["unique_string_values"] == 1, "repeated unique count mismatch")
    require(make_cell_value(numeric, 0.25, 1, 2, 3) == 1002003,
        "numeric cell value mismatch")
    require(make_cell_value(unique, 0.25, 1, 2, 3) == "s1-r2-c3",
        "unique cell value mismatch")

    report = build_report(Path("out"), 2, 3, 1, 0.25, [{
        "name": "xlsxwriter-constant-memory-numeric",
        "library": "xlsxwriter",
        "library_mode": "constant-memory",
        "case": "numeric",
        "status": "completed",
        "elapsed_ms": 10,
        "million_cells_per_second": 0.0006,
        "peak_memory_mb": 20.0,
        "output_mib": 0.01,
        "openpyxl": {"status": "opened"},
    }, {
        "name": "openpyxl-write-only-numeric",
        "library": "openpyxl",
        "library_mode": "write-only",
        "case": "numeric",
        "status": "skipped",
        "skip_reason": "missing",
    }])
    markdown = render_markdown(report)
    require("Python XLSX Writer Benchmark Summary" in markdown, "markdown title mismatch")
    require("xlsxwriter" in markdown, "markdown completed case mismatch")
    require("missing" in markdown, "markdown skipped reason mismatch")
    with tempfile.TemporaryDirectory() as temp:
        markdown_path = Path(temp) / "summary.md"
        write_text(markdown_path, markdown)
        require(markdown_path.read_text(encoding="utf-8") == markdown,
            "markdown write mismatch")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--self-test", action="store_true",
        help="Run lightweight internal checks without invoking Python writer libraries.")
    parser.add_argument("--worker", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--worker-result", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--library", action="append", dest="libraries",
        help="Python writer library: xlsxwriter-constant-memory or openpyxl-write-only.")
    parser.add_argument("--case", action="append", dest="cases",
        help="Case in scenario:string_pattern format. May be passed multiple times.")
    parser.add_argument("--output", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--output-dir", type=Path, default=Path("build/qa/python-writer-benchmarks"),
        help="Directory for generated workbooks and reports.")
    parser.add_argument("--output-markdown", type=Path,
        help="Write a Markdown summary to this path.")
    parser.add_argument("--rows", type=int, default=1000, help="Rows per worksheet.")
    parser.add_argument("--cols", type=int, default=10, help="Columns per worksheet.")
    parser.add_argument("--sheets", type=int, default=1, help="Worksheet count.")
    parser.add_argument("--mixed-string-ratio", type=float, default=0.25,
        help="String ratio used by mixed scenario cases.")
    parser.add_argument("--verify-openpyxl", action="store_true",
        help="Open generated workbooks with openpyxl and verify Sheet1 first/last cells.")
    parser.add_argument("--strict-missing", action="store_true",
        help="Fail if a requested Python writer module is not installed.")
    parser.add_argument("--sample-interval-ms", type=int, default=50,
        help="Worker process memory sampling interval in milliseconds.")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        print("OK: run_python_writer_benchmarks.py self-test passed")
        return 0

    if args.worker:
        require(args.worker_result is not None, "--worker-result is required in worker mode")
        require(args.output is not None, "--output is required in worker mode")
        require(args.libraries and len(args.libraries) == 1, "--library is required in worker mode")
        args.library = args.libraries[0]
        require(args.cases and len(args.cases) == 1, "--case is required in worker mode")
        args.case = args.cases[0]
        return run_worker(args)

    require(args.rows > 0 and args.cols > 0 and args.sheets > 0,
        "rows, cols, and sheets must be positive")
    require(0.0 <= args.mixed_string_ratio <= 1.0,
        "mixed string ratio must be between 0 and 1")
    require(args.sample_interval_ms > 0, "sample interval must be positive")

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    libraries = [parse_library(value) for value in (args.libraries or DEFAULT_LIBRARIES)]
    cases = [parse_case(value) for value in (args.cases or DEFAULT_CASES)]
    script_path = Path(__file__).resolve()

    results = []
    for library in libraries:
        for case in cases:
            results.append(run_case(
                script_path=script_path,
                library=library,
                case=case,
                output_dir=output_dir,
                rows=args.rows,
                cols=args.cols,
                sheets=args.sheets,
                mixed_string_ratio=args.mixed_string_ratio,
                verify_openpyxl=args.verify_openpyxl,
                strict_missing=args.strict_missing,
                sample_interval_seconds=args.sample_interval_ms / 1000.0,
            ))

    report = build_report(
        output_dir=output_dir,
        rows=args.rows,
        cols=args.cols,
        sheets=args.sheets,
        mixed_string_ratio=args.mixed_string_ratio,
        cases=results,
    )
    report_path = output_dir / "python-writer-benchmark-report.json"
    report_path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    markdown = render_markdown(report)
    markdown_path = args.output_markdown.resolve() if args.output_markdown else (
        output_dir / "python-writer-benchmark-summary.md")
    write_text(markdown_path, markdown)
    print(markdown)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI should emit one concise failure line.
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
